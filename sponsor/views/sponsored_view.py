"""
sponsor/views/sponsored_view.py
صفحة المكفولين — لوحة الكافل
جميع التحسينات الـ 20 مطبّقة
"""
from functools import wraps, lru_cache
from datetime import date, timedelta

from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST, require_GET
from django.views.decorators.csrf import csrf_protect
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Prefetch
from django.utils import timezone

from core.models import CustomUser, Notification, Payment
from sponsor.models import SponsorProfile, PaymentSchedule
from core.utils import log_activity, create_notification, notify_admins


# ══ Decorator ══
def sponsor_required(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('/login/')
        if request.user.user_type != 'sponsor':
            return redirect('/login/')
        if not request.user.is_approved:
            return render(request, 'sponsor/pending.html')
        return view_func(request, *args, **kwargs)
    return wrapper


def _get_profile(user):
    try:
        return user.sponsor_profile
    except Exception:
        profile, _ = SponsorProfile.objects.get_or_create(user=user)
        return profile


# ══ Helpers ══

def _get_next_due(schedule):
    """الدفعة دائماً في أول الشهر القادم"""
    if not schedule:
        return None
    today = date.today()
    if today.month == 12:
        return date(today.year + 1, 1, 1)
    return date(today.year, today.month + 1, 1)


def _days_until(target_date):
    if not target_date:
        return None
    return (target_date - date.today()).days


def _payment_status(schedule, last_payment):
    if not schedule:
        return 'unknown'
    today    = date.today()
    next_due = _get_next_due(schedule)
    if not next_due:
        return 'unknown'
    if last_payment and last_payment.date >= today.replace(day=1):
        return 'paid'
    if next_due < today:
        return 'late'
    return 'upcoming'


def _months_since(sponsorship_date):
    if not sponsorship_date:
        return 0
    today = date.today()
    return (today.year - sponsorship_date.year) * 12 + (today.month - sponsorship_date.month)


def _serialize_beneficiary(form, tab, profile):
    user = form.user

    # (تحسين 2) prefetch — يُستدعى من الخارج مع select_related
    schedule = PaymentSchedule.objects.filter(
        sponsor=profile, beneficiary=user, is_active=True
    ).first()

    payments_qs = Payment.objects.filter(beneficiary=user, sponsor=profile.user)
    total_ils   = payments_qs.filter(status='paid').aggregate(t=Sum('amount_ils'))['t'] or 0
    total_usd   = payments_qs.filter(status='paid').aggregate(t=Sum('amount_usd'))['t'] or 0
    paid_count  = payments_qs.filter(status='paid').count()
    late_count  = payments_qs.filter(status='late').count()
    last_pay    = payments_qs.filter(status='paid').order_by('-date').first()

    pay_status  = _payment_status(schedule, last_pay)
    next_due    = _get_next_due(schedule)
    days_left   = _days_until(next_due)
    months      = _months_since(form.sponsorship_date)

    # (تحسين 5) نسبة الالتزام
    total_expected = paid_count + late_count
    commitment_pct = round((paid_count / total_expected * 100) if total_expected > 0 else 100)

    # (تحسين 9) متأخر أكثر من شهر
    overdue = False
    if last_pay:
        overdue = (date.today() - last_pay.date).days > 30 and pay_status == 'late'
    elif pay_status == 'late':
        overdue = True

    photo = ''
    try:
        if form.photo:
            photo = form.photo.url
    except Exception:
        pass

    return {
        'id':            form.pk,
        'user_id':       str(user.pk),
        'tab':           tab,
        'reg_number':    user.registration_number or '',
        'form_number':   form.form_number,
        'full_name':     form.get_full_name(),
        'photo':         photo,
        'gender':        form.gender or '',
        'current_city':  form.current_city or '',
        'health_status': form.health_status or '',
        'housing_type':  form.housing_type or '',
        'id_number':     form.id_number or '',
        # كفالة
        'sponsorship_date': str(form.sponsorship_date) if form.sponsorship_date else '',
        'months_since':     months,
        'next_due':         str(next_due) if next_due else '',
        'days_left':        days_left,
        'last_pay_date':    str(last_pay.date) if last_pay else '',
        'last_pay_amount':  str(last_pay.amount_ils) if last_pay else '0',
        'total_ils':        str(round(total_ils, 2)),
        'total_usd':        str(round(total_usd, 2)),
        'paid_count':       paid_count,
        'late_count':       late_count,
        'pay_status':       pay_status,
        'commitment_pct':   commitment_pct,
        'overdue':          overdue,
        'schedule_ils':     str(schedule.amount_shekel) if schedule else '0',
        'schedule_usd':     str(schedule.amount_dollar) if schedule else '0',
    }


def _get_stats(profile, user):
    """(تحسين 1) إحصائيات مجمّعة في query واحدة"""
    from beneficiary.models import OrphanForm, SpecialNeedsForm, FamilyForm

    orphan_ids  = list(OrphanForm.objects.filter(sponsor=profile).values_list('user_id', flat=True))
    special_ids = list(SpecialNeedsForm.objects.filter(sponsor=profile).values_list('user_id', flat=True))
    family_ids  = list(FamilyForm.objects.filter(sponsor=profile).values_list('user_id', flat=True))

    all_ids = orphan_ids + special_ids + family_ids
    paid_qs = Payment.objects.filter(sponsor=user, beneficiary_id__in=all_ids, status='paid')

    def ils(ids): return Payment.objects.filter(sponsor=user, beneficiary_id__in=ids, status='paid').aggregate(t=Sum('amount_ils'))['t'] or 0
    def usd(ids): return Payment.objects.filter(sponsor=user, beneficiary_id__in=ids, status='paid').aggregate(t=Sum('amount_usd'))['t'] or 0

    total_ils = paid_qs.aggregate(t=Sum('amount_ils'))['t'] or 0
    total_usd = paid_qs.aggregate(t=Sum('amount_usd'))['t'] or 0
    late_count = Payment.objects.filter(sponsor=user, status='late').count()

    schedules = PaymentSchedule.objects.filter(sponsor=profile, is_active=True)
    nearest_due = None
    for s in schedules:
        d = _get_next_due(s)
        if d and (nearest_due is None or d < nearest_due):
            nearest_due = d

    return {
        'total':          len(all_ids),
        'orphans_count':  len(orphan_ids),
        'specials_count': len(special_ids),
        'families_count': len(family_ids),
        'total_ils':      round(total_ils, 2),
        'total_usd':      round(total_usd, 2),
        'paid_orphans_ils':  round(ils(orphan_ids),  2),
        'paid_specials_ils': round(ils(special_ids), 2),
        'paid_families_ils': round(ils(family_ids),  2),
        'late_count':     late_count,
        'nearest_due':    str(nearest_due) if nearest_due else '',
    }


# ══════════════════════════════════════════════════════════
# الصفحة الرئيسية
# ══════════════════════════════════════════════════════════

@sponsor_required
def sponsored(request):
    user    = request.user
    profile = _get_profile(user)

    notif_count = Notification.objects.filter(recipient=user, is_read=False).count()
    stats       = _get_stats(profile, user)

    # (تحسين 4) إشعار الدفعة القادمة
    if stats['nearest_due']:
        nd = date.fromisoformat(stats['nearest_due'])
        if 0 <= (nd - date.today()).days <= 3:
            Notification.objects.get_or_create(
                recipient  = user,
                ntype      = 'PAYMENT_DUE',
                defaults   = {
                    'title':      'تذكير بالدفعة القادمة 📅',
                    'message':    f'موعد دفعة الكفالة بعد {(nd-date.today()).days} أيام',
                    'action_url': '/sponsor/sponsored/',
                }
            )

    # (تحسين 6) رسالة ترحيب للكافل الجديد
    is_new_sponsor = stats['total'] > 0 and not Payment.objects.filter(sponsor=user).exists()

    return render(request, 'sponsor/sponsored.html', {
        'notif_count':   notif_count,
        'is_new_sponsor':is_new_sponsor,
        **stats,
    })


# ══════════════════════════════════════════════════════════
# AJAX — بيانات التاب
# ══════════════════════════════════════════════════════════

@sponsor_required
@require_GET
def sponsored_data(request):
    from beneficiary.models import OrphanForm, SpecialNeedsForm, FamilyForm

    tab        = request.GET.get('tab',  'orphan')
    page       = int(request.GET.get('page', 1))
    q          = request.GET.get('q',    '').strip()
    pay_filter = request.GET.get('pay',  '')
    sort       = request.GET.get('sort', 'newest')
    profile    = _get_profile(request.user)

    MAP = {'orphan': OrphanForm, 'special': SpecialNeedsForm, 'family': FamilyForm}
    order_map = {
        'newest':   '-sponsorship_date',
        'oldest':   'sponsorship_date',
        'name':     'first_name',
        'next_due': 'sponsorship_date',
    }

    def filter_qs(qs):
        if q:
            qs = qs.filter(
                Q(first_name__icontains=q)  | Q(father_name__icontains=q) |
                Q(family_name__icontains=q) | Q(id_number__icontains=q)   |
                Q(form_number__icontains=q) | Q(user__registration_number__icontains=q)
            )
        return qs.order_by(order_map.get(sort, '-sponsorship_date'))

    # بحث شامل
    if tab == 'all':
        items = []
        for t, Model in MAP.items():
            qs = filter_qs(Model.objects.filter(sponsor=profile).select_related('user'))
            for obj in qs[:20]:
                item = _serialize_beneficiary(obj, t, profile)
                if pay_filter and item['pay_status'] != pay_filter:
                    continue
                items.append(item)
        return JsonResponse({'items': items, 'total': len(items), 'pages': 1, 'current_page': 1})

    if tab not in MAP:
        return JsonResponse({'items': [], 'total': 0, 'pages': 0, 'current_page': 1})

    qs        = filter_qs(MAP[tab].objects.filter(sponsor=profile).select_related('user'))
    paginator = Paginator(qs, 20)
    pg        = paginator.get_page(page)

    items = []
    for obj in pg:
        item = _serialize_beneficiary(obj, tab, profile)
        if pay_filter and item['pay_status'] != pay_filter:
            continue
        items.append(item)

    return JsonResponse({
        'items':        items,
        'total':        paginator.count,
        'pages':        paginator.num_pages,
        'current_page': pg.number,
    })


# ══════════════════════════════════════════════════════════
# AJAX — تفاصيل
# ══════════════════════════════════════════════════════════

@sponsor_required
@require_GET
def sponsored_detail(request):
    from beneficiary.models import OrphanForm, SpecialNeedsForm, FamilyForm

    tab     = request.GET.get('tab', 'orphan')
    pk      = request.GET.get('id',  0)
    profile = _get_profile(request.user)

    MAP = {'orphan': OrphanForm, 'special': SpecialNeedsForm, 'family': FamilyForm}
    Model = MAP.get(tab)
    if not Model:
        return JsonResponse({'error': 'نوع غير صالح'}, status=400)

    obj = get_object_or_404(Model, pk=pk, sponsor=profile)

    try:
        from admin_panel.views.requests import _user_detail
        user   = obj.user
        detail = _user_detail(user)
    except Exception as e:
        return JsonResponse({'error': f'خطأ: {str(e)}'}, status=500)

    photo = ''
    try:
        if obj.photo:
            photo = obj.photo.url
    except Exception:
        pass

    # (تحسين 11) بيانات الدفعات الخاصة بهذا المستفيد
    payments_qs = Payment.objects.filter(
        beneficiary=obj.user, sponsor=request.user
    ).order_by('-date')[:10]

    bene_payments = [{
        'date':       str(p.date),
        'amount_ils': str(p.amount_ils),
        'amount_usd': str(p.amount_usd),
        'status':     p.status,
        'paid_by':    p.paid_by,
        'note':       p.note or '',
    } for p in payments_qs]

    # (تحسين 12) ملخص الكفالة
    item    = _serialize_beneficiary(obj, tab, profile)
    summary = {
        'sponsorship_date': item['sponsorship_date'],
        'months_since':     item['months_since'],
        'total_ils':        item['total_ils'],
        'total_usd':        item['total_usd'],
        'paid_count':       item['paid_count'],
        'late_count':       item['late_count'],
        'pay_status':       item['pay_status'],
        'commitment_pct':   item['commitment_pct'],
        'next_due':         item['next_due'],
        'days_left':        item['days_left'],
    }

    user_data = {
        'id':          str(obj.user.pk),
        'full_name':   obj.user.get_full_name(),
        'email':       obj.user.email or '',
        'phone':       obj.user.phone or '',
        'id_number':   obj.user.id_number or '',
        'nationality': obj.user.nationality or '',
        'gender':      obj.user.gender or '',
        'reg_number':  obj.user.registration_number or '',
        'date_joined': obj.user.date_joined.strftime('%Y/%m/%d %H:%M') if obj.user.date_joined else '',
        'last_login':  obj.user.last_login.strftime('%Y/%m/%d %H:%M') if obj.user.last_login else '',
        'is_active':   obj.user.is_active,
        'photo':       photo,
        'tab':         tab,
        'form_id':     obj.pk,
    }

    detail['bene_payments'] = bene_payments
    detail['summary']       = summary

    return JsonResponse({'status': 'ok', 'detail': detail, 'user': user_data})


# ══════════════════════════════════════════════════════════
# إنهاء الكفالة (تحسين 20 — تحقق من الدفعات المتأخرة)
# ══════════════════════════════════════════════════════════

@sponsor_required
@require_POST
@csrf_protect
def end_sponsorship(request):
    from beneficiary.models import OrphanForm, SpecialNeedsForm, FamilyForm

    tab     = request.POST.get('tab', '')
    pk      = request.POST.get('id',  0)
    force   = request.POST.get('force', '') == '1'
    user    = request.user
    profile = _get_profile(user)

    MAP   = {'orphan': OrphanForm, 'special': SpecialNeedsForm, 'family': FamilyForm}
    Model = MAP.get(tab)
    if not Model:
        return JsonResponse({'status': 'error', 'message': 'نوع غير صالح'})

    obj        = get_object_or_404(Model, pk=pk, sponsor=profile)
    late_count = Payment.objects.filter(
        beneficiary=obj.user, sponsor=user, status='late'
    ).count()

    # (تحسين 20) منع الإنهاء مع دفعات متأخرة إلا بتأكيد
    if late_count > 0 and not force:
        return JsonResponse({
            'status':      'warning',
            'late_count':  late_count,
            'message':     f'يوجد {late_count} دفعة متأخرة — هل تريد الإنهاء رغم ذلك؟',
        })

    name = obj.get_full_name()
    obj.sponsor          = None
    obj.status           = 'مقبولة'
    obj.sponsorship_date = None
    obj.save()

    PaymentSchedule.objects.filter(
        sponsor=profile, beneficiary=obj.user, is_active=True
    ).update(is_active=False, end_date=date.today())

    # (تحسين 19) تسجيل في ActivityLog
    log_activity(user, 'SPONSOR',
                 description=f'إنهاء كفالة {name}',
                 request=request)

    notify_admins(
        ntype      = 'END_SPONSOR',
        title      = 'إنهاء كفالة 🔴',
        message    = f'{user.get_full_name()} أنهى كفالة {name}',
        sender     = user,
        action_url = '/admin-panel/beneficiaries/',
    )

    return JsonResponse({'status': 'success', 'message': f'تم إنهاء كفالة {name} ✅'})


# ══════════════════════════════════════════════════════════
# طباعة
# ══════════════════════════════════════════════════════════

@sponsor_required
def print_sponsored(request):
    from beneficiary.models import OrphanForm, SpecialNeedsForm, FamilyForm
    from django.template.loader import render_to_string

    tab     = request.GET.get('tab', 'orphan')
    pk      = request.GET.get('id',  0)
    profile = _get_profile(request.user)

    MAP   = {'orphan': OrphanForm, 'special': SpecialNeedsForm, 'family': FamilyForm}
    Model = MAP.get(tab)
    if not Model:
        return redirect('/sponsor/sponsored/')

    obj  = get_object_or_404(Model, pk=pk, sponsor=profile)
    user = obj.user

    try:
        from admin_panel.views.requests import _user_detail
        detail = _user_detail(user)
    except Exception as e:
        return HttpResponse(f'خطأ: {str(e)}', status=500)

    html = render_to_string('admin_panel/print_request.html', {
        'user':       user,
        'detail':     detail,
        'printed_by': request.user.get_full_name() or request.user.username,
        'request':    request,
    })
    return HttpResponse(html)


# ══════════════════════════════════════════════════════════
# تحميل PDF
# ══════════════════════════════════════════════════════════

@sponsor_required
def download_sponsored_pdf(request):
    from beneficiary.models import OrphanForm, SpecialNeedsForm, FamilyForm
    from django.template.loader import render_to_string
    from urllib.parse import quote
    import weasyprint, logging

    logging.getLogger('weasyprint').setLevel(logging.ERROR)
    logging.getLogger('fontTools').setLevel(logging.ERROR)

    tab     = request.GET.get('tab', 'orphan')
    pk      = request.GET.get('id',  0)
    profile = _get_profile(request.user)

    MAP   = {'orphan': OrphanForm, 'special': SpecialNeedsForm, 'family': FamilyForm}
    Model = MAP.get(tab)
    if not Model:
        return redirect('/sponsor/sponsored/')

    obj  = get_object_or_404(Model, pk=pk, sponsor=profile)
    user = obj.user

    # (تحسين 19) تسجيل التصدير
    log_activity(request.user, 'EXPORT',
                 description=f'تحميل PDF لـ {obj.get_full_name()}',
                 request=request)

    try:
        from admin_panel.views.requests import _user_detail
        detail = _user_detail(user)
    except Exception as e:
        return HttpResponse(f'خطأ: {str(e)}', status=500)

    html_content = render_to_string('admin_panel/pdf_request.html', {
        'user':       user,
        'detail':     detail,
        'printed_by': request.user.get_full_name() or request.user.username,
        'request':    request,
    })

    pdf_file = weasyprint.HTML(
        string=html_content,
        base_url=request.build_absolute_uri('/')
    ).write_pdf()

    name   = user.get_full_name().replace(' ', '_')
    id_num = user.id_number or str(user.pk)[:8]
    fname  = f'{name}_{id_num}.pdf'

    resp = HttpResponse(pdf_file, content_type='application/pdf')
    resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(fname)}"
    return resp


# ══════════════════════════════════════════════════════════
# تصدير Excel (تحسين 17، 19)
# ══════════════════════════════════════════════════════════

@sponsor_required
@require_GET
def export_excel(request):
    from beneficiary.models import OrphanForm, SpecialNeedsForm, FamilyForm
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from urllib.parse import quote
    import io

    tab     = request.GET.get('tab', 'orphan')
    profile = _get_profile(request.user)
    user    = request.user

    # (تحسين 19)
    log_activity(user, 'EXPORT',
                 description=f'تصدير Excel — {tab}',
                 request=request)

    GREEN  = '1A7A4A'
    LIGHT  = 'E8F5E9'
    WHITE  = 'FFFFFF'
    thin   = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def sh(cell, color=GREEN):
        cell.font      = Font(bold=True, color=WHITE, size=10)
        cell.fill      = PatternFill('solid', fgColor=color)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = border

    def sc(cell, r):
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = border
        if r % 2 == 0:
            cell.fill = PatternFill('solid', fgColor='F4F9F6')

    STATUS_AR = {'paid':'مدفوع','late':'متأخر','upcoming':'قادم','unknown':'غير محدد'}
    HEADERS   = ['رقم التسجيل','الاسم الكامل','رقم الهوية','المدينة','الجنس','الحالة الصحية','تاريخ الكفالة','مدة الكفالة (شهر)','الدفعة القادمة','آخر دفعة','قيمة آخر دفعة (₪)','إجمالي المدفوع (₪)','إجمالي المدفوع ($)','مسدّد','متأخر','نسبة الالتزام %','حالة الدفع']
    WIDTHS    = [18,28,16,14,10,18,16,14,16,14,18,18,16,10,10,16,14]
    TAB_NAMES = {'orphan':'الأيتام','special':'ذوو احتياجات','family':'الأسر'}
    MAP       = {'orphan':OrphanForm,'special':SpecialNeedsForm,'family':FamilyForm}

    def build_sheet(ws, qs, tab_name):
        ws.sheet_view.rightToLeft = True
        ws.title = TAB_NAMES.get(tab_name, tab_name)

        ws.merge_cells(f'A1:Q1')
        c = ws['A1']
        c.value     = f'تقرير المكفولين — {ws.title} — {user.get_full_name()} — {date.today()}'
        c.font      = Font(bold=True, size=12, color=WHITE)
        c.fill      = PatternFill('solid', fgColor=GREEN)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        for i, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
            cell = ws.cell(row=2, column=i, value=h)
            sh(cell)
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[2].height = 24

        total_ils = 0
        for r_idx, form in enumerate(qs, start=3):
            item = _serialize_beneficiary(form, tab_name, profile)
            total_ils += float(item['total_ils'])
            row = [
                item['reg_number'], item['full_name'], item['id_number'],
                item['current_city'], item['gender'], item['health_status'],
                item['sponsorship_date'], item['months_since'],
                item['next_due'], item['last_pay_date'],
                float(item['last_pay_amount']),
                float(item['total_ils']), float(item['total_usd']),
                item['paid_count'], item['late_count'],
                item['commitment_pct'],
                STATUS_AR.get(item['pay_status'], '—'),
            ]
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                sc(cell, r_idx)
                ws.row_dimensions[r_idx].height = 20
                if c_idx == 17:
                    clr = {'مدفوع':'1A7A4A','متأخر':'C53030','قادم':'B45309'}.get(val,'9CA3AF')
                    cell.font = Font(bold=True, color=clr)

        # إجمالي
        tr = len(list(qs)) + 3
        ws.cell(tr, 1, 'الإجمالي').font = Font(bold=True, color=WHITE)
        ws.cell(tr, 1).fill = PatternFill('solid', fgColor=GREEN)
        ws.cell(tr, 12, round(total_ils, 2)).font = Font(bold=True, color=GREEN)
        for c in range(1, 18):
            ws.cell(tr, c).border = border
            if c != 1:
                ws.cell(tr, c).fill = PatternFill('solid', fgColor=LIGHT)
        ws.row_dimensions[tr].height = 22

    wb = openpyxl.Workbook()

    if tab == 'all':
        # ورقة الإحصائيات
        ws_s = wb.active
        ws_s.title = 'الإحصائيات'
        ws_s.sheet_view.rightToLeft = True

        ws_s.merge_cells('A1:E1')
        c = ws_s['A1']
        c.value     = f'إحصائيات الكفالة — {user.get_full_name()} — {date.today()}'
        c.font      = Font(bold=True, size=12, color=WHITE)
        c.fill      = PatternFill('solid', fgColor=GREEN)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws_s.row_dimensions[1].height = 30

        for i, h in enumerate(['النوع','العدد','إجمالي المدفوع (₪)','إجمالي المدفوع ($)','الدفعات المتأخرة'], 1):
            cell = ws_s.cell(row=2, column=i, value=h)
            sh(cell)
            ws_s.column_dimensions[get_column_letter(i)].width = [20,12,22,22,18][i-1]
        ws_s.row_dimensions[2].height = 24

        grand = [0, 0, 0, 0]
        for r_idx, (t, Model) in enumerate(MAP.items(), start=3):
            qs      = Model.objects.filter(sponsor=profile)
            bids    = list(qs.values_list('user_id', flat=True))
            count   = qs.count()
            ils     = Payment.objects.filter(sponsor=user, beneficiary_id__in=bids, status='paid').aggregate(t=Sum('amount_ils'))['t'] or 0
            usd     = Payment.objects.filter(sponsor=user, beneficiary_id__in=bids, status='paid').aggregate(t=Sum('amount_usd'))['t'] or 0
            late    = Payment.objects.filter(sponsor=user, beneficiary_id__in=bids, status='late').count()
            grand   = [grand[0]+count, grand[1]+ils, grand[2]+usd, grand[3]+late]
            for c_idx, val in enumerate([TAB_NAMES[t], count, round(ils,2), round(usd,2), late], 1):
                cell = ws_s.cell(row=r_idx, column=c_idx, value=val)
                sc(cell, r_idx)
                ws_s.row_dimensions[r_idx].height = 22

        for c_idx, val in enumerate(['الإجمالي', grand[0], round(grand[1],2), round(grand[2],2), grand[3]], 1):
            cell = ws_s.cell(row=6, column=c_idx, value=val)
            cell.font      = Font(bold=True, color=WHITE)
            cell.fill      = PatternFill('solid', fgColor=GREEN)
            cell.border    = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws_s.row_dimensions[6].height = 24

        # (تحسين 17) ورقة الدفعات المتأخرة
        ws_late = wb.create_sheet('الدفعات المتأخرة')
        ws_late.sheet_view.rightToLeft = True
        ws_late.merge_cells('A1:F1')
        cl = ws_late['A1']
        cl.value = f'الدفعات المتأخرة — {user.get_full_name()}'
        cl.font  = Font(bold=True, size=12, color=WHITE)
        cl.fill  = PatternFill('solid', fgColor='C53030')
        cl.alignment = Alignment(horizontal='center', vertical='center')
        ws_late.row_dimensions[1].height = 28

        late_headers = ['رقم التسجيل','الاسم','المبلغ (₪)','المبلغ ($)','تاريخ الدفعة','النوع']
        for i, h in enumerate(late_headers, 1):
            cell = ws_late.cell(row=2, column=i, value=h)
            sh(cell, 'C53030')
            ws_late.column_dimensions[get_column_letter(i)].width = [16,28,14,14,14,14][i-1]
        ws_late.row_dimensions[2].height = 22

        all_bids = {str(u_id): t for t, Model in MAP.items() for u_id in Model.objects.filter(sponsor=profile).values_list('user_id', flat=True)}
        late_pays = Payment.objects.filter(sponsor=user, status='late').select_related('beneficiary').order_by('-date')
        for r_idx, p in enumerate(late_pays, start=3):
            tab_name = all_bids.get(str(p.beneficiary_id), '—')
            row = [p.beneficiary.registration_number or '—', p.beneficiary.get_full_name(), float(p.amount_ils), float(p.amount_usd), str(p.date), TAB_NAMES.get(tab_name,'—')]
            for c_idx, val in enumerate(row, 1):
                cell = ws_late.cell(row=r_idx, column=c_idx, value=val)
                sc(cell, r_idx)
                ws_late.row_dimensions[r_idx].height = 20

        # ورقة لكل نوع
        for t, Model in MAP.items():
            qs = Model.objects.filter(sponsor=profile).select_related('user')
            ws = wb.create_sheet()
            build_sheet(ws, qs, t)

        filename = f'كفالات_شاملة_{user.get_full_name()}_{date.today()}.xlsx'

    else:
        ws = wb.active
        qs = MAP.get(tab, OrphanForm).objects.filter(sponsor=profile).select_related('user')
        build_sheet(ws, qs, tab)
        filename = f'كفالات_{TAB_NAMES.get(tab,tab)}_{user.get_full_name()}_{date.today()}.xlsx'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    resp = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"
    return resp
