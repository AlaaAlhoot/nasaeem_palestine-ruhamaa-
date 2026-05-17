"""
logs.py — فيوز سجل النشاط
"""

from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_GET, require_POST
from django.views.decorators.csrf import csrf_protect
from django.utils import timezone
from django.db.models import Count, Q
from datetime import date, timedelta
import io

from core.models import ActivityLog, CustomUser
from core.utils import log_activity
from .decorators import admin_required


# ── اسم المستخدم الصحيح ──
def _user_name(user):
    if not user: return 'النظام'
    name = user.get_full_name()
    return name if name.strip() else (user.username or 'مجهول')


# ── تنسيق التاريخ ──
def _fmt_dt(dt):
    if not dt: return '—'
    try:
        return timezone.localtime(dt).strftime('%Y/%m/%d %H:%M:%S')
    except Exception:
        return str(dt)


# ── إحداثيات الإجراءات ──
ACTION_CONFIG = {
    'LOGIN':        {'icon': '🔑', 'color': '#1a7a4a', 'label': 'دخول'},
    'LOGOUT':       {'icon': '🚪', 'color': '#b45309', 'label': 'خروج'},
    'FAILED_LOGIN': {'icon': '⚠️', 'color': '#c53030', 'label': 'دخول فاشل'},
    'CREATE':       {'icon': '➕', 'color': '#2b6cb0', 'label': 'إنشاء'},
    'UPDATE':       {'icon': '✏️', 'color': '#7c3aed', 'label': 'تعديل'},
    'DELETE':       {'icon': '🗑️', 'color': '#c53030', 'label': 'حذف'},
    'APPROVE':      {'icon': '✅', 'color': '#1a7a4a', 'label': 'قبول'},
    'REJECT':       {'icon': '❌', 'color': '#c53030', 'label': 'رفض'},
    'EXPORT':       {'icon': '📊', 'color': '#2b6cb0', 'label': 'تصدير'},
    'MESSAGE':      {'icon': '✉️', 'color': '#7c3aed', 'label': 'رسالة'},
    'PAYMENT':      {'icon': '💰', 'color': '#1a7a4a', 'label': 'دفعة'},
    'VIEW':         {'icon': '👁',  'color': '#6b7280', 'label': 'عرض'},
    'BLOCKED':      {'icon': '🚫', 'color': '#c53030', 'label': 'حظر'},
}

DANGER_ACTIONS = {'DELETE', 'BLOCKED', 'FAILED_LOGIN', 'REJECT'}


def _build_log(log):
    cfg = ACTION_CONFIG.get(log.action, {'icon': '📌', 'color': '#6b7280', 'label': log.action})
    return {
        'id':           log.pk,
        'user':         _user_name(log.user),
        'user_type':    log.user.user_type if log.user else '',
        'action':       log.action,
        'action_label': cfg['label'],
        'action_icon':  cfg['icon'],
        'action_color': cfg['color'],
        'is_danger':    log.action in DANGER_ACTIONS,
        'description':  log.description or '—',
        'target_model': log.target_model or '',
        'target_id':    str(log.target_id) if log.target_id else '',
        'ip_address':   log.ip_address or '',
        'created_at':   log.created_at.isoformat(),
        'created_fmt':  _fmt_dt(log.created_at),
    }


# ══════════════════════════════════════════════
# الصفحة الرئيسية
# ══════════════════════════════════════════════
@admin_required
def logs_list(request):
    return render(request, 'admin_panel/logs.html')


# ══════════════════════════════════════════════
# جلب البيانات — AJAX
# ══════════════════════════════════════════════
@admin_required
@require_GET
def logs_data(request):
    PER_PAGE = 30
    page     = int(request.GET.get('page',   1))
    action   = request.GET.get('action',     'all').strip()
    q        = request.GET.get('q',          '').strip()
    utype    = request.GET.get('user_type',  '').strip()
    date_f   = request.GET.get('date',       '').strip()
    ip_f     = request.GET.get('ip',         '').strip()
    sort     = request.GET.get('sort',       '-date').strip()

    qs = ActivityLog.objects.select_related('user').all()

    # فلتر الإجراء
    if action != 'all':
        qs = qs.filter(action=action)

    # فلتر نوع المستخدم
    if utype:
        qs = qs.filter(user__user_type=utype)

    # فلتر IP
    if ip_f:
        qs = qs.filter(ip_address=ip_f)

    # فلتر التاريخ
    today = date.today()
    if date_f == 'today':
        qs = qs.filter(created_at__date=today)
    elif date_f == 'week':
        qs = qs.filter(created_at__date__gte=today - timedelta(days=7))
    elif date_f == 'month':
        qs = qs.filter(created_at__date__gte=today - timedelta(days=30))

    # البحث
    if q:
        qs = qs.filter(
            Q(description__icontains=q) |
            Q(user__first_name__icontains=q) |
            Q(user__family_name__icontains=q) |
            Q(user__username__icontains=q) |
            Q(ip_address__icontains=q) |
            Q(target_model__icontains=q)
        ).distinct()

    # ترتيب
    if sort == 'date':
        qs = qs.order_by('created_at')
    elif sort == '-action':
        qs = qs.order_by('-action', '-created_at')
    else:
        qs = qs.order_by('-created_at')

    total      = qs.count()
    total_pages= max(1, (total + PER_PAGE - 1) // PER_PAGE)
    page       = max(1, min(page, total_pages))
    start      = (page - 1) * PER_PAGE
    logs       = qs[start: start + PER_PAGE]

    return JsonResponse({
        'logs':        [_build_log(l) for l in logs],
        'pagination': {
            'page':        page,
            'total_pages': total_pages,
            'total':       total,
            'per_page':    PER_PAGE,
        },
        'stats':  _get_stats(request),
        'charts': _get_chart_data(),
        'top_users': _get_top_users(),
    })


# ══════════════════════════════════════════════
# تفاصيل سجل واحد
# ══════════════════════════════════════════════
@admin_required
@require_GET
def log_detail(request):
    log_id = request.GET.get('id', '').strip()
    try:
        log = ActivityLog.objects.select_related('user').get(pk=log_id)
        return JsonResponse({'status': 'success', 'log': _build_log(log)})
    except ActivityLog.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'السجل غير موجود'})


# ══════════════════════════════════════════════
# حذف السجلات القديمة
# ══════════════════════════════════════════════
@admin_required
@require_POST
@csrf_protect
def delete_old_logs(request):
    days = int(request.POST.get('days', 30))
    cutoff = timezone.now() - timedelta(days=days)
    deleted, _ = ActivityLog.objects.filter(created_at__lt=cutoff).delete()
    log_activity(
        request.user, 'DELETE',
        description=f'حذف {deleted} سجل نشاط أقدم من {days} يوم',
        request=request,
    )
    return JsonResponse({'status': 'success', 'deleted': deleted, 'message': f'تم حذف {deleted} سجل'})


# ══════════════════════════════════════════════
# تصدير Excel
# ══════════════════════════════════════════════
@admin_required
@require_GET
def export_logs(request):
    action = request.GET.get('action', 'all').strip()
    q      = request.GET.get('q',     '').strip()
    date_f = request.GET.get('date',  '').strip()
    utype  = request.GET.get('user_type', '').strip()

    qs = ActivityLog.objects.select_related('user').order_by('-created_at')
    if action != 'all': qs = qs.filter(action=action)
    if utype:           qs = qs.filter(user__user_type=utype)
    if q:
        qs = qs.filter(
            Q(description__icontains=q) |
            Q(user__first_name__icontains=q) |
            Q(user__username__icontains=q) |
            Q(ip_address__icontains=q)
        ).distinct()
    today = date.today()
    if date_f == 'today': qs = qs.filter(created_at__date=today)
    elif date_f == 'week': qs = qs.filter(created_at__date__gte=today - timedelta(days=7))
    elif date_f == 'month': qs = qs.filter(created_at__date__gte=today - timedelta(days=30))
    qs = qs[:2000]

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return HttpResponse('openpyxl غير مثبّتة', status=500)

    PURPLE = '7C3AED'
    GREEN  = '1A7A4A'
    RED    = 'C53030'
    thin   = Side(style='thin', color='E2E8F0')
    border = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)
    C_ALG  = Alignment(horizontal='center', vertical='center')
    R_ALG  = Alignment(horizontal='right',  vertical='center')

    wb  = openpyxl.Workbook()

    # ══ ورقة 1: السجلات ══
    ws1 = wb.active
    ws1.title = 'سجل النشاط'
    ws1.sheet_view.rightToLeft = True

    headers = ['المستخدم', 'نوع المستخدم', 'الإجراء', 'التفاصيل', 'النموذج', 'IP', 'التاريخ والوقت']
    for i, h in enumerate(headers, 1):
        c = ws1.cell(row=1, column=i, value=h)
        c.font      = Font(bold=True, color='FFFFFF')
        c.fill      = PatternFill('solid', fgColor=PURPLE)
        c.alignment = C_ALG
        c.border    = border
    ws1.row_dimensions[1].height = 26

    TYPE_LABELS = {'admin': 'أدمن', 'orphan': 'يتيم', 'family': 'أسرة', 'special': 'ذوو احتياجات', 'sponsor': 'كافل'}
    for r_idx, log in enumerate(qs, 2):
        cfg  = ACTION_CONFIG.get(log.action, {'label': log.action})
        uname = _user_name(log.user)
        row  = [
            uname,
            TYPE_LABELS.get(log.user.user_type, log.user.user_type) if log.user else '—',
            cfg['label'],
            log.description or '—',
            log.target_model or '—',
            log.ip_address   or '—',
            _fmt_dt(log.created_at),
        ]
        is_danger = log.action in DANGER_ACTIONS
        for c_idx, val in enumerate(row, 1):
            cell           = ws1.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = R_ALG
            cell.border    = border
            if is_danger:
                cell.fill = PatternFill('solid', fgColor='FEF2F2')
            elif r_idx % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='F5F3FF')
        ws1.row_dimensions[r_idx].height = 20

    for col, w in zip(['A','B','C','D','E','F','G'], [22, 14, 14, 44, 16, 16, 20]):
        ws1.column_dimensions[col].width = w

    # ══ ورقة 2: إحصائيات ══
    ws2 = wb.create_sheet(title='الإحصائيات')
    ws2.sheet_view.rightToLeft = True
    ws2.cell(row=1, column=1, value='الإجراء').font  = Font(bold=True, color='FFFFFF')
    ws2.cell(row=1, column=1).fill = PatternFill('solid', fgColor=GREEN)
    ws2.cell(row=1, column=2, value='العدد').font    = Font(bold=True, color='FFFFFF')
    ws2.cell(row=1, column=2).fill = PatternFill('solid', fgColor=GREEN)
    for cell in [ws2.cell(row=1,column=1), ws2.cell(row=1,column=2)]:
        cell.alignment = C_ALG; cell.border = border
    ws2.row_dimensions[1].height = 26

    stats = ActivityLog.objects.values('action').annotate(c=Count('id')).order_by('-c')
    for r_idx, s in enumerate(stats, 2):
        cfg = ACTION_CONFIG.get(s['action'], {'label': s['action'], 'icon': '📌'})
        ws2.cell(row=r_idx, column=1, value=f"{cfg.get('icon','')} {cfg['label']}").alignment = R_ALG
        ws2.cell(row=r_idx, column=2, value=s['c']).alignment = C_ALG
        for col in [1, 2]:
            ws2.cell(row=r_idx, column=col).border = border
        ws2.row_dimensions[r_idx].height = 20
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 12

    # ══ ورقة 3: الأكثر نشاطاً ══
    ws3 = wb.create_sheet(title='الأكثر نشاطاً')
    ws3.sheet_view.rightToLeft = True
    for i, h in enumerate(['المستخدم', 'نوع المستخدم', 'عدد الإجراءات'], 1):
        c = ws3.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='2B6CB0')
        c.alignment = C_ALG; c.border = border
    ws3.row_dimensions[1].height = 26
    top = ActivityLog.objects.filter(user__isnull=False).values(
        'user__id','user__first_name','user__family_name','user__username','user__user_type'
    ).annotate(c=Count('id')).order_by('-c')[:20]
    for r_idx, u in enumerate(top, 2):
        name = f"{u['user__first_name'] or ''} {u['user__family_name'] or ''}".strip() or u['user__username']
        ws3.cell(row=r_idx, column=1, value=name).alignment = R_ALG
        ws3.cell(row=r_idx, column=2, value=TYPE_LABELS.get(u['user__user_type'], u['user__user_type'])).alignment = R_ALG
        ws3.cell(row=r_idx, column=3, value=u['c']).alignment = C_ALG
        for col in [1,2,3]: ws3.cell(row=r_idx, column=col).border = border
        ws3.row_dimensions[r_idx].height = 20
    ws3.column_dimensions['A'].width = 22
    ws3.column_dimensions['B'].width = 14
    ws3.column_dimensions['C'].width = 14

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    log_activity(request.user, 'EXPORT', description='تصدير سجل النشاط', request=request)

    resp = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="activity_log_{date.today()}.xlsx"'
    return resp


# ══════════════════════════════════════════════
# مساعدات
# ══════════════════════════════════════════════
def _get_stats(request):
    today = date.today()
    qs    = ActivityLog.objects.all()
    yesterday_count = qs.filter(created_at__date=today - timedelta(days=1)).count()
    today_count     = qs.filter(created_at__date=today).count()
    diff = today_count - yesterday_count
    return {
        'total':          qs.count(),
        'today':          today_count,
        'yesterday':      yesterday_count,
        'today_diff':     diff,
        'week':           qs.filter(created_at__date__gte=today - timedelta(days=7)).count(),
        'danger':         qs.filter(action__in=list(DANGER_ACTIONS)).count(),
        'failed_logins':  qs.filter(action='FAILED_LOGIN', created_at__date=today).count(),
        'by_action':      {
            a: qs.filter(action=a).count()
            for a in ACTION_CONFIG.keys()
        },
    }


def _get_chart_data():
    today  = date.today()
    result = []
    for i in range(6, -1, -1):
        d     = today - timedelta(days=i)
        count = ActivityLog.objects.filter(created_at__date=d).count()
        result.append({'date': d.strftime('%m/%d'), 'count': count})
    return result


def _get_top_users():
    top = ActivityLog.objects.filter(user__isnull=False).values(
        'user__first_name', 'user__family_name', 'user__username', 'user__user_type'
    ).annotate(c=Count('id')).order_by('-c')[:8]
    TYPE_LABELS = {'admin':'أدمن','orphan':'يتيم','family':'أسرة','special':'ذوو احتياجات','sponsor':'كافل'}
    return [{
        'name':  (f"{u['user__first_name'] or ''} {u['user__family_name'] or ''}".strip() or u['user__username']),
        'type':  TYPE_LABELS.get(u['user__user_type'], u['user__user_type']),
        'count': u['c'],
    } for u in top]
