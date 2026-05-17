"""
beneficiary/views/payments.py
صفحة المدفوعات للمستفيد + APIs + تصدير Excel/PDF
"""
import os
import hashlib
from io import BytesIO
from datetime import datetime
from decimal import Decimal
from urllib.parse import quote

from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_GET, require_POST
from django.views.decorators.csrf import csrf_protect
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Q
from django.utils import timezone

from core.models import Payment, CustomUser
from core.utils import log_activity
import hashlib
from decimal import Decimal
from django.utils import timezone
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_GET
from django.db.models import Sum
ALLOWED_BENE_TYPES = ('orphan', 'special', 'family')

PAID_BY_LABELS = dict(Payment.PAID_BY)
STATUS_LABELS  = dict(Payment.STATUS)


# ════════════════════════════════════════════
#  Helpers
# ════════════════════════════════════════════

def _bene_required(request):
    """تأكيد أن المستخدم مستفيد."""
    return request.user.is_authenticated and request.user.user_type in ALLOWED_BENE_TYPES


def _serialize_payment(p):
    """تحويل Payment إلى dict."""
    return {
        'id':            p.pk,
        'date':          p.date.isoformat() if p.date else '',
        'amount_ils':    str(p.amount_ils),
        'amount_usd':    str(p.amount_usd),
        'paid_by':       p.paid_by,
        'paid_by_label': PAID_BY_LABELS.get(p.paid_by, p.paid_by),
        'paid_by_note':  p.paid_by_note or '',
        'status':        p.status,
        'status_label':  STATUS_LABELS.get(p.status, p.status),
        'note':          p.note or '',
        'sponsor':       p.sponsor.get_full_name() if p.sponsor else '',
        'sponsor_reg':   getattr(p.sponsor, 'registration_number', '') if p.sponsor else '',
        'created_at':    p.created_at.isoformat() if p.created_at else '',
        'created_by':    p.created_by.get_full_name() if p.created_by else '',
    }


def _filter_queryset(qs, filters):
    """تطبيق الفلاتر على الـ queryset."""
    paid_by = filters.get('paid_by', '').strip()
    status  = filters.get('status', '').strip()
    df, dt  = filters.get('date_from', '').strip(), filters.get('date_to', '').strip()
    q       = filters.get('q', '').strip()

    if paid_by: qs = qs.filter(paid_by=paid_by)
    if status:  qs = qs.filter(status=status)
    if df:      qs = qs.filter(date__gte=df)
    if dt:      qs = qs.filter(date__lte=dt)
    if q:
        qs = qs.filter(
            Q(note__icontains=q) | Q(paid_by_note__icontains=q) |
            Q(sponsor__first_name__icontains=q) | Q(sponsor__family_name__icontains=q)
        )
    return qs


# ════════════════════════════════════════════
#  Page View
# ════════════════════════════════════════════

@login_required(login_url='/login/')
def payments_view(request):
    if not _bene_required(request):
        return redirect('/')
    return render(request, 'beneficiary/payments.html', {})


# ════════════════════════════════════════════
#  Data API (للجدول والإحصائيات)
# ════════════════════════════════════════════

@login_required(login_url='/login/')
@require_GET
def payments_data_api(request):
    if not _bene_required(request):
        return JsonResponse({'status': 'error', 'message': 'غير مسموح'}, status=403)

    qs = Payment.objects.filter(beneficiary=request.user).select_related('sponsor')
    qs = _filter_queryset(qs, request.GET).order_by('-date', '-created_at')

    # الإحصائيات (على الكل بدون فلتر)
    all_qs = Payment.objects.filter(beneficiary=request.user)
    total_ils = all_qs.aggregate(s=Sum('amount_ils'))['s'] or Decimal('0')
    total_usd = all_qs.aggregate(s=Sum('amount_usd'))['s'] or Decimal('0')

    # هذا الشهر
    now = timezone.now()
    month_qs = all_qs.filter(date__year=now.year, date__month=now.month)
    month_ils = month_qs.aggregate(s=Sum('amount_ils'))['s'] or Decimal('0')

    # توزيع المصادر
    by_source = {}
    for code, label in Payment.PAID_BY:
        cnt = all_qs.filter(paid_by=code).count()
        amt = all_qs.filter(paid_by=code).aggregate(s=Sum('amount_ils'))['s'] or Decimal('0')
        by_source[code] = {'label': label, 'count': cnt, 'amount': str(amt)}

    # توزيع الحالة
    by_status = {}
    for code, label in Payment.STATUS:
        cnt = all_qs.filter(status=code).count()
        by_status[code] = {'label': label, 'count': cnt}

    # آخر دفعة
    last = all_qs.order_by('-date').first()

    return JsonResponse({
        'status': 'success',
        'data': {
            'payments':     [_serialize_payment(p) for p in qs],
            'count':        qs.count(),
            'total_ils':    str(total_ils),
            'total_usd':    str(total_usd),
            'month_ils':    str(month_ils),
            'month_count':  month_qs.count(),
            'by_source':    by_source,
            'by_status':    by_status,
            'last_date':    last.date.isoformat() if last and last.date else None,
            'last_amount':  str(last.amount_ils) if last else '0',
            'paid_by_choices': [{'value': c, 'label': l} for c, l in Payment.PAID_BY],
            'status_choices':  [{'value': c, 'label': l} for c, l in Payment.STATUS],
        }
    })


# ════════════════════════════════════════════
#  Detail API (لمودل التفاصيل)
# ════════════════════════════════════════════

@login_required(login_url='/login/')
@require_GET
def payment_detail_api(request):
    if not _bene_required(request):
        return JsonResponse({'status': 'error', 'message': 'غير مسموح'}, status=403)

    pid = request.GET.get('id', '').strip()
    try:
        p = Payment.objects.select_related('sponsor', 'created_by').get(
            pk=pid, beneficiary=request.user
        )
    except Payment.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'الدفعة غير موجودة'})

    return JsonResponse({'status': 'success', 'data': _serialize_payment(p)})


# ════════════════════════════════════════════
#  Export Excel
# ════════════════════════════════════════════

@login_required(login_url='/login/')
@require_GET
def export_excel_api(request):
    if not _bene_required(request):
        return JsonResponse({'status': 'error', 'message': 'غير مسموح'}, status=403)

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    qs = Payment.objects.filter(beneficiary=request.user).select_related('sponsor')
    qs = _filter_queryset(qs, request.GET).order_by('-date')

    wb = Workbook()
    ws = wb.active
    ws.title = 'سجل الدفعات'
    ws.sheet_view.rightToLeft = True

    # تنسيقات
    HDR_FILL = PatternFill('solid', fgColor='1A7A4A')
    HDR_FONT = Font(name='Tajawal', size=12, bold=True, color='FFFFFF')
    LBL_FONT = Font(name='Tajawal', size=11, bold=True, color='1A7A4A')
    VAL_FONT = Font(name='Tajawal', size=11)
    ALIGN_C  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ALIGN_R  = Alignment(horizontal='right', vertical='center', wrap_text=True)
    THIN     = Side(style='thin', color='C8E6CF')
    BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    ALT_FILL = PatternFill('solid', fgColor='F4F9F6')

    # ترويسة
    user = request.user
    ws.merge_cells('A1:H1')
    c = ws.cell(row=1, column=1, value=f'💰 سجل الدفعات — {user.get_full_name()}')
    c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = ALIGN_C
    ws.row_dimensions[1].height = 32

    ws.merge_cells('A2:H2')
    c2 = ws.cell(row=2, column=1, value=f'تاريخ التقرير: {timezone.now().strftime("%Y-%m-%d %H:%M")}')
    c2.font = Font(name='Tajawal', size=10, italic=True, color='6B7280')
    c2.alignment = ALIGN_C
    ws.row_dimensions[2].height = 22

    # رؤوس الأعمدة
    headers = ['#', 'التاريخ', 'المبلغ ₪', 'المبلغ $', 'مصدر الدفع', 'الجهة/الكافل', 'الحالة', 'الملاحظة']
    widths  = [6, 14, 14, 14, 16, 25, 12, 30]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = LBL_FONT; cell.fill = ALT_FILL
        cell.alignment = ALIGN_C; cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = w

    # البيانات
    total_ils = total_usd = Decimal('0')
    for i, p in enumerate(qs, 1):
        row = i + 4
        source = PAID_BY_LABELS.get(p.paid_by, p.paid_by)
        if p.paid_by == 'external' and p.paid_by_note:
            source = f'{source} ({p.paid_by_note})'
        sponsor = p.sponsor.get_full_name() if p.sponsor else '—'

        values = [
            i,
            p.date.strftime('%Y-%m-%d') if p.date else '—',
            float(p.amount_ils),
            float(p.amount_usd),
            source,
            sponsor,
            STATUS_LABELS.get(p.status, p.status),
            p.note or '—',
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = VAL_FONT
            cell.alignment = ALIGN_C if col != 8 else ALIGN_R
            cell.border = BORDER
            if i % 2 == 0: cell.fill = ALT_FILL
            if col in (3, 4):
                cell.number_format = '#,##0.00'

        total_ils += p.amount_ils
        total_usd += p.amount_usd

    # صف الإجمالي
    if qs.exists():
        last = len(qs) + 5
        ws.cell(row=last, column=1, value='الإجمالي').font = LBL_FONT
        ws.merge_cells(start_row=last, start_column=1, end_row=last, end_column=2)
        c3 = ws.cell(row=last, column=3, value=float(total_ils))
        c4 = ws.cell(row=last, column=4, value=float(total_usd))
        for c in (c3, c4):
            c.font = LBL_FONT; c.alignment = ALIGN_C; c.border = BORDER
            c.fill = PatternFill('solid', fgColor='E8F5EE')
            c.number_format = '#,##0.00'
        ws.row_dimensions[last].height = 26

    # الإرسال
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'payments_{user.id_number or user.pk}_{timezone.now().strftime("%Y%m%d")}.xlsx'
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"

    log_activity(user, 'EXPORT', description='تصدير سجل الدفعات Excel', request=request)
    return response


# ════════════════════════════════════════════
#  Export PDF (مع علامة مائية وحماية)
# ════════════════════════════════════════════

@login_required(login_url='/login/')
@require_GET
def export_pdf_api(request):
    if not _bene_required(request):
        return JsonResponse({'status': 'error', 'message': 'غير مسموح'}, status=403)

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm, mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    )
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.pdfgen import canvas

    user = request.user
    qs = Payment.objects.filter(beneficiary=user).select_related('sponsor')
    qs = _filter_queryset(qs, request.GET).order_by('-date')

    # ─── محاولة تحميل خط عربي ───
    font_name = 'Helvetica'
    try:
        from django.conf import settings
        candidates = [
            os.path.join(settings.BASE_DIR, 'static', 'fonts', 'Tajawal-Regular.ttf'),
            os.path.join(settings.BASE_DIR, 'static', 'fonts', 'Amiri-Regular.ttf'),
            os.path.join(settings.BASE_DIR, 'static', 'fonts', 'Cairo-Regular.ttf'),
        ]
        for p in candidates:
            if os.path.exists(p):
                pdfmetrics.registerFont(TTFont('ArabicFont', p))
                font_name = 'ArabicFont'
                break
    except Exception:
        pass

    # ─── تشكيل النص العربي ───
    def reshape(text):
        try:
            import arabic_reshaper
            from bidi.algorithm import get_display
            return get_display(arabic_reshaper.reshape(str(text or '')))
        except ImportError:
            return str(text or '')

    # ─── معلومات الأمان ───
    serial = hashlib.sha256(
        f'{user.pk}-{timezone.now().isoformat()}-{user.id_number or ""}'.encode()
    ).hexdigest()[:16].upper()
    timestamp = timezone.now().strftime('%Y-%m-%d %H:%M:%S')

    # ─── Canvas المخصص (للعلامة المائية والترقيم) ───
    class WatermarkCanvas(canvas.Canvas):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self.pages = []

        def showPage(self):
            self.pages.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            n = len(self.pages)
            for state in self.pages:
                self.__dict__.update(state)
                self._draw_watermark()
                self._draw_footer(self._pageNumber, n)
                super().showPage()
            super().save()

        def _draw_watermark(self):
            """علامة مائية مكررة بشفافية."""
            self.saveState()
            self.setFont(font_name, 50)
            self.setFillColor(colors.HexColor('#1a7a4a'))
            self.setFillAlpha(0.06)
            text = reshape('رُحَمَاء — وثيقة رسمية')
            # طبقات متقاطعة
            for y in range(50, 850, 180):
                for x_off in (-100, 200, 500):
                    self.saveState()
                    self.translate(x_off, y)
                    self.rotate(30)
                    self.drawString(0, 0, text)
                    self.restoreState()
            self.restoreState()

        def _draw_footer(self, page_num, total):
            """تذييل: الترقيم + الرقم التسلسلي + التحقق."""
            self.saveState()
            self.setFont(font_name, 8)
            self.setFillColor(colors.HexColor('#6b7280'))

            # خط فاصل
            self.setStrokeColor(colors.HexColor('#1a7a4a'))
            self.setLineWidth(0.5)
            self.line(40, 35, A4[0] - 40, 35)

            # يمين: الرقم التسلسلي
            self.drawRightString(A4[0] - 40, 22,
                reshape(f'الرقم التسلسلي: {serial}'))

            # وسط: الصفحة
            self.drawCentredString(A4[0] / 2, 22,
                reshape(f'صفحة {page_num} من {total}'))

            # يسار: الزمن
            self.drawString(40, 22, f'Generated: {timestamp}')

            # أسفل: تحذير
            self.setFont(font_name, 7)
            self.setFillColor(colors.HexColor('#c53030'))
            self.drawCentredString(A4[0] / 2, 10,
                reshape('⚠ هذه وثيقة محمية — أي تعديل عليها يُعدّ تزويراً'))
            self.restoreState()

    # ─── إعداد المستند ───
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        rightMargin=1.8*cm,
        leftMargin=1.8*cm,
        topMargin=1.5*cm,
        bottomMargin=2*cm,
        title=f'سجل دفعات - {user.get_full_name()}',
        author='Ruhamaa System',
        subject='Payment Statement',
        creator='Ruhamaa Charitable Platform',
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleAR', parent=styles['Heading1'],
        fontName=font_name, fontSize=18, alignment=TA_CENTER,
        textColor=colors.HexColor('#1a7a4a'), spaceAfter=10,
    )
    subtitle_style = ParagraphStyle(
        'SubtitleAR', parent=styles['Normal'],
        fontName=font_name, fontSize=11, alignment=TA_CENTER,
        textColor=colors.HexColor('#6b7280'), spaceAfter=6,
    )
    label_style = ParagraphStyle(
        'LabelAR', parent=styles['Normal'],
        fontName=font_name, fontSize=10, alignment=TA_RIGHT,
        textColor=colors.HexColor('#1a7a4a'),
    )
    value_style = ParagraphStyle(
        'ValueAR', parent=styles['Normal'],
        fontName=font_name, fontSize=10, alignment=TA_RIGHT,
        textColor=colors.HexColor('#1a1a2e'),
    )
    cell_style = ParagraphStyle(
        'CellAR', parent=styles['Normal'],
        fontName=font_name, fontSize=9, alignment=TA_CENTER,
    )

    story = []

    # ─── العنوان ───
    story.append(Paragraph(reshape('💰 كشف حساب الدفعات'), title_style))
    story.append(Paragraph(reshape('جمعية رُحَمَاء الخيرية'), subtitle_style))
    story.append(Spacer(1, 0.4*cm))

    # ─── معلومات المستفيد ───
    info_data = [
        [Paragraph(reshape(user.get_full_name() or '—'), value_style),
         Paragraph(reshape('الاسم:'), label_style)],
        [Paragraph(reshape(user.id_number or '—'), value_style),
         Paragraph(reshape('رقم الهوية:'), label_style)],
        [Paragraph(reshape(user.registration_number or '—'), value_style),
         Paragraph(reshape('رقم التسجيل:'), label_style)],
        [Paragraph(timestamp, value_style),
         Paragraph(reshape('تاريخ التقرير:'), label_style)],
    ]
    info_tbl = Table(info_data, colWidths=[10*cm, 5*cm])
    info_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f4f9f6')),
        ('GRID',       (0, 0), (-1, -1), 0.4, colors.HexColor('#c8e6cf')),
        ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING',(0, 0), (-1, -1), 8),
        ('RIGHTPADDING',(0,0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING',(0,0),(-1, -1), 6),
    ]))
    story.append(info_tbl)
    story.append(Spacer(1, 0.5*cm))

    # ─── الإحصائيات ───
    total_ils = qs.aggregate(s=Sum('amount_ils'))['s'] or Decimal('0')
    total_usd = qs.aggregate(s=Sum('amount_usd'))['s'] or Decimal('0')
    paid_count = qs.filter(status='paid').count()
    pending_count = qs.filter(status='pending').count()
    late_count = qs.filter(status='late').count()

    stats_data = [
        [
            Paragraph(reshape(f'<b>{qs.count()}</b>'), cell_style),
            Paragraph(reshape(f'<b>{total_ils:,.0f} ₪</b>'), cell_style),
            Paragraph(reshape(f'<b>{total_usd:,.2f} $</b>'), cell_style),
            Paragraph(reshape(f'<b>{paid_count}</b>'), cell_style),
            Paragraph(reshape(f'<b>{pending_count}</b>'), cell_style),
            Paragraph(reshape(f'<b>{late_count}</b>'), cell_style),
        ],
        [
            Paragraph(reshape('عدد الدفعات'), cell_style),
            Paragraph(reshape('الإجمالي ₪'), cell_style),
            Paragraph(reshape('الإجمالي $'), cell_style),
            Paragraph(reshape('مدفوعة'), cell_style),
            Paragraph(reshape('معلقة'), cell_style),
            Paragraph(reshape('متأخرة'), cell_style),
        ],
    ]
    stats_tbl = Table(stats_data, colWidths=[2.7*cm]*6)
    stats_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e8f5ee')),
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#1a7a4a')),
        ('TEXTCOLOR',  (0, 1), (-1, 1), colors.white),
        ('GRID',       (0, 0), (-1, -1), 0.4, colors.HexColor('#1a7a4a')),
        ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING',(0,0),(-1, -1), 8),
    ]))
    story.append(stats_tbl)
    story.append(Spacer(1, 0.6*cm))

    # ─── الجدول ───
    if qs.exists():
        headers = ['الملاحظة', 'الحالة', 'الجهة', 'المصدر', 'المبلغ $', 'المبلغ ₪', 'التاريخ', '#']
        table_data = [[Paragraph(reshape(h), cell_style) for h in headers]]

        for i, p in enumerate(qs, 1):
            sponsor = p.sponsor.get_full_name() if p.sponsor else '—'
            source = PAID_BY_LABELS.get(p.paid_by, '')
            if p.paid_by == 'external' and p.paid_by_note:
                source = f'{source}: {p.paid_by_note[:15]}'

            note_text = (p.note or '—')[:40]
            row = [
                Paragraph(reshape(note_text), cell_style),
                Paragraph(reshape(STATUS_LABELS.get(p.status, p.status)), cell_style),
                Paragraph(reshape(sponsor[:25]), cell_style),
                Paragraph(reshape(source[:25]), cell_style),
                Paragraph(f'{p.amount_usd:,.2f}', cell_style),
                Paragraph(f'{p.amount_ils:,.0f}', cell_style),
                Paragraph(p.date.strftime('%Y-%m-%d') if p.date else '—', cell_style),
                Paragraph(str(i), cell_style),
            ]
            table_data.append(row)

        # صف الإجمالي
        table_data.append([
            Paragraph('', cell_style),
            Paragraph('', cell_style),
            Paragraph('', cell_style),
            Paragraph(reshape('<b>الإجمالي</b>'), cell_style),
            Paragraph(f'<b>{total_usd:,.2f}</b>', cell_style),
            Paragraph(f'<b>{total_ils:,.0f}</b>', cell_style),
            Paragraph('', cell_style),
            Paragraph('', cell_style),
        ])

        col_widths = [3.5*cm, 1.6*cm, 2.5*cm, 2.5*cm, 1.8*cm, 1.8*cm, 1.8*cm, 0.8*cm]
        tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            # الترويسة
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a7a4a')),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
            ('FONTSIZE',   (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING',(0,0), (-1, 0), 8),
            # الجسم
            ('GRID',       (0, 0), (-1, -1), 0.3, colors.HexColor('#c8e6cf')),
            ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE',   (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 5),
            ('BOTTOMPADDING',(0,1), (-1, -1), 5),
            # صفوف بديلة
            ('ROWBACKGROUNDS', (0, 1), (-1, -2),
             [colors.white, colors.HexColor('#f4f9f6')]),
            # صف الإجمالي
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8f5ee')),
            ('LINEABOVE',  (0, -1), (-1, -1), 1.5, colors.HexColor('#1a7a4a')),
        ]))
        story.append(tbl)
    else:
        story.append(Spacer(1, 1*cm))
        story.append(Paragraph(
            reshape('لا توجد دفعات مسجّلة'),
            ParagraphStyle('Empty', parent=styles['Normal'],
                fontName=font_name, fontSize=14, alignment=TA_CENTER,
                textColor=colors.HexColor('#9ca3af'))
        ))

    # ─── ختم رقمي / توقيع ───
    story.append(Spacer(1, 0.8*cm))
    sign_style = ParagraphStyle(
        'Sign', parent=styles['Normal'],
        fontName=font_name, fontSize=8, alignment=TA_CENTER,
        textColor=colors.HexColor('#6b7280'),
    )
    sign_data = [[
        Paragraph(reshape(
            f'<b>الرقم التسلسلي:</b> {serial}<br/>'
            f'<b>تم التوليد في:</b> {timestamp}<br/>'
            f'<b>للتحقق من صحة الوثيقة، يرجى التواصل مع إدارة الجمعية</b>'
        ), sign_style)
    ]]
    sign_tbl = Table(sign_data, colWidths=[16*cm])
    sign_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#fff8e1')),
        ('BOX',        (0, 0), (-1, -1), 1, colors.HexColor('#f59e0b')),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING',(0,0),(-1, -1), 10),
    ]))
    story.append(sign_tbl)

    # ─── البناء ───
    doc.build(story, canvasmaker=WatermarkCanvas)

    # ─── الحماية: كلمة مرور + منع النسخ ───
    pdf_data = buf.getvalue()
    buf.close()

    try:
        from PyPDF2 import PdfReader, PdfWriter
        reader = PdfReader(BytesIO(pdf_data))
        writer = PdfWriter()
        for page in reader.pages:
            writer.add_page(page)

        # كلمة المرور = آخر 4 أرقام من الهوية أو "0000"
        password = (user.id_number or '0000')[-4:]
        writer.encrypt(
            user_password=password,
            owner_password=f'RUH-{serial}',
            use_128bit=True,
            permissions_flag=4  # طباعة فقط، منع النسخ والتعديل
        )

        out = BytesIO()
        writer.write(out)
        pdf_data = out.getvalue()
        out.close()
    except ImportError:
        pass  # لو PyPDF2 غير مثبت، يتم إرجاع الـ PDF بدون حماية

    # ─── الإرسال ───
    filename = f'payments_{user.id_number or user.pk}_{timezone.now().strftime("%Y%m%d")}.pdf'
    response = HttpResponse(pdf_data, content_type='application/pdf')
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"

    # رسالة بكلمة المرور في الـ header (يقرأها الفرونت)
    response['X-PDF-Password'] = (user.id_number or '0000')[-4:]

    log_activity(user, 'EXPORT',
                 description=f'تصدير سجل الدفعات PDF (Serial: {serial})',
                 request=request)
    return response

# ════════════════════════════════════════════
#  دالة الـ hash لكلمة السر (مطابقة للجافاسكريبت)
# ════════════════════════════════════════════

def _js_hash(s: str) -> str:
    """
    نفس خوارزمية simpleHash في الجافاسكريبت:
        h = (31 * h + char_code) | 0   (32-bit signed int)
    ثم تحويله إلى unsigned hex.
    """
    import ctypes
    h = ctypes.c_int32(0)
    for ch in s:
        h = ctypes.c_int32((31 * h.value + ord(ch)))
    # unsigned
    return format(ctypes.c_uint32(h.value).value, 'x')


# ════════════════════════════════════════════
#  View: كشف الدفعات HTML قابل للطباعة
# ════════════════════════════════════════════

@login_required(login_url='/login/')
@require_GET
def export_pdf_html_api(request):
    """
    يعرض صفحة HTML مصمّمة للطباعة كـ PDF من المتصفح.
    محمية بكلمة سر (آخر 4 أرقام من رقم الهوية).
    """
    if not _bene_required(request):
        return redirect('/')

    user = request.user

    # ── البيانات ──
    qs = Payment.objects.filter(beneficiary=user).select_related('sponsor')
    qs = _filter_queryset(qs, request.GET).order_by('-date')

    # ── الإحصائيات ──
    total_ils     = qs.aggregate(s=Sum('amount_ils'))['s'] or Decimal('0')
    total_usd     = qs.aggregate(s=Sum('amount_usd'))['s'] or Decimal('0')
    paid_count    = qs.filter(status='paid').count()
    pending_count = qs.filter(status='pending').count()
    late_count    = qs.filter(status='late').count()

    # ── تحضير بيانات الدفعات للقالب ──
    payments_list = []
    for p in qs:
        payments_list.append({
            'date':          p.date.strftime('%Y-%m-%d') if p.date else '—',
            'amount_ils':    p.amount_ils,
            'amount_usd':    p.amount_usd,
            'paid_by':       p.paid_by,
            'paid_by_label': PAID_BY_LABELS.get(p.paid_by, p.paid_by),
            'paid_by_note':  p.paid_by_note or '',
            'status':        p.status,
            'status_label':  STATUS_LABELS.get(p.status, p.status),
            'note':          p.note or '',
            'sponsor':       p.sponsor.get_full_name() if p.sponsor else '',
        })

    # ── معلومات الكافل ──
    sponsor_info = None
    last_sponsor = qs.filter(sponsor__isnull=False).first()
    if last_sponsor and last_sponsor.sponsor:
        sp = last_sponsor.sponsor
        sponsor_info = {
            'name':       sp.get_full_name(),
            'reg_number': getattr(sp, 'registration_number', ''),
            'phone':      getattr(sp, 'phone', ''),
        }

    # ── الرقم التسلسلي ──
    serial = hashlib.sha256(
        f'{user.pk}-{timezone.now().isoformat()}-{user.id_number or ""}'.encode()
    ).hexdigest()[:16].upper()

    # ── كلمة السر: آخر 4 أرقام من الهوية ──
    raw_pwd = (user.id_number or '0000')[-4:]
    pwd_hash = _js_hash(raw_pwd)

    # ── الفترة الزمنية للفلتر (للعرض في القالب) ──
    filter_from = request.GET.get('date_from', '').strip()
    filter_to   = request.GET.get('date_to', '').strip()

    context = {
        'user':         user,
        'payments':     payments_list,
        'stats': {
            'count':         qs.count(),
            'total_ils':     total_ils,
            'total_usd':     total_usd,
            'paid_count':    paid_count,
            'pending_count': pending_count,
            'late_count':    late_count,
        },
        'sponsor_info': sponsor_info,
        'serial':       serial,
        'report_date':  timezone.now().strftime('%Y-%m-%d %H:%M'),
        'pwd_hash':     pwd_hash,
        'filter_from':  filter_from,
        'filter_to':    filter_to,
    }

    log_activity(user, 'EXPORT',
                 description=f'فتح كشف الدفعات HTML للطباعة (Serial: {serial})',
                 request=request)

    return render(request, 'beneficiary/payments_pdf.html', context)


@login_required(login_url='/login/')
@require_GET
def download_pdf_html_api(request):
    if not _bene_required(request):
        return redirect('/')

    try:
        from weasyprint import HTML as WeasyprintHTML
        from django.template.loader import render_to_string
    except ImportError:
        return JsonResponse({'status': 'error', 'message': 'WeasyPrint غير مثبت'}, status=500)

    user = request.user
    qs = Payment.objects.filter(beneficiary=user).select_related('sponsor')
    qs = _filter_queryset(qs, request.GET).order_by('-date')

    total_ils     = qs.aggregate(s=Sum('amount_ils'))['s'] or Decimal('0')
    total_usd     = qs.aggregate(s=Sum('amount_usd'))['s'] or Decimal('0')

    payments_list = []
    for p in qs:
        payments_list.append({
            'date':          p.date.strftime('%Y-%m-%d') if p.date else '—',
            'amount_ils':    p.amount_ils,
            'amount_usd':    p.amount_usd,
            'paid_by':       p.paid_by,
            'paid_by_label': PAID_BY_LABELS.get(p.paid_by, p.paid_by),
            'paid_by_note':  p.paid_by_note or '',
            'status':        p.status,
            'status_label':  STATUS_LABELS.get(p.status, p.status),
            'note':          p.note or '',
            'sponsor':       p.sponsor.get_full_name() if p.sponsor else '',
        })

    sponsor_info = None
    last_sponsor = qs.filter(sponsor__isnull=False).first()
    if last_sponsor and last_sponsor.sponsor:
        sp = last_sponsor.sponsor
        sponsor_info = {
            'name':       sp.get_full_name(),
            'reg_number': getattr(sp, 'registration_number', ''),
            'phone':      getattr(sp, 'phone', ''),
        }

    serial = hashlib.sha256(
        f'{user.pk}-{timezone.now().isoformat()}-{user.id_number or ""}'.encode()
    ).hexdigest()[:16].upper()

    context = {
        'user':         user,
        'payments':     payments_list,
        'stats': {
            'count':         qs.count(),
            'total_ils':     total_ils,
            'total_usd':     total_usd,
            'paid_count':    qs.filter(status='paid').count(),
            'pending_count': qs.filter(status='pending').count(),
            'late_count':    qs.filter(status='late').count(),
        },
        'sponsor_info':  sponsor_info,
        'serial':        serial,
        'report_date':   timezone.now().strftime('%Y-%m-%d %H:%M'),
        'filter_from':   request.GET.get('date_from', ''),
        'filter_to':     request.GET.get('date_to', ''),
        'is_download':   True,   # ← يخبر القالب أن هذا تحميل وليس طباعة
    }

    # تحويل HTML إلى PDF
    html_string = render_to_string('beneficiary/payments_pdf.html', context, request=request)
    pdf_file = WeasyprintHTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()

    # كلمة السر: آخر 4 أرقام من الهوية
    try:
        from PyPDF2 import PdfReader, PdfWriter
        from io import BytesIO

        reader = PdfReader(BytesIO(pdf_file))
        writer = PdfWriter()
        for page in reader.pages:
            writer.add_page(page)

        password = (user.id_number or '0000')[-4:]
        writer.encrypt(
            user_password=password,
            owner_password=f'RUH-{serial}',
            use_128bit=True,
            permissions_flag=4
        )

        out = BytesIO()
        writer.write(out)
        pdf_file = out.getvalue()
    except ImportError:
        pass  # بدون حماية إن لم يكن PyPDF2 مثبتاً

    filename = f'payments_{user.id_number or user.pk}_{timezone.now().strftime("%Y%m%d")}.pdf'
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"
    response['X-PDF-Password'] = (user.id_number or '0000')[-4:]

    log_activity(user, 'EXPORT',
                 description=f'تحميل كشف الدفعات PDF مصمم (Serial: {serial})',
                 request=request)
    return response