from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_protect

from core.models import CustomUser, Notification
from core.utils import log_activity, get_exchange_rates
from .decorators import admin_required


@admin_required
def orphans_list(request):
    from beneficiary.models import OrphanForm, SpecialNeedsForm, FamilyForm

    bene_type = request.GET.get('type',   'orphan')
    status    = request.GET.get('status', 'all')
    search    = request.GET.get('q',      '').strip()

    model_map = {
        'orphan':  OrphanForm,
        'special': SpecialNeedsForm,
        'family':  FamilyForm,
    }
    Model = model_map.get(bene_type, OrphanForm)
    qs    = Model.objects.select_related('user', 'sponsor__user').order_by('-created_at')

    if status != 'all':
        qs = qs.filter(status=status)
    if search:
        qs = qs.filter(first_name__icontains=search) | \
             qs.filter(family_name__icontains=search) | \
             qs.filter(id_number__icontains=search)

    notif_count = Notification.objects.filter(
        recipient=request.user, is_read=False
    ).count()

    context = {
        'beneficiaries': qs.distinct(),
        'bene_type':     bene_type,
        'status':        status,
        'search':        search,
        'notif_count':   notif_count,
        'counts': {
            'orphan':  OrphanForm.objects.count(),
            'special': SpecialNeedsForm.objects.count(),
            'family':  FamilyForm.objects.count(),
        }
    }
    return render(request, 'admin_panel/orphans.html', context)


@admin_required
@require_POST
@csrf_protect
def update_status(request):
    """تحديث حالة الاستمارة"""
    from beneficiary.models import OrphanForm, SpecialNeedsForm, FamilyForm
    from core.utils import create_notification

    bene_type = request.POST.get('bene_type', '').strip()
    bene_id   = request.POST.get('bene_id',   '').strip()
    new_status= request.POST.get('status',    '').strip()

    model_map = {
        'orphan':  OrphanForm,
        'special': SpecialNeedsForm,
        'family':  FamilyForm,
    }
    Model = model_map.get(bene_type)
    if not Model:
        return JsonResponse({'status': 'error', 'message': 'نوع غير صالح'})

    try:
        form = Model.objects.get(pk=bene_id)
        old  = form.status
        form.status = new_status
        form.save()

        # إشعار المستفيد
        create_notification(
            recipient  = form.user,
            ntype      = 'SYSTEM',
            title      = f'تم تحديث حالة استمارتك',
            message    = f'تم تغيير الحالة من "{old}" إلى "{new_status}"',
            sender     = request.user,
        )

        log_activity(
            request.user, 'UPDATE',
            description  = f'تحديث حالة استمارة {form.form_number}: {old} → {new_status}',
            target_model = type(form).__name__,
            target_id    = form.pk,
            old_value    = {'status': old},
            new_value    = {'status': new_status},
            request      = request,
        )
        return JsonResponse({'status': 'success', 'message': 'تم التحديث ✅'})

    except Exception:
        return JsonResponse({'status': 'error', 'message': 'حدث خطأ'})


@admin_required
def export_orphans(request):
    """تصدير Excel"""
    from beneficiary.models import OrphanForm, SpecialNeedsForm, FamilyForm
    bene_type = request.GET.get('type', 'orphan')
    model_map = {'orphan': OrphanForm, 'special': SpecialNeedsForm, 'family': FamilyForm}
    Model     = model_map.get(bene_type, OrphanForm)
    qs        = Model.objects.all().order_by('-created_at')

    log_activity(
        request.user, 'EXPORT',
        description = f'تصدير قائمة {bene_type}',
        request     = request,
    )

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import io

        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = 'المستفيدون'
        ws.sheet_view.rightToLeft = True

        GREEN = '1A7A4A'
        LIGHT = 'E8F5E9'
        thin  = Side(style='thin', color='CCCCCC')
        border= openpyxl.styles.Border(left=thin,right=thin,top=thin,bottom=thin)

        headers = ['رقم الاستمارة','الاسم الكامل','رقم الهوية',
                   'المدينة','الحالة الصحية','حالة الاستمارة','تاريخ التسجيل']

        for i, h in enumerate(headers, 1):
            c           = ws.cell(row=1, column=i, value=h)
            c.font      = Font(bold=True, color='FFFFFF')
            c.fill      = PatternFill('solid', fgColor=GREEN)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border    = border
        ws.row_dimensions[1].height = 24

        for r_idx, obj in enumerate(qs, start=2):
            row = [
                obj.form_number,
                obj.get_full_name(),
                obj.id_number or '—',
                obj.current_city or '—',
                obj.health_status or '—',
                obj.status,
                str(obj.created_at.date()),
            ]
            for c_idx, val in enumerate(row, 1):
                cell           = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.border    = border
                ws.row_dimensions[r_idx].height = 20
                if r_idx % 2 == 0:
                    cell.fill = PatternFill('solid', fgColor='F4F9F6')

        for col, w in zip(['A','B','C','D','E','F','G'], [16,28,14,14,16,16,14]):
            ws.column_dimensions[col].width = w

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        resp = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        resp['Content-Disposition'] = f'attachment; filename="{bene_type}_list.xlsx"'
        return resp

    except ImportError:
        return HttpResponse('مكتبة openpyxl غير مثبّتة', status=500)