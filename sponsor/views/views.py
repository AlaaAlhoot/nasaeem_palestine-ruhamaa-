from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_protect
from django.utils import timezone
from django.db.models import Sum, Q
import re

from core.models import CustomUser, Notification, DirectMessage
from core.utils import (log_activity, get_client_ip, get_exchange_rates,
                        create_notification, notify_admins, compress_image)
from sponsor.models import SponsorProfile, PaymentReceipt, PaymentSchedule
from sponsor.views import *
from sponsor.views import base
from sponsor.views.base import _get_profile,sponsor_required


# ==================== Decorator ====================
# views.py
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required





def _get_beneficiaries(profile):
    """يرجع قائمة المستفيدين المرتبطين بالكافل"""
    from beneficiary.models import OrphanForm, SpecialNeedsForm, FamilyForm
    orphans  = OrphanForm.objects.filter(sponsor=profile).select_related('user')
    specials = SpecialNeedsForm.objects.filter(sponsor=profile).select_related('user')
    families = FamilyForm.objects.filter(sponsor=profile).select_related('user')
    return list(orphans) + list(specials) + list(families)





# ==================== قائمة الأيتام ====================

import json

@sponsor_required
def orphans(request):
    user    = request.user
    profile = _get_profile(user)

    from beneficiary.models import OrphanForm, SpecialNeedsForm, FamilyForm

    my_orphans  = OrphanForm.objects.filter(sponsor=profile)
    my_specials = SpecialNeedsForm.objects.filter(sponsor=profile)
    my_families = FamilyForm.objects.filter(sponsor=profile)
    free_orphans  = OrphanForm.objects.filter(sponsor__isnull=True, status='مقبولة')
    free_specials = SpecialNeedsForm.objects.filter(sponsor__isnull=True, status='مقبولة')
    free_families = FamilyForm.objects.filter(sponsor__isnull=True, status='مقبولة')

    def serialize(qs):
        result = []
        for b in qs:
            result.append({
                'id':            b.pk,
                'full_name':     b.get_full_name(),
                'form_number':   b.form_number,
                'id_number':     b.id_number or '',
                'birth_date':    str(b.birth_date) if b.birth_date else '',
                'nationality':   b.nationality or '',
                'gender':        b.gender or '',
                'health_status': b.health_status or '',
                'current_city':  b.current_city or '',
                'current_street':b.current_street or '',
                'current_landmark': getattr(b, 'current_landmark', '') or '',
                'housing_type':  b.housing_type or '',
                'phone1':        b.phone1 or '',
                'whatsapp':      getattr(b, 'whatsapp', '') or '',
                'story':         getattr(b, 'story', '') or '',
                'case_details':  getattr(b, 'case_details', '') or '',
                'general_status':getattr(b, 'general_status', '') or '',
                'education_level': getattr(b, 'education_level', '') or '',
                'sponsorship_date': str(b.sponsorship_date) if b.sponsorship_date else '',
            })
        return result

    notif_count = Notification.objects.filter(
        recipient=user, is_read=False
    ).count()

    context = {
        'my_orphans':    my_orphans,
        'my_specials':   my_specials,
        'my_families':   my_families,
        'free_orphans':  free_orphans,
        'free_specials': free_specials,
        'free_families': free_families,
        'orphans_json':  json.dumps(serialize(list(my_orphans)  + list(free_orphans)),  ensure_ascii=False),
        'specials_json': json.dumps(serialize(list(my_specials) + list(free_specials)), ensure_ascii=False),
        'families_json': json.dumps(serialize(list(my_families) + list(free_families)), ensure_ascii=False),
        'notif_count':   notif_count,
    }
    return render(request, 'sponsor/orphans.html', context)


# ==================== رفع وصل مالي ====================

@sponsor_required
def receipt(request):
    user    = request.user
    profile = _get_profile(user)
    rates   = get_exchange_rates()

    # آخر 5 وصولات
    receipts = PaymentReceipt.objects.filter(
        sponsor=profile
    ).order_by('-created_at')[:5]

    notif_count = Notification.objects.filter(
        recipient=user, is_read=False
    ).count()

    context = {
        'receipts':   receipts,
        'rates':      rates,
        'notif_count':notif_count,
    }
    return render(request, 'sponsor/receipt.html', context)


@sponsor_required
@require_POST
@csrf_protect
def submit_receipt(request):
    user    = request.user
    profile = _get_profile(user)

    receipt_date  = request.POST.get('receipt_date',  '').strip()
    sender_name   = request.POST.get('sender_name',   '').strip()
    amount_shekel = request.POST.get('amount_shekel', '0').strip()
    amount_dollar = request.POST.get('amount_dollar', '0').strip()
    unique_number = request.POST.get('unique_number', '').strip()
    notes         = request.POST.get('notes',         '').strip()
    errors        = {}

    # التحقق
    if not receipt_date:
        errors['receipt_date'] = 'تاريخ الوصل مطلوب'
    if not sender_name or len(sender_name) > 200:
        errors['sender_name'] = 'اسم المُرسِل مطلوب'
    if re.search(r'<script|javascript:', sender_name, re.I):
        errors['sender_name'] = 'مدخل غير صالح'
    try:
        amount_shekel = float(amount_shekel)
        if amount_shekel <= 0:
            errors['amount_shekel'] = 'المبلغ يجب أن يكون أكبر من صفر'
    except ValueError:
        errors['amount_shekel'] = 'مبلغ غير صالح'
    try:
        amount_dollar = float(amount_dollar)
    except ValueError:
        amount_dollar = 0
    if not unique_number:
        errors['unique_number'] = 'الرقم الفريد للوصل مطلوب'
    if 'receipt_image' not in request.FILES:
        errors['receipt_image'] = 'صورة الوصل مطلوبة'

    if errors:
        return JsonResponse({'status': 'error', 'errors': errors,
                             'message': 'يرجى تصحيح الأخطاء'})

    # التحقق من الصورة
    img = request.FILES['receipt_image']
    if img.size > 4 * 1024 * 1024:
        return JsonResponse({'status': 'error',
                             'errors': {'receipt_image': 'الصورة تتجاوز 4MB'}})
    ext = img.name.split('.')[-1].lower()
    if ext not in ['jpg', 'jpeg', 'png', 'pdf']:
        return JsonResponse({'status': 'error',
                             'errors': {'receipt_image': 'نوع الملف غير مسموح'}})

    # حفظ
    rec = PaymentReceipt.objects.create(
        sponsor       = profile,
        receipt_date  = receipt_date,
        sender_name   = sender_name,
        amount_shekel = amount_shekel,
        amount_dollar = amount_dollar,
        unique_number = unique_number,
        receipt_image = img,
        notes         = notes,
        status        = 'بانتظار المراجعة',
    )

    # إشعار الأدمن
    notify_admins(
        ntype      = 'NEW_RECEIPT',
        title      = 'وصل مالي جديد 💳',
        message    = f'وصل من {user.get_full_name()} — {amount_shekel} ₪',
        sender     = user,
        action_url = '/admin-panel/receipts/',
    )

    log_activity(user, 'RECEIPT',
                 description=f'رفع وصل مالي {amount_shekel}₪',
                 request=request)

    return JsonResponse({
        'status':  'success',
        'message': 'تم إرسال الوصل للمراجعة ✅ سيتم إشعارك بعد المراجعة'
    })


# ==================== المحفظة ====================

@sponsor_required
def wallet(request):
    user    = request.user
    profile = _get_profile(user)
    rates   = get_exchange_rates()

    # فلتر الحالة
    status_filter = request.GET.get('status', 'all')
    qs = PaymentReceipt.objects.filter(sponsor=profile).order_by('-receipt_date')
    if status_filter != 'all':
        qs = qs.filter(status=status_filter)

    # الإجماليات
    approved = PaymentReceipt.objects.filter(sponsor=profile, status='موافق')
    total_shekel  = approved.aggregate(t=Sum('amount_shekel'))['t'] or 0
    total_dollar  = approved.aggregate(t=Sum('amount_dollar'))['t'] or 0
    pending_count = PaymentReceipt.objects.filter(
        sponsor=profile, status='بانتظار المراجعة'
    ).count()

    notif_count = Notification.objects.filter(
        recipient=user, is_read=False
    ).count()

    context = {
        'receipts':     qs,
        'total_shekel': total_shekel,
        'total_dollar': total_dollar,
        'pending_count':pending_count,
        'status_filter':status_filter,
        'rates':        rates,
        'notif_count':  notif_count,
    }
    return render(request, 'sponsor/wallet.html', context)


# ==================== التواصل ====================











# ==================== الإعدادات ====================

@sponsor_required
def settings_view(request):
    user    = request.user
    profile = _get_profile(user)
    notif_count = Notification.objects.filter(
        recipient=user, is_read=False
    ).count()
    return render(request, 'sponsor/settings.html', {
        'profile':    profile,
        'notif_count':notif_count,
    })


@sponsor_required
@require_POST
@csrf_protect
def update_settings(request):
    user    = request.user
    profile = _get_profile(user)
    action  = request.POST.get('action', '').strip()

    # تغيير كلمة المرور
    if action == 'change_password':
        old_pass  = request.POST.get('old_password',  '').strip()
        new_pass  = request.POST.get('new_password',  '').strip()
        new_pass2 = request.POST.get('new_password2', '').strip()

        if not user.check_password(old_pass):
            return JsonResponse({'status': 'error',
                                 'message': 'كلمة المرور الحالية غير صحيحة'})
        if len(new_pass) < 8:
            return JsonResponse({'status': 'error',
                                 'message': 'كلمة المرور الجديدة قصيرة جداً'})
        if new_pass != new_pass2:
            return JsonResponse({'status': 'error',
                                 'message': 'كلمتا المرور غير متطابقتين'})

        user.set_password(new_pass)
        user.save()
        log_activity(user, 'UPDATE',
                     description='تغيير كلمة المرور',
                     request=request)
        return JsonResponse({'status': 'success',
                             'message': 'تم تغيير كلمة المرور ✅'})

    # تحديث البيانات الأساسية
    if action == 'update_profile':
        gender    = request.POST.get('gender',    '').strip()
        job       = request.POST.get('job',       '').strip()
        country   = request.POST.get('country',   '').strip()
        city      = request.POST.get('city',      '').strip()

        # منع XSS
        for v in [gender, job, country, city]:
            if re.search(r'<script|javascript:', v, re.I):
                return JsonResponse({'status': 'error', 'message': 'مدخل غير صالح'})

        profile.gender  = gender
        profile.job     = job
        profile.country = country
        profile.city    = city
        profile.save()

        # الصورة الشخصية
        if 'photo' in request.FILES:
            img = request.FILES['photo']
            if img.size <= 4 * 1024 * 1024:
                compressed = compress_image(img)
                profile.photo.save(img.name, compressed, save=True)

        log_activity(user, 'UPDATE',
                     description='تحديث بيانات الحساب',
                     request=request)
        return JsonResponse({'status': 'success',
                             'message': 'تم تحديث البيانات ✅'})

    # تبديل الوضع الليلي
    if action == 'toggle_dark':
        user.dark_mode = not user.dark_mode
        user.save()
        return JsonResponse({'status': 'success', 'dark': user.dark_mode})

    return JsonResponse({'status': 'error', 'message': 'إجراء غير معروف'})


# ==================== helpers Excel/PDF ====================

def _export_excel(user, profile, receipts, rates):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import io
        from datetime import date

        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = 'سجل الوصولات'
        ws.sheet_view.rightToLeft = True

        GREEN = '1A7A4A'
        LIGHT = 'E8F5E9'
        WHITE = 'FFFFFF'
        thin  = Side(style='thin', color='CCCCCC')
        border= openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)

        # Header
        ws.merge_cells('A1:F1')
        c = ws['A1']
        c.value     = f'سجل الوصولات المالية — {user.get_full_name()}'
        c.font      = Font(bold=True, size=13, color=WHITE)
        c.fill      = PatternFill('solid', fgColor=GREEN)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 32

        ws.merge_cells('A2:F2')
        d = ws['A2']
        d.value     = f'تاريخ التصدير: {date.today()}'
        d.font      = Font(size=10, color=GREEN)
        d.fill      = PatternFill('solid', fgColor=LIGHT)
        d.alignment = Alignment(horizontal='center')

        headers = ['التاريخ', 'اسم المُرسِل', 'الرقم الفريد',
                   'المبلغ (₪)', 'المبلغ ($)', 'الحالة']
        for i, h in enumerate(headers, 1):
            cell        = ws.cell(row=3, column=i, value=h)
            cell.font   = Font(bold=True, color=WHITE)
            cell.fill   = PatternFill('solid', fgColor='2ECC71')
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        ws.row_dimensions[3].height = 24

        total = 0
        for r_idx, rec in enumerate(receipts, start=4):
            row = [str(rec.receipt_date), rec.sender_name, rec.unique_number,
                   float(rec.amount_shekel), float(rec.amount_dollar), rec.status]
            total += float(rec.amount_shekel)
            for c_idx, val in enumerate(row, 1):
                cell           = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border    = border
                ws.row_dimensions[r_idx].height = 20
                if r_idx % 2 == 0:
                    cell.fill = PatternFill('solid', fgColor='F4F9F6')

        # Total
        tr = len(list(receipts)) + 4
        ws.cell(tr, 1, 'الإجمالي').font = Font(bold=True)
        ws.cell(tr, 4, total).font = Font(bold=True, color=GREEN)
        ws.cell(tr, 5, round(total * rates['ILS_TO_USD'], 2)).font = Font(bold=True, color=GREEN)
        for c in range(1, 7):
            ws.cell(tr, c).fill = PatternFill('solid', fgColor=LIGHT)
            ws.cell(tr, c).border = border

        for col, w in zip(['A','B','C','D','E','F'], [14,24,18,14,14,16]):
            ws.column_dimensions[col].width = w

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        resp = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        resp['Content-Disposition'] = 'attachment; filename="sponsor_report.xlsx"'
        return resp
    except ImportError:
        return HttpResponse('مكتبة openpyxl غير مثبّتة', status=500)


def _export_pdf(user, profile, receipts, rates):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import (SimpleDocTemplate, Table,
                                        TableStyle, Paragraph, Spacer)
        from reportlab.lib.styles import ParagraphStyle
        import io
        from datetime import date

        buffer = io.BytesIO()
        doc    = SimpleDocTemplate(buffer, pagesize=A4,
                                   rightMargin=40, leftMargin=40,
                                   topMargin=40, bottomMargin=40)
        story  = []
        green  = colors.HexColor('#1A7A4A')

        title_s = ParagraphStyle('t', fontSize=16, textColor=green,
                                 alignment=1, fontName='Helvetica-Bold',
                                 spaceAfter=6)
        sub_s   = ParagraphStyle('s', fontSize=10, textColor=colors.grey,
                                 alignment=1, spaceAfter=20)

        story.append(Paragraph('تقرير الوصولات المالية', title_s))
        story.append(Paragraph(
            f'{user.get_full_name()} — {date.today()}', sub_s
        ))

        data = [['التاريخ', 'الرقم الفريد', 'المبلغ (₪)', 'المبلغ ($)', 'الحالة']]
        total = 0
        for r in receipts:
            data.append([
                str(r.receipt_date), r.unique_number,
                str(r.amount_shekel), str(r.amount_dollar), r.status,
            ])
            total += float(r.amount_shekel)
        data.append(['الإجمالي', '', str(total),
                     str(round(total * rates['ILS_TO_USD'], 2)), ''])

        t = Table(data, colWidths=[80, 100, 80, 80, 80])
        t.setStyle(TableStyle([
            ('BACKGROUND',    (0,0), (-1,0), green),
            ('TEXTCOLOR',     (0,0), (-1,0), colors.white),
            ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
            ('GRID',          (0,0), (-1,-1), 0.5, colors.grey),
            ('ROWBACKGROUNDS',(0,1), (-1,-2),
             [colors.HexColor('#F4F9F6'), colors.white]),
            ('BACKGROUND',    (0,-1), (-1,-1), colors.HexColor('#E8F5E9')),
            ('FONTNAME',      (0,-1), (-1,-1), 'Helvetica-Bold'),
        ]))
        story.append(t)
        doc.build(story)
        buffer.seek(0)

        resp = HttpResponse(buffer, content_type='application/pdf')
        resp['Content-Disposition'] = 'attachment; filename="sponsor_report.pdf"'
        return resp
    except ImportError:
        return HttpResponse('مكتبة reportlab غير مثبّتة', status=500)