"""
receipts.py — فيوز الوصولات المالية
"""

from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST, require_GET
from django.views.decorators.csrf import csrf_protect
from django.utils import timezone
from django.db.models import Sum, Q
from decimal import Decimal
import io, json

from core.models import CustomUser, Payment, Notification
from core.utils import log_activity, create_notification
from .decorators import admin_required


# ── مساعد اسم المستخدم ──
def _uname(user):
    if not user: return '—'
    return user.get_full_name() or user.username or '—'


# ── تحويل العملة إلى شيقل ودولار ──
def _convert_to_ils_usd(amount, currency):
    """تحويل المبلغ إلى شيقل ودولار باستخدام أسعار الصرف"""
    try:
        from core.utils import get_exchange_rates
        rates = get_exchange_rates()
        amount = Decimal(str(amount))

        RATES_TO_ILS = {
            'ILS': Decimal('1'),
            'USD': Decimal(str(rates.get('USD_TO_ILS', 3.7))),
            'JOD': Decimal(str(rates.get('JOD_TO_ILS', 5.2))),
            'SAR': Decimal(str(rates.get('SAR_TO_ILS', 0.99))),
            'EGP': Decimal(str(rates.get('EGP_TO_ILS', 0.075))),
        }
        USD_RATE = Decimal(str(rates.get('USD_TO_ILS', 3.7)))

        rate_to_ils = RATES_TO_ILS.get(currency, Decimal('1'))
        amount_ils  = round(amount * rate_to_ils, 2)
        amount_usd  = round(amount_ils / USD_RATE, 2)
        return amount_ils, amount_usd
    except Exception:
        return Decimal(str(amount)), Decimal('0')


# ── بناء بيانات وصل ──
def _build_receipt_data(r):
    return {
        'id':            r.pk,
        'system_ref':    r.system_ref or '',
        'unique_number': r.unique_number or '',
        'sponsor_name':  _uname(r.sponsor.user) if r.sponsor else '—',
        'sponsor_id':    str(r.sponsor.user.pk) if r.sponsor else '',
        'sponsor_reg':   r.sponsor.user.registration_number or '' if r.sponsor else '',
        'beneficiary_name': _uname(r.beneficiary) if r.beneficiary else '—',
        'beneficiary_id':   str(r.beneficiary.pk) if r.beneficiary else '',
        'beneficiary_type': r.beneficiary.user_type if r.beneficiary else '',
        'amount_original':  str(r.amount_original or 0),
        'currency':         r.currency or 'USD',
        'amount_ils':       str(r.amount_shekel or 0),
        'amount_usd':       str(r.amount_dollar or 0),
        'receipt_date':     str(r.receipt_date) if r.receipt_date else '',
        'sender_name':      r.sender_name or '',
        'status':           r.status,
        'notes':            r.notes or '',
        'reject_reason':    r.reject_reason or '',
        'receipt_image':    r.receipt_image.url if r.receipt_image else '',
        'reviewed_by':      _uname(r.reviewed_by) if r.reviewed_by else '',
        'reviewed_at':      timezone.localtime(r.reviewed_at).strftime('%Y/%m/%d %H:%M') if r.reviewed_at else '',
        'submitted_at':     timezone.localtime(r.submitted_at).strftime('%Y/%m/%d %H:%M') if r.submitted_at else '',
        'created_at':       timezone.localtime(r.created_at).strftime('%Y/%m/%d %H:%M') if r.created_at else '',
    }


# ══════════════════════════════════════════════
# الصفحة الرئيسية
# ══════════════════════════════════════════════
@admin_required
def receipts_list(request):
    return render(request, 'admin_panel/receipts.html')


# ══════════════════════════════════════════════
# جلب البيانات AJAX
# ══════════════════════════════════════════════
@admin_required
@require_GET
def receipts_data(request):
    from sponsor.models import PaymentReceipt

    status   = request.GET.get('status',   'pending')
    q        = request.GET.get('q',        '').strip()
    currency = request.GET.get('currency', '').strip()
    date_f   = request.GET.get('date',     '').strip()
    page     = int(request.GET.get('page', 1))
    PER_PAGE = 20

    STATUS_MAP = {
        'pending':  'بانتظار المراجعة',
        'approved': 'موافق',
        'rejected': 'مرفوض',
        'all':      None,
    }

    qs = PaymentReceipt.objects.select_related(
        'sponsor__user', 'beneficiary', 'reviewed_by'
    ).order_by('-created_at')

    s = STATUS_MAP.get(status)
    if s: qs = qs.filter(status=s)

    if q:
        qs = qs.filter(
            Q(sponsor__user__first_name__icontains=q) |
            Q(sponsor__user__family_name__icontains=q) |
            Q(unique_number__icontains=q) |
            Q(system_ref__icontains=q) |
            Q(sender_name__icontains=q)
        ).distinct()

    if currency:
        qs = qs.filter(currency=currency)

    from datetime import date, timedelta
    today = date.today()
    if date_f == 'today':  qs = qs.filter(created_at__date=today)
    elif date_f == 'week': qs = qs.filter(created_at__date__gte=today - timedelta(days=7))
    elif date_f == 'month':qs = qs.filter(created_at__date__gte=today - timedelta(days=30))

    total      = qs.count()
    total_pages= max(1, (total + PER_PAGE - 1) // PER_PAGE)
    receipts   = qs[(page-1)*PER_PAGE: page*PER_PAGE]

    # إحصائيات
    all_qs     = PaymentReceipt.objects.all()
    approved_qs= all_qs.filter(status='موافق')
    stats = {
        'pending':      all_qs.filter(status='بانتظار المراجعة').count(),
        'approved':     all_qs.filter(status='موافق').count(),
        'rejected':     all_qs.filter(status='مرفوض').count(),
        'total':        all_qs.count(),
        'total_ils':    str(round(approved_qs.aggregate(s=Sum('amount_shekel'))['s'] or 0, 2)),
        'total_usd':    str(round(approved_qs.aggregate(s=Sum('amount_dollar'))['s'] or 0, 2)),
        'month_ils':    str(round(approved_qs.filter(
            reviewed_at__year=today.year,
            reviewed_at__month=today.month,
        ).aggregate(s=Sum('amount_shekel'))['s'] or 0, 2)),
    }

    return JsonResponse({
        'receipts':   [_build_receipt_data(r) for r in receipts],
        'stats':      stats,
        'pagination': {'page': page, 'total_pages': total_pages, 'total': total},
    })


# ══════════════════════════════════════════════
# تفاصيل وصل واحد
# ══════════════════════════════════════════════
@admin_required
@require_GET
def receipt_detail(request):
    from sponsor.models import PaymentReceipt
    rid = request.GET.get('id','').strip()
    try:
        r = PaymentReceipt.objects.select_related(
            'sponsor__user','beneficiary','reviewed_by'
        ).get(pk=rid)
        return JsonResponse({'status':'success','receipt': _build_receipt_data(r)})
    except PaymentReceipt.DoesNotExist:
        return JsonResponse({'status':'error','message':'غير موجود'})


# ══════════════════════════════════════════════
# الموافقة على الوصل
# ══════════════════════════════════════════════
@admin_required
@require_POST
@csrf_protect
def approve_receipt(request):
    from sponsor.models import PaymentReceipt

    rid          = request.POST.get('receipt_id', '').strip()
    notes        = request.POST.get('notes',      '').strip()
    # حقول قابلة للتعديل عند الموافقة
    amount_edit  = request.POST.get('amount_original', '').strip()
    currency_edit= request.POST.get('currency', '').strip()
    date_edit    = request.POST.get('receipt_date', '').strip()
    bene_edit    = request.POST.get('beneficiary_id', '').strip()

    try:
        receipt = PaymentReceipt.objects.select_related(
            'sponsor__user','beneficiary'
        ).get(pk=rid)

        if receipt.status == 'موافق':
            return JsonResponse({'status':'error','message':'الوصل مقبول مسبقاً'})

        was_edited = False

        # تطبيق التعديلات إن وجدت
        if amount_edit and amount_edit != str(receipt.amount_original):
            receipt.amount_original = Decimal(amount_edit)
            was_edited = True
        if currency_edit and currency_edit != receipt.currency:
            receipt.currency = currency_edit
            was_edited = True
        if date_edit and date_edit != str(receipt.receipt_date):
            from datetime import date as ddate
            receipt.receipt_date = ddate.fromisoformat(date_edit)
            was_edited = True
        if bene_edit and bene_edit != str(receipt.beneficiary_id if receipt.beneficiary else ''):
            try:
                receipt.beneficiary = CustomUser.objects.get(pk=bene_edit)
                was_edited = True
            except CustomUser.DoesNotExist:
                pass

        # تحويل المبلغ
        amount_ils, amount_usd = _convert_to_ils_usd(
            receipt.amount_original, receipt.currency
        )
        receipt.amount_shekel = amount_ils
        receipt.amount_dollar = amount_usd
        receipt.notes         = notes
        receipt.status        = 'موافق'
        receipt.reviewed_by   = request.user
        receipt.reviewed_at   = timezone.now()
        receipt.save()

        # إنشاء Payment تلقائياً
        if receipt.beneficiary:
            try:
                Payment.objects.create(
                    beneficiary = receipt.beneficiary,
                    sponsor     = receipt.sponsor.user if receipt.sponsor else None,
                    amount_ils  = amount_ils,
                    amount_usd  = amount_usd,
                    paid_by     = 'sponsor',
                    status      = 'paid',
                    note        = f'وصل رقم {receipt.system_ref} — {receipt.unique_number}',
                    date        = receipt.receipt_date,
                    created_by  = request.user,
                )
            except Exception as e:
                pass  # لا نوقف العملية بسبب خطأ في الـ Payment

        # بريد الموافقة
        try:
            from django.core.mail import send_mail
            from django.conf import settings as django_settings
            sponsor_user = receipt.sponsor.user if receipt.sponsor else None
            if sponsor_user and sponsor_user.email:
                edit_note = f'\n\nملاحظة: تم إجراء بعض التعديلات على بيانات الوصل.' if was_edited else ''
                send_mail(
                    subject   = 'تم قبول وصلك المالي ✅ — منصة رُحَمَاء',
                    message   = (
                        f'السلام عليكم {_uname(sponsor_user)},\n\n'
                        f'يسعدنا إبلاغك بأنه تم قبول وصلك المالي.\n\n'
                        f'تفاصيل الوصل:\n'
                        f'• الرقم المرجعي: {receipt.system_ref}\n'
                        f'• رقم الوصل: {receipt.unique_number}\n'
                        f'• المبلغ: {receipt.amount_original} {receipt.currency}\n'
                        f'• ما يعادل: {receipt.amount_shekel}₪ / ${receipt.amount_usd}\n'
                        f'• المستفيد: {_uname(receipt.beneficiary)}\n'
                        f'• تاريخ الوصل: {receipt.receipt_date}\n'
                        f'• تاريخ المراجعة: {timezone.localtime(receipt.reviewed_at).strftime("%Y/%m/%d %H:%M")}\n'
                        f'• اسم المرسل: {receipt.sender_name}\n'
                        + (f'• ملاحظات الإدارة: {notes}\n' if notes else '')
                        + edit_note +
                        f'\n\nشكراً لكرمكم وتبرعكم.\n\nفريق منصة رُحَمَاء'
                    ),
                    from_email    = django_settings.DEFAULT_FROM_EMAIL,
                    recipient_list= [sponsor_user.email],
                    fail_silently = True,
                )
        except Exception:
            pass

        # إشعار
        try:
            create_notification(
                recipient = receipt.sponsor.user,
                ntype     = 'RECEIPT_OK',
                title     = 'تم قبول وصلك المالي ✅',
                message   = f'تم قبول الوصل بمبلغ {receipt.amount_shekel}₪',
                sender    = request.user,
            )
        except Exception:
            pass

        log_activity(
            request.user, 'APPROVE',
            description  = f'موافقة على وصل {receipt.system_ref} — {receipt.amount_shekel}₪ — المستفيد: {_uname(receipt.beneficiary)}',
            target_model = 'PaymentReceipt',
            target_id    = str(receipt.pk),
            request      = request,
        )

        return JsonResponse({
            'status':  'success',
            'message': 'تمت الموافقة وإرسال البريد ✅',
            'receipt': _build_receipt_data(receipt),
        })

    except PaymentReceipt.DoesNotExist:
        return JsonResponse({'status':'error','message':'الوصل غير موجود'})
    except Exception as e:
        return JsonResponse({'status':'error','message':str(e)})


# ══════════════════════════════════════════════
# رفض الوصل
# ══════════════════════════════════════════════
@admin_required
@require_POST
@csrf_protect
def reject_receipt(request):
    from sponsor.models import PaymentReceipt

    rid    = request.POST.get('receipt_id',    '').strip()
    reason = request.POST.get('reject_reason', '').strip()

    if not reason:
        return JsonResponse({'status':'error','message':'سبب الرفض مطلوب'})

    try:
        receipt               = PaymentReceipt.objects.select_related('sponsor__user').get(pk=rid)
        receipt.status        = 'مرفوض'
        receipt.reject_reason = reason
        receipt.reviewed_by   = request.user
        receipt.reviewed_at   = timezone.now()
        receipt.save()

        # بريد الرفض
        try:
            from django.core.mail import send_mail
            from django.conf import settings as django_settings
            sponsor_user = receipt.sponsor.user if receipt.sponsor else None
            if sponsor_user and sponsor_user.email:
                send_mail(
                    subject   = 'تم رفض وصلك المالي ❌ — منصة رُحَمَاء',
                    message   = (
                        f'السلام عليكم {_uname(sponsor_user)},\n\n'
                        f'نأسف لإبلاغك بأنه تم رفض وصلك المالي.\n\n'
                        f'تفاصيل الوصل:\n'
                        f'• الرقم المرجعي: {receipt.system_ref}\n'
                        f'• رقم الوصل: {receipt.unique_number}\n'
                        f'• المبلغ: {receipt.amount_original} {receipt.currency}\n'
                        f'• تاريخ الوصل: {receipt.receipt_date}\n\n'
                        f'سبب الرفض: {reason}\n\n'
                        f'يمكنك إعادة إرسال الوصل بعد التصحيح.\n\n'
                        f'فريق منصة رُحَمَاء'
                    ),
                    from_email    = django_settings.DEFAULT_FROM_EMAIL,
                    recipient_list= [sponsor_user.email],
                    fail_silently = True,
                )
        except Exception:
            pass

        try:
            create_notification(
                recipient = receipt.sponsor.user,
                ntype     = 'RECEIPT_REJ',
                title     = 'تم رفض وصلك المالي ❌',
                message   = f'سبب الرفض: {reason}',
                sender    = request.user,
            )
        except Exception:
            pass

        log_activity(
            request.user, 'REJECT',
            description  = f'رفض وصل {receipt.system_ref} — السبب: {reason}',
            target_model = 'PaymentReceipt',
            target_id    = str(receipt.pk),
            request      = request,
        )

        return JsonResponse({'status':'success','message':'تم الرفض وإرسال البريد'})

    except PaymentReceipt.DoesNotExist:
        return JsonResponse({'status':'error','message':'الوصل غير موجود'})


# ══════════════════════════════════════════════
# حذف الوصل
# ══════════════════════════════════════════════
@admin_required
@require_POST
@csrf_protect
def delete_receipt(request):
    from sponsor.models import PaymentReceipt

    rid = request.POST.get('receipt_id','').strip()
    try:
        r = PaymentReceipt.objects.select_related('sponsor__user').get(pk=rid)
        ref = r.system_ref or str(r.pk)
        r.delete()
        log_activity(
            request.user, 'DELETE',
            description  = f'حذف وصل {ref}',
            target_model = 'PaymentReceipt',
            target_id    = rid,
            request      = request,
        )
        return JsonResponse({'status':'success','message':'تم الحذف'})
    except PaymentReceipt.DoesNotExist:
        return JsonResponse({'status':'error','message':'غير موجود'})


# ══════════════════════════════════════════════
# تصدير Excel
# ══════════════════════════════════════════════
@admin_required
@require_GET
def export_receipts(request):
    from sponsor.models import PaymentReceipt

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return HttpResponse('openpyxl غير مثبّتة', status=500)

    status   = request.GET.get('status', 'all')
    STATUS_MAP = {'pending':'بانتظار المراجعة','approved':'موافق','rejected':'مرفوض'}

    qs = PaymentReceipt.objects.select_related(
        'sponsor__user','beneficiary','reviewed_by'
    ).order_by('-created_at')
    s = STATUS_MAP.get(status)
    if s: qs = qs.filter(status=s)

    PURPLE = '7C3AED'
    thin   = Side(style='thin', color='E2E8F0')
    border = openpyxl.styles.Border(left=thin,right=thin,top=thin,bottom=thin)
    C_ALG  = Alignment(horizontal='center',vertical='center')
    R_ALG  = Alignment(horizontal='right', vertical='center')

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = 'الوصولات'
    ws.sheet_view.rightToLeft = True

    headers = ['الرقم المرجعي','رقم الوصل','اسم الكافل','المستفيد','المبلغ الأصلي','العملة','بالشيقل','بالدولار','اسم المرسل','تاريخ الوصل','الحالة','مراجَع من','تاريخ المراجعة','ملاحظات']
    for i,h in enumerate(headers,1):
        c = ws.cell(row=1,column=i,value=h)
        c.font=Font(bold=True,color='FFFFFF')
        c.fill=PatternFill('solid',fgColor=PURPLE)
        c.alignment=C_ALG; c.border=border
    ws.row_dimensions[1].height=26

    for r_idx,r in enumerate(qs,2):
        row=[
            r.system_ref or '', r.unique_number or '',
            _uname(r.sponsor.user) if r.sponsor else '—',
            _uname(r.beneficiary) if r.beneficiary else '—',
            str(r.amount_original or 0), r.currency or '',
            str(r.amount_shekel or 0), str(r.amount_dollar or 0),
            r.sender_name or '', str(r.receipt_date or ''),
            r.status,
            _uname(r.reviewed_by) if r.reviewed_by else '',
            timezone.localtime(r.reviewed_at).strftime('%Y/%m/%d %H:%M') if r.reviewed_at else '',
            r.notes or '',
        ]
        for c_idx,val in enumerate(row,1):
            cell=ws.cell(row=r_idx,column=c_idx,value=val)
            cell.alignment=R_ALG; cell.border=border
            if r_idx%2==0: cell.fill=PatternFill('solid',fgColor='F5F3FF')
        ws.row_dimensions[r_idx].height=20

    for col,w in zip('ABCDEFGHIJKLMN',[16,16,20,20,12,8,12,12,18,14,14,18,18,24]):
        ws.column_dimensions[col].width=w

    output=io.BytesIO()
    wb.save(output); output.seek(0)
    resp=HttpResponse(output.read(),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition']=f'attachment; filename="receipts_{status}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    log_activity(request.user,'EXPORT',description=f'تصدير الوصولات — {status}',request=request)
    return resp


# ══════════════════════════════════════════════
# PDF السند المالي
# ══════════════════════════════════════════════
@admin_required
@require_GET
def receipt_pdf(request):
    import os, datetime
    from sponsor.models import PaymentReceipt

    # ── كتابة فورية في الـ log ──
    def write_log(msg):
        try:
            log_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'logs', 'ruhamaa.log')
            with open(log_path, 'a', encoding='utf-8') as lf:
                lf.write(f"[{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}\n")
        except Exception as ex:
            pass

    write_log(f"receipt_pdf CALLED — id={request.GET.get('id')} — user={request.user}")

    rid = request.GET.get('id', '').strip()

    if not rid:
        write_log("receipt_pdf ERROR — no id provided")
        return HttpResponse('معرف الوصل مطلوب', status=400)

    try:
        r = PaymentReceipt.objects.select_related(
            'sponsor__user', 'beneficiary', 'reviewed_by'
        ).get(pk=rid)

        write_log(f"receipt_pdf FOUND — ref={r.system_ref} | shekel={r.amount_shekel} | dollar={r.amount_dollar}")

        # ── جلب معلومات الموقع ──
        try:
            from core.models import SystemSettings
            site_name  = SystemSettings.get('site_name',  'منصة رُحَمَاء')
            site_phone = SystemSettings.get('site_phone', '')
            site_email = SystemSettings.get('site_email', '')
        except Exception as e:
            write_log(f"receipt_pdf SystemSettings ERROR — {e}")
            site_name  = 'منصة رُحَمَاء'
            site_phone = ''
            site_email = ''

        # ── المبالغ مباشرة من قاعدة البيانات ──
        amount_ils = r.amount_shekel if r.amount_shekel is not None else 0
        amount_usd = r.amount_dollar if r.amount_dollar is not None else 0

        # ── تاريخ الطباعة ──
        print_date = timezone.localtime(timezone.now()).strftime('%Y/%m/%d %H:%M')

        write_log(f"receipt_pdf CONTEXT — print_date={print_date} | amount_ils={amount_ils} | amount_usd={amount_usd}")

        context = {
            'receipt':    r,
            'site_name':  site_name,
            'site_phone': site_phone,
            'site_email': site_email,
            'print_date': print_date,
            'amount_ils': amount_ils,
            'amount_usd': amount_usd,
        }

        return render(request, 'admin_panel/receipt_pdf.html', context)

    except PaymentReceipt.DoesNotExist:
        write_log(f"receipt_pdf NOT FOUND — id={rid}")
        return HttpResponse('الوصل غير موجود', status=404)

    except Exception as e:
        write_log(f"receipt_pdf EXCEPTION — {e}")
        return HttpResponse(f'خطأ: {e}', status=500)


# ══════════════════════════════════════════════
# قائمة المستفيدين للكافل (للمودل)
# ══════════════════════════════════════════════
@admin_required
@require_GET
def sponsor_beneficiaries(request):
    sponsor_id = request.GET.get('sponsor_id','').strip()
    if not sponsor_id:
        return JsonResponse({'beneficiaries':[]})
    try:
        from beneficiary.models import OrphanForm, FamilyForm, SpecialNeedsForm
        from sponsor.models import SponsorProfile
        sp = SponsorProfile.objects.get(user__pk=sponsor_id)
        benes = []
        TYPE_LABELS = {'orphan':'يتيم','family':'أسرة','special':'ذوو احتياجات'}
        for Model in [OrphanForm, FamilyForm, SpecialNeedsForm]:
            for f in Model.objects.filter(sponsor=sp).select_related('user'):
                benes.append({
                    'id':    str(f.user.pk),
                    'name':  f.user.get_full_name(),
                    'type':  TYPE_LABELS.get(f.user.user_type,''),
                    'reg':   f.user.registration_number or '',
                })
        return JsonResponse({'beneficiaries': benes})
    except Exception:
        return JsonResponse({'beneficiaries':[]})
@admin_required
@require_GET
def receipt_pdf_download(request):
    import weasyprint, logging
    from urllib.parse import quote
    from django.template.loader import render_to_string
    from sponsor.models import PaymentReceipt

    logging.getLogger('weasyprint').setLevel(logging.ERROR)
    logging.getLogger('fontTools').setLevel(logging.ERROR)

    rid = request.GET.get('id', '').strip()
    try:
        r = PaymentReceipt.objects.select_related(
            'sponsor__user', 'beneficiary', 'reviewed_by'
        ).get(pk=rid, status='موافق')

        try:
            from core.models import SystemSettings
            site_name  = SystemSettings.get('site_name',  'منصة رُحَمَاء')
            site_phone = SystemSettings.get('site_phone', '')
            site_email = SystemSettings.get('site_email', '')
        except Exception:
            site_name  = 'منصة رُحَمَاء'
            site_phone = ''
            site_email = ''

        amount_ils = r.amount_shekel if r.amount_shekel is not None else 0
        amount_usd = r.amount_dollar if r.amount_dollar is not None else 0

        html_content = render_to_string('admin_panel/receipt_pdf.html', {
            'receipt':    r,
            'site_name':  site_name,
            'site_phone': site_phone,
            'site_email': site_email,
            'print_date': timezone.localtime(timezone.now()).strftime('%Y/%m/%d %H:%M'),
            'amount_ils': amount_ils,
            'amount_usd': amount_usd,
        }, request=request)

        pdf_file = weasyprint.HTML(
            string=html_content,
            base_url=request.build_absolute_uri('/')
        ).write_pdf()

        fname = f'وصل_{r.system_ref}.pdf'
        resp  = HttpResponse(pdf_file, content_type='application/pdf')
        resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(fname)}"
        return resp

    except PaymentReceipt.DoesNotExist:
        return HttpResponse('الوصل غير موجود', status=404)
    except Exception as e:
        return HttpResponse(f'خطأ: {e}', status=500)