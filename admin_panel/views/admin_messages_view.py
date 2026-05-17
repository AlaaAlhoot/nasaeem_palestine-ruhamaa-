
import os
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST, require_GET
from django.views.decorators.csrf import csrf_protect
from django.db.models import Q, Max, Count, OuterRef, Subquery
from django.utils import timezone
from urllib.parse import quote

from core.models import CustomUser, DirectMessage, Notification
from core.utils import log_activity, create_notification, fmt_dt
from .decorators import admin_required
from django.db.models import Max

# ══════════════════════════════════════════════════════════
# الصفحة الرئيسية
# ══════════════════════════════════════════════════════════

@admin_required
def messages_view(request):
    admin       = request.user
    notif_count = Notification.objects.filter(recipient=admin, is_read=False).count()

    # محادثات الأدمن
    users_with_msgs = CustomUser.objects.filter(
        Q(sent_direct_messages__recipient__user_type='admin') |
        Q(received_direct_messages__sender__user_type='admin')
    ).exclude(user_type='admin').distinct().order_by('first_name')

    conversations = []
    for u in users_with_msgs:
        last_msg = DirectMessage.objects.filter(
            Q(sender=u, recipient=admin) | Q(sender=admin, recipient=u)
        ).order_by('-created_at').first()

        unread = DirectMessage.objects.filter(
            sender=u, recipient=admin, is_read=False
        ).count()

        conversations.append({
            'user':      u,
            'last_msg':  last_msg,
            'unread':    unread,
            'last_time': fmt_dt(last_msg.created_at) if last_msg else '—',
        })

    conversations.sort(
        key=lambda x: x['last_msg'].created_at if x['last_msg'] else timezone.now().replace(year=2000),
        reverse=True
    )

    # كل المستخدمين للمراقبة
    monitor_users = CustomUser.objects.exclude(
        user_type='admin'
    ).filter(is_active=True).order_by('first_name')

    # آخر ID للرسائل الواردة — لمنع الـ badge من إحصاء القديمة
    from django.db.models import Max
    last_msg_id = DirectMessage.objects.filter(
        recipient=admin
    ).aggregate(m=Max('id'))['m'] or 0

    return render(request, 'admin_panel/messages.html', {
        'notif_count':   notif_count,
        'conversations': conversations,
        'monitor_users': monitor_users,
        'last_msg_id':   last_msg_id,
    })


@admin_required
@require_GET
def get_user_conversations(request):
    user_id = request.GET.get('user_id', '').strip()
    if not user_id:
        return JsonResponse({'pairs': []})

    # كل الأشخاص اللي راسلهم أو راسلوه
    sent_to = DirectMessage.objects.filter(
        sender_id=user_id
    ).values_list('recipient_id', flat=True).distinct()

    received_from = DirectMessage.objects.filter(
        recipient_id=user_id
    ).values_list('sender_id', flat=True).distinct()

    partner_ids = set(list(sent_to) + list(received_from))
    partner_ids.discard(user_id)  # أزل نفسه

    pairs = []
    for pid in partner_ids:
        try:
            partner = CustomUser.objects.get(pk=pid)
            pairs.append({
                'id':   str(partner.pk),
                'name': partner.get_full_name(),
                'type': partner.user_type,
            })
        except CustomUser.DoesNotExist:
            pass

    return JsonResponse({'pairs': pairs})
# ══════════════════════════════════════════════════════════
# جلب رسائل محادثة معينة
# ══════════════════════════════════════════════════════════

@admin_required
@require_GET

def get_conversation(request):
    admin   = request.user
    user_id = request.GET.get('user_id', '').strip()
    tab     = request.GET.get('tab', 'mine')

    if tab == 'monitor':
        user1_id = request.GET.get('user1_id', '').strip()
        user2_id = request.GET.get('user2_id', '').strip()

        # لو جاء user_id بدون طرف ثانٍ
        if not user1_id and user_id:
            user1_id = user_id

        if not user1_id:
            return JsonResponse({'messages': []})

        if user2_id:
            msgs = DirectMessage.objects.filter(
                Q(sender_id=user1_id, recipient_id=user2_id) |
                Q(sender_id=user2_id, recipient_id=user1_id)
            ).select_related('sender', 'recipient').order_by('created_at')
        else:
            # كل محادثات هذا المستخدم مع أي شخص
            msgs = DirectMessage.objects.filter(
                Q(sender_id=user1_id) | Q(recipient_id=user1_id)
            ).select_related('sender', 'recipient').order_by('created_at')

    else:
        if not user_id:
            return JsonResponse({'messages': []})
        msgs = DirectMessage.objects.filter(
            Q(sender_id=user_id, recipient=admin) |
            Q(sender=admin, recipient_id=user_id)
        ).select_related('sender', 'recipient').order_by('created_at')

        DirectMessage.objects.filter(
            sender_id=user_id, recipient=admin, is_read=False
        ).update(is_read=True, is_delivered=True)

    def att_info(msg):
        if not msg.attachment:
            return None, None
        ext = os.path.splitext(msg.attachment.name)[1].lower()
        t   = 'image' if ext in ['.jpg','.jpeg','.png','.gif','.webp','.bmp'] else 'file'
        return msg.attachment.url, t

    result = []
    for m in msgs:
        url, att_type = att_info(m)
        result.append({
            'id':           m.pk,
            'text':         m.message,
            'sender_id':    m.sender.pk,
            'sender_name':  m.sender.get_full_name(),
            'sender_type':  m.sender.user_type,
            'is_mine':      m.sender == admin,
            'is_read':      m.is_read,
            'is_delivered': m.is_delivered,
            'time':         m.created_at.strftime('%H:%M'),
            'date':         m.created_at.strftime('%Y/%m/%d'),
            'attachment':   url,
            'att_type':     att_type,
        })

    return JsonResponse({'messages': result})

# ══════════════════════════════════════════════════════════
# إرسال رسالة
# ══════════════════════════════════════════════════════════

@admin_required
@require_POST
@csrf_protect
def send_message(request):
    admin        = request.user
    recipient_id = request.POST.get('recipient_id', '').strip()
    text         = request.POST.get('message', '').strip()
    attachment   = request.FILES.get('attachment')

    if not text and not attachment:
        return JsonResponse({'status': 'error', 'message': 'الرسالة فارغة'})
    if attachment and attachment.size > 3 * 1024 * 1024:
        return JsonResponse({'status': 'error', 'message': 'حجم الملف يتجاوز 3MB'})

    try:
        recipient = CustomUser.objects.get(pk=recipient_id)
    except CustomUser.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'المستلم غير موجود'})

    msg = DirectMessage(sender=admin, recipient=recipient, message=text, is_delivered=True)
    if attachment:
        msg.attachment = attachment
    msg.save()

    create_notification(
        recipient  = recipient,
        ntype      = 'NEW_MSG',
        title      = 'رسالة من الإدارة 💬',
        message    = f'رسالة من {admin.get_full_name()}',
        sender     = admin,
        action_url = '/sponsor/messages/',
    )

    log_activity(admin, 'MESSAGE',
                 description=f'رسالة إلى {recipient.get_full_name()}',
                 request=request)

    att_url  = msg.attachment.url if msg.attachment else None
    att_type = None
    if msg.attachment:
        ext      = os.path.splitext(msg.attachment.name)[1].lower()
        att_type = 'image' if ext in ['.jpg','.jpeg','.png','.gif','.webp'] else 'file'

    return JsonResponse({
        'status':     'success',
        'id':         msg.pk,
        'time':       msg.created_at.strftime('%H:%M'),
        'attachment': att_url,
        'att_type':   att_type,
    })


# ══════════════════════════════════════════════════════════
# قائمة المستخدمين للمراقبة
# ══════════════════════════════════════════════════════════

@admin_required
@require_GET
def monitor_users(request):
    q = request.GET.get('q', '').strip()
    users = CustomUser.objects.exclude(user_type='admin').filter(is_active=True)
    if q:
        users = users.filter(
            Q(first_name__icontains=q) | Q(family_name__icontains=q) |
            Q(username__icontains=q)
        )
    result = [{
        'id':        u.pk,
        'name':      u.get_full_name(),
        'username':  u.username,
        'type':      u.user_type,
    } for u in users[:50]]
    return JsonResponse({'users': result})


# ══════════════════════════════════════════════════════════
# تصدير محادثة كـ Excel
# ══════════════════════════════════════════════════════════

@admin_required
@require_GET
def export_conversation(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    admin    = request.user
    user_id  = request.GET.get('user_id', '').strip()
    user1_id = request.GET.get('user1_id', '').strip()
    user2_id = request.GET.get('user2_id', '').strip()

    if user1_id and user2_id:
        msgs = DirectMessage.objects.filter(
            Q(sender_id=user1_id, recipient_id=user2_id) |
            Q(sender_id=user2_id, recipient_id=user1_id)
        ).select_related('sender','recipient').order_by('created_at')
        title = 'مراقبة محادثة'
    elif user_id:
        msgs = DirectMessage.objects.filter(
            Q(sender_id=user_id, recipient=admin) |
            Q(sender=admin, recipient_id=user_id)
        ).select_related('sender','recipient').order_by('created_at')
        try:
            u     = CustomUser.objects.get(pk=user_id)
            title = f'محادثة مع {u.get_full_name()}'
        except Exception:
            title = 'محادثة'
    else:
        return HttpResponse('معرف المستخدم مطلوب', status=400)

    log_activity(admin, 'EXPORT', description=f'تصدير محادثة Excel', request=request)

    GREEN='1A7A4A'; BLUE='185FA5'; WHITE='FFFFFF'; LIGHT='E8F5E9'
    thin  = Side(style='thin', color='CCCCCC')
    brd   = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = 'المحادثة'
    ws.sheet_view.rightToLeft = True

    ws.merge_cells('A1:F1')
    c = ws['A1']
    c.value     = title
    c.font      = Font(bold=True, size=12, color=WHITE)
    c.fill      = PatternFill('solid', fgColor=GREEN)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    hdrs   = ['#','المرسل','نوع المستخدم','الرسالة','مرفق','التاريخ والوقت']
    widths = [6, 22, 14, 50, 20, 20]
    for i,(h,w) in enumerate(zip(hdrs,widths),1):
        cell = ws.cell(2,i,h)
        cell.font      = Font(bold=True, color=WHITE, size=10)
        cell.fill      = PatternFill('solid', fgColor=BLUE)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = brd
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 22

    TYPE_AR = {'admin':'أدمن','sponsor':'كافل','orphan':'يتيم','family':'أسرة','special':'ذو احتياج'}
    for ri,m in enumerate(msgs, start=3):
        row = [
            ri-2,
            m.sender.get_full_name(),
            TYPE_AR.get(m.sender.user_type, m.sender.user_type),
            m.message or '',
            m.attachment.url if m.attachment else '—',
            m.created_at.strftime('%Y/%m/%d %H:%M'),
        ]
        for ci,val in enumerate(row,1):
            cell = ws.cell(ri,ci,val)
            cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
            cell.border    = brd
            if ri%2==0: cell.fill = PatternFill('solid', fgColor=LIGHT)
        ws.row_dimensions[ri].height = 20

    output = io.BytesIO()
    wb.save(output); output.seek(0)
    fname = f'محادثة_{title}.xlsx'
    resp  = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(fname)}"
    return resp


# ══════════════════════════════════════════════════════════
# Polling
# ══════════════════════════════════════════════════════════

@admin_required
@require_GET
def poll_messages(request):
    admin   = request.user
    last_id = int(request.GET.get('last_id', 0))
    user_id = request.GET.get('user_id', '').strip()

    qs = DirectMessage.objects.filter(
        id__gt=last_id
    ).select_related('sender')

    if user_id:
        # فقط رسائل المحادثة الحالية
        from django.db.models import Q
        qs = qs.filter(
            Q(sender_id=user_id, recipient=admin) |
            Q(sender=admin, recipient_id=user_id)
        )
    else:
        qs = qs.filter(recipient=admin)

    result = [{
        'id':          m.pk,
        'text':        m.message,
        'sender_id':   m.sender.pk,
        'sender_name': m.sender.get_full_name(),
        'is_mine':     m.sender == admin,
        'time':        m.created_at.strftime('%H:%M'),
        'is_read':     m.is_read,
        'is_delivered':m.is_delivered,
        'attachment':  m.attachment.url if m.attachment else None,
        'att_type':    None,
    } for m in qs.order_by('id')]

    qs.filter(is_delivered=False).update(is_delivered=True)
    return JsonResponse({'messages': result})