"""
beneficiary/views/aids.py
"""
import hashlib
import ctypes
from urllib.parse import quote
from decimal import Decimal
from io import BytesIO
from datetime import date

from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_GET
from django.db.models import Sum, Count, Q
from django.utils import timezone
from django.template.loader import render_to_string

from core.models import Aid  # ← مسار الموديل

# ── اضبط هذا حسب مشروعك ──
def _bene_required(request):
    return request.user.is_authenticated

def log_activity(user, action, description='', request=None):
    """استبدل بدالة الـ log الحقيقية في مشروعك"""
    pass

# ────────────────────────────
AID_TYPE_LABELS = dict(Aid.AID_TYPES)

# ════════════════════════════════════════════
#  مساعدة: hash لكلمة السر (مطابقة JS)
# ════════════════════════════════════════════
def _js_hash(s: str) -> str:
    h = ctypes.c_int32(0)
    for ch in s:
        h = ctypes.c_int32(31 * h.value + ord(ch))
    return format(ctypes.c_uint32(h.value).value, 'x')


# ════════════════════════════════════════════
#  فلترة الـ QuerySet
# ════════════════════════════════════════════
def _filter_qs(qs, params):
    q         = params.get('q', '').strip()
    aid_type  = params.get('aid_type', '').strip()
    provider  = params.get('provider', '').strip()
    date_from = params.get('date_from', '').strip()
    date_to   = params.get('date_to', '').strip()

    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(provider__icontains=q) | Q(note__icontains=q))
    if aid_type:
        qs = qs.filter(aid_type=aid_type)
    if provider:
        qs = qs.filter(provider__icontains=provider)
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)
    return qs


# ════════════════════════════════════════════
#  تحضير قائمة المساعدات للـ context
# ════════════════════════════════════════════
def _build_aids_list(qs):
    result = []
    for a in qs:
        result.append({
            'id':             a.pk,
            'name':           a.name,
            'aid_type':       a.aid_type,
            'aid_type_label': AID_TYPE_LABELS.get(a.aid_type, a.aid_type),
            'quantity':       a.quantity,
            'provider':       a.provider,
            'date':           a.date.strftime('%Y-%m-%d') if a.date else '—',
            'note':           a.note or '',
            'created_by':     a.created_by.get_full_name() if a.created_by else '',
            'created_at':     a.created_at.isoformat() if a.created_at else '',
        })
    return result


# ════════════════════════════════════════════
#  1. صفحة المساعدات
# ════════════════════════════════════════════
@login_required(login_url='/login/')
def aids_view(request):
    if not _bene_required(request):
        return redirect('/')
    return render(request, 'beneficiary/aid.html')


# ════════════════════════════════════════════
#  2. API: بيانات المساعدات
# ════════════════════════════════════════════
@login_required(login_url='/login/')
@require_GET
def aids_data_api(request):
    if not _bene_required(request):
        return JsonResponse({'status': 'error', 'message': 'غير مصرح'}, status=403)

    user = request.user
    qs   = Aid.objects.filter(beneficiary=user)
    qs   = _filter_qs(qs, request.GET).order_by('-date')

    # إحصائيات الأنواع
    type_counts = {t: 0 for t, _ in Aid.AID_TYPES}
    for row in qs.values('aid_type').annotate(c=Count('id')):
        type_counts[row['aid_type']] = row['c']

    # إجمالي الكميات
    total_qty = qs.aggregate(s=Sum('quantity'))['s'] or 0

    # هذا الشهر
    now = timezone.now()
    month_count = qs.filter(date__year=now.year, date__month=now.month).count()

    # قائمة الجهات المقدمة (للفلتر)
    providers = list(qs.values_list('provider', flat=True).distinct().order_by('provider'))

    return JsonResponse({
        'status': 'success',
        'data': {
            'aids':        _build_aids_list(qs),
            'count':       qs.count(),
            'total_quantity': total_qty,
            'month_count': month_count,
            'type_counts': type_counts,
            'type_choices': [{'value': v, 'label': l} for v, l in Aid.AID_TYPES],
            'providers':   providers,
        }
    })


# ════════════════════════════════════════════
#  3. API: تفاصيل مساعدة واحدة
# ════════════════════════════════════════════
@login_required(login_url='/login/')
@require_GET
def aid_detail_api(request):
    if not _bene_required(request):
        return JsonResponse({'status': 'error', 'message': 'غير مصرح'}, status=403)

    try:
        aid = Aid.objects.get(pk=request.GET.get('id'), beneficiary=request.user)
    except Aid.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'المساعدة غير موجودة'}, status=404)

    return JsonResponse({
        'status': 'success',
        'data': {
            'id':             aid.pk,
            'name':           aid.name,
            'aid_type':       aid.aid_type,
            'aid_type_label': AID_TYPE_LABELS.get(aid.aid_type, aid.aid_type),
            'quantity':       aid.quantity,
            'provider':       aid.provider,
            'date':           aid.date.strftime('%Y-%m-%d') if aid.date else '—',
            'note':           aid.note or '',
            'created_by':     aid.created_by.get_full_name() if aid.created_by else '',
            'created_at':     aid.created_at.isoformat() if aid.created_at else '',
        }
    })


# ════════════════════════════════════════════
#  4. تصدير Excel
# ════════════════════════════════════════════
@login_required(login_url='/login/')
@require_GET
def export_excel_api(request):
    if not _bene_required(request):
        return JsonResponse({'status': 'error', 'message': 'غير مصرح'}, status=403)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return JsonResponse({'status': 'error', 'message': 'openpyxl غير مثبت'}, status=500)

    user = request.user
    qs   = Aid.objects.filter(beneficiary=user)
    qs   = _filter_qs(qs, request.GET).order_by('-date')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'المساعدات'
    ws.sheet_view.rightToLeft = True

    # ألوان
    GREEN  = '1A7A4A'
    LGREEN = 'F0FDF4'
    WHITE  = 'FFFFFF'
    GRAY   = 'F9FAFB'

    # العنوان الرئيسي
    ws.merge_cells('A1:H1')
    c = ws['A1']
    c.value = f'كشف المساعدات — {user.get_full_name()}'
    c.font  = Font(bold=True, size=14, color=WHITE)
    c.fill  = PatternFill('solid', fgColor=GREEN)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32

    # معلومات
    ws.merge_cells('A2:H2')
    c2 = ws['A2']
    c2.value = f'تاريخ التقرير: {timezone.now().strftime("%Y-%m-%d %H:%M")}  |  الهوية: {user.id_number or "—"}  |  العدد: {qs.count()}'
    c2.font  = Font(size=9, color=GREEN)
    c2.fill  = PatternFill('solid', fgColor=LGREEN.replace('#',''))
    c2.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 18

    # رؤوس الأعمدة
    HEADERS = ['#', 'اسم المساعدة', 'النوع', 'الكمية', 'الجهة المقدمة', 'التاريخ', 'أضافها', 'ملاحظة']
    WIDTHS  = [5, 28, 14, 10, 24, 14, 18, 32]
    thin = Side(style='thin', color='BBBBBB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, (hdr, w) in enumerate(zip(HEADERS, WIDTHS), 1):
        cell = ws.cell(row=3, column=col, value=hdr)
        cell.font      = Font(bold=True, color=WHITE, size=10)
        cell.fill      = PatternFill('solid', fgColor=GREEN)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 22

    # البيانات
    for idx, a in enumerate(qs, 1):
        row = idx + 3
        row_data = [
            idx,
            a.name,
            AID_TYPE_LABELS.get(a.aid_type, a.aid_type),
            a.quantity,
            a.provider,
            a.date.strftime('%Y-%m-%d') if a.date else '—',
            a.created_by.get_full_name() if a.created_by else '—',
            a.note or '—',
        ]
        fill = PatternFill('solid', fgColor='F0FDF4' if idx % 2 == 0 else WHITE)
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill      = fill
            cell.border    = border
            cell.alignment = Alignment(horizontal='center' if col in (1,4) else 'right', vertical='center')
            cell.font      = Font(size=9)
        ws.row_dimensions[row].height = 18

    # الإجمالي
    total_row = qs.count() + 4
    ws.merge_cells(f'A{total_row}:C{total_row}')
    ws[f'A{total_row}'].value = f'الإجمالي ({qs.count()} مساعدة)'
    ws[f'A{total_row}'].font  = Font(bold=True, color=WHITE, size=10)
    ws[f'A{total_row}'].fill  = PatternFill('solid', fgColor=GREEN)
    ws[f'A{total_row}'].alignment = Alignment(horizontal='center')
    ws[f'D{total_row}'].value = qs.aggregate(s=Sum('quantity'))['s'] or 0
    ws[f'D{total_row}'].font  = Font(bold=True, color=WHITE)
    ws[f'D{total_row}'].fill  = PatternFill('solid', fgColor=GREEN)
    ws[f'D{total_row}'].alignment = Alignment(horizontal='center')

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'aids_{user.id_number or user.pk}_{timezone.now().strftime("%Y%m%d")}.xlsx'
    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"

    log_activity(user, 'EXPORT', description='تصدير المساعدات Excel', request=request)
    return response


# ════════════════════════════════════════════
#  مساعدة مشتركة: بيانات context للـ PDF
# ════════════════════════════════════════════
def _pdf_context(user, qs, request):
    aids_list = _build_aids_list(qs)
    total_qty = qs.aggregate(s=Sum('quantity'))['s'] or 0

    # إحصائيات الأنواع للـ PDF
    type_stats = {}
    for row in qs.values('aid_type').annotate(c=Count('id')):
        type_stats[row['aid_type']] = row['c']

    serial = hashlib.sha256(
        f'{user.pk}-{timezone.now().isoformat()}-{user.id_number or ""}'.encode()
    ).hexdigest()[:16].upper()

    raw_pwd  = (user.id_number or '0000')[-4:]
    pwd_hash = _js_hash(raw_pwd)

    return {
        'user':        user,
        'aids':        aids_list,
        'stats': {
            'count':       qs.count(),
            'total_qty':   total_qty,
            'type_stats':  type_stats,
            'type_labels': AID_TYPE_LABELS,
        },
        'serial':      serial,
        'report_date': timezone.now().strftime('%Y-%m-%d %H:%M'),
        'pwd_hash':    pwd_hash,
        'filter_from': request.GET.get('date_from', ''),
        'filter_to':   request.GET.get('date_to', ''),
    }


# ════════════════════════════════════════════
#  5. طباعة HTML (متصفح)
# ════════════════════════════════════════════
@login_required(login_url='/login/')
@require_GET
def export_pdf_html_api(request):
    if not _bene_required(request):
        return redirect('/')

    user = request.user
    qs   = _filter_qs(Aid.objects.filter(beneficiary=user), request.GET).order_by('-date')
    ctx  = _pdf_context(user, qs, request)
    ctx['is_download'] = False

    log_activity(user, 'EXPORT', description=f'فتح كشف المساعدات للطباعة (Serial: {ctx["serial"]})', request=request)
    return render(request, 'beneficiary/aid_pdf.html', ctx)


# ════════════════════════════════════════════
#  6. تحميل PDF (WeasyPrint)
# ════════════════════════════════════════════
@login_required(login_url='/login/')
@require_GET
def download_pdf_api(request):
    if not _bene_required(request):
        return redirect('/')

    try:
        from weasyprint import HTML as WP
    except ImportError:
        return JsonResponse({'status': 'error', 'message': 'WeasyPrint غير مثبت'}, status=500)

    user = request.user
    qs   = _filter_qs(Aid.objects.filter(beneficiary=user), request.GET).order_by('-date')
    ctx  = _pdf_context(user, qs, request)
    ctx['is_download'] = True

    html_str = render_to_string('beneficiary/aid_pdf.html', ctx, request=request)
    # ضمان ظهور المحتوى في WeasyPrint
    html_str = html_str.replace(
        'id="protected-content" style="display:none"',
        'id="protected-content" style="display:block"'
    )

    pdf_bytes = WP(string=html_str, base_url=request.build_absolute_uri('/')).write_pdf()

    # حماية بكلمة السر
    try:
        from PyPDF2 import PdfReader, PdfWriter
        reader = PdfReader(BytesIO(pdf_bytes))
        writer = PdfWriter()
        for page in reader.pages:
            writer.add_page(page)
        raw_pwd = (user.id_number or '0000')[-4:]
        writer.encrypt(user_password=raw_pwd, owner_password=f'AID-{ctx["serial"]}', use_128bit=True)
        out = BytesIO()
        writer.write(out)
        pdf_bytes = out.getvalue()
    except ImportError:
        raw_pwd = (user.id_number or '0000')[-4:]

    filename = f'aids_{user.id_number or user.pk}_{timezone.now().strftime("%Y%m%d")}.pdf'
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"
    response['X-PDF-Password'] = raw_pwd

    log_activity(user, 'EXPORT', description=f'تحميل كشف المساعدات PDF (Serial: {ctx["serial"]})', request=request)
    return response
