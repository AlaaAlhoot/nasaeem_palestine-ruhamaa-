"""
sponsor/views/reports_view.py — صفحة التقارير الشاملة
"""
from functools import wraps
from datetime import date, timedelta

from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_GET
from django.db.models import Sum, Count, Avg, Q
from django.utils import timezone

from core.models import CustomUser, Notification, Payment
from sponsor.models import SponsorProfile, PaymentReceipt
from core.utils import log_activity, get_exchange_rates, fmt_dt


# ══ Decorator ══
def sponsor_required(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('/login/')
        if request.user.user_type != 'sponsor':
            return redirect('/login/')
        if not request.user.is_approved:
            return render(request, 'sponsor/pending.html')
        return view_func(request, *args, **kwargs)
    return wrapper


def _get_profile(user):
    try:
        return user.sponsor_profile
    except Exception:
        profile, _ = SponsorProfile.objects.get_or_create(user=user)
        return profile


def _date_range(period):
    today = date.today()
    if period == 'today':      return today, today
    elif period == 'week':     return today - timedelta(days=7), today
    elif period == 'month':    return today - timedelta(days=30), today
    elif period == 'six_months': return today - timedelta(days=180), today
    elif period == 'year':     return today - timedelta(days=365), today
    return None, None


def _period_label(period):
    return {
        'today': 'اليوم', 'week': 'آخر أسبوع', 'month': 'آخر شهر',
        'six_months': 'آخر 6 أشهر', 'year': 'آخر سنة', 'all': 'كل الوقت',
    }.get(period, 'كل الوقت')


# ══════════════════════════════════════════════════════════
# إحصائيات شاملة
# ══════════════════════════════════════════════════════════

def _get_stats(user, profile, period='all'):
    from beneficiary.models import OrphanForm, SpecialNeedsForm, FamilyForm
    d_from, d_to = _date_range(period)

    orphans  = OrphanForm.objects.filter(sponsor=profile).count()
    specials = SpecialNeedsForm.objects.filter(sponsor=profile).count()
    families = FamilyForm.objects.filter(sponsor=profile).count()
    total    = orphans + specials + families

    # الدفعات
    pays_qs = Payment.objects.filter(sponsor=user)
    if d_from:
        pays_qs = pays_qs.filter(date__range=(d_from, d_to))

    paid_qs    = pays_qs.filter(status='paid')
    late_qs    = pays_qs.filter(status='late')
    pending_qs = pays_qs.filter(status='pending')

    total_ils  = paid_qs.aggregate(t=Sum('amount_ils'))['t'] or 0
    total_usd  = paid_qs.aggregate(t=Sum('amount_usd'))['t'] or 0
    paid_count = paid_qs.count()
    late_count = late_qs.count()
    pend_count = pending_qs.count()

    total_expected = paid_count + late_count
    commitment_pct = round((paid_count / total_expected * 100) if total_expected > 0 else 100)

    # متوسط الدفعة
    avg_ils = paid_qs.aggregate(a=Avg('amount_ils'))['a'] or 0

    # آخر دفعة
    last_pay = Payment.objects.filter(sponsor=user, status='paid').order_by('-date').first()
    last_pay_date = str(last_pay.date) if last_pay else '—'
    last_pay_ils  = float(last_pay.amount_ils) if last_pay else 0

    # الوصولات
    receipts_qs = PaymentReceipt.objects.filter(sponsor=profile)
    if d_from:
        receipts_qs = receipts_qs.filter(submitted_at__date__range=(d_from, d_to))

    rec_approved = receipts_qs.filter(status='موافق').count()
    rec_rejected = receipts_qs.filter(status='مرفوض').count()
    rec_pending  = receipts_qs.filter(status='بانتظار المراجعة').count()
    rec_total    = receipts_qs.count()

    rec_ils = receipts_qs.filter(status='موافق').aggregate(t=Sum('amount_shekel'))['t'] or 0
    rec_usd = receipts_qs.filter(status='موافق').aggregate(t=Sum('amount_dollar'))['t'] or 0

    return {
        'total':          total,
        'orphans':        orphans,
        'specials':       specials,
        'families':       families,
        'total_ils':      round(float(total_ils), 2),
        'total_usd':      round(float(total_usd), 2),
        'paid_count':     paid_count,
        'late_count':     late_count,
        'pend_count':     pend_count,
        'commitment_pct': commitment_pct,
        'avg_ils':        round(float(avg_ils), 2),
        'last_pay_date':  last_pay_date,
        'last_pay_ils':   last_pay_ils,
        'rec_total':      rec_total,
        'rec_approved':   rec_approved,
        'rec_rejected':   rec_rejected,
        'rec_pending':    rec_pending,
        'rec_ils':        round(float(rec_ils), 2),
        'rec_usd':        round(float(rec_usd), 2),
    }


# ══════════════════════════════════════════════════════════
# بيانات المخططات
# ══════════════════════════════════════════════════════════

def _get_chart_data(user, profile, period='all'):
    from beneficiary.models import OrphanForm, SpecialNeedsForm, FamilyForm
    from django.db.models.functions import TruncMonth
    from django.db import connection
    d_from, d_to = _date_range(period)

    # مخطط الدفعات الشهرية — بدون TruncMonth
    pays_qs = Payment.objects.filter(sponsor=user, status='paid')
    if d_from:
        pays_qs = pays_qs.filter(date__range=(d_from, d_to))

    # تجميع يدوي حسب السنة والشهر
    from collections import defaultdict
    monthly_dict = defaultdict(lambda: {'total_ils': 0, 'count': 0})
    for p in pays_qs.values('date', 'amount_ils'):
        key = p['date'].strftime('%Y/%m')
        monthly_dict[key]['total_ils'] += float(p['amount_ils'] or 0)
        monthly_dict[key]['count']     += 1

    months_data = [
        {'month': k, 'total_ils': round(v['total_ils'], 2), 'count': v['count']}
        for k, v in sorted(monthly_dict.items())
    ]

    # مخطط مقارنة الدفع لكل مستفيد
    bene_pays = []
    for Model, label in [(OrphanForm,'يتيم'),(SpecialNeedsForm,'ذو احتياج'),(FamilyForm,'أسرة')]:
        for f in Model.objects.filter(sponsor=profile).select_related('user'):
            q = Payment.objects.filter(sponsor=user, beneficiary=f.user, status='paid')
            if d_from:
                q = q.filter(date__range=(d_from, d_to))
            total = q.aggregate(t=Sum('amount_ils'))['t'] or 0
            bene_pays.append({
                'name':  f.get_full_name(),
                'total': round(float(total), 2),
                'type':  label,
            })

    # مخطط الوصولات الشهرية — تجميع يدوي
    rec_qs = PaymentReceipt.objects.filter(sponsor=profile)
    if d_from:
        rec_qs = rec_qs.filter(submitted_at__date__range=(d_from, d_to))

    rec_dict = defaultdict(lambda: {'approved': 0, 'rejected': 0, 'pending': 0})
    for r in rec_qs.values('submitted_at', 'status'):
        key = r['submitted_at'].strftime('%Y/%m') if r['submitted_at'] else '—'
        if r['status'] == 'موافق':              rec_dict[key]['approved'] += 1
        elif r['status'] == 'مرفوض':            rec_dict[key]['rejected'] += 1
        else:                                   rec_dict[key]['pending']  += 1

    rec_months = [
        {'month': k, 'approved': v['approved'], 'rejected': v['rejected'], 'pending': v['pending']}
        for k, v in sorted(rec_dict.items())
    ]

    return {
        'monthly':    months_data,
        'bene_pays':  bene_pays,
        'rec_months': rec_months,
    }

# ══════════════════════════════════════════════════════════
# كروت الدفعات
# ══════════════════════════════════════════════════════════

def _get_payments_cards(user, period='all', bene_id='', status_filter=''):
    d_from, d_to = _date_range(period)
    pays_qs = Payment.objects.filter(sponsor=user).select_related('beneficiary').order_by('-date')

    if d_from:       pays_qs = pays_qs.filter(date__range=(d_from, d_to))
    if bene_id:      pays_qs = pays_qs.filter(beneficiary_id=bene_id)
    if status_filter: pays_qs = pays_qs.filter(status=status_filter)

    STATUS_AR = {'paid': 'مسدّد', 'late': 'متأخر', 'pending': 'معلّق'}
    PAID_AR   = {'sponsor': 'كافل', 'admin': 'إدارة', 'external': 'جهة خارجية'}

    return [{
        'id':         p.pk,
        'bene_name':  p.beneficiary.get_full_name() if p.beneficiary else '—',
        'amount_ils': float(p.amount_ils),
        'amount_usd': float(p.amount_usd),
        'status':     p.status,
        'status_ar':  STATUS_AR.get(p.status, p.status),
        'date':       str(p.date),
        'paid_by':    PAID_AR.get(p.paid_by, p.paid_by),
        'note':       p.note or '',
    } for p in pays_qs[:100]]


# ══════════════════════════════════════════════════════════
# كروت الوصولات
# ══════════════════════════════════════════════════════════

def _get_receipts_cards(profile, period='all'):
    d_from, d_to = _date_range(period)
    rec_qs = PaymentReceipt.objects.filter(sponsor=profile).select_related('beneficiary').order_by('-submitted_at')
    if d_from:
        rec_qs = rec_qs.filter(submitted_at__date__range=(d_from, d_to))

    return [{
        'id':          r.pk,
        'system_ref':  r.system_ref or r.unique_number,
        'bene_name':   r.beneficiary.get_full_name() if r.beneficiary else '—',
        'amount_orig': float(r.amount_original),
        'currency':    r.currency,
        'amount_ils':  float(r.amount_shekel),
        'amount_usd':  float(r.amount_dollar),
        'status':      r.status,
        'date':        str(r.receipt_date) if r.receipt_date else '—',
        'submitted':   fmt_dt(r.submitted_at),
    } for r in rec_qs[:50]]


# ══════════════════════════════════════════════════════════
# ملخص المستفيدين
# ══════════════════════════════════════════════════════════

def _get_benes_summary(user, profile, period='all'):
    from beneficiary.models import OrphanForm, SpecialNeedsForm, FamilyForm
    d_from, d_to = _date_range(period)
    result = []

    for Model, label in [(OrphanForm,'يتيم'),(SpecialNeedsForm,'ذو احتياج'),(FamilyForm,'أسرة')]:
        for f in Model.objects.filter(sponsor=profile).select_related('user'):
            pays = Payment.objects.filter(sponsor=user, beneficiary=f.user)
            if d_from:
                pays = pays.filter(date__range=(d_from, d_to))
            paid  = pays.filter(status='paid')
            late  = pays.filter(status='late')
            total = paid.aggregate(t=Sum('amount_ils'))['t'] or 0
            last  = paid.order_by('-date').first()
            result.append({
                'name':      f.get_full_name(),
                'type':      label,
                'reg':       f.user.registration_number or f.form_number,
                'city':      f.current_city or '—',
                'paid_cnt':  paid.count(),
                'late_cnt':  late.count(),
                'total_ils': round(float(total), 2),
                'last_date': str(last.date) if last else '—',
                'sp_date':   str(f.sponsorship_date) if f.sponsorship_date else '—',
            })
    return result


# ══════════════════════════════════════════════════════════
# الصفحة الرئيسية
# ══════════════════════════════════════════════════════════

@sponsor_required
def reports(request):
    user    = request.user
    profile = _get_profile(user)

    notif_count = Notification.objects.filter(recipient=user, is_read=False).count()
    stats       = _get_stats(user, profile)

    from beneficiary.models import OrphanForm, SpecialNeedsForm, FamilyForm
    benes = []
    for Model, label in [(OrphanForm,'يتيم'),(SpecialNeedsForm,'ذو احتياج'),(FamilyForm,'أسرة')]:
        for f in Model.objects.filter(sponsor=profile).select_related('user'):
            benes.append({'id': str(f.user.pk), 'name': f.get_full_name(), 'type': label})

    return render(request, 'sponsor/reports.html', {
        'notif_count': notif_count,
        'stats':       stats,
        'benes':       benes,
    })


# ══════════════════════════════════════════════════════════
# AJAX
# ══════════════════════════════════════════════════════════

@sponsor_required
@require_GET
def reports_data(request):
    user    = request.user
    profile = _get_profile(user)
    period  = request.GET.get('period', 'all')
    bene_id = request.GET.get('bene',   '')
    status  = request.GET.get('status', '')
    section = request.GET.get('section','all')  # all | payments | receipts | benes

    data = {'stats': _get_stats(user, profile, period)}

    if section in ('all', 'charts'):
        data['charts'] = _get_chart_data(user, profile, period)

    if section in ('all', 'payments'):
        data['payments'] = _get_payments_cards(user, period, bene_id, status)

    if section in ('all', 'receipts'):
        data['receipts'] = _get_receipts_cards(profile, period)

    if section in ('all', 'benes'):
        data['benes_summary'] = _get_benes_summary(user, profile, period)

    return JsonResponse(data)


# ══════════════════════════════════════════════════════════
# تصدير Excel — شامل
# ══════════════════════════════════════════════════════════

@sponsor_required
@require_GET
def export_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from urllib.parse import quote
    import io

    user    = request.user
    profile = _get_profile(user)
    period  = request.GET.get('period', 'all')
    d_from, d_to = _date_range(period)

    log_activity(user, 'EXPORT', description=f'تصدير تقرير Excel — {period}', request=request)

    GREEN = '1A7A4A'; LIGHT = 'E8F5E9'; WHITE = 'FFFFFF'; BLUE = '2B6CB0'; RED = 'C53030'
    thin  = Side(style='thin', color='CCCCCC')
    brd   = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_cell(cell, color=GREEN):
        cell.font      = Font(bold=True, color=WHITE, size=10)
        cell.fill      = PatternFill('solid', fgColor=color)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = brd

    def data_cell(cell, r):
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border    = brd
        if r % 2 == 0:
            cell.fill = PatternFill('solid', fgColor=LIGHT)

    def title_row(ws, text, cols):
        ws.merge_cells(f'A1:{get_column_letter(cols)}1')
        c = ws['A1']
        c.value     = text
        c.font      = Font(bold=True, size=13, color=WHITE)
        c.fill      = PatternFill('solid', fgColor=GREEN)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

    wb = openpyxl.Workbook()

    # ══ ورقة 1: ملخص ══
    from beneficiary.models import OrphanForm, SpecialNeedsForm, FamilyForm
    stats = _get_stats(user, profile, period)

    ws0 = wb.active
    ws0.title = 'الملخص'
    ws0.sheet_view.rightToLeft = True
    title_row(ws0, f'ملخص تقرير الكافل — {user.get_full_name()} — {date.today()}', 2)

    summary_data = [
        ('بيانات الكافل', ''),
        ('الاسم الكامل',       user.get_full_name()),
        ('رقم التسجيل',        user.registration_number or '—'),
        ('البريد الإلكتروني',  user.email or '—'),
        ('تاريخ الانضمام',     str(user.date_joined.date())),
        ('المهنة',             profile.job or '—'),
        ('الدولة / المدينة',   f'{profile.country or "—"} / {profile.city or "—"}'),
        ('', ''),
        ('إحصائيات الكفالة', ''),
        ('إجمالي المكفولين',   stats['total']),
        ('أيتام',              stats['orphans']),
        ('ذوو احتياجات',       stats['specials']),
        ('أسر',                stats['families']),
        ('', ''),
        ('إحصائيات الدفعات', ''),
        ('إجمالي المدفوع (₪)', stats['total_ils']),
        ('إجمالي المدفوع ($)', stats['total_usd']),
        ('دفعات مسددة',        stats['paid_count']),
        ('دفعات متأخرة',       stats['late_count']),
        ('دفعات معلقة',        stats['pend_count']),
        ('نسبة الالتزام',      f'{stats["commitment_pct"]}%'),
        ('متوسط الدفعة (₪)',   stats['avg_ils']),
        ('آخر دفعة',           stats['last_pay_date']),
        ('', ''),
        ('إحصائيات الوصولات', ''),
        ('إجمالي الوصولات',    stats['rec_total']),
        ('وصولات موافقة',      stats['rec_approved']),
        ('وصولات مرفوضة',      stats['rec_rejected']),
        ('وصولات انتظار',      stats['rec_pending']),
        ('إجمالي الوصولات (₪)',stats['rec_ils']),
    ]

    ws0.column_dimensions['A'].width = 24
    ws0.column_dimensions['B'].width = 28

    for r, (k, v) in enumerate(summary_data, start=2):
        if not k:
            ws0.row_dimensions[r].height = 8
            continue
        if not v:
            c = ws0.cell(r, 1, k)
            c.font = Font(bold=True, color=WHITE, size=10)
            c.fill = PatternFill('solid', fgColor='2ECC71')
            c.border = brd
            ws0.merge_cells(f'A{r}:B{r}')
            ws0.row_dimensions[r].height = 22
            continue
        ck = ws0.cell(r, 1, k)
        cv = ws0.cell(r, 2, v)
        ck.font   = Font(bold=True, size=9)
        ck.fill   = PatternFill('solid', fgColor='F4F9F6') if r%2==0 else PatternFill()
        ck.border = brd
        cv.border = brd
        cv.fill   = PatternFill('solid', fgColor='F4F9F6') if r%2==0 else PatternFill()
        ws0.row_dimensions[r].height = 20

    # ══ ورقة 2: الدفعات ══
    pays_qs = Payment.objects.filter(sponsor=user).select_related('beneficiary').order_by('-date')
    if d_from:
        pays_qs = pays_qs.filter(date__range=(d_from, d_to))

    ws1 = wb.create_sheet('الدفعات')
    ws1.sheet_view.rightToLeft = True
    title_row(ws1, f'سجل الدفعات — {_period_label(period)}', 7)

    hdrs1 = ['#','المستفيد','النوع','المبلغ (₪)','المبلغ ($)','الحالة','التاريخ']
    widths1 = [6,28,14,14,14,12,14]
    for i,(h,w) in enumerate(zip(hdrs1,widths1),1):
        hdr_cell(ws1.cell(2,i,h))
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.row_dimensions[2].height = 22

    STATUS_AR = {'paid':'مسدّد','late':'متأخر','pending':'معلّق'}
    total_ils = total_usd = 0

    for ri,p in enumerate(pays_qs, start=3):
        bene_type = '—'
        for Model,label in [(OrphanForm,'يتيم'),(SpecialNeedsForm,'ذو احتياج'),(FamilyForm,'أسرة')]:
            if p.beneficiary and Model.objects.filter(user=p.beneficiary).exists():
                bene_type = label; break

        row = [ri-2, p.beneficiary.get_full_name() if p.beneficiary else '—', bene_type,
               float(p.amount_ils), float(p.amount_usd),
               STATUS_AR.get(p.status,p.status), str(p.date)]

        if p.status == 'paid':
            total_ils += float(p.amount_ils)
            total_usd += float(p.amount_usd)

        for ci,val in enumerate(row,1):
            cell = ws1.cell(ri,ci,val)
            data_cell(cell,ri)
            ws1.row_dimensions[ri].height = 20
            if ci == 6:
                clr = {'مسدّد':GREEN,'متأخر':RED,'معلّق':'B45309'}.get(val,'9CA3AF')
                cell.font = Font(bold=True,color=clr)
            if ci in (4,5):
                cell.font = Font(bold=True,color=GREEN if ci==4 else BLUE)

    tr = pays_qs.count()+3
    ws1.cell(tr,1,'الإجمالي').fill = PatternFill('solid',fgColor=GREEN)
    ws1.cell(tr,1).font = Font(bold=True,color=WHITE)
    ws1.cell(tr,4,round(total_ils,2)).font = Font(bold=True,color=WHITE)
    ws1.cell(tr,4).fill = PatternFill('solid',fgColor=GREEN)
    ws1.cell(tr,5,round(total_usd,2)).font = Font(bold=True,color=WHITE)
    ws1.cell(tr,5).fill = PatternFill('solid',fgColor=BLUE)
    for ci in range(1,8): ws1.cell(tr,ci).border = brd
    ws1.row_dimensions[tr].height = 22

    # ══ ورقة 3: الوصولات ══
    rec_qs = PaymentReceipt.objects.filter(sponsor=profile).select_related('beneficiary').order_by('-submitted_at')
    if d_from:
        rec_qs = rec_qs.filter(submitted_at__date__range=(d_from, d_to))

    ws2 = wb.create_sheet('الوصولات المالية')
    ws2.sheet_view.rightToLeft = True
    title_row(ws2, f'سجل الوصولات — {_period_label(period)}', 8)

    hdrs2 = ['#','الرقم المرجعي','المستفيد','المبلغ الأصلي','العملة','بالشيقل','بالدولار','الحالة']
    widths2 = [6,20,28,14,10,14,14,12]
    for i,(h,w) in enumerate(zip(hdrs2,widths2),1):
        hdr_cell(ws2.cell(2,i,h))
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.row_dimensions[2].height = 22

    for ri,r in enumerate(rec_qs, start=3):
        row2 = [ri-2, r.system_ref or r.unique_number,
                r.beneficiary.get_full_name() if r.beneficiary else '—',
                float(r.amount_original), r.currency,
                float(r.amount_shekel), float(r.amount_dollar), r.status]
        for ci,val in enumerate(row2,1):
            cell = ws2.cell(ri,ci,val)
            data_cell(cell,ri)
            ws2.row_dimensions[ri].height = 20
            if ci == 8:
                clr = {'موافق':GREEN,'مرفوض':RED,'بانتظار المراجعة':'B45309'}.get(val,'9CA3AF')
                cell.font = Font(bold=True,color=clr)

    # ══ ورقة 4: المكفولون ══
    ws3 = wb.create_sheet('المكفولون')
    ws3.sheet_view.rightToLeft = True
    title_row(ws3, f'قائمة المكفولين — {user.get_full_name()}', 7)

    hdrs3 = ['#','الاسم','رقم التسجيل','النوع','المدينة','تاريخ الكفالة','إجمالي المدفوع (₪)']
    widths3 = [6,28,18,14,14,16,18]
    for i,(h,w) in enumerate(zip(hdrs3,widths3),1):
        hdr_cell(ws3.cell(2,i,h))
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.row_dimensions[2].height = 22

    idx = 0
    for Model,label in [(OrphanForm,'يتيم'),(SpecialNeedsForm,'ذو احتياج'),(FamilyForm,'أسرة')]:
        for f in Model.objects.filter(sponsor=profile).select_related('user'):
            idx += 1
            total_f = Payment.objects.filter(
                sponsor=user, beneficiary=f.user, status='paid'
            ).aggregate(t=Sum('amount_ils'))['t'] or 0
            row3 = [idx, f.get_full_name(), f.user.registration_number or f.form_number,
                    label, f.current_city or '—',
                    str(f.sponsorship_date) if f.sponsorship_date else '—',
                    round(float(total_f),2)]
            for ci,val in enumerate(row3,1):
                cell = ws3.cell(idx+2,ci,val)
                data_cell(cell,idx)
                ws3.row_dimensions[idx+2].height = 20
                if ci == 7:
                    cell.font = Font(bold=True,color=GREEN)

    output = io.BytesIO()
    wb.save(output); output.seek(0)
    fname = f'تقرير_كافل_{user.get_full_name()}_{date.today()}.xlsx'
    resp  = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(fname)}"
    return resp


# ══════════════════════════════════════════════════════════
# تصدير PDF — شامل
# ══════════════════════════════════════════════════════════

@sponsor_required
def export_pdf_blob(request):
    from django.template.loader import render_to_string
    import weasyprint, logging

    logging.getLogger('weasyprint').setLevel(logging.ERROR)
    logging.getLogger('fontTools').setLevel(logging.ERROR)

    user    = request.user
    profile = _get_profile(user)
    period  = request.GET.get('period', 'all')
    d_from, d_to = _date_range(period)

    log_activity(user, 'EXPORT', description=f'تصدير تقرير PDF — {period}', request=request)

    from beneficiary.models import OrphanForm, SpecialNeedsForm, FamilyForm

    pays_qs = Payment.objects.filter(sponsor=user).select_related('beneficiary').order_by('-date')
    if d_from:
        pays_qs = pays_qs.filter(date__range=(d_from, d_to))

    rec_qs = PaymentReceipt.objects.filter(sponsor=profile).select_related('beneficiary').order_by('-submitted_at')
    if d_from:
        rec_qs = rec_qs.filter(submitted_at__date__range=(d_from, d_to))

    benes = []
    for Model,label in [(OrphanForm,'يتيم'),(SpecialNeedsForm,'ذو احتياج'),(FamilyForm,'أسرة')]:
        for f in Model.objects.filter(sponsor=profile).select_related('user'):
            total_f = Payment.objects.filter(
                sponsor=user, beneficiary=f.user, status='paid'
            ).aggregate(t=Sum('amount_ils'))['t'] or 0
            benes.append({
                'name':    f.get_full_name(),
                'reg':     f.user.registration_number or f.form_number,
                'type':    label,
                'city':    f.current_city or '—',
                'sp_date': str(f.sponsorship_date) if f.sponsorship_date else '—',
                'total':   round(float(total_f),2),
            })

    stats = _get_stats(user, profile, period)

    context = {
        'user':        user,
        'profile':     profile,
        'stats':       stats,
        'payments':    pays_qs[:60],
        'receipts':    rec_qs[:40],
        'benes':       benes,
        'period':      period,
        'period_label': _period_label(period),
        'print_date':  date.today().strftime('%Y/%m/%d'),
        'request':     request,
    }

    html_content = render_to_string('sponsor/report_pdf.html', context)
    pdf_file = weasyprint.HTML(
        string=html_content,
        base_url=request.build_absolute_uri('/')
    ).write_pdf()

    resp = HttpResponse(pdf_file, content_type='application/pdf')
    resp['Content-Length'] = str(len(pdf_file))
    resp['Cache-Control']  = 'no-store'
    return resp
