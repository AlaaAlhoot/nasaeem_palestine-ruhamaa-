from django.db.models import Count, Sum, Avg, Q
from django.utils import timezone
from django.contrib.auth import get_user_model
from django.core.cache import cache
from django.conf import settings
import json, os
from datetime import timedelta, datetime
from typing import Dict, List, Optional, Any

from .models import ActivityLog, SystemHealth
from projects.models import Project
from contact.models import ContactMessage, Newsletter
from main.models import SiteSettings
import logging

User = get_user_model()
logger = logging.getLogger(__name__)


# ========== دوال مساعدة للطلبات ==========

def get_client_ip(request) -> str:
    """الحصول على عنوان IP الحقيقي للمستخدم"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0].strip()
    else:
        ip = request.META.get('REMOTE_ADDR', '0.0.0.0')
    return ip


def get_user_agent(request) -> str:
    """الحصول على معلومات المتصفح"""
    user_agent = request.META.get('HTTP_USER_AGENT', '')
    return user_agent[:500] if user_agent else ''


# ========== دوال تسجيل الأنشطة ==========

def log_user_activity(user, action: str, title: str, description: str = '', **kwargs):
    """تسجيل نشاط المستخدم"""
    try:
        activity_data = {
            'user':        user,
            'username':    user.username if user else 'System',
            'action':      action,
            'title':       title,
            'description': description,
            'timestamp':   timezone.now(),
            'user_agent':  kwargs.get('user_agent', ''),
            'ip_address':  kwargs.get('ip_address', ''),
            'session_key': kwargs.get('session_key', ''),
            'level':       kwargs.get('level', 'info'),
            'extra_data':  kwargs.get('extra_data', {}),
        }

        if 'content_object' in kwargs and kwargs['content_object']:
            activity_data['content_object'] = kwargs['content_object']

        ActivityLog.objects.create(**activity_data)

    except Exception as e:
        logger.error(f"خطأ في تسجيل النشاط: {e}")


def log_create_activity(user, title: str, model_name: str, instance=None, request=None):
    """تسجيل نشاط الإنشاء — دالة مختصرة"""
    try:
        ActivityLog.log_activity(
            user=user,
            action='create',
            title=f'إضافة {model_name}: {title}',
            description=f'تم إضافة {model_name} بنجاح',
            content_object=instance,
            ip_address=get_client_ip(request) if request else '',
            user_agent=get_user_agent(request) if request else '',
            session_key=request.session.session_key if request else '',
        )
    except Exception as e:
        logger.error(f"خطأ في log_create_activity: {e}")


def log_update_activity(user, title: str, model_name: str, instance=None, request=None):
    """تسجيل نشاط التحديث — دالة مختصرة"""
    try:
        ActivityLog.log_activity(
            user=user,
            action='update',
            title=f'تحديث {model_name}: {title}',
            description=f'تم تحديث {model_name} بنجاح',
            content_object=instance,
            ip_address=get_client_ip(request) if request else '',
            user_agent=get_user_agent(request) if request else '',
            session_key=request.session.session_key if request else '',
        )
    except Exception as e:
        logger.error(f"خطأ في log_update_activity: {e}")


def log_delete_activity(user, title: str, model_name: str, request=None):
    """تسجيل نشاط الحذف — دالة مختصرة"""
    try:
        ActivityLog.log_activity(
            user=user,
            action='delete',
            title=f'حذف {model_name}: {title}',
            description=f'تم حذف {model_name} بنجاح',
            ip_address=get_client_ip(request) if request else '',
            user_agent=get_user_agent(request) if request else '',
            session_key=request.session.session_key if request else '',
        )
    except Exception as e:
        logger.error(f"خطأ في log_delete_activity: {e}")


# ========== إحصائيات لوحة التحكم ==========

def get_dashboard_statistics() -> Dict[str, Any]:
    """الحصول على إحصائيات لوحة التحكم مع التخزين المؤقت"""
    cache_key = 'dashboard_stats'
    stats = cache.get(cache_key)

    if not stats:
        try:
            # ✅ CustomUser لا يحتوي profile__last_activity — استخدم last_seen
            users_stats = User.objects.aggregate(
                total=Count('id'),
                active=Count('id', filter=Q(is_active=True)),
                staff=Count('id', filter=Q(is_staff=True)),
                online=Count('id', filter=Q(
                    last_seen__gte=timezone.now() - timedelta(minutes=15)
                ))
            )

            projects_stats = Project.objects.aggregate(
                total=Count('id'),
                active=Count('id', filter=Q(status='active')),
                completed=Count('id', filter=Q(status='completed')),
                featured=Count('id', filter=Q(is_featured=True))
            )

            messages_stats = ContactMessage.objects.aggregate(
                total=Count('id'),
                new=Count('id', filter=Q(status='new')),
                replied=Count('id', filter=Q(status='replied'))
            )

            newsletter_stats = Newsletter.objects.aggregate(
                total=Count('id'),
                active=Count('id', filter=Q(is_active=True)),
                confirmed=Count('id', filter=Q(confirmed_at__isnull=False))
            )

            stats = {
                'users':       users_stats,
                'projects':    projects_stats,
                'messages':    messages_stats,
                'newsletter':  newsletter_stats,
                'last_updated': timezone.now().isoformat()
            }

            cache.set(cache_key, stats, 300)

        except Exception as e:
            logger.error(f"خطأ في get_dashboard_statistics: {e}")
            stats = {'error': str(e)}

    return stats


# ========== صحة النظام ==========

def get_system_health() -> dict:
    """الحصول على معلومات صحة النظام"""
    try:
        import psutil
        import platform
        from django.db import connection

        cpu_percent    = psutil.cpu_percent(interval=1)
        cpu_count      = psutil.cpu_count()
        memory         = psutil.virtual_memory()
        disk           = psutil.disk_usage('/')

        if cpu_percent > 80 or memory.percent > 80 or disk.percent > 80:
            status = 'critical'
        elif cpu_percent > 60 or memory.percent > 60 or disk.percent > 60:
            status = 'warning'
        else:
            status = 'healthy'

        try:
            db_connections = len(connection.queries)
        except Exception:
            db_connections = 0

        # ✅ CustomUser يحتوي last_seen بدل profile__last_activity
        try:
            active_users = User.objects.filter(
                last_seen__gte=timezone.now() - timedelta(minutes=15)
            ).count()
        except Exception:
            active_users = 0

        health_data = {
            'status':               status,
            'cpu_usage_percent':    round(cpu_percent, 2),
            'cpu_count':            cpu_count,
            'memory_usage_percent': round(memory.percent, 2),
            'memory_total':         memory.total,
            'memory_used':          memory.used,
            'disk_usage_percent':   round(disk.percent, 2),
            'disk_total':           disk.total,
            'disk_used':            disk.used,
            'db_connections':       db_connections,
            'active_users':         active_users,
            'platform':             platform.system(),
            'platform_version':     platform.version(),
            'architecture':         platform.machine(),
            'timestamp':            timezone.now(),
        }

        # حفظ في قاعدة البيانات
        try:
            SystemHealth.objects.create(
                status=status,
                response_time=0.1,
                memory_usage_percent=round(memory.percent, 2),
                memory_total=memory.total,
                memory_used=memory.used,
                disk_usage_percent=round(disk.percent, 2),
                disk_total=disk.total,
                disk_used=disk.used,
                cpu_usage_percent=round(cpu_percent, 2),
                db_connections=db_connections,
                db_query_time=0,
                active_users=active_users,
                errors_count=0,
                warnings_count=0,
            )
        except Exception as save_error:
            logger.warning(f"فشل حفظ صحة النظام: {save_error}")

        return health_data

    except ImportError:
        logger.warning("psutil غير مثبت")
        return _empty_health('unknown')
    except Exception as e:
        logger.error(f"خطأ في get_system_health: {e}")
        return _empty_health('error')


def _empty_health(status: str) -> dict:
    """إرجاع بيانات صحة فارغة"""
    return {
        'status': status,
        'cpu_usage_percent': 0, 'cpu_count': 0,
        'memory_usage_percent': 0, 'memory_total': 0, 'memory_used': 0,
        'disk_usage_percent': 0, 'disk_total': 0, 'disk_used': 0,
        'db_connections': 0, 'active_users': 0,
        'timestamp': timezone.now(),
    }


# ========== التقارير ==========

def generate_report(report_type: str, format_type: str = 'pdf', user=None) -> Dict[str, Any]:
    """توليد التقارير بصيغ مختلفة"""
    try:
        reporters = {
            'monthly_summary': _get_monthly_summary_data,
            'projects_report': _get_projects_report_data,
            'users_activity':  _get_users_activity_data,
            'messages_report': _get_messages_report_data,
        }

        if report_type not in reporters:
            return {'success': False, 'error': 'نوع تقرير غير مدعوم'}

        data = reporters[report_type]()

        generators = {
            'pdf':   _generate_pdf_report,
            'excel': _generate_excel_report,
            'csv':   _generate_csv_report,
        }

        if format_type not in generators:
            return {'success': False, 'error': 'صيغة غير مدعومة'}

        filename = generators[format_type](report_type, data)

        if user:
            log_user_activity(
                user=user,
                action='export',
                title=f'توليد تقرير {report_type}',
                description=f'تم توليد تقرير {report_type} بصيغة {format_type}'
            )

        return {'success': True, 'filename': filename, 'url': f'/media/reports/{filename}'}

    except Exception as e:
        logger.error(f"خطأ في generate_report: {e}")
        return {'success': False, 'error': str(e)}


# ========== النسخ الاحتياطي ==========

def backup_database() -> Dict[str, Any]:
    """إنشاء نسخة احتياطية من قاعدة البيانات"""
    try:
        timestamp   = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_name = f'backup_{timestamp}.sql'
        backup_dir  = os.path.join(settings.MEDIA_ROOT, 'backups')
        backup_path = os.path.join(backup_dir, backup_name)
        os.makedirs(backup_dir, exist_ok=True)

        engine = settings.DATABASES['default']['ENGINE']

        if 'postgresql' in engine:
            import subprocess
            db = settings.DATABASES['default']
            env = os.environ.copy()
            env['PGPASSWORD'] = db['PASSWORD']
            subprocess.run([
                'pg_dump',
                '-h', db.get('HOST', 'localhost'),
                '-p', str(db.get('PORT', 5432)),
                '-U', db['USER'],
                '-d', db['NAME'],
                '-f', backup_path
            ], env=env, check=True)

        elif 'sqlite' in engine:
            import shutil
            shutil.copy2(settings.DATABASES['default']['NAME'], backup_path)

        return {'success': True, 'filename': backup_name, 'path': backup_path}

    except Exception as e:
        logger.error(f"خطأ في backup_database: {e}")
        return {'success': False, 'error': str(e)}


# ========== تنظيف السجلات ==========

def clean_old_logs(days: int = 30):
    """تنظيف السجلات القديمة"""
    try:
        cutoff = timezone.now() - timedelta(days=days)
        deleted_count, _ = ActivityLog.objects.filter(timestamp__lt=cutoff).delete()
        return {'success': True, 'deleted_count': deleted_count}
    except Exception as e:
        return {'success': False, 'error': str(e)}


# ========== المحتوى الشائع ==========

def get_popular_content() -> Dict[str, List]:
    """الحصول على المحتوى الشائع"""
    try:
        return {
            'projects': list(
                Project.objects.filter(is_active=True)
                .order_by('-views_count')[:5]
                .values('title_ar', 'views_count', 'likes_count')
            ),
            'messages_this_week': ContactMessage.objects.filter(
                created_at__gte=timezone.now() - timedelta(days=7)
            ).count()
        }
    except Exception as e:
        logger.error(f"خطأ في get_popular_content: {e}")
        return {'projects': [], 'messages_this_week': 0}


# ========== صلاحيات المستخدم ==========

def check_user_permissions(user) -> dict:
    """الحصول على جميع صلاحيات المستخدم"""
    if not user.is_authenticated:
        return {
            'authenticated':    False,
            'is_dashboard_user': False,
            'is_admin':         False,
            'can_edit_content': False,
            'role':             None,
            'resources':        {},
            'user_info':        None
        }

    if user.is_superuser:
        return {
            'authenticated':    True,
            'is_dashboard_user': True,
            'is_admin':         True,
            'can_edit_content': True,
            'role':             'super_admin',
            'role_display':     'مدير عام',
            'resources': {key: True for key in [
                'users_management', 'system_settings', 'content_editing',
                'reports_viewing', 'statistics_viewing', 'projects_management',
                'messages_management', 'newsletter_management', 'backup_restore',
                'system_health', 'activity_logs', 'site_settings'
            ]},
            'user_info': {
                'id':          str(user.id),  # ✅ UUID → str
                'username':    user.username,
                'email':       user.email,
                'full_name':   user.get_full_name(),
                'is_superuser': True,
            }
        }

    try:
        profile = user.profile

        role_permissions = {
            'super_admin': {
                'users_management': True, 'system_settings': True, 'content_editing': True,
                'reports_viewing': True, 'statistics_viewing': True, 'projects_management': True,
                'messages_management': True, 'newsletter_management': True, 'backup_restore': True,
                'system_health': True, 'activity_logs': True, 'site_settings': True
            },
            'admin': {
                'users_management': True, 'system_settings': True, 'content_editing': True,
                'reports_viewing': True, 'statistics_viewing': True, 'projects_management': True,
                'messages_management': True, 'newsletter_management': True, 'backup_restore': True,
                'system_health': True, 'activity_logs': True, 'site_settings': False
            },
            'editor': {
                'users_management': False, 'system_settings': False, 'content_editing': True,
                'reports_viewing': True, 'statistics_viewing': True, 'projects_management': True,
                'messages_management': True, 'newsletter_management': True, 'backup_restore': False,
                'system_health': False, 'activity_logs': False, 'site_settings': False
            },
            'moderator': {
                'users_management': False, 'system_settings': False, 'content_editing': False,
                'reports_viewing': True, 'statistics_viewing': True, 'projects_management': False,
                'messages_management': True, 'newsletter_management': False, 'backup_restore': False,
                'system_health': False, 'activity_logs': True, 'site_settings': False
            },
            'viewer': {
                'users_management': False, 'system_settings': False, 'content_editing': False,
                'reports_viewing': False, 'statistics_viewing': True, 'projects_management': False,
                'messages_management': False, 'newsletter_management': False, 'backup_restore': False,
                'system_health': False, 'activity_logs': False, 'site_settings': False
            }
        }

        user_resources = role_permissions.get(profile.role, role_permissions['viewer'])

        return {
            'authenticated':    True,
            'is_dashboard_user': profile.is_active_staff and user.is_active,
            'is_admin':         profile.is_admin(),
            'can_edit_content': profile.can_edit_content(),
            'role':             profile.role,
            'role_display':     profile.get_role_display(),
            'resources':        user_resources,
            'user_info': {
                'id':            str(user.id),  # ✅ UUID → str
                'username':      user.username,
                'email':         user.email,
                'full_name':     user.get_full_name(),
                'display_name':  profile.get_display_name(),
                'avatar_url':    profile.avatar.url if hasattr(profile, 'avatar') and profile.avatar else None,
                'last_activity': getattr(profile, 'last_activity', user.last_seen),
                'is_superuser':  False,
            }
        }

    except Exception as e:
        logger.error(f"خطأ في check_user_permissions للمستخدم {user.username}: {e}")
        return {
            'authenticated':    True,
            'is_dashboard_user': user.is_staff,
            'is_admin':         user.is_superuser,
            'can_edit_content': user.is_staff,
            'role':             'legacy_user',
            'role_display':     'مستخدم قديم',
            'resources': {key: user.is_staff for key in [
                'users_management', 'system_settings', 'content_editing',
                'reports_viewing', 'statistics_viewing'
            ]},
            'user_info': {
                'id':        str(user.id),  # ✅ UUID → str
                'username':  user.username,
                'email':     user.email,
                'full_name': user.get_full_name(),
                'error':     'فشل في تحميل بيانات المستخدم'
            }
        }


# ========== دوال البيانات الداخلية ==========

def _get_monthly_summary_data():
    start_date = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    return {
        'period':     start_date.strftime('%Y-%m'),
        'projects':   Project.objects.filter(created_at__gte=start_date).count(),
        'messages':   ContactMessage.objects.filter(created_at__gte=start_date).count(),
        'users':      User.objects.filter(date_joined__gte=start_date).count(),
        'activities': ActivityLog.objects.filter(timestamp__gte=start_date).count(),
    }


def _get_projects_report_data():
    return {
        'total':      Project.objects.count(),
        'by_status':  dict(Project.objects.values('status').annotate(count=Count('id')).values_list('status', 'count')),
        'by_category': dict(
            Project.objects.values('category__name_ar').annotate(count=Count('id')).values_list('category__name_ar', 'count')
        ),
        'top_viewed': list(Project.objects.order_by('-views_count')[:10].values('title_ar', 'views_count', 'status')),
        'recent':     list(Project.objects.order_by('-created_at')[:10].values('title_ar', 'created_at', 'status')),
    }


def _get_users_activity_data():
    # ✅ CustomUser لا يحتوي profile__role — استخدم user_type
    return {
        'total_users':         User.objects.count(),
        'active_users':        User.objects.filter(is_active=True).count(),
        'staff_users':         User.objects.filter(is_staff=True).count(),
        'active_last_30_days': User.objects.filter(
            last_login__gte=timezone.now() - timedelta(days=30)
        ).count(),
        'by_type': dict(
            User.objects.values('user_type').annotate(count=Count('id')).values_list('user_type', 'count')
        ),
        'login_stats': list(
            ActivityLog.objects.filter(
                action='login',
                timestamp__gte=timezone.now() - timedelta(days=30)
            ).extra({'date': 'DATE(timestamp)'})
            .values('date').annotate(count=Count('id')).order_by('date')
        )
    }


def _get_messages_report_data():
    return {
        'total_messages': ContactMessage.objects.count(),
        'by_status':   dict(ContactMessage.objects.values('status').annotate(count=Count('id')).values_list('status', 'count')),
        'by_priority': dict(ContactMessage.objects.values('priority').annotate(count=Count('id')).values_list('priority', 'count')),
        'recent_messages': list(
            ContactMessage.objects.order_by('-created_at')[:20]
            .values('name', 'subject', 'status', 'created_at')
        )
    }


def _generate_pdf_report(report_type: str, data: Dict) -> str:
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename  = f'{report_type}_{timestamp}.pdf'
        filepath  = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
        os.makedirs(os.path.dirname(filepath), exist_ok=True)

        c = canvas.Canvas(filepath, pagesize=letter)
        c.drawString(100, 750, f"تقرير {report_type}")
        c.drawString(100, 730, f"تاريخ الإنشاء: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        y = 700
        for key, value in data.items():
            c.drawString(100, y, f"{key}: {value}")
            y -= 20
            if y < 50:
                c.showPage()
                y = 750
        c.save()
        return filename

    except Exception:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename  = f'{report_type}_{timestamp}.txt'
        filepath  = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(f"تقرير {report_type}\n")
            f.write(f"تاريخ الإنشاء: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n")
            for key, value in data.items():
                f.write(f"{key}: {value}\n")
        return filename


def _generate_excel_report(report_type: str, data: Dict) -> str:
    try:
        from openpyxl import Workbook

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename  = f'{report_type}_{timestamp}.xlsx'
        filepath  = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
        os.makedirs(os.path.dirname(filepath), exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = report_type
        ws['A1'] = f'تقرير {report_type}'
        ws['A2'] = f'تاريخ الإنشاء: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
        row = 4
        for key, value in data.items():
            ws[f'A{row}'] = str(key)
            ws[f'B{row}'] = str(value)
            row += 1
        wb.save(filepath)
        return filename

    except ImportError:
        return _generate_csv_report(report_type, data)


def _generate_csv_report(report_type: str, data: Dict) -> str:
    import csv

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename  = f'{report_type}_{timestamp}.csv'
    filepath  = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['البيان', 'القيمة'])
        for key, value in data.items():
            writer.writerow([str(key), str(value)])

    return filename


# في نهاية utils.py
get_system_health_data = get_system_health