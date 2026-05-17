"""
settings.py — فيوز صفحة الإعدادات الشاملة
"""

from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST, require_GET
from django.views.decorators.csrf import csrf_protect
from django.utils import timezone
from django.core.mail import send_mail, EmailMessage
from django.conf import settings as django_settings
from django.db import connection
import re, json, io, sys
from datetime import timedelta

from core.models import CustomUser, ActivityLog, Complaint
from core.utils import log_activity
from .decorators import admin_required, settings_required


# ══════════════════════════════════════════════
# مساعدات SystemSettings (key/value)
# ══════════════════════════════════════════════
def _gs(key, default=''):
    """get setting"""
    try:
        from core.models import SystemSettings
        return SystemSettings.get(key, default)
    except Exception:
        return default

def _ss(key, value, user=None):
    """set setting"""
    try:
        from core.models import SystemSettings
        SystemSettings.set(key, value, user)
    except Exception:
        pass

def _uname(user):
    if not user: return '—'
    return user.get_full_name() or user.username or '—'

def _build_sys_ctx():
    """بناء context الإعدادات"""
    return {
        'site_name':             _gs('site_name',       'منصة رُحَمَاء'),
        'site_address':          _gs('site_address'),
        'site_email':            _gs('site_email'),
        'site_phone':            _gs('site_phone'),
        'site_whatsapp':         _gs('site_whatsapp'),
        'work_days_type':        _gs('work_days_type',  'week'),
        'work_days_from':        _gs('work_days_from',  'السبت'),
        'work_days_to':          _gs('work_days_to',    'الخميس'),
        'work_days_custom':      _gs('work_days_custom',''),
        'work_time_from_h':      _gs('work_time_from_h','9'),
        'work_time_from_m':      _gs('work_time_from_m','00'),
        'work_time_from_p':      _gs('work_time_from_p','صباحاً'),
        'work_time_to_h':        _gs('work_time_to_h',  '4'),
        'work_time_to_m':        _gs('work_time_to_m',  '00'),
        'work_time_to_p':        _gs('work_time_to_p',  'مساءً'),
        'maintenance_mode':      _gs('maintenance_mode','false') == 'true',
        'notifications_enabled': _gs('notifications_enabled','true') == 'true',
        'max_login_attempts':    _gs('max_login_attempts','5'),
        'session_timeout':       _gs('session_timeout','60'),
        'blocked_ips':           json.loads(_gs('blocked_ips','[]') or '[]'),
    }


# ══════════════════════════════════════════════
# الصفحة الرئيسية
# ══════════════════════════════════════════════
@settings_required
@admin_required
def settings_view(request):
    user   = request.user
    admins = CustomUser.objects.filter(user_type='admin').order_by('-date_joined')
    log_count = ActivityLog.objects.filter(user=user).count()

    try:
        import django
        django_ver = django.__version__
    except Exception:
        django_ver = '—'

    py_ver = f'{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}'

    # إعدادات البريد من settings.py المشروع
    email_config = {
        'host':     getattr(django_settings, 'EMAIL_HOST',     '—'),
        'port':     getattr(django_settings, 'EMAIL_PORT',     '—'),
        'user':     getattr(django_settings, 'EMAIL_HOST_USER','—'),
        'use_tls':  getattr(django_settings, 'EMAIL_USE_TLS',  False),
        'from':     getattr(django_settings, 'DEFAULT_FROM_EMAIL','—'),
    }

    context = {
        'user':         user,
        'admins':       admins,
        'sys_s':        _build_sys_ctx(),
        'log_count':    log_count,
        'django_ver':   django_ver,
        'py_ver':       py_ver,
        'db_engine':    connection.settings_dict.get('ENGINE','').split('.')[-1],
        'db_name':      str(connection.settings_dict.get('NAME','')),
        'email_config': email_config,
        'perms_list': [
            ('beneficiaries', 'إدارة المستفيدين'),
            ('sponsors',      'إدارة الكفلاء'),
            ('payments',      'إدارة الدفعات'),
            ('aids',          'إدارة المساعدات'),
            ('logs',          'عرض السجلات'),
            ('settings',      'الوصول للإعدادات'),
        ],
        'days_list': ['السبت','الأحد','الاثنين','الثلاثاء','الأربعاء','الخميس','الجمعة'],
    }
    return render(request, 'admin_panel/settings.html', context)


# ══════════════════════════════════════════════
# الملف الشخصي
# ══════════════════════════════════════════════
@admin_required
@require_POST
@csrf_protect
def update_profile(request):
    import traceback, os
    from django.core.files.base import ContentFile
    from django.core.files.storage import default_storage
    try:
        user = request.user
        fields = ['first_name', 'father_name', 'grand_name', 'family_name',
                  'email', 'phone', 'phone_country', 'whatsapp', 'whatsapp_country']
        changed = []
        for f in fields:
            if f in request.POST:
                val = request.POST.get(f, '').strip()
                if getattr(user, f, '') != val:
                    setattr(user, f, val)
                    changed.append(f)

        if 'photo' in request.FILES:
            original_file = request.FILES['photo']
            try:
                from core.utils import compress_image
                compressed = compress_image(original_file)
                ext = os.path.splitext(original_file.name)[1] or '.jpg'
                filename = f"profiles/profile_{user.id}{ext}"
                if user.profile_image and default_storage.exists(user.profile_image.name):
                    default_storage.delete(user.profile_image.name)
                saved_path = default_storage.save(filename, ContentFile(compressed.read()))
                user.profile_image.name = saved_path
            except Exception as e:
                print("photo error:", e)
                user.profile_image = original_file
            changed.append('profile_image')

        if changed:
            # احفظ بشكل صريح مع update_fields
            user.save(update_fields=changed)
            log_activity(user, 'UPDATE',
                        description=f'تعديل الملف الشخصي: {", ".join(changed)}',
                        request=request)

        photo_url = user.profile_image.url if user.profile_image else ''
        return JsonResponse({
            'status': 'success',
            'message': 'تم حفظ البيانات ✅',
            'photo_url': photo_url,
            'full_name': user.get_full_name() or user.username
        })

    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e),
            'traceback': traceback.format_exc()
        }, status=500)

@admin_required
@require_POST
@csrf_protect
def change_password(request):
    user      = request.user
    old_pass  = request.POST.get('old_password',  '').strip()
    new_pass  = request.POST.get('new_password',  '').strip()
    new_pass2 = request.POST.get('new_password2', '').strip()

    if not user.check_password(old_pass):
        return JsonResponse({'status':'error','message':'كلمة المرور الحالية غير صحيحة'})
    if len(new_pass) < 8:
        return JsonResponse({'status':'error','message':'كلمة المرور يجب أن تكون 8 أحرف على الأقل'})
    if new_pass != new_pass2:
        return JsonResponse({'status':'error','message':'كلمتا المرور غير متطابقتان'})

    user.set_password(new_pass)
    user.save()
    log_activity(user, 'UPDATE', description='تغيير كلمة مرور الأدمن', request=request)
    return JsonResponse({'status':'success','message':'تم تغيير كلمة المرور ✅'})


# ══════════════════════════════════════════════
# إدارة الأدمن
# ══════════════════════════════════════════════
@admin_required
@require_GET
def check_username(request):
    username   = request.GET.get('username','').strip()
    exclude_id = request.GET.get('exclude_id','').strip()
    if not username:
        return JsonResponse({'available':True})
    # تحقق النمط
    if not re.match(r'^[a-zA-Z0-9_@.+-]{4,}$', username):
        return JsonResponse({'available':False,'message':'4 أحرف على الأقل — إنجليزي/أرقام فقط'})
    if re.search(r'[\u0600-\u06FF\s]', username):
        return JsonResponse({'available':False,'message':'لا يُسمح بالعربية أو المسافات'})
    qs = CustomUser.objects.filter(username=username)
    if exclude_id: qs = qs.exclude(pk=exclude_id)
    return JsonResponse({'available': not qs.exists(),
                         'message': 'مستخدم مسبقاً' if qs.exists() else 'متاح ✅'})


@admin_required
@require_GET
def check_email_available(request):
    email      = request.GET.get('email','').strip()
    exclude_id = request.GET.get('exclude_id','').strip()
    if not email:
        return JsonResponse({'available':True})
    qs = CustomUser.objects.filter(email=email)
    if exclude_id: qs = qs.exclude(pk=exclude_id)
    return JsonResponse({'available': not qs.exists(),
                         'message': 'مستخدم مسبقاً' if qs.exists() else 'متاح ✅'})


@admin_required
@require_POST
@csrf_protect
def create_admin(request):
    username    = request.POST.get('username',   '').strip()
    first_name  = request.POST.get('first_name', '').strip()
    father_name = request.POST.get('father_name','').strip()
    grand_name  = request.POST.get('grand_name', '').strip()
    family_name = request.POST.get('family_name','').strip()
    email       = request.POST.get('email',      '').strip()
    phone       = request.POST.get('phone',      '').strip()
    password    = request.POST.get('password',   '').strip()

    # تحقق
    if not username or not password or not email or not first_name:
        return JsonResponse({'status':'error','message':'الحقول المطلوبة ناقصة'})
    if not re.match(r'^[a-zA-Z0-9_@.+-]{4,}$', username):
        return JsonResponse({'status':'error','message':'اسم المستخدم غير صحيح'})
    if len(password) < 8:
        return JsonResponse({'status':'error','message':'كلمة المرور يجب 8 أحرف على الأقل'})
    if CustomUser.objects.filter(username=username).exists():
        return JsonResponse({'status':'error','message':'اسم المستخدم مستخدم مسبقاً'})
    if CustomUser.objects.filter(email=email).exists():
        return JsonResponse({'status':'error','message':'البريد الإلكتروني مستخدم مسبقاً'})

    try:
        new_admin = CustomUser(
            username    = username,
            first_name  = first_name,
            father_name = father_name,
            grand_name  = grand_name,
            family_name = family_name,
            email       = email,
            phone       = phone,
            user_type   = 'admin',
            is_approved = True,
            is_active   = True,
        )
        new_admin.set_password(password)
        new_admin.save()

        log_activity(request.user, 'CREATE',
                     description=f'إنشاء أدمن جديد: {username}',
                     target_model='CustomUser', target_id=str(new_admin.pk),
                     request=request)
        return JsonResponse({'status':'success','message':f'تم إنشاء الأدمن {username} ✅',
                             'admin': {
                                 'id':       str(new_admin.pk),
                                 'name':     new_admin.get_full_name() or username,
                                 'username': username,
                                 'email':    email,
                             }})
    except Exception as e:
        return JsonResponse({'status':'error','message':str(e)})


@admin_required
@require_POST
@csrf_protect
def update_admin(request):

    admin_id    = request.POST.get('admin_id',   '').strip()
    first_name  = request.POST.get('first_name', '').strip()
    father_name = request.POST.get('father_name','').strip()
    grand_name  = request.POST.get('grand_name', '').strip()
    family_name = request.POST.get('family_name','').strip()
    email       = request.POST.get('email',      '').strip()
    phone       = request.POST.get('phone',      '').strip()
    is_active   = request.POST.get('is_active',  '1') == '1'

    try:
        admin = CustomUser.objects.get(pk=admin_id, user_type='admin')
        admin.whatsapp = request.POST.get('whatsapp', '').strip()
        new_pass = request.POST.get('new_password', '').strip()
        if new_pass and not admin.is_superuser:
            if len(new_pass) < 8:
                return JsonResponse({'status': 'error', 'message': 'كلمة المرور يجب 8 أحرف على الأقل'})
            admin.set_password(new_pass)

        admin.first_name  = first_name
        admin.father_name = father_name
        admin.grand_name  = grand_name
        admin.family_name = family_name
        admin.email       = email
        admin.phone       = phone

        if admin != request.user:
            admin.is_active = is_active

        # صلاحيات الإدارة
        perm_fields = {
            'can_manage_beneficiaries': 'perm_beneficiaries',
            'can_manage_sponsors':      'perm_sponsors',
            'can_manage_payments':      'perm_payments',
            'can_manage_aids':          'perm_aids',
            'can_view_logs':            'perm_logs',
            'can_access_settings':      'perm_settings',
        }

        for field, post_key in perm_fields.items():
            if hasattr(admin, field):
                setattr(
                    admin,
                    field,
                    request.POST.get(post_key, '') == '1'
                )

        admin.save()

        log_activity(
            request.user,
            'UPDATE',
            description=f'تعديل بيانات الأدمن: {_uname(admin)}',
            request=request
        )

        return JsonResponse({
            'status': 'success',
            'message': 'تم تعديل البيانات ✅',
            'name': admin.get_full_name() or admin.username
        })

    except CustomUser.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'الأدمن غير موجود'
        })


@admin_required
@require_POST
@csrf_protect
def toggle_admin(request):
    admin_id = request.POST.get('admin_id', '').strip()
    try:
        admin = CustomUser.objects.get(pk=admin_id, user_type='admin')
        if admin == request.user:
            return JsonResponse({'status': 'error', 'message': 'لا يمكنك تعطيل حسابك الخاص'})

        admin.is_active = not admin.is_active
        admin.save()

        action = 'تفعيل' if admin.is_active else 'تعطيل'
        log_activity(
            request.user, 'UPDATE',
            description=f'{action} حساب الأدمن: {_uname(admin)}',
            target_model='CustomUser',
            target_id=str(admin.pk),
            request=request,
        )
        return JsonResponse({
            'status':    'success',
            'is_active': admin.is_active,
            'message':   f'تم {action} الحساب ✅',
        })
    except CustomUser.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'الأدمن غير موجود'})


@admin_required
@require_POST
@csrf_protect
def toggle_admin_settings(request):
    admin_id = request.POST.get('admin_id', '').strip()
    try:
        admin = CustomUser.objects.get(pk=admin_id, user_type='admin')

        if not hasattr(admin, 'can_access_settings'):
            return JsonResponse({
                'status':  'error',
                'message': 'حقل can_access_settings غير موجود في الموديل — يرجى إضافته',
            })

        admin.can_access_settings = not admin.can_access_settings
        admin.save()

        action = 'منح' if admin.can_access_settings else 'سحب'
        log_activity(
            request.user, 'UPDATE',
            description=f'{action} صلاحية الوصول للإعدادات — الأدمن: {_uname(admin)}',
            target_model='CustomUser',
            target_id=str(admin.pk),
            request=request,
        )
        return JsonResponse({
            'status':     'success',
            'can_access': admin.can_access_settings,
            'message':    f'تم {action} صلاحية الإعدادات ✅',
        })
    except CustomUser.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'الأدمن غير موجود'})


# ══════════════════════════════════════════════
# معلومات الموقع
# ══════════════════════════════════════════════
@admin_required
@require_POST
@csrf_protect
def update_site_info(request):
    fields = ['site_name','site_address','site_email','site_phone','site_whatsapp',
              'work_days_type','work_days_from','work_days_to','work_days_custom',
              'work_time_from_h','work_time_from_m','work_time_from_p',
              'work_time_to_h','work_time_to_m','work_time_to_p']
    for f in fields:
        if f in request.POST:
            _ss(f, request.POST.get(f,'').strip(), request.user)
    log_activity(request.user, 'UPDATE', description='تعديل معلومات الموقع', request=request)
    return JsonResponse({'status':'success','message':'تم حفظ معلومات الموقع ✅'})


# ══════════════════════════════════════════════
# وضع الصيانة
# ══════════════════════════════════════════════
@admin_required
@require_POST
@csrf_protect
def toggle_maintenance(request):
    current = _gs('maintenance_mode','false') == 'true'
    new_val = not current
    _ss('maintenance_mode', 'true' if new_val else 'false', request.user)
    action = 'تفعيل' if new_val else 'إيقاف'
    log_activity(request.user, 'UPDATE', description=f'{action} وضع الصيانة', request=request)
    return JsonResponse({'status':'success','maintenance':new_val,
                         'message':f'تم {action} وضع الصيانة ✅'})


# ══════════════════════════════════════════════
# أسعار الصرف
# ══════════════════════════════════════════════
@admin_required
@require_GET
def get_rates(request):
    try:
        import urllib.request
        url = 'https://api.exchangerate-api.com/v4/latest/ILS'
        with urllib.request.urlopen(url, timeout=5) as resp:
            data = json.loads(resp.read())
        rates = data.get('rates', {})
        def safe(v, d=1): return rates.get(v, d)
        result = {
            'ILS_TO_USD': round(safe('USD'),    4),
            'ILS_TO_EUR': round(safe('EUR'),    4),
            'ILS_TO_JOD': round(safe('JOD'),    4),
            'ILS_TO_SAR': round(safe('SAR'),    4),
            'ILS_TO_EGP': round(safe('EGP'),    4),
            'USD_TO_ILS': round(1/safe('USD'),  4),
            'EUR_TO_ILS': round(1/safe('EUR'),  4),
            'JOD_TO_ILS': round(1/safe('JOD'),  4),
            'SAR_TO_ILS': round(1/safe('SAR'),  4),
            'EGP_TO_ILS': round(1/safe('EGP'),  4),
            'updated_at': timezone.localtime(timezone.now()).strftime('%Y/%m/%d %H:%M'),
        }
        return JsonResponse({'status':'success','rates':result})
    except Exception as e:
        try:
            from core.utils import get_exchange_rates
            r = get_exchange_rates()
            r['updated_at'] = timezone.localtime(timezone.now()).strftime('%Y/%m/%d %H:%M')
            return JsonResponse({'status':'success','rates':r})
        except Exception:
            return JsonResponse({'status':'error','message':str(e)})


# ══════════════════════════════════════════════
# قاعدة البيانات
# ══════════════════════════════════════════════
@admin_required
@require_GET
def db_stats(request):
    tables = []
    try:
        engine = connection.settings_dict.get('ENGINE','')
        with connection.cursor() as cursor:
            if 'sqlite' in engine:
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
                for (name,) in cursor.fetchall():
                    try:
                        cursor.execute(f'SELECT COUNT(*) FROM "{name}"')
                        tables.append({'name':name,'count':cursor.fetchone()[0],'size':'—'})
                    except Exception:
                        continue
            elif 'mysql' in engine:
                db = connection.settings_dict.get('NAME','')
                cursor.execute(f"""
                    SELECT table_name,IFNULL(table_rows,0),
                    ROUND((data_length+index_length)/1024,1)
                    FROM information_schema.tables
                    WHERE table_schema='{db}' ORDER BY table_rows DESC""")
                for r in cursor.fetchall():
                    tables.append({'name':r[0],'count':r[1],'size':f'{r[2]} KB'})
            elif 'postgresql' in engine:
                cursor.execute("""
                    SELECT relname,n_live_tup,
                    pg_size_pretty(pg_total_relation_size(relid))
                    FROM pg_stat_user_tables ORDER BY n_live_tup DESC""")
                for r in cursor.fetchall():
                    tables.append({'name':r[0],'count':r[1],'size':r[2]})
    except Exception as e:
        return JsonResponse({'status':'error','message':str(e)})
    return JsonResponse({'status':'success','tables':tables})


@admin_required
@require_POST
@csrf_protect
def clear_cache(request):
    results = []
    try:
        from django.core.cache import cache
        cache.clear()
        results.append('Django cache ✅')
    except Exception as e:
        results.append(f'Django cache خطأ: {e}')
    try:
        import redis
        r = redis.Redis(
            host=getattr(django_settings,'REDIS_HOST','localhost'),
            port=getattr(django_settings,'REDIS_PORT',6379))
        r.flushdb()
        results.append('Redis ✅')
    except Exception:
        results.append('Redis: غير متاح')
    log_activity(request.user,'UPDATE',description='مسح الكاش',request=request)
    return JsonResponse({'status':'success','message':' | '.join(results)})


@admin_required
@require_POST
@csrf_protect
def optimize_db(request):
    try:
        engine = connection.settings_dict.get('ENGINE','')
        with connection.cursor() as cursor:
            if 'sqlite' in engine:
                cursor.execute('VACUUM')
                cursor.execute('ANALYZE')
            elif 'postgresql' in engine:
                connection.set_autocommit(True)
                cursor.execute('VACUUM ANALYZE')
            elif 'mysql' in engine:
                cursor.execute('SHOW TABLES')
                for (t,) in cursor.fetchall()[:20]:
                    try: cursor.execute(f'OPTIMIZE TABLE `{t}`')
                    except Exception: continue
        log_activity(request.user,'UPDATE',description='تحسين قاعدة البيانات',request=request)
        return JsonResponse({'status':'success','message':'تم تحسين قاعدة البيانات ✅'})
    except Exception as e:
        return JsonResponse({'status':'error','message':str(e)})


# ══════════════════════════════════════════════
# النسخ الاحتياطية
# ══════════════════════════════════════════════
@admin_required
@require_GET
def backup_data(request):
    fmt = request.GET.get('format','json').strip()
    if fmt == 'json':   return _backup_json(request)
    if fmt == 'excel':  return _backup_excel(request)
    if fmt == 'sql':    return _backup_sql(request)
    return HttpResponse('صيغة غير مدعومة', status=400)


def _get_all_models():
    from django.apps import apps
    result = []
    for model in apps.get_models():
        if model._meta.app_label in ['admin','contenttypes','sessions','auth']:
            continue
        result.append(model)
    return result


def _backup_json(request):
    from django.core import serializers
    data = {}
    for Model in _get_all_models():
        try:
            data[f'{Model._meta.app_label}.{Model.__name__}'] = json.loads(
                serializers.serialize('json', Model.objects.all()))
        except Exception:
            pass
    output   = json.dumps(data, ensure_ascii=False, indent=2)
    response = HttpResponse(output, content_type='application/json; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="backup_{timezone.now().strftime("%Y%m%d_%H%M%S")}.json"'
    log_activity(request.user,'EXPORT',description='تصدير نسخة احتياطية JSON',request=request)
    return response


def _backup_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return HttpResponse('openpyxl غير مثبّتة', status=500)

    PURPLE = '7C3AED'
    GREEN  = '1A7A4A'
    thin   = Side(style='thin', color='E2E8F0')
    border = openpyxl.styles.Border(left=thin,right=thin,top=thin,bottom=thin)
    C_ALG  = Alignment(horizontal='center',vertical='center',wrap_text=True)
    R_ALG  = Alignment(horizontal='right', vertical='center',wrap_text=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for Model in _get_all_models():
        try:
            fields = [f for f in Model._meta.concrete_fields]
            if not fields: continue
            sheet_name = f'{Model._meta.app_label}_{Model.__name__}'[:31]
            ws = wb.create_sheet(title=sheet_name)
            ws.sheet_view.rightToLeft = True

            # رأس الجدول
            for i, f in enumerate(fields, 1):
                c = ws.cell(row=1, column=i, value=f.verbose_name or f.name)
                c.font      = Font(bold=True, color='FFFFFF', size=10)
                c.fill      = PatternFill('solid', fgColor=PURPLE)
                c.alignment = C_ALG
                c.border    = border
            ws.row_dimensions[1].height = 24

            # البيانات
            for r_idx, obj in enumerate(Model.objects.all()[:3000], 2):
                for c_idx, f in enumerate(fields, 1):
                    try:
                        val = getattr(obj, f.attname, '')
                        if hasattr(val,'isoformat'): val = str(val)
                        elif val is None: val = ''
                        cell = ws.cell(row=r_idx, column=c_idx, value=str(val))
                        cell.alignment = R_ALG
                        cell.border    = border
                        if r_idx % 2 == 0:
                            cell.fill = PatternFill('solid', fgColor='F5F3FF')
                    except Exception:
                        ws.cell(row=r_idx, column=c_idx, value='—')
                ws.row_dimensions[r_idx].height = 18

            # عرض الأعمدة
            for i, f in enumerate(fields, 1):
                col_letter = openpyxl.utils.get_column_letter(i)
                ws.column_dimensions[col_letter].width = max(14, min(40, len(str(f.verbose_name or f.name)) + 6))
        except Exception:
            continue

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="backup_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    log_activity(request.user,'EXPORT',description='تصدير نسخة احتياطية Excel',request=request)
    return response


def _backup_sql(request):
    lines = [f'-- Backup: {timezone.now().isoformat()}\n-- Platform: ruhamaa\n\n']
    for Model in _get_all_models():
        try:
            table  = Model._meta.db_table
            fields = [f for f in Model._meta.concrete_fields]
            f_names= [f.column for f in fields]
            lines.append(f'\n-- ══ {Model._meta.app_label}.{Model.__name__} ══\n')
            for obj in Model.objects.all()[:3000]:
                vals = []
                for f in fields:
                    v = getattr(obj, f.attname, None)
                    if v is None:           vals.append('NULL')
                    elif isinstance(v,bool): vals.append('1' if v else '0')
                    elif isinstance(v,(int,float)): vals.append(str(v))
                    else: vals.append("'"+str(v).replace("'","''")+"'")
                lines.append(f"INSERT INTO `{table}` ({','.join(f_names)}) VALUES ({','.join(vals)});\n")
        except Exception:
            continue
    content  = ''.join(lines)
    response = HttpResponse(content, content_type='application/sql')
    response['Content-Disposition'] = f'attachment; filename="backup_{timezone.now().strftime("%Y%m%d_%H%M%S")}.sql"'
    log_activity(request.user,'EXPORT',description='تصدير نسخة احتياطية SQL',request=request)
    return response


# ══════════════════════════════════════════════
# البريد
# ══════════════════════════════════════════════
@admin_required
@require_POST
@csrf_protect
def test_email(request):
    to_email = request.POST.get('email','').strip()
    if not to_email:
        return JsonResponse({'status':'error','message':'البريد مطلوب'})
    try:
        send_mail(
            subject   = 'اختبار البريد — منصة رُحَمَاء 🌿',
            message   = (
                'السلام عليكم،\n\n'
                'هذا بريد تجريبي من منصة رُحَمَاء للتأكد من صحة إعدادات البريد.\n\n'
                'إذا وصلتك هذه الرسالة فإن إعدادات البريد تعمل بشكل صحيح ✅\n\n'
                'فريق منصة رُحَمَاء'
            ),
            from_email = django_settings.DEFAULT_FROM_EMAIL,
            recipient_list=[to_email],
            fail_silently=False,
        )
        return JsonResponse({'status':'success','message':f'تم الإرسال إلى {to_email} ✅'})
    except Exception as e:
        return JsonResponse({'status':'error','message':str(e)})


@admin_required
@require_POST
@csrf_protect
def toggle_notifications(request):
    current = _gs('notifications_enabled','true') == 'true'
    new_val = not current
    _ss('notifications_enabled','true' if new_val else 'false', request.user)
    return JsonResponse({'status':'success','enabled':new_val,
                         'message':f'{"تم تفعيل" if new_val else "تم إيقاف"} الإشعارات'})


# ══════════════════════════════════════════════
# الأمان
# ══════════════════════════════════════════════
@admin_required
@require_POST
@csrf_protect
def update_security(request):
    if 'max_login_attempts' in request.POST:
        _ss('max_login_attempts', request.POST.get('max_login_attempts','5'), request.user)
    if 'session_timeout' in request.POST:
        _ss('session_timeout', request.POST.get('session_timeout','60'), request.user)
    log_activity(request.user,'UPDATE',description='تعديل إعدادات الأمان',request=request)
    return JsonResponse({'status':'success','message':'تم حفظ إعدادات الأمان ✅'})


@admin_required
@require_POST
@csrf_protect
def manage_blocked_ips(request):
    action = request.POST.get('action','').strip()
    ip     = request.POST.get('ip','').strip()
    try:
        blocked = json.loads(_gs('blocked_ips','[]') or '[]')
        if action == 'add' and ip and ip not in blocked:
            blocked.append(ip)
            log_activity(request.user,'UPDATE',description=f'حظر IP: {ip}',request=request)
        elif action == 'remove' and ip in blocked:
            blocked.remove(ip)
            log_activity(request.user,'UPDATE',description=f'رفع حظر IP: {ip}',request=request)
        _ss('blocked_ips', json.dumps(blocked), request.user)
        return JsonResponse({'status':'success','blocked_ips':blocked})
    except Exception as e:
        return JsonResponse({'status':'error','message':str(e)})


# ══════════════════════════════════════════════
# معلومات النظام
# ══════════════════════════════════════════════
@admin_required
@require_GET
def system_info(request):
    try:
        import django, platform, sys
        from django.utils import timezone
        from django.db import connection

        # الجلسات النشطة حسب النوع
        try:
            from django.contrib.sessions.models import Session
            from django.utils import timezone as tz

            active_sessions_raw = Session.objects.filter(expire_date__gte=tz.now())
            active_sessions = active_sessions_raw.count()

            sessions_by_type = {
                'admin': 0,
                'orphan': 0,
                'family': 0,
                'special': 0,
                'sponsor': 0,
                'unknown': 0
            }

            for session in active_sessions_raw:
                try:
                    data = session.get_decoded()
                    user_id = data.get('_auth_user_id')

                    if user_id:
                        user = CustomUser.objects.filter(pk=user_id).values('user_type').first()
                        utype = user['user_type'] if user else 'unknown'

                        if utype in sessions_by_type:
                            sessions_by_type[utype] += 1
                        else:
                            sessions_by_type['unknown'] += 1

                except Exception:
                    sessions_by_type['unknown'] += 1

        except Exception:
            active_sessions = '—'
            sessions_by_type = {}

        # الذاكرة
        try:
            import psutil
            mem = psutil.virtual_memory()
            mem_used  = f'{mem.percent}%'
            mem_total = f'{round(mem.total/1024**3,1)} GB'
        except Exception:
            mem_used = mem_total = '—'

        # آخر السجلات
        last_logs = ActivityLog.objects.select_related('user').order_by('-created_at')[:5]

        return JsonResponse({
            'status':          'success',
            'django_version':  django.__version__,
            'python_version':  f'{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}',
            'platform':        platform.system() + ' ' + platform.release(),
            'db_engine':       connection.settings_dict.get('ENGINE','').split('.')[-1],
            'db_name':         str(connection.settings_dict.get('NAME','')),
            'active_sessions': active_sessions,
            'sessions_by_type': sessions_by_type,
            'mem_used':        mem_used,
            'mem_total':       mem_total,
            'total_users':     CustomUser.objects.count(),
            'total_logs':      ActivityLog.objects.count(),
            'last_logs': [{
                'user':   _uname(l.user),
                'action': l.action,
                'desc':   l.description,
                'time':   timezone.localtime(l.created_at).strftime('%Y/%m/%d %H:%M'),
            } for l in last_logs],
        })

    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })

@admin_required
@require_POST
@csrf_protect
def delete_admin(request):
    admin_id = request.POST.get('admin_id','').strip()
    try:
        admin = CustomUser.objects.get(pk=admin_id, user_type='admin')
        if admin == request.user:
            return JsonResponse({'status':'error','message':'لا يمكنك حذف حسابك الخاص'})
        if admin.is_superuser:
            return JsonResponse({'status':'error','message':'لا يمكن حذف السوبر يوزر'})
        name = _uname(admin)
        admin.delete()
        log_activity(request.user,'DELETE',
                     description=f'حذف حساب الأدمن: {name}',
                     request=request)
        return JsonResponse({'status':'success','message':f'تم حذف {name} ✅'})
    except CustomUser.DoesNotExist:
        return JsonResponse({'status':'error','message':'غير موجود'})