"""
admin_panel/views/reports.py
لوحة التقارير الشهرية الشاملة
"""

from datetime import date, timedelta
from dateutil.relativedelta import relativedelta

from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_GET
from django.db.models import Q, Sum, Count, Avg
from django.utils import timezone

from core.models import CustomUser, ActivityLog, Notification, Payment, Aid, UserNote
from .decorators import admin_required


# ══════════════════════════════════════════════
#  مساعدات
# ══════════════════════════════════════════════

def _date_range(period, date_from=None, date_to=None):
    """حساب نطاق التاريخ حسب الفترة"""
    today = date.today()
    if period == 'week':
        return today - timedelta(days=7), today
    elif period == 'month':
        return today.replace(day=1), today
    elif period == 'quarter':
        return today - relativedelta(months=3), today
    elif period == 'year':
        return today.replace(month=1, day=1), today
    elif period == 'custom' and date_from and date_to:
        from datetime import datetime
        return (
            datetime.strptime(date_from, '%Y-%m-%d').date(),
            datetime.strptime(date_to,   '%Y-%m-%d').date(),
        )
    else:  # all
        return None, None


def _filter_by_sponsor(qs, sponsor_id, user_field='beneficiary'):
    """فلترة حسب الكافل"""
    if not sponsor_id:
        return qs
    from sponsor.models import SponsorProfile
    try:
        sp = SponsorProfile.objects.get(pk=sponsor_id)
        from beneficiary.models import OrphanForm, FamilyForm, SpecialNeedsForm
        bene_ids = set()
        for Model in [OrphanForm, FamilyForm, SpecialNeedsForm]:
            bene_ids |= set(
                Model.objects.filter(sponsor=sp).values_list('user_id', flat=True)
            )
        return qs.filter(**{f'{user_field}__in': list(bene_ids)})
    except SponsorProfile.DoesNotExist:
        return qs.none()

def _apply_date_filter(qs, d_from, d_to, date_field='date_joined'):
    from django.utils import timezone as tz
    import datetime
    if d_from:
        dt_from = tz.make_aware(datetime.datetime.combine(d_from, datetime.time.min))
        qs = qs.filter(**{f'{date_field}__gte': dt_from})
    if d_to:
        dt_to = tz.make_aware(datetime.datetime.combine(d_to, datetime.time.max))
        qs = qs.filter(**{f'{date_field}__lte': dt_to})
    return qs


# ══════════════════════════════════════════════
#  صفحة التقارير
# ══════════════════════════════════════════════

@admin_required
def reports_view(request):
    from sponsor.models import SponsorProfile
    from core.models import Notification

    sponsors = SponsorProfile.objects.select_related('user').filter(
        user__is_approved=True
    ).order_by('user__first_name')

    notif_count = Notification.objects.filter(
        recipient=request.user, is_read=False
    ).count()

    return render(request, 'admin_panel/reports.html', {
        'sponsors':    sponsors,
        'notif_count': notif_count,
    })


# ══════════════════════════════════════════════
#  API — بيانات التقارير
# ══════════════════════════════════════════════

@admin_required
@require_GET
def reports_data(request):
    period     = request.GET.get('period',     'month')
    date_from  = request.GET.get('date_from',  '')
    date_to    = request.GET.get('date_to',    '')
    sponsor_id = request.GET.get('sponsor_id', '')

    d_from, d_to = _date_range(period, date_from, date_to)

    # ── المستفيدون ──
    users_qs = CustomUser.objects.filter(is_approved=True).exclude(user_type='admin')
    users_date_qs = _apply_date_filter(users_qs, d_from, d_to, 'date_joined')

    total_users   = users_qs.count()
    new_users     = users_date_qs.count()
    orphans_total = users_qs.filter(user_type='orphan').count()
    families_total= users_qs.filter(user_type='family').count()
    specials_total= users_qs.filter(user_type='special').count()
    sponsors_total= users_qs.filter(user_type='sponsor').count()
    new_orphans   = users_date_qs.filter(user_type='orphan').count()
    new_families  = users_date_qs.filter(user_type='family').count()
    new_specials  = users_date_qs.filter(user_type='special').count()
    new_sponsors  = users_date_qs.filter(user_type='sponsor').count()
    disabled_users= users_qs.filter(is_active=False).count()

    # ── الكفالة ──
    from beneficiary.models import OrphanForm, FamilyForm, SpecialNeedsForm
    sponsored_orphans  = OrphanForm.objects.filter(sponsor__isnull=False).count()
    sponsored_families = FamilyForm.objects.filter(sponsor__isnull=False).count()
    sponsored_specials = SpecialNeedsForm.objects.filter(sponsor__isnull=False).count()
    total_sponsored    = sponsored_orphans + sponsored_families + sponsored_specials
    total_beneficiaries= orphans_total + families_total + specials_total
    unsponsored        = total_beneficiaries - total_sponsored

    # ── المدفوعات ──
    pays_qs = Payment.objects.all()
    if sponsor_id:
        pays_qs = _filter_by_sponsor(pays_qs, sponsor_id, 'beneficiary')
    pays_date_qs = _apply_date_filter(pays_qs, d_from, d_to, 'date')

    total_pays_ils  = pays_date_qs.aggregate(s=Sum('amount_ils'))['s'] or 0
    total_pays_usd  = pays_date_qs.aggregate(s=Sum('amount_usd'))['s'] or 0
    paid_count      = pays_date_qs.filter(status='paid').count()
    pending_count   = pays_date_qs.filter(status='pending').count()
    late_count      = pays_date_qs.filter(status='late').count()
    paid_ils        = pays_date_qs.filter(status='paid').aggregate(s=Sum('amount_ils'))['s'] or 0
    pending_ils     = pays_date_qs.filter(status='pending').aggregate(s=Sum('amount_ils'))['s'] or 0
    late_ils        = pays_date_qs.filter(status='late').aggregate(s=Sum('amount_ils'))['s'] or 0
    avg_pay_ils     = pays_date_qs.aggregate(a=Avg('amount_ils'))['a'] or 0

    # مصادر الدفع
    sponsor_pays = pays_date_qs.filter(paid_by='sponsor').aggregate(s=Sum('amount_ils'))['s'] or 0
    admin_pays   = pays_date_qs.filter(paid_by='admin').aggregate(s=Sum('amount_ils'))['s'] or 0
    ext_pays     = pays_date_qs.filter(paid_by='external').aggregate(s=Sum('amount_ils'))['s'] or 0

    # ── المساعدات ──
    aids_qs = Aid.objects.all()
    if sponsor_id:
        aids_qs = _filter_by_sponsor(aids_qs, sponsor_id, 'beneficiary')
    aids_date_qs = _apply_date_filter(aids_qs, d_from, d_to, 'date')

    total_aids  = aids_date_qs.count()
    food_aids   = aids_date_qs.filter(aid_type='food').count()
    medical_aids= aids_date_qs.filter(aid_type='medical').count()
    fin_aids    = aids_date_qs.filter(aid_type='financial').count()
    cloth_aids  = aids_date_qs.filter(aid_type='clothing').count()
    furn_aids   = aids_date_qs.filter(aid_type='furniture').count()
    edu_aids    = aids_date_qs.filter(aid_type='education').count()
    other_aids  = aids_date_qs.filter(aid_type='other').count()

    # ── سجل النشاط ──
    logs_qs     = ActivityLog.objects.all()
    logs_date_qs= _apply_date_filter(logs_qs, d_from, d_to, 'created_at')
    total_logs  = logs_date_qs.count()
    login_count = logs_date_qs.filter(action='LOGIN').count()

    # ── بيانات الرسوم البيانية ──

    # المدفوعات الشهرية (آخر 12 شهر)
    monthly_pays = []
    for i in range(11, -1, -1):
        m_date   = date.today() - relativedelta(months=i)
        m_total  = Payment.objects.filter(
            date__year=m_date.year,
            date__month=m_date.month,
        )
        if sponsor_id:
            m_total = _filter_by_sponsor(m_total, sponsor_id, 'beneficiary')
        monthly_pays.append({
            'month': f'{m_date.year}/{m_date.month:02d}',
            'ils':   float(m_total.aggregate(s=Sum('amount_ils'))['s'] or 0),
            'count': m_total.count(),
        })

    # المستفيدون الجدد شهرياً (آخر 12 شهر)
    monthly_new_users = []
    for i in range(11, -1, -1):
        m_date = date.today() - relativedelta(months=i)
        count  = CustomUser.objects.filter(
            is_approved=True,
            date_joined__year=m_date.year,
            date_joined__month=m_date.month,
        ).exclude(user_type='admin').count()
        monthly_new_users.append({
            'month': f'{m_date.year}/{m_date.month:02d}',
            'count': count,
        })

    # توزيع المستفيدين
    distribution = {
        'orphan':  orphans_total,
        'family':  families_total,
        'special': specials_total,
        'sponsor': sponsors_total,
    }

    # أفضل 5 كفلاء من حيث المدفوعات
    from sponsor.models import SponsorProfile
    top_sponsors = []
    for sp in SponsorProfile.objects.select_related('user').all():
        total = Payment.objects.filter(
            sponsor=sp.user
        ).aggregate(s=Sum('amount_ils'))['s'] or 0
        if total > 0:
            top_sponsors.append({
                'name':  sp.user.get_full_name(),
                'total': float(total),
            })
    top_sponsors = sorted(top_sponsors, key=lambda x: x['total'], reverse=True)[:5]

    # توزيع المساعدات حسب النوع
    aids_distribution = {
        'food':      food_aids,
        'medical':   medical_aids,
        'financial': fin_aids,
        'clothing':  cloth_aids,
        'furniture': furn_aids,
        'education': edu_aids,
        'other':     other_aids,
    }

    return JsonResponse({
        'period': {
            'from': str(d_from) if d_from else 'الكل',
            'to':   str(d_to)   if d_to   else 'الكل',
        },

        # ملخص عام
        'summary': {
            'total_users':        total_users,
            'new_users':          new_users,
            'disabled_users':     disabled_users,
            'total_beneficiaries':total_beneficiaries,
            'total_sponsored':    total_sponsored,
            'unsponsored':        unsponsored,
            'total_pays_ils':     float(round(total_pays_ils, 2)),
            'total_pays_usd':     float(round(total_pays_usd, 2)),
            'total_aids':         total_aids,
            'avg_pay_ils':        float(round(avg_pay_ils, 2)),
        },

        # المستفيدون
        'beneficiaries': {
            'orphans_total':   orphans_total,
            'families_total':  families_total,
            'specials_total':  specials_total,
            'sponsors_total':  sponsors_total,
            'new_orphans':     new_orphans,
            'new_families':    new_families,
            'new_specials':    new_specials,
            'new_sponsors':    new_sponsors,
            'sponsored':       total_sponsored,
            'unsponsored':     unsponsored,
        },

        # المدفوعات
        'payments': {
            'total_ils':   float(round(total_pays_ils, 2)),
            'total_usd':   float(round(total_pays_usd, 2)),
            'paid_count':  paid_count,
            'paid_ils':    float(round(paid_ils, 2)),
            'pending_count':pending_count,
            'pending_ils': float(round(pending_ils, 2)),
            'late_count':  late_count,
            'late_ils':    float(round(late_ils, 2)),
            'sponsor_pays':float(round(sponsor_pays, 2)),
            'admin_pays':  float(round(admin_pays, 2)),
            'ext_pays':    float(round(ext_pays, 2)),
            'avg_ils':     float(round(avg_pay_ils, 2)),
        },

        # المساعدات
        'aids': {
            'total':    total_aids,
            'types':    aids_distribution,
        },

        # النشاط
        'activity': {
            'total_logs':  total_logs,
            'login_count': login_count,
        },

        # الرسوم البيانية
        'charts': {
            'monthly_pays':      monthly_pays,
            'monthly_new_users': monthly_new_users,
            'distribution':      distribution,
            'top_sponsors':      top_sponsors,
            'aids_distribution': aids_distribution,
        },
    })


# ══════════════════════════════════════════════
#  تصدير Excel
# ══════════════════════════════════════════════

@admin_required
@require_GET
def export_report_excel(request):
    period     = request.GET.get('period',     'month')
    date_from  = request.GET.get('date_from',  '')
    date_to    = request.GET.get('date_to',    '')
    sponsor_id = request.GET.get('sponsor_id', '')

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return HttpResponse('openpyxl غير مثبّتة', status=500)

    d_from, d_to = _date_range(period, date_from, date_to)

    wb  = openpyxl.Workbook()
    PURPLE = '7C3AED'
    GREEN  = '1A7A4A'
    ORANGE = 'B45309'
    BLUE   = '2B6CB0'
    thin   = Side(style='thin', color='E2E8F0')
    border = openpyxl.styles.Border(left=thin,right=thin,top=thin,bottom=thin)

    def _hdr(ws, headers, color):
        ws.sheet_view.rightToLeft = True
        for i, h in enumerate(headers, 1):
            c           = ws.cell(row=1, column=i, value=h)
            c.font      = Font(bold=True, color='FFFFFF')
            c.fill      = PatternFill('solid', fgColor=color)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border    = border
        ws.row_dimensions[1].height = 26

    def _row(ws, r_idx, values, alt='F5F3FF'):
        for c_idx, val in enumerate(values, 1):
            cell           = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border    = border
            if r_idx % 2 == 0:
                cell.fill = PatternFill('solid', fgColor=alt)
        ws.row_dimensions[r_idx].height = 20

    period_label = f'{d_from} → {d_to}' if d_from else 'كل الوقت'

    # ── ورقة 1: الملخص العام ──
    ws1       = wb.active
    ws1.title = 'الملخص العام'
    _hdr(ws1, ['المؤشر', 'القيمة'], PURPLE)

    users_qs     = CustomUser.objects.filter(is_approved=True).exclude(user_type='admin')
    users_date_qs= _apply_date_filter(users_qs, d_from, d_to, 'date_joined')
    pays_qs      = Payment.objects.all()
    pays_date_qs = _apply_date_filter(pays_qs, d_from, d_to, 'date')
    aids_date_qs = _apply_date_filter(Aid.objects.all(), d_from, d_to, 'date')

    if sponsor_id:
        pays_date_qs = _filter_by_sponsor(pays_date_qs, sponsor_id, 'beneficiary')
        aids_date_qs = _filter_by_sponsor(aids_date_qs, sponsor_id, 'beneficiary')

    total_ils = pays_date_qs.aggregate(s=Sum('amount_ils'))['s'] or 0
    total_usd = pays_date_qs.aggregate(s=Sum('amount_usd'))['s'] or 0

    summary_rows = [
        ('الفترة الزمنية',          period_label),
        ('إجمالي المستخدمين',       users_qs.count()),
        ('مستفيدون جدد في الفترة',  users_date_qs.count()),
        ('إجمالي الأيتام',          users_qs.filter(user_type='orphan').count()),
        ('إجمالي الأسر',            users_qs.filter(user_type='family').count()),
        ('إجمالي ذوو الاحتياجات',   users_qs.filter(user_type='special').count()),
        ('إجمالي الكفلاء',          users_qs.filter(user_type='sponsor').count()),
        ('إجمالي المدفوعات ₪',      str(round(total_ils, 2))),
        ('إجمالي المدفوعات $',      str(round(total_usd, 2))),
        ('إجمالي المساعدات',        aids_date_qs.count()),
        ('مدفوعات مكتملة',          pays_date_qs.filter(status='paid').count()),
        ('مدفوعات معلّقة',           pays_date_qs.filter(status='pending').count()),
        ('مدفوعات متأخرة',          pays_date_qs.filter(status='late').count()),
    ]
    for r_idx, (k, v) in enumerate(summary_rows, 2):
        _row(ws1, r_idx, [k, v])
    ws1.column_dimensions['A'].width = 28
    ws1.column_dimensions['B'].width = 20

    # ── ورقة 2: المدفوعات ──
    ws2       = wb.create_sheet(title='المدفوعات')
    _hdr(ws2, ['التاريخ','المبلغ ₪','المبلغ $','مصدر الدفع','المستفيد','الحالة','ملاحظة'], GREEN)
    STATUS_MAP  = {'paid':'مدفوعة','pending':'معلّقة','late':'متأخرة'}
    PAID_BY_MAP = {'sponsor':'كافل','admin':'إدارة','external':'جهة خارجية'}
    for r_idx, p in enumerate(pays_date_qs.select_related('beneficiary','created_by','sponsor').order_by('-date'), 2):
        _row(ws2, r_idx, [
            str(p.date), str(p.amount_ils), str(p.amount_usd),
            PAID_BY_MAP.get(p.paid_by, p.paid_by),
            p.beneficiary.get_full_name() if p.beneficiary else '—',
            STATUS_MAP.get(p.status, p.status),
            p.note or '',
        ], 'F0FDF4')
    for col, w in zip(['A','B','C','D','E','F','G'], [14,14,14,14,24,12,20]):
        ws2.column_dimensions[col].width = w

    # ── ورقة 3: المساعدات ──
    ws3       = wb.create_sheet(title='المساعدات')
    _hdr(ws3, ['التاريخ','الاسم','النوع','الكمية','الجهة','المستفيد','معتمد الطلب','ملاحظة'], ORANGE)
    AID_TYPES = {'food':'غذائية','medical':'طبية','financial':'مالية','clothing':'ملابس','furniture':'أثاث','education':'تعليمية','other':'أخرى'}
    for r_idx, a in enumerate(aids_date_qs.select_related('beneficiary','created_by').order_by('-date'), 2):
        created_by = (a.created_by.get_full_name() or a.created_by.username) if a.created_by else 'النظام'
        _row(ws3, r_idx, [
            str(a.date), a.name, AID_TYPES.get(a.aid_type, a.aid_type),
            a.quantity, a.provider,
            a.beneficiary.get_full_name() if a.beneficiary else '—',
            created_by, a.note or '',
        ], 'FEF3C7')
    for col, w in zip(['A','B','C','D','E','F','G','H'], [14,22,12,10,20,22,18,20]):
        ws3.column_dimensions[col].width = w

    # ── ورقة 4: المستفيدون الجدد ──
    ws4       = wb.create_sheet(title='المستفيدون الجدد')
    _hdr(ws4, ['رقم الاستمارة','الاسم الكامل','النوع','رقم الهوية','تاريخ التسجيل','الكافل'], BLUE)
    TYPE_NAMES = {'orphan':'يتيم','family':'أسرة','special':'ذوو احتياجات','sponsor':'كافل'}
    for r_idx, u in enumerate(users_date_qs.order_by('-date_joined'), 2):
        from beneficiary.models import OrphanForm, FamilyForm, SpecialNeedsForm
        model_map = {'orphan': OrphanForm, 'family': FamilyForm, 'special': SpecialNeedsForm}
        Model      = model_map.get(u.user_type)
        sponsor_name = '—'
        if Model:
            try:
                form = Model.objects.get(user=u)
                if hasattr(form, 'sponsor') and form.sponsor:
                    sponsor_name = form.sponsor.user.get_full_name()
            except Exception:
                pass
        _row(ws4, r_idx, [
            u.registration_number or '—',
            u.get_full_name(),
            TYPE_NAMES.get(u.user_type, u.user_type),
            u.id_number or '—',
            str(u.date_joined.date()),
            sponsor_name,
        ], 'EFF6FF')
    for col, w in zip(['A','B','C','D','E','F'], [18,26,14,14,14,22]):
        ws4.column_dimensions[col].width = w

    from urllib.parse import quote
    import io
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    fname = f'report_{period}_{date.today()}.xlsx'
    resp  = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(fname)}"
    return resp


# ══════════════════════════════════════════════
#  تصدير PDF
# ══════════════════════════════════════════════

@admin_required
@require_GET
def export_report_pdf(request):
    from urllib.parse import quote
    from django.template.loader import render_to_string
    import weasyprint
    import logging
    logging.getLogger('weasyprint').setLevel(logging.ERROR)
    logging.getLogger('fontTools').setLevel(logging.ERROR)

    period     = request.GET.get('period',     'month')
    date_from  = request.GET.get('date_from',  '')
    date_to    = request.GET.get('date_to',    '')
    sponsor_id = request.GET.get('sponsor_id', '')

    d_from, d_to = _date_range(period, date_from, date_to)

    # جمع البيانات
    users_qs     = CustomUser.objects.filter(is_approved=True).exclude(user_type='admin')
    users_date_qs= _apply_date_filter(users_qs, d_from, d_to, 'date_joined')
    pays_date_qs = _apply_date_filter(Payment.objects.all(), d_from, d_to, 'date')
    aids_date_qs = _apply_date_filter(Aid.objects.all(), d_from, d_to, 'date')

    if sponsor_id:
        pays_date_qs = _filter_by_sponsor(pays_date_qs, sponsor_id, 'beneficiary')
        aids_date_qs = _filter_by_sponsor(aids_date_qs, sponsor_id, 'beneficiary')

    context = {
        'request':    request,
        'period':     period,
        'date_from':  str(d_from) if d_from else 'الكل',
        'date_to':    str(d_to)   if d_to   else 'الكل',
        'printed_by': request.user.get_full_name() or request.user.username,
        'summary': {
            'total_users':   users_qs.count(),
            'new_users':     users_date_qs.count(),
            'orphans':       users_qs.filter(user_type='orphan').count(),
            'families':      users_qs.filter(user_type='family').count(),
            'specials':      users_qs.filter(user_type='special').count(),
            'sponsors':      users_qs.filter(user_type='sponsor').count(),
            'total_pays_ils':str(round(pays_date_qs.aggregate(s=Sum('amount_ils'))['s'] or 0, 2)),
            'total_pays_usd':str(round(pays_date_qs.aggregate(s=Sum('amount_usd'))['s'] or 0, 2)),
            'paid_ils':      str(round(pays_date_qs.filter(status='paid').aggregate(s=Sum('amount_ils'))['s'] or 0, 2)),
            'pending_ils':   str(round(pays_date_qs.filter(status='pending').aggregate(s=Sum('amount_ils'))['s'] or 0, 2)),
            'late_ils':      str(round(pays_date_qs.filter(status='late').aggregate(s=Sum('amount_ils'))['s'] or 0, 2)),
            'total_aids':    aids_date_qs.count(),
        },
        'payments': pays_date_qs.select_related('beneficiary').order_by('-date')[:50],
        'aids':     aids_date_qs.select_related('beneficiary','created_by').order_by('-date')[:50],
    }

    html_content = render_to_string('admin_panel/report_pdf.html', context)
    pdf_file     = weasyprint.HTML(
        string=html_content,
        base_url=request.build_absolute_uri('/')
    ).write_pdf()

    fname = f'report_{period}_{date.today()}.pdf'
    resp  = HttpResponse(pdf_file, content_type='application/pdf')
    resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(fname)}"
    return resp
