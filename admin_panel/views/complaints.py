"""
complaints.py — فيوز الشكاوى والتواصل
"""

from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_GET, require_POST
from django.views.decorators.csrf import csrf_protect
from django.utils import timezone
from django.core.mail import send_mail
from django.conf import settings
from urllib.parse import quote
import re
import io

from core.models import Complaint, Notification
from core.utils import log_activity
from .decorators import admin_required


# ══════════════════════════════════════════════
# الصفحة الرئيسية
# ══════════════════════════════════════════════
@admin_required
def complaints_list(request):
    return render(request, 'admin_panel/complaints.html')


# ══════════════════════════════════════════════
# جلب البيانات بـ AJAX
# ══════════════════════════════════════════════
@admin_required
@require_GET
def complaints_data(request):
    complaints = Complaint.objects.order_by('-created_at')

    data = [{
        'id':          c.pk,
        'name':        c.name,
        'email':       c.email,
        'phone':       f'{c.phone_country or ""}{c.phone or ""}',
        'subject':     c.subject,
        'message':     c.message,
        'ip_address':  c.ip_address or '',
        'is_seen':     c.is_seen,
        'is_replied':  c.is_replied,
        'reply_text':  c.reply_text or '',
        'replied_by': c.replied_by.get_full_name() or c.replied_by.username if c.replied_by else '',
        'replied_at':  c.replied_at.isoformat() if c.replied_at else None,
        'created_at':  c.created_at.isoformat(),
        'priority':    getattr(c, 'priority', '') or '',
    } for c in complaints]

    return JsonResponse({
        'complaints': data,
        'counts':     _get_counts(),
    })


# ══════════════════════════════════════════════
# تفاصيل شكوى واحدة
# ══════════════════════════════════════════════
@admin_required
@require_GET
def complaint_detail(request):
    comp_id = request.GET.get('id', '').strip()
    try:
        c = Complaint.objects.get(pk=comp_id)
    except Complaint.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'غير موجود'})

    # تحديد كمشاهدة تلقائياً
    if not c.is_seen:
        c.is_seen = True
        c.save(update_fields=['is_seen'])

    # بناء Timeline
    events = []
    events.append({
        'icon':  '📝',
        'color': '#7c3aed',
        'title': 'استلام الشكوى',
        'time':  _fmt_dt(c.created_at),
    })
    if c.is_seen:
        events.append({
            'icon':  '👁',
            'color': '#2b6cb0',
            'title': 'تم المشاهدة',
            'time':  _fmt_dt(c.replied_at) if c.replied_at else '—',
        })
    if c.is_replied:
        events.append({
            'icon':  '✉️',
            'color': '#1a7a4a',
            'title': f'تم الرد — بواسطة: {c.replied_by.get_full_name() if c.replied_by else "الإدارة"}',
            'time':  _fmt_dt(c.replied_at),
        })

    # الملاحظات الداخلية
    internal_notes = []
    try:
        from core.models import ComplaintNote
        notes = ComplaintNote.objects.filter(complaint=c).order_by('created_at')
        internal_notes = [{
            'text':       n.note,
            'admin':      n.admin.get_full_name() if n.admin else 'الإدارة',
            'created_at': _fmt_dt(n.created_at),
        } for n in notes]
    except Exception:
        pass

    return JsonResponse({
        'status': 'success',
        'complaint': {
            'id':          c.pk,
            'name':        c.name,
            'email':       c.email,
            'phone':       f'{c.phone_country or ""}{c.phone or ""}',
            'subject':     c.subject,
            'message':     c.message,
            'ip_address':  c.ip_address or '',
            'is_seen':     c.is_seen,
            'is_replied':  c.is_replied,
            'reply_text':  c.reply_text or '',
            'replied_by':  c.replied_by.get_full_name() if c.replied_by else '',
            'replied_at':  c.replied_at.isoformat() if c.replied_at else None,
            'created_at':  c.created_at.isoformat(),
            'priority':    getattr(c, 'priority', '') or '',
        },
        'events':         events,
        'internal_notes': internal_notes,
    })


# ══════════════════════════════════════════════
# تحديد كمشاهدة
# ══════════════════════════════════════════════
@admin_required
@require_POST
@csrf_protect
def mark_seen(request):
    comp_id = request.POST.get('complaint_id', '').strip()
    try:
        c         = Complaint.objects.get(pk=comp_id)
        c.is_seen = True
        c.save(update_fields=['is_seen'])
        return JsonResponse({'status': 'success', 'counts': _get_counts()})
    except Complaint.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'غير موجود'})


# ══════════════════════════════════════════════
# الرد على الشكوى
# ══════════════════════════════════════════════
@admin_required
@require_POST
@csrf_protect
def reply_complaint(request):
    comp_id       = request.POST.get('complaint_id',  '').strip()
    reply_text    = request.POST.get('reply_text',    '').strip()
    reply_subject = request.POST.get('reply_subject', '').strip()

    if not reply_text:
        return JsonResponse({'status': 'error', 'message': 'نص الرد مطلوب'})

    # منع XSS
    if re.search(r'<script|javascript:', reply_text, re.I):
        return JsonResponse({'status': 'error', 'message': 'مدخل غير صالح'}, status=400)

    try:
        c             = Complaint.objects.get(pk=comp_id)
        c.is_replied  = True
        c.is_seen     = True
        c.reply_text  = reply_text
        c.replied_by  = request.user
        c.replied_at  = timezone.now()
        c.save()

        # إرسال البريد
        subject = reply_subject or f'رد على شكواك — منصة رُحَمَاء'
        try:
            send_mail(
                subject    = subject,
                message    = (
                    f'السلام عليكم {c.name},\n\n'
                    f'شكراً لتواصلك معنا.\n\n'
                    f'ردنا على شكواك ({c.subject}):\n\n'
                    f'{reply_text}\n\n'
                    f'فريق منصة رُحَمَاء'
                ),
                from_email = settings.DEFAULT_FROM_EMAIL,
                recipient_list=[c.email],
                fail_silently=True,
            )
        except Exception:
            pass

        log_activity(
            request.user, 'MESSAGE',
            description  = f'رد على شكوى: {c.subject}',
            target_model = 'Complaint',
            target_id    = str(c.pk),
            request      = request,
        )

        return JsonResponse({
            'status':     'success',
            'message':    'تم الرد وإرساله على البريد ✅',
            'replied_by': request.user.get_full_name() or request.user.username,
            'replied_at': c.replied_at.isoformat(),
            'counts':     _get_counts(),
        })

    except Complaint.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'الشكوى غير موجودة'})


# ══════════════════════════════════════════════
# تحديد الأولوية
# ══════════════════════════════════════════════
@admin_required
@require_POST
@csrf_protect
def set_priority(request):
    comp_id  = request.POST.get('complaint_id', '').strip()
    priority = request.POST.get('priority',     '').strip()

    if priority not in ['high', 'medium', 'low']:
        return JsonResponse({'status': 'error', 'message': 'أولوية غير صحيحة'})

    try:
        c = Complaint.objects.get(pk=comp_id)
        # تحقق من وجود حقل priority في الموديل
        if hasattr(c, 'priority'):
            c.priority = priority
            c.save(update_fields=['priority'])
        return JsonResponse({'status': 'success', 'message': 'تم تحديد الأولوية'})
    except Complaint.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'غير موجود'})


# ══════════════════════════════════════════════
# ملاحظة داخلية
# ══════════════════════════════════════════════
@admin_required
@require_POST
@csrf_protect
def add_internal_note(request):
    comp_id = request.POST.get('complaint_id', '').strip()
    note    = request.POST.get('note',         '').strip()

    if not note:
        return JsonResponse({'status': 'error', 'message': 'الملاحظة فارغة'})

    try:
        c = Complaint.objects.get(pk=comp_id)
        try:
            from core.models import ComplaintNote
            ComplaintNote.objects.create(
                complaint  = c,
                note       = note,
                admin      = request.user,
                created_at = timezone.now(),
            )
            notes = ComplaintNote.objects.filter(complaint=c).order_by('created_at')
            notes_data = [{
                'text':       n.note,
                'admin': _admin_name(n.admin) if n.admin else 'الإدارة',
                'created_at': _fmt_dt(n.created_at),
            } for n in notes]
        except Exception:
            # إذا لم يكن الموديل موجوداً
            notes_data = [{'text': note, 'admin': request.user.get_full_name(), 'created_at': _fmt_dt(timezone.now())}]

        return JsonResponse({'status': 'success', 'notes': notes_data})
    except Complaint.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'غير موجود'})


# ══════════════════════════════════════════════
# حذف الشكوى
# ══════════════════════════════════════════════
@admin_required
@require_POST
@csrf_protect
def delete_complaint(request):
    comp_id = request.POST.get('complaint_id', '').strip()
    try:
        Complaint.objects.get(pk=comp_id).delete()
        return JsonResponse({'status': 'success', 'message': 'تم الحذف', 'counts': _get_counts()})
    except Complaint.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'غير موجود'})


# ══════════════════════════════════════════════
# تصدير Excel
# ══════════════════════════════════════════════
@admin_required
@require_GET
def export_complaints(request):
    status   = request.GET.get('status',   '').strip()
    q        = request.GET.get('q',        '').strip()
    priority = request.GET.get('priority', '').strip()

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return HttpResponse('openpyxl غير مثبّتة', status=500)

    qs = Complaint.objects.order_by('-created_at')
    if status == 'unseen':    qs = qs.filter(is_seen=False)
    elif status == 'unreplied': qs = qs.filter(is_replied=False)
    elif status == 'replied':   qs = qs.filter(is_replied=True)
    if q:        qs = qs.filter(subject__icontains=q) | qs.filter(name__icontains=q) | qs.filter(email__icontains=q)
    if priority and hasattr(Complaint, 'priority'): qs = qs.filter(priority=priority)

    PURPLE = '7C3AED'
    thin   = Side(style='thin', color='E2E8F0')
    border = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)
    C_ALG  = Alignment(horizontal='center', vertical='center')
    R_ALG  = Alignment(horizontal='right',  vertical='center')

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = 'الشكاوى'
    ws.sheet_view.rightToLeft = True

    headers = ['الاسم', 'البريد', 'الجوال', 'الموضوع', 'الرسالة', 'الأولوية', 'مشاهدة', 'تم الرد', 'الرد', 'المُجيب', 'التاريخ']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font      = Font(bold=True, color='FFFFFF')
        c.fill      = PatternFill('solid', fgColor=PURPLE)
        c.alignment = C_ALG
        c.border    = border
    ws.row_dimensions[1].height = 26

    PRIORITY_MAP = {'high': 'عالية', 'medium': 'متوسطة', 'low': 'منخفضة'}
    for r_idx, comp in enumerate(qs, 2):
        row = [
            comp.name,
            comp.email,
            f'{comp.phone_country or ""}{comp.phone or ""}',
            comp.subject,
            comp.message,
            PRIORITY_MAP.get(getattr(comp,'priority',''), '—'),
            '✅' if comp.is_seen    else '❌',
            '✅' if comp.is_replied else '❌',
            comp.reply_text or '',
            (comp.replied_by.get_full_name() or comp.replied_by.username) if comp.replied_by else '',
            comp.created_at.strftime('%Y/%m/%d %H:%M') if comp.created_at else '',
        ]
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = R_ALG
            cell.border    = border
            if r_idx % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='F5F3FF')
        ws.row_dimensions[r_idx].height = 20

    for col, w in zip(['A','B','C','D','E','F','G','H','I','J','K'], [18,22,14,24,40,12,10,10,40,18,16]):
        ws.column_dimensions[col].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    resp = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f"attachment; filename*=UTF-8''complaints.xlsx"
    return resp


# ══════════════════════════════════════════════
# مساعدات
# ══════════════════════════════════════════════
def _get_counts():
    return {
        'all':       Complaint.objects.count(),
        'unseen':    Complaint.objects.filter(is_seen=False).count(),
        'unreplied': Complaint.objects.filter(is_replied=False).count(),
        'replied':   Complaint.objects.filter(is_replied=True).count(),
    }

def _fmt_dt(dt):
    if not dt: return '—'
    try:
        from django.utils import timezone as tz
        return tz.localtime(dt).strftime('%Y/%m/%d %H:%M')
    except Exception:
        return str(dt)

def _admin_name(user):
    if not user: return '—'
    return user.get_full_name() or user.username or '—'



from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
@require_POST
def public_submit(request):
    """استقبال رسائل صفحة الصيانة"""
    Complaint.objects.create(
        name    = request.POST.get('name',    '').strip(),
        email   = request.POST.get('email',   '').strip(),
        phone   = request.POST.get('phone',   '').strip(),
        subject = request.POST.get('subject', 'رسالة من صفحة الصيانة'),
        message = request.POST.get('message', '').strip(),
        ip_address = request.META.get('REMOTE_ADDR',''),
    )
    return JsonResponse({'status': 'success'})