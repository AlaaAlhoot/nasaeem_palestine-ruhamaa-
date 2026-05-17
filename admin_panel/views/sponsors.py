

import io
from datetime import date
from urllib.parse import quote

from django.http import JsonResponse, HttpResponse
from django.shortcuts import render
from django.utils import timezone
from django.views.decorators.http import require_POST, require_GET
from django.views.decorators.csrf import csrf_protect
from django.db.models import Q, Sum, Count

import sponsor
from core.models import CustomUser, Notification, ActivityLog, UserNote, Payment
from core.utils import (log_activity, get_client_ip,
                        create_notification, send_email, fmt_dt)
from .decorators import admin_required


# ══════════════════════════════════════════════
#  مساعدات
# ══════════════════════════════════════════════

def _get_sponsor_profile(user):
    from sponsor.models import SponsorProfile
    try:
        return SponsorProfile.objects.get(user=user)
    except SponsorProfile.DoesNotExist:
        return None


def _build_sponsor_data(profile):
    """بناء بيانات كافل واحد للـ API"""
    from beneficiary.models import OrphanForm, SpecialNeedsForm, FamilyForm

    user = profile.user

    # عدد المكفولين
    orphan_count  = OrphanForm.objects.filter(sponsor=profile).count()
    family_count  = FamilyForm.objects.filter(sponsor=profile).count()
    special_count = SpecialNeedsForm.objects.filter(sponsor=profile).count()
    total_bene    = orphan_count + family_count + special_count

    # إجمالي المدفوعات
    total_paid = Payment.objects.filter(
        beneficiary__user_type__in=['orphan','family','special'],
        sponsor=profile.user,
    ).aggregate(s=Sum('amount_ils'))['s'] or 0

    total_paid_usd = Payment.objects.filter(
        sponsor=profile.user,
    ).aggregate(s=Sum('amount_usd'))['s'] or 0

    # المدفوعات الشهرية
    today      = timezone.now()
    month_paid = Payment.objects.filter(
        sponsor=profile.user,
        date__year=today.year,
        date__month=today.month,
    ).aggregate(s=Sum('amount_ils'))['s'] or 0

    photo = None
    if hasattr(profile, 'photo') and profile.photo:
        photo = profile.photo.url

    return {
        'id':              str(user.pk),
        'profile_id':      str(profile.pk),
        'full_name':       user.get_full_name(),
        'username':        user.username,
        'first_name':      user.first_name,
        'father_name':     user.father_name,
        'grand_name':      user.grand_name,
        'family_name':     user.family_name,
        'email':           user.email,
        'phone':           f'{user.phone_country}{user.phone}',
        'phone_country':   user.phone_country,
        'id_number':       user.id_number or '',
        'nationality':     user.nationality or '',
        'gender':          user.gender or '',
        'reg_number':      user.registration_number or '',
        'date_joined':     user.date_joined.isoformat(),
        'last_login':      user.last_login.isoformat() if user.last_login else None,
        'is_active':       user.is_active,
        'photo':           photo,
        'can_contact':     getattr(profile, 'can_contact', False),
        'allow_comm_all': getattr(profile, 'allow_comm_all', False),
        'job':             getattr(profile, 'job',     '') or '',
        'country':         getattr(profile, 'country', '') or '',
        'city':            getattr(profile, 'city',    '') or '',
        'benes_count':     total_bene,      # ← اسم موحّد مع الـ JS
        'orphan_count':    orphan_count,
        'family_count':    family_count,
        'special_count':   special_count,
        'total_bene':      total_bene,
        'total_paid_ils':  str(round(total_paid,     2)),
        'total_paid_usd':  str(round(total_paid_usd, 2)),
        'month_paid_ils':  str(round(month_paid,     2)),
        'whatsapp': (user.whatsapp_country or '') + (user.whatsapp or '') if user.whatsapp else '',
        'whatsapp_country': user.whatsapp_country or '+970',
        'nationality_code': user.nationality_code or '',
        'phone_num': user.phone or '',
        'sp_phone': profile.phone or '',
        'sp_phone_country': profile.phone_country or '+970',
        'sp_whatsapp': profile.whatsapp or '',
        'sp_whatsapp_country': profile.whatsapp_country or '+970',
    }


# ══════════════════════════════════════════════
#  صفحة العرض
# ══════════════════════════════════════════════

@admin_required
def sponsors_list(request):
    notif_count = Notification.objects.filter(
        recipient=request.user, is_read=False
    ).count()
    return render(request, 'admin_panel/sponsors.html', {
        'notif_count': notif_count,

    })


# ══════════════════════════════════════════════
#  API — بيانات الجدول
# ══════════════════════════════════════════════

@admin_required
@require_GET
def sponsors_data(request):
    from sponsor.models import SponsorProfile

    profiles = SponsorProfile.objects.select_related('user').filter(
        user__is_approved=True
    ).order_by('-user__date_joined')

    data  = [_build_sponsor_data(p) for p in profiles]
    today = date.today()

    total        = len(data)
    active_sp    = sum(1 for d in data if d['total_bene'] > 0)
    no_bene      = total - active_sp
    total_bene   = sum(d['total_bene'] for d in data)
    disabled     = sum(1 for d in data if not d['is_active'])
    this_month   = profiles.filter(
        user__date_joined__year=today.year,
        user__date_joined__month=today.month,
    ).count()

    return JsonResponse({
        'sponsors': data,
        'stats': {
            'total':       total,
            'active':      active_sp,
            'no_bene':     no_bene,
            'total_bene':  total_bene,
            'disabled':    disabled,
            'this_month':  this_month,

        }
    })


# ══════════════════════════════════════════════
#  API — تفاصيل كافل
# ══════════════════════════════════════════════

@admin_required
@require_GET
def sponsor_detail(request):
    user_id = request.GET.get('id', '').strip()
    if not user_id:
        return JsonResponse({'status': 'error', 'message': 'معرف غير صالح'})

    try:
        user    = CustomUser.objects.get(pk=user_id)
        profile = _get_sponsor_profile(user)
        if not profile:
            return JsonResponse({'status': 'error', 'message': 'الملف الشخصي غير موجود'})

        from admin_panel.views.requests import _user_detail
        detail = _user_detail(user)

        # المكفولون
        from beneficiary.models import OrphanForm, SpecialNeedsForm, FamilyForm
        orphans  = OrphanForm.objects.filter(sponsor=profile).select_related('user')
        families = FamilyForm.objects.filter(sponsor=profile).select_related('user')
        specials = SpecialNeedsForm.objects.filter(sponsor=profile).select_related('user')

        def _bene_item(form):
            return {
                'id':       str(form.user.pk),
                'name':     form.user.get_full_name(),
                'reg':      form.user.registration_number or '',
                'status':   form.status,
                'id_number':form.user.id_number or '',
            }

        return JsonResponse({
            'status':  'success',
            'detail':  detail,
            'profile': _build_sponsor_data(profile),
            'beneficiaries': {
                'orphans':  [_bene_item(f) for f in orphans],
                'families': [_bene_item(f) for f in families],
                'specials': [_bene_item(f) for f in specials],
            }
        })
    except CustomUser.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'المستخدم غير موجود'})


# ══════════════════════════════════════════════
#  تعطيل / تفعيل
# ══════════════════════════════════════════════

@admin_required
@require_POST
@csrf_protect
def toggle_sponsor(request):
    user_id = request.POST.get('user_id', '').strip()
    active  = request.POST.get('active', '1') == '1'

    if not user_id:
        return JsonResponse({'status': 'error', 'message': 'معرف غير صالح'})

    try:
        user           = CustomUser.objects.get(pk=user_id)
        user.is_active = active
        user.save(update_fields=['is_active'])

        action = 'تفعيل' if active else 'تعطيل'
        log_activity(request.user, 'UPDATE',
                     description=f'{action} حساب كافل: {user.get_full_name()}',
                     target_model='CustomUser', target_id=user.pk, request=request)
        create_notification(
            recipient=user, ntype='SECURITY',
            title=f'تم {action} حسابك',
            message=f'تم {action} حسابك من قِبل الإدارة',
            sender=request.user,
        )
        return JsonResponse({'status': 'success', 'message': f'تم {action} الحساب ✅'})
    except CustomUser.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'المستخدم غير موجود'})


# ══════════════════════════════════════════════
#  حذف
# ══════════════════════════════════════════════

@admin_required
@require_POST
@csrf_protect
def delete_sponsor(request):
    user_id = request.POST.get('user_id', '').strip()

    if not user_id:
        return JsonResponse({'status': 'error', 'message': 'معرف غير صالح'})

    try:
        user = CustomUser.objects.get(pk=user_id)
        name = user.get_full_name()
        log_activity(request.user, 'DELETE',
                     description=f'حذف كافل: {name}',
                     target_model='CustomUser', target_id=user.pk, request=request)
        user.delete()
        return JsonResponse({'status': 'success', 'message': f'تم حذف {name} نهائياً'})
    except CustomUser.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'المستخدم غير موجود'})


# ══════════════════════════════════════════════
#  الملاحظات
# ══════════════════════════════════════════════

@admin_required
@require_GET
def get_sponsor_notes(request):
    user_id = request.GET.get('user_id', '').strip()
    if not user_id:
        return JsonResponse({'status': 'error', 'message': 'معرف غير صالح'})

    notes = UserNote.objects.filter(
        user_id=user_id
    ).select_related('admin').order_by('-created_at')

    return JsonResponse({'status': 'success', 'notes': [{
        'id':         n.pk,
        'note':       n.note,
        'admin':      n.admin.get_short_name() if n.admin else 'النظام',
        'created_at': fmt_dt(n.created_at),
    } for n in notes]})


@admin_required
@require_POST
@csrf_protect
def add_sponsor_note(request):
    user_id = request.POST.get('user_id', '').strip()
    note    = request.POST.get('note', '').strip()

    if not user_id or not note:
        return JsonResponse({'status': 'error', 'message': 'بيانات ناقصة'})

    try:
        user = CustomUser.objects.get(pk=user_id)
        UserNote.objects.create(user=user, admin=request.user, note=note)
        log_activity(request.user, 'UPDATE',
                     description=f'ملاحظة على كافل: {user.get_full_name()}',
                     target_model='CustomUser', target_id=user.pk, request=request)
        return JsonResponse({'status': 'success', 'message': 'تم حفظ الملاحظة 📝'})
    except CustomUser.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'المستخدم غير موجود'})


# ══════════════════════════════════════════════
#  التواصل
# ══════════════════════════════════════════════

@admin_required
@require_POST
@csrf_protect
def toggle_contact(request):
    user_id = request.POST.get('user_id', '').strip()

    try:
        user    = CustomUser.objects.get(pk=user_id)
        profile = _get_sponsor_profile(user)
        if not profile:
            return JsonResponse({'status': 'error', 'message': 'الملف الشخصي غير موجود'})

        profile.can_contact = not profile.can_contact
        profile.save(update_fields=['can_contact'])

        action = 'تفعيل' if profile.can_contact else 'إلغاء'
        log_activity(request.user, 'UPDATE',
                     description=f'{action} تواصل كافل: {user.get_full_name()}',
                     target_model='SponsorProfile', target_id=profile.pk, request=request)

        return JsonResponse({
            'status':      'success',
            'can_contact': profile.can_contact,
            'message':     f'تم {action} صلاحية التواصل ✅',
        })
    except CustomUser.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'المستخدم غير موجود'})


# ══════════════════════════════════════════════
#  سجل النشاط
# ══════════════════════════════════════════════

@admin_required
@require_GET
def get_sponsor_logs(request):
    user_id = request.GET.get('user_id', '').strip()
    if not user_id:
        return JsonResponse({'status': 'error', 'message': 'معرف غير صالح'})

    logs = ActivityLog.objects.filter(
        Q(user_id=user_id) | Q(target_id=user_id, target_model='CustomUser')
    ).select_related('user').order_by('-created_at')[:200]

    return JsonResponse({'status': 'success', 'logs': [{
        'id':          l.pk,
        'action':      l.action,
        'description': l.description,
        'created_at': fmt_dt(l.created_at),
        'admin':       l.user.get_short_name() if l.user else 'النظام',
        'ip':          l.ip_address or '',
    } for l in logs]})


@admin_required
@require_GET
def export_sponsor_logs(request):
    user_id = request.GET.get('user_id', '').strip()
    ids     = request.GET.get('ids', '')

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return HttpResponse('openpyxl غير مثبّتة', status=500)

    if ids and ids != 'all':
        id_list = [i.strip() for i in ids.split(',') if i.strip().isdigit()]
        logs    = ActivityLog.objects.filter(pk__in=id_list).order_by('-created_at')
    else:
        logs = ActivityLog.objects.filter(
            Q(user_id=user_id) | Q(target_id=user_id, target_model='CustomUser')
        ).order_by('-created_at')

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = 'سجل النشاط'
    ws.sheet_view.rightToLeft = True

    GREEN  = '7C3AED'
    thin   = Side(style='thin', color='E2E8F0')
    border = openpyxl.styles.Border(left=thin,right=thin,top=thin,bottom=thin)

    headers = ['التاريخ', 'الإجراء', 'التفاصيل', 'الأدمن', 'IP']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font      = Font(bold=True, color='FFFFFF')
        c.fill      = PatternFill('solid', fgColor=GREEN)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = border

    for r_idx, l in enumerate(logs, 2):
        row = [fmt_dt(l.created_at), l.action,
               l.description, l.user.get_short_name() if l.user else 'النظام',
               l.ip_address or '']
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border    = border
            if r_idx % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='F5F3FF')

    for col, w in zip(['A','B','C','D','E'], [16, 14, 40, 16, 14]):
        ws.column_dimensions[col].width = w

    output = io.BytesIO()
    wb.save(output); output.seek(0)
    resp = HttpResponse(output.read(),
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = "attachment; filename*=UTF-8''sponsor_activity_log.xlsx"
    return resp


# ══════════════════════════════════════════════
#  تصدير Excel للقائمة
# ══════════════════════════════════════════════

@admin_required
@require_GET
def export_sponsors(request):
    from sponsor.models import SponsorProfile

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return HttpResponse('openpyxl غير مثبّتة', status=500)

    profiles = SponsorProfile.objects.select_related('user').filter(
        user__is_approved=True
    ).order_by('-user__date_joined')

    data = [_build_sponsor_data(p) for p in profiles]

    filter_type = request.GET.get('filter', 'all')
    if filter_type == 'active':
        data = [d for d in data if d['total_bene'] > 0]
    elif filter_type == 'no_bene':
        data = [d for d in data if d['total_bene'] == 0]

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = 'الكفلاء'
    ws.sheet_view.rightToLeft = True

    GREEN  = '1A7A4A'
    thin   = Side(style='thin', color='E2E8F0')
    border = openpyxl.styles.Border(left=thin,right=thin,top=thin,bottom=thin)

    headers = ['رقم الاستمارة', 'الاسم الكامل', 'البريد', 'الجوال',
               'الدولة', 'المدينة', 'المهنة', 'عدد المكفولين',
               'إجمالي المدفوع ₪', 'حالة الحساب', 'تاريخ التسجيل']

    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font      = Font(bold=True, color='FFFFFF')
        c.fill      = PatternFill('solid', fgColor=GREEN)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = border
    ws.row_dimensions[1].height = 24

    for r_idx, d in enumerate(data, 2):
        row = [d['reg_number'], d['full_name'], d['email'], d['phone'],
               d['country'], d['city'], d['job'], d['total_bene'],
               d['total_paid_ils'], 'نشط' if d['is_active'] else 'معطّل',
               d['date_joined'][:10]]
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border    = border
            if r_idx % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='F4F9F6')

    for col, w in zip(['A','B','C','D','E','F','G','H','I','J','K'],
                      [18, 26, 24, 16, 14, 14, 16, 12, 16, 12, 14]):
        ws.column_dimensions[col].width = w

    output = io.BytesIO()
    wb.save(output); output.seek(0)
    resp = HttpResponse(output.read(),
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f"attachment; filename*=UTF-8''sponsors_{filter_type}.xlsx"

    log_activity(request.user, 'EXPORT',
                 description=f'تصدير قائمة الكفلاء — {filter_type}',
                 request=request)
    return resp
@admin_required
@require_GET
def get_benes(request):
    sponsor_id = request.GET.get('sponsor_id', '').strip()
    bene_type  = request.GET.get('type', 'orphan')

    if not sponsor_id:
        return JsonResponse({'status': 'error', 'message': 'معرف غير صالح'})

    try:
        sponsor = CustomUser.objects.get(pk=sponsor_id, user_type='sponsor')
    except CustomUser.DoesNotExist:
        return JsonResponse({'benes': []})

    from beneficiary.models import OrphanForm, FamilyForm, SpecialNeedsForm
    from sponsor.models import SponsorProfile
    from django.db.models import Sum
    from datetime import date

    try:
        sp_profile = SponsorProfile.objects.get(user=sponsor)
    except SponsorProfile.DoesNotExist:
        return JsonResponse({'benes': []})

    MODEL_MAP = {'orphan': OrphanForm, 'family': FamilyForm, 'special': SpecialNeedsForm}
    Model     = MODEL_MAP.get(bene_type)
    if not Model:
        return JsonResponse({'benes': []})

    forms = Model.objects.filter(sponsor=sp_profile).select_related('user')

    result = []
    for f in forms:
        pays      = Payment.objects.filter(beneficiary=f.user, sponsor=sponsor)
        total     = pays.aggregate(s=Sum('amount_ils'))['s'] or 0
        late      = pays.filter(status='late').count()
        pending   = pays.filter(status='pending').count()
        last_pay  = pays.order_by('-date').first()

        result.append({
            'id':           str(f.user.pk),
            'full_name':    f.user.get_full_name(),
            'reg_number':   f.user.registration_number or '—',
            'id_number':    f.user.id_number or '—',
            'status':       f.status,
            'is_active':    f.user.is_active,
            'total_pays':   str(round(total, 2)),
            'late_count':   late,
            'pending_count':pending,
            'last_pay':     str(last_pay.date) if last_pay else None,
        })

    return JsonResponse({'benes': result})
"""
دوال المدفوعات في sponsors.py
get_bene_pays, export_bene_pays, export_sponsor_pays
"""

from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_GET
from django.db.models import Sum, Count
from urllib.parse import quote
import io

from core.models import CustomUser, Payment
from .decorators import admin_required


# ══════════════════════════════════════════════
# جلب دفعات مستفيد واحد من كافل محدد
# ══════════════════════════════════════════════

@admin_required
@require_GET
def get_bene_pays(request):
    user_id    = request.GET.get('user_id',    '').strip()
    sponsor_id = request.GET.get('sponsor_id', '').strip()

    if not user_id or not sponsor_id:
        return JsonResponse({'status': 'error', 'message': 'بيانات ناقصة'})

    try:
        user    = CustomUser.objects.get(pk=user_id)
        sponsor = CustomUser.objects.get(pk=sponsor_id)
    except CustomUser.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'مستخدم غير موجود'})

    pays = Payment.objects.filter(
        beneficiary=user,
        sponsor=sponsor,
    ).order_by('-date')

    total_ils    = pays.aggregate(s=Sum('amount_ils'))['s'] or 0
    total_usd    = pays.aggregate(s=Sum('amount_usd'))['s'] or 0
    late_count   = pays.filter(status='late').count()
    pending_count= pays.filter(status='pending').count()
    paid_count   = pays.filter(status='paid').count()

    STATUS_MAP = {'paid': 'مدفوعة', 'pending': 'معلّقة', 'late': 'متأخرة'}

    return JsonResponse({
        'status': 'success',
        'summary': {
            'total_ils':    str(round(total_ils, 2)),
            'total_usd':    str(round(total_usd, 2)),
            'paid_count':   paid_count,
            'late_count':   late_count,
            'pending_count':pending_count,
            'pays_count':   pays.count(),
        },
        'payments': [{
            'id':         p.pk,
            'date':       str(p.date),
            'amount_ils': str(p.amount_ils),
            'amount_usd': str(p.amount_usd),
            'status':     p.status,
            'status_label': STATUS_MAP.get(p.status, p.status),
            'note':       p.note or '',
        } for p in pays],
    })


# ══════════════════════════════════════════════
# تصدير Excel — دفعات مستفيد واحد من كافل محدد
# ══════════════════════════════════════════════

@admin_required
@require_GET
def export_bene_pays(request):
    user_id    = request.GET.get('user_id',    '').strip()
    sponsor_id = request.GET.get('sponsor_id', '').strip()

    if not user_id or not sponsor_id:
        return HttpResponse('بيانات ناقصة', status=400)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return HttpResponse('openpyxl غير مثبّتة', status=500)

    try:
        user    = CustomUser.objects.get(pk=user_id)
        sponsor = CustomUser.objects.get(pk=sponsor_id)
    except CustomUser.DoesNotExist:
        return HttpResponse('مستخدم غير موجود', status=404)

    pays = Payment.objects.filter(
        beneficiary=user,
        sponsor=sponsor,
    ).order_by('-date')

    GREEN  = '1A7A4A'
    thin   = Side(style='thin', color='E2E8F0')
    border = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)
    C_ALIGN= Alignment(horizontal='center', vertical='center')
    R_ALIGN= Alignment(horizontal='right',  vertical='center')

    STATUS_MAP = {'paid': 'مدفوعة', 'pending': 'معلّقة', 'late': 'متأخرة'}

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = 'الدفعات'
    ws.sheet_view.rightToLeft = True

    # رأس الجدول
    headers = ['التاريخ', 'المبلغ ₪', 'المبلغ $', 'الحالة', 'ملاحظة']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font      = Font(bold=True, color='FFFFFF')
        c.fill      = PatternFill('solid', fgColor=GREEN)
        c.alignment = C_ALIGN
        c.border    = border
    ws.row_dimensions[1].height = 26

    # معلومات الكافل والمستفيد
    ws.cell(row=1, column=6, value=f'الكافل: {sponsor.get_full_name()}').font = Font(bold=True, color='1A7A4A')
    ws.cell(row=1, column=7, value=f'المستفيد: {user.get_full_name()}').font  = Font(bold=True, color='1A7A4A')

    # البيانات
    for r_idx, p in enumerate(pays, 2):
        row = [
            str(p.date),
            str(p.amount_ils),
            str(p.amount_usd),
            STATUS_MAP.get(p.status, p.status),
            p.note or '',
        ]
        for c_idx, val in enumerate(row, 1):
            cell           = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = R_ALIGN
            cell.border    = border
            if r_idx % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='F0FDF4')
        ws.row_dimensions[r_idx].height = 20

    # إجماليات
    last_row = pays.count() + 2
    total_ils = pays.aggregate(s=Sum('amount_ils'))['s'] or 0
    total_usd = pays.aggregate(s=Sum('amount_usd'))['s'] or 0
    ws.cell(row=last_row, column=1, value='الإجمالي').font = Font(bold=True)
    ws.cell(row=last_row, column=2, value=str(round(total_ils, 2))).font = Font(bold=True, color='1A7A4A')
    ws.cell(row=last_row, column=3, value=str(round(total_usd, 2))).font = Font(bold=True, color='2B6CB0')

    for col, w in zip(['A','B','C','D','E','F','G'], [14, 14, 14, 12, 24, 22, 22]):
        ws.column_dimensions[col].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    fname = f'pays_{user.get_full_name()}_{sponsor.get_full_name()}.xlsx'.replace(' ', '_')
    resp  = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(fname)}"
    return resp


# ══════════════════════════════════════════════
# تصدير Excel شامل — كل مدفوعات الكافل (4 ورقات)
# ══════════════════════════════════════════════

@admin_required
@require_GET
def export_sponsor_pays(request):
    sponsor_id = request.GET.get('sponsor_id', '').strip()
    if not sponsor_id:
        return HttpResponse('معرف غير صالح', status=400)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return HttpResponse('openpyxl غير مثبّتة', status=500)

    try:
        sponsor = CustomUser.objects.get(pk=sponsor_id, user_type='sponsor')
    except CustomUser.DoesNotExist:
        return HttpResponse('الكافل غير موجود', status=404)

    from beneficiary.models import OrphanForm, FamilyForm, SpecialNeedsForm
    from sponsor.models import SponsorProfile
    from django.utils import timezone
    from dateutil.relativedelta import relativedelta
    from datetime import date

    try:
        sp_profile = SponsorProfile.objects.get(user=sponsor)
    except SponsorProfile.DoesNotExist:
        return HttpResponse('ملف الكافل غير موجود', status=404)

    # جلب المكفولين
    orphans  = OrphanForm.objects.filter(sponsor=sp_profile).select_related('user')
    families = FamilyForm.objects.filter(sponsor=sp_profile).select_related('user')
    specials = SpecialNeedsForm.objects.filter(sponsor=sp_profile).select_related('user')

    all_benes = (
        [(f.user, 'يتيم')    for f in orphans] +
        [(f.user, 'أسرة')    for f in families] +
        [(f.user, 'ذوو احتياجات') for f in specials]
    )

    all_pays = Payment.objects.filter(sponsor=sponsor).order_by('-date')

    STATUS_MAP  = {'paid': 'مدفوعة', 'pending': 'معلّقة', 'late': 'متأخرة'}
    PURPLE = '7C3AED'
    GREEN  = '1A7A4A'
    ORANGE = 'B45309'
    RED    = 'C53030'
    BLUE   = '2B6CB0'
    thin   = Side(style='thin', color='E2E8F0')
    border = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)

    def _hdr(ws, headers, color):
        ws.sheet_view.rightToLeft = True
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font      = Font(bold=True, color='FFFFFF')
            c.fill      = PatternFill('solid', fgColor=color)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border    = border
        ws.row_dimensions[1].height = 26

    def _row(ws, r_idx, values, alt='F5F3FF'):
        for c_idx, val in enumerate(values, 1):
            cell           = ws.cell(row=r_idx, column=c_idx, value=str(val) if val is not None else '—')
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border    = border
            if r_idx % 2 == 0:
                cell.fill = PatternFill('solid', fgColor=alt)
        ws.row_dimensions[r_idx].height = 20

    wb = openpyxl.Workbook()

    # ══ ورقة 1: ملخص الكافل ══
    ws1       = wb.active
    ws1.title = 'ملخص الكافل'
    _hdr(ws1, ['البيان', 'القيمة'], PURPLE)

    orphan_pays  = all_pays.filter(beneficiary__user_type='orphan')
    family_pays  = all_pays.filter(beneficiary__user_type='family')
    special_pays = all_pays.filter(beneficiary__user_type='special')
    today        = date.today()
    month_pays   = all_pays.filter(date__year=today.year, date__month=today.month)

    summary_rows = [
        ('اسم الكافل',                  sponsor.get_full_name()),
        ('رقم الاستمارة',               sponsor.registration_number or '—'),
        ('البريد الإلكتروني',            sponsor.email),
        ('',                            ''),
        ('إجمالي المكفولين',             len(all_benes)),
        ('عدد الأيتام',                 orphans.count()),
        ('عدد الأسر',                   families.count()),
        ('عدد ذوو الاحتياجات',          specials.count()),
        ('',                            ''),
        ('إجمالي الدفعات الكلي ₪',      str(round(all_pays.aggregate(s=Sum('amount_ils'))['s'] or 0, 2))),
        ('إجمالي الدفعات الكلي $',      str(round(all_pays.aggregate(s=Sum('amount_usd'))['s'] or 0, 2))),
        ('دفعات الأيتام ₪',             str(round(orphan_pays.aggregate(s=Sum('amount_ils'))['s'] or 0, 2))),
        ('دفعات الأسر ₪',              str(round(family_pays.aggregate(s=Sum('amount_ils'))['s'] or 0, 2))),
        ('دفعات ذوو الاحتياجات ₪',      str(round(special_pays.aggregate(s=Sum('amount_ils'))['s'] or 0, 2))),
        ('مدفوعات هذا الشهر ₪',         str(round(month_pays.aggregate(s=Sum('amount_ils'))['s'] or 0, 2))),
        ('',                            ''),
        ('عدد الدفعات الكلي',           all_pays.count()),
        ('دفعات مكتملة',               all_pays.filter(status='paid').count()),
        ('دفعات معلّقة',                all_pays.filter(status='pending').count()),
        ('دفعات متأخرة',               all_pays.filter(status='late').count()),
    ]
    for r_idx, (k, v) in enumerate(summary_rows, 2):
        if not k:
            ws1.row_dimensions[r_idx].height = 8
            continue
        _row(ws1, r_idx, [k, v], 'F5F3FF')
    ws1.column_dimensions['A'].width = 26
    ws1.column_dimensions['B'].width = 22

    # ══ ورقة 2: تفاصيل المكفولين والدفعات ══
    ws2       = wb.create_sheet(title='تفاصيل المكفولين')
    _hdr(ws2, ['المستفيد', 'النوع', 'إجمالي الدفعات ₪', 'إجمالي الدفعات $',
               'مكتملة', 'معلّقة', 'متأخرة', 'عدد الدفعات', 'آخر دفعة'], GREEN)

    for r_idx, (user, type_label) in enumerate(all_benes, 2):
        u_pays    = all_pays.filter(beneficiary=user)
        total_ils = u_pays.aggregate(s=Sum('amount_ils'))['s'] or 0
        total_usd = u_pays.aggregate(s=Sum('amount_usd'))['s'] or 0
        last_pay  = u_pays.order_by('-date').first()
        _row(ws2, r_idx, [
            user.get_full_name(), type_label,
            str(round(total_ils, 2)), str(round(total_usd, 2)),
            u_pays.filter(status='paid').count(),
            u_pays.filter(status='pending').count(),
            u_pays.filter(status='late').count(),
            u_pays.count(),
            str(last_pay.date) if last_pay else '—',
        ], 'F0FDF4')
    for col, w in zip(['A','B','C','D','E','F','G','H','I'], [24, 14, 16, 16, 10, 10, 10, 12, 14]):
        ws2.column_dimensions[col].width = w

    # ══ ورقة 3: كل الدفعات التفصيلية ══
    ws3       = wb.create_sheet(title='كل الدفعات')
    _hdr(ws3, ['التاريخ', 'المستفيد', 'النوع', 'المبلغ ₪', 'المبلغ $', 'الحالة', 'ملاحظة'], BLUE)

    TYPE_LABELS = {'orphan': 'يتيم', 'family': 'أسرة', 'special': 'ذوو احتياجات'}
    for r_idx, p in enumerate(all_pays.select_related('beneficiary'), 2):
        _row(ws3, r_idx, [
            str(p.date),
            p.beneficiary.get_full_name() if p.beneficiary else '—',
            TYPE_LABELS.get(p.beneficiary.user_type, '—') if p.beneficiary else '—',
            str(p.amount_ils),
            str(p.amount_usd),
            STATUS_MAP.get(p.status, p.status),
            p.note or '',
        ], 'EFF6FF')
    for col, w in zip(['A','B','C','D','E','F','G'], [14, 24, 14, 14, 14, 12, 20]):
        ws3.column_dimensions[col].width = w

    # ══ ورقة 4: الملخص الشهري (آخر 12 شهر) ══
    ws4       = wb.create_sheet(title='الملخص الشهري')
    _hdr(ws4, ['الشهر', 'عدد الدفعات', 'إجمالي ₪', 'إجمالي $', 'مكتملة', 'معلّقة', 'متأخرة'], ORANGE)

    from dateutil.relativedelta import relativedelta
    for i, r_idx in enumerate(range(11, -1, -1), 2):
        m_date   = today - relativedelta(months=r_idx)
        m_pays   = all_pays.filter(date__year=m_date.year, date__month=m_date.month)
        m_ils    = m_pays.aggregate(s=Sum('amount_ils'))['s'] or 0
        m_usd    = m_pays.aggregate(s=Sum('amount_usd'))['s'] or 0
        _row(ws4, i, [
            f'{m_date.year}/{m_date.month:02d}',
            m_pays.count(),
            str(round(m_ils, 2)),
            str(round(m_usd, 2)),
            m_pays.filter(status='paid').count(),
            m_pays.filter(status='pending').count(),
            m_pays.filter(status='late').count(),
        ], 'FEF3C7')
    for col, w in zip(['A','B','C','D','E','F','G'], [14, 14, 16, 16, 10, 10, 10]):
        ws4.column_dimensions[col].width = w

    # ══ ورقة 5: الدفعات المتأخرة والمعلّقة ══
    ws5       = wb.create_sheet(title='دفعات تحتاج متابعة')
    _hdr(ws5, ['التاريخ', 'المستفيد', 'النوع', 'المبلغ ₪', 'الحالة', 'أيام التأخير', 'ملاحظة'], RED)

    problem_pays = all_pays.filter(
        status__in=['late', 'pending']
    ).select_related('beneficiary').order_by('status', 'date')

    for r_idx, p in enumerate(problem_pays, 2):
        days_late = (today - p.date).days if p.status == 'late' else 0
        _row(ws5, r_idx, [
            str(p.date),
            p.beneficiary.get_full_name() if p.beneficiary else '—',
            TYPE_LABELS.get(p.beneficiary.user_type, '—') if p.beneficiary else '—',
            str(p.amount_ils),
            STATUS_MAP.get(p.status, p.status),
            str(days_late) if days_late else '—',
            p.note or '',
        ], 'FEF2F2')
    for col, w in zip(['A','B','C','D','E','F','G'], [14, 24, 14, 14, 12, 12, 20]):
        ws5.column_dimensions[col].width = w

    # ══ حفظ وإرسال ══
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    fname = f'sponsor_pays_{sponsor.get_full_name()}.xlsx'.replace(' ', '_')
    resp  = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(fname)}"
    return resp


@admin_required
@require_POST
@csrf_protect
def edit_sponsor(request):
    user_id = request.POST.get('user_id', '').strip()
    if not user_id:
        return JsonResponse({'status': 'error', 'message': 'معرف غير صالح'})

    try:
        user = CustomUser.objects.get(pk=user_id, user_type='sponsor')
    except CustomUser.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'الكافل غير موجود'})

    errors = {}

    # ── حقول CustomUser ──
    USER_FIELDS = [
        'first_name', 'father_name', 'grand_name', 'family_name',
        'id_number', 'nationality', 'nationality_code', 'gender',
        'phone_country', 'phone', 'whatsapp_country', 'whatsapp',
    ]

    for field in USER_FIELDS:
        if field not in request.POST:
            continue
        val = request.POST.get(field, '').strip()

        # تحقق التكرار
        if field == 'id_number' and val:
            if CustomUser.objects.filter(id_number=val).exclude(pk=user_id).exists():
                errors['id_number'] = 'رقم الهوية مستخدم مسبقاً'
                continue
            # تحقق صيغة الهوية
            if not val.isdigit() or len(val) != 9 or val[0] not in ['4','7','8','9']:
                errors['id_number'] = 'رقم الهوية غير صحيح'
                continue

        if field == 'phone' and val:
            if CustomUser.objects.filter(phone=val).exclude(pk=user_id).exists():
                errors['phone'] = 'رقم الجوال مستخدم مسبقاً'
                continue

        # تحقق المطلوب
        if field in ['first_name', 'father_name', 'family_name'] and not val:
            errors[field] = 'هذا الحقل مطلوب'
            continue

        setattr(user, field, val)

    # ── كلمة المرور ──
    password = request.POST.get('password', '').strip()
    if password:
        if len(password) < 8:
            errors['password'] = 'كلمة المرور يجب أن تكون 8 أحرف على الأقل'
        else:
            user.set_password(password)

    if errors:
        return JsonResponse({'status': 'error', 'errors': errors})

    user.save()

    # ── SponsorProfile ──
    SP_FIELDS = ['job', 'country', 'city', 'phone_country', 'phone', 'whatsapp_country', 'whatsapp']
    try:
        from sponsor.models import SponsorProfile
        sp = SponsorProfile.objects.get(user=user)
        for field in SP_FIELDS:
            if field in request.POST:
                setattr(sp, field, request.POST.get(field, '').strip())
        sp.save()
    except SponsorProfile.DoesNotExist:
        pass
    except Exception as e:
        pass

    # ── تسجيل النشاط ──
    try:
        log_activity(
            request.user, 'UPDATE',
            description=f'تعديل بيانات الكافل: {user.get_full_name()} — بواسطة: {request.user.get_full_name() or request.user.username}',
            target_model='CustomUser',
            target_id=str(user.pk),
            request=request,
        )
    except Exception:
        pass

    return JsonResponse({'status': 'success', 'message': 'تم حفظ التعديلات ✅'})



@admin_required
@require_POST
def toggle_comm_all(request):
    from sponsor.models import SponsorProfile
    try:
        sponsor_id = request.POST.get('sponsor_id')
        profile = SponsorProfile.objects.get(user__id=sponsor_id)
        profile.can_contact = not profile.can_contact
        profile.save()
        return JsonResponse({'success': True, 'new_state': profile.can_contact})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)