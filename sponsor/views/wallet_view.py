"""
sponsor/views/wallet_view.py
صفحة المحفظة — لوحة الكافل
"""
from functools import wraps
from datetime import date

from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST, require_GET
from django.views.decorators.csrf import csrf_protect
from django.core.paginator import Paginator
from django.db.models import Q, Sum
from django.utils import timezone

from core.models import CustomUser, Notification
from sponsor.models import SponsorProfile, PaymentReceipt
from core.utils import log_activity, notify_admins, get_exchange_rates


# ══ Decorator ══
def sponsor_required(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('/login/')
        if request.user.user_type != 'sponsor':
            return redirect('/login/')
        if not request.user.is_approved:
            return render(request, 'sponsor/pending.html')
        return view_func(request, *args, **kwargs)
    return wrapper


def _get_profile(user):
    try:
        return user.sponsor_profile
    except Exception:
        profile, _ = SponsorProfile.objects.get_or_create(user=user)
        return profile


def _get_beneficiaries(profile):
    """قائمة المستفيدين المكفولين من هذا الكافل"""
    from beneficiary.models import OrphanForm, SpecialNeedsForm, FamilyForm
    result = []
    for Model, tab in [(OrphanForm,'orphan'),(SpecialNeedsForm,'special'),(FamilyForm,'family')]:
        for f in Model.objects.filter(sponsor=profile).select_related('user'):
            result.append({
                'id':       str(f.user.pk),
                'name':     f.get_full_name(),
                'reg':      f.user.registration_number or f.form_number,
                'tab':      tab,
            })
    return result


def _serialize_receipt(r):
    return {
        'id':           r.pk,
        'system_ref':   r.system_ref or '',
        'unique_number':r.unique_number or '',
        'sender_name':  r.sender_name or '',
        'amount_original': str(r.amount_original),
        'currency':     r.currency,
        'amount_shekel':str(r.amount_shekel),
        'amount_dollar':str(r.amount_dollar),
        'receipt_date': str(r.receipt_date) if r.receipt_date else '',
        'receipt_image':r.receipt_image.url if r.receipt_image else '',
        'status':       r.status,
        'reject_reason':r.reject_reason or '',
        'notes':        r.notes or '',
        'submitted_at': r.submitted_at.strftime('%Y/%m/%d %H:%M') if r.submitted_at else '',
        'reviewed_at':  r.reviewed_at.strftime('%Y/%m/%d %H:%M') if r.reviewed_at else '',
        'beneficiary':  r.beneficiary.get_full_name() if r.beneficiary else '—',
        'beneficiary_id': str(r.beneficiary_id) if r.beneficiary_id else '',
    }


# ══════════════════════════════════════════════════════════
# الصفحة الرئيسية
# ══════════════════════════════════════════════════════════

@sponsor_required
def wallet(request):
    user    = request.user
    profile = _get_profile(user)
    rates   = get_exchange_rates()

    receipts = PaymentReceipt.objects.filter(sponsor=profile)
    approved = receipts.filter(status='موافق')
    pending  = receipts.filter(status='بانتظار المراجعة')
    rejected = receipts.filter(status='مرفوض')

    total_ils    = approved.aggregate(t=Sum('amount_shekel'))['t'] or 0
    total_usd    = approved.aggregate(t=Sum('amount_dollar'))['t'] or 0
    pending_ils  = pending.aggregate(t=Sum('amount_shekel'))['t']  or 0
    pending_usd  = pending.aggregate(t=Sum('amount_dollar'))['t']  or 0
    rejected_ils = rejected.aggregate(t=Sum('amount_shekel'))['t'] or 0
    rejected_usd = rejected.aggregate(t=Sum('amount_dollar'))['t'] or 0

    notif_count   = Notification.objects.filter(recipient=user, is_read=False).count()
    beneficiaries = _get_beneficiaries(profile)

    CURRENCY_CHOICES = [
        ('ILS', 'شيقل إسرائيلي'),
        ('USD', 'دولار أمريكي'),
        ('JOD', 'دينار أردني'),
        ('SAR', 'ريال سعودي'),
        ('EGP', 'جنيه مصري'),
    ]

    return render(request, 'sponsor/wallet.html', {
        'notif_count':    notif_count,
        'total_count':    receipts.count(),
        'approved_count': approved.count(),
        'pending_count':  pending.count(),
        'rejected_count': rejected.count(),
        'total_ils':      round(total_ils,    2),
        'total_usd':      round(total_usd,    2),
        'pending_ils':    round(pending_ils,  2),
        'pending_usd':    round(pending_usd,  2),
        'rejected_ils':   round(rejected_ils, 2),
        'rejected_usd':   round(rejected_usd, 2),
        'beneficiaries':  beneficiaries,
        'currency_choices': CURRENCY_CHOICES,
        'rates':          rates,
    })


# ══════════════════════════════════════════════════════════
# AJAX — بيانات الوصولات
# ══════════════════════════════════════════════════════════

@sponsor_required
@require_GET
def receipts_data(request):
    profile = _get_profile(request.user)
    status  = request.GET.get('status', 'بانتظار المراجعة')
    page    = int(request.GET.get('page', 1))
    q       = request.GET.get('q', '').strip()

    qs = PaymentReceipt.objects.filter(
        sponsor=profile, status=status
    ).select_related('beneficiary').order_by('-submitted_at')

    if q:
        qs = qs.filter(
            Q(unique_number__icontains=q) |
            Q(sender_name__icontains=q)   |
            Q(system_ref__icontains=q)    |
            Q(beneficiary__first_name__icontains=q) |
            Q(beneficiary__family_name__icontains=q)
        )

    paginator = Paginator(qs, 20)
    pg        = paginator.get_page(page)

    return JsonResponse({
        'items':        [_serialize_receipt(r) for r in pg],
        'total':        paginator.count,
        'pages':        paginator.num_pages,
        'current_page': pg.number,

    })


# ══════════════════════════════════════════════════════════
# رفع وصل جديد
# ══════════════════════════════════════════════════════════

@sponsor_required
@require_POST
@csrf_protect
def submit_receipt(request):
    import re
    user    = request.user
    profile = _get_profile(user)

    beneficiary_id = request.POST.get('beneficiary_id', '').strip()
    unique_number  = request.POST.get('unique_number',  '').strip()
    sender_name    = request.POST.get('sender_name',    '').strip()
    amount_str     = request.POST.get('amount_original','0').strip()
    currency       = request.POST.get('currency',       'USD').strip()
    amount_shekel  = request.POST.get('amount_shekel',  '0').strip()
    amount_dollar  = request.POST.get('amount_dollar',  '0').strip()
    receipt_date   = request.POST.get('receipt_date',   '').strip()
    notes          = request.POST.get('notes',          '').strip()
    errors         = {}

    # تحقق
    if not beneficiary_id:
        errors['beneficiary_id'] = 'يرجى اختيار المستفيد'
    if not unique_number:
        errors['unique_number'] = 'رقم الوصل مطلوب'
    elif PaymentReceipt.objects.filter(sponsor=profile, unique_number=unique_number).exists():
        errors['unique_number'] = 'رقم الوصل مكرر'
    if not sender_name:
        errors['sender_name'] = 'اسم المرسل مطلوب'
    if re.search(r'<script|javascript:', sender_name, re.I):
        errors['sender_name'] = 'مدخل غير صالح'
    try:
        amount_original = float(amount_str)
        if amount_original <= 0:
            errors['amount_original'] = 'المبلغ يجب أن يكون أكبر من صفر'
    except ValueError:
        errors['amount_original'] = 'مبلغ غير صالح'
        amount_original = 0
    if not receipt_date:
        errors['receipt_date'] = 'تاريخ الوصل مطلوب'
    if 'receipt_image' not in request.FILES:
        errors['receipt_image'] = 'صورة الوصل مطلوبة'

    if errors:
        return JsonResponse({'status': 'error', 'errors': errors})

    # التحقق من الصورة
    img = request.FILES['receipt_image']
    if img.size > 5 * 1024 * 1024:
        return JsonResponse({'status': 'error', 'errors': {'receipt_image': 'الصورة تتجاوز 5MB'}})
    if img.name.split('.')[-1].lower() not in ['jpg', 'jpeg', 'png', 'pdf']:
        return JsonResponse({'status': 'error', 'errors': {'receipt_image': 'نوع الملف غير مسموح'}})

    # المستفيد
    beneficiary = None
    try:
        beneficiary = CustomUser.objects.get(pk=beneficiary_id)
    except Exception:
        pass

    try:
        a_shekel = float(amount_shekel) if amount_shekel else 0
        a_dollar = float(amount_dollar) if amount_dollar else 0
    except ValueError:
        a_shekel = a_dollar = 0

    rec = PaymentReceipt.objects.create(
        sponsor         = profile,
        beneficiary     = beneficiary,
        unique_number   = unique_number,
        sender_name     = sender_name,
        amount_original = amount_original,
        currency        = currency,
        amount_shekel   = a_shekel,
        amount_dollar   = a_dollar,
        receipt_date    = receipt_date,
        receipt_image   = img,
        notes           = notes,
        status          = 'بانتظار المراجعة',
    )

    notify_admins(
        ntype      = 'NEW_RECEIPT',
        title      = 'وصل مالي جديد 💳',
        message    = f'{user.get_full_name()} — {amount_original} {currency}',
        sender     = user,
        action_url = '/admin-panel/receipts/',
    )

    log_activity(user, 'RECEIPT',
                 description=f'رفع وصل مالي {amount_original} {currency}',
                 request=request)

    return JsonResponse({
        'status':  'success',
        'message': 'تم إرسال الوصل للمراجعة ✅ سيتم إشعارك بعد المراجعة',
        'receipt': _serialize_receipt(rec),
    })


# ══════════════════════════════════════════════════════════
# فحص رقم الوصل
# ══════════════════════════════════════════════════════════

@sponsor_required
@require_GET
def check_unique(request):
    profile = _get_profile(request.user)
    number  = request.GET.get('number', '').strip()
    rid     = request.GET.get('exclude', '')  # للتعديل

    qs = PaymentReceipt.objects.filter(sponsor=profile, unique_number=number)
    if rid:
        qs = qs.exclude(pk=rid)

    return JsonResponse({'exists': qs.exists()})


# ══════════════════════════════════════════════════════════
# إعادة إرسال المرفوض
# ══════════════════════════════════════════════════════════

@sponsor_required
@require_POST
@csrf_protect
def resubmit_receipt(request):
    import re
    user    = request.user
    profile = _get_profile(user)
    rid     = request.POST.get('receipt_id', '').strip()

    rec = get_object_or_404(PaymentReceipt, pk=rid, sponsor=profile, status='مرفوض')
    # تحديث المستفيد إذا تم إرساله
    beneficiary_id = request.POST.get('beneficiary_id', '').strip()
    if beneficiary_id:
        try:
            rec_beneficiary = CustomUser.objects.get(pk=beneficiary_id)
            rec.beneficiary = rec_beneficiary
        except CustomUser.DoesNotExist:
            pass

    sender_name   = request.POST.get('sender_name',    rec.sender_name).strip()
    unique_number = request.POST.get('unique_number',  rec.unique_number).strip()
    amount_str    = request.POST.get('amount_original',str(rec.amount_original)).strip()
    currency      = request.POST.get('currency',       rec.currency).strip()
    amount_shekel = request.POST.get('amount_shekel',  str(rec.amount_shekel)).strip()
    amount_dollar = request.POST.get('amount_dollar',  str(rec.amount_dollar)).strip()
    receipt_date  = request.POST.get('receipt_date',   str(rec.receipt_date)).strip()
    notes         = request.POST.get('notes',          rec.notes or '').strip()
    errors        = {}

    if not unique_number:
        errors['unique_number'] = 'رقم الوصل مطلوب'
    elif PaymentReceipt.objects.filter(
        sponsor=profile, unique_number=unique_number
    ).exclude(pk=rid).exists():
        errors['unique_number'] = 'رقم الوصل مكرر'

    if not sender_name:
        errors['sender_name'] = 'اسم المرسل مطلوب'

    try:
        amount_original = float(amount_str)
        if amount_original <= 0:
            errors['amount_original'] = 'المبلغ يجب أن يكون أكبر من صفر'
    except ValueError:
        errors['amount_original'] = 'مبلغ غير صالح'
        amount_original = rec.amount_original

    if errors:
        return JsonResponse({'status': 'error', 'errors': errors})

    rec.sender_name    = sender_name
    rec.unique_number  = unique_number
    rec.amount_original= amount_original
    rec.currency       = currency
    rec.receipt_date   = receipt_date
    rec.notes          = notes
    rec.status         = 'بانتظار المراجعة'
    rec.reject_reason  = None

    try:
        rec.amount_shekel = float(amount_shekel)
        rec.amount_dollar = float(amount_dollar)
    except ValueError:
        pass

    if 'receipt_image' in request.FILES:
        img = request.FILES['receipt_image']
        if img.size <= 5 * 1024 * 1024:
            rec.receipt_image = img

    rec.save()

    notify_admins(
        ntype      = 'NEW_RECEIPT',
        title      = 'وصل معاد إرساله 🔄',
        message    = f'{user.get_full_name()} أعاد إرسال وصل {rec.system_ref}',
        sender     = user,
        action_url = '/admin-panel/receipts/',
    )

    log_activity(user, 'RECEIPT',
                 description=f'إعادة إرسال وصل {rec.system_ref}',
                 request=request)

    return JsonResponse({
        'status':  'success',
        'message': 'تم إعادة إرسال الوصل للمراجعة ✅',
        'receipt': _serialize_receipt(rec),
    })


# ══════════════════════════════════════════════════════════
# تصدير Excel
# ══════════════════════════════════════════════════════════

@sponsor_required
@require_GET
def export_receipts_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from urllib.parse import quote
    import io

    user    = request.user
    profile = _get_profile(user)
    status  = request.GET.get('status', 'all')

    log_activity(user, 'EXPORT',
                 description=f'تصدير Excel وصولات — {status}',
                 request=request)

    GREEN  = '1A7A4A'
    LIGHT  = 'E8F5E9'
    WHITE  = 'FFFFFF'
    thin   = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.sheet_view.rightToLeft = True
    ws.title = 'الوصولات المالية'

    ws.merge_cells('A1:I1')
    c = ws['A1']
    c.value     = f'الوصولات المالية — {user.get_full_name()} — {date.today()}'
    c.font      = Font(bold=True, size=12, color=WHITE)
    c.fill      = PatternFill('solid', fgColor=GREEN)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    headers = ['رقم النظام','رقم الوصل','اسم المرسل','المستفيد','المبلغ الأصلي','العملة','بالشيقل (₪)','بالدولار ($)','الحالة']
    widths  = [20, 18, 22, 24, 14, 10, 14, 14, 14]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font      = Font(bold=True, color=WHITE, size=10)
        cell.fill      = PatternFill('solid', fgColor=GREEN)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = border
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 22

    qs = PaymentReceipt.objects.filter(sponsor=profile)
    if status != 'all':
        qs = qs.filter(status=status)
    qs = qs.select_related('beneficiary').order_by('-submitted_at')

    total_ils = total_usd = 0
    for r_idx, r in enumerate(qs, start=3):
        row = [
            r.system_ref or '—', r.unique_number, r.sender_name,
            r.beneficiary.get_full_name() if r.beneficiary else '—',
            float(r.amount_original), r.currency,
            float(r.amount_shekel), float(r.amount_dollar), r.status,
        ]
        if r.status == 'موافق':
            total_ils += float(r.amount_shekel)
            total_usd += float(r.amount_dollar)
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = border
            ws.row_dimensions[r_idx].height = 20
            if r_idx % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='F4F9F6')
            if c_idx == 9:
                clr = {'موافق':'1A7A4A','مرفوض':'C53030','بانتظار المراجعة':'B45309'}.get(val,'9CA3AF')
                cell.font = Font(bold=True, color=clr)

    # إجمالي
    tr = qs.count() + 3
    ws.cell(tr, 1, 'الإجمالي المقبول').font = Font(bold=True, color=WHITE)
    ws.cell(tr, 1).fill  = PatternFill('solid', fgColor=GREEN)
    ws.cell(tr, 7, round(total_ils, 2)).font = Font(bold=True, color=GREEN)
    ws.cell(tr, 8, round(total_usd, 2)).font = Font(bold=True, color=GREEN)
    for c in range(1, 10):
        ws.cell(tr, c).border = border
        if c != 1:
            ws.cell(tr, c).fill = PatternFill('solid', fgColor=LIGHT)
    ws.row_dimensions[tr].height = 22

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    fname = f'وصولات_{user.get_full_name()}_{date.today()}.xlsx'
    resp  = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(fname)}"
    return resp


@sponsor_required
def download_receipt_pdf(request):
    from django.template.loader import render_to_string
    from urllib.parse import quote
    import weasyprint, logging, datetime

    logging.getLogger('weasyprint').setLevel(logging.ERROR)
    logging.getLogger('fontTools').setLevel(logging.ERROR)

    rid     = request.GET.get('id', '')
    profile = _get_profile(request.user)
    rec     = get_object_or_404(PaymentReceipt, pk=rid, sponsor=profile, status='موافق')

    # ── معلومات الموقع ──
    try:
        from core.models import SystemSettings
        site_name  = SystemSettings.get('site_name',  'منصة رُحَمَاء')
        site_phone = SystemSettings.get('site_phone', '')
        site_email = SystemSettings.get('site_email', '')
    except Exception:
        site_name  = 'منصة رُحَمَاء'
        site_phone = ''
        site_email = ''

    # ── المبالغ من قاعدة البيانات مباشرة ──
    amount_ils = rec.amount_shekel if rec.amount_shekel is not None else 0
    amount_usd = rec.amount_dollar if rec.amount_dollar is not None else 0

    log_activity(request.user, 'EXPORT',
                 description=f'تحميل PDF وصل {rec.system_ref}',
                 request=request)

    html_content = render_to_string('admin_panel/receipt_pdf.html', {
        'receipt':    rec,
        'site_name':  site_name,
        'site_phone': site_phone,
        'site_email': site_email,
        'print_date': timezone.localtime(timezone.now()).strftime('%Y/%m/%d %H:%M'),
        'amount_ils': amount_ils,
        'amount_usd': amount_usd,
        'printed_by': request.user.get_full_name(),
        'request':    request,
    })

    pdf_file = weasyprint.HTML(
        string=html_content,
        base_url=request.build_absolute_uri('/')
    ).write_pdf()

    fname = f'وصل_{rec.system_ref}.pdf'
    resp  = HttpResponse(pdf_file, content_type='application/pdf')
    resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(fname)}"
    return resp
