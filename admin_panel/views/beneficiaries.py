"""
admin_panel/views/beneficiaries.py
فيوز موحد للأيتام والأسر وذوي الاحتياجات الخاصة
"""

import io
from datetime import date
from urllib.parse import quote

from django.http import JsonResponse, HttpResponse
from django.shortcuts import render
from django.utils import timezone
from django.views.decorators.http import require_POST, require_GET
from django.views.decorators.csrf import csrf_protect
from django.db.models import Q, Sum

from core.models import CustomUser, Notification, ActivityLog, UserNote, Payment, Aid
from core.utils import (log_activity, get_client_ip, get_exchange_rates,
                        create_notification, send_email, fmt_dt)
from .decorators import admin_required

# ══════════════════════════════════════════════
#  مساعدات
# ══════════════════════════════════════════════

MODEL_MAP_FUNC = None  # يُحمَّل عند الاستدعاء لتجنب circular imports

def _get_model_map():
    from beneficiary.models import OrphanForm, SpecialNeedsForm, FamilyForm
    return {
        'orphan':  OrphanForm,
        'special': SpecialNeedsForm,
        'family':  FamilyForm,
    }

def _get_form(user, bene_type):
    """جلب استمارة المستفيد"""
    model_map = _get_model_map()
    Model = model_map.get(bene_type)
    if not Model:
        return None
    try:
        return Model.objects.get(user=user)
    except Model.DoesNotExist:
        return None

def _user_detail_bene(user):
    """جمع بيانات المستفيد التفصيلية — يعيد استخدام نفس منطق requests.py"""
    from admin_panel.views.requests import _user_detail
    return _user_detail(user)

def _build_user_data(user, bene_type):
    """بناء بيانات مستخدم واحد للـ API"""
    from beneficiary.models import OrphanForm, SpecialNeedsForm, FamilyForm
    from sponsor.models import SponsorProfile

    model_map = _get_model_map()
    Model     = model_map.get(bene_type)
    form      = None
    sponsor_name = None
    sponsor_id   = None
    sponsor_reg  = None
    bene_id      = None
    status       = '—'

    if Model:
        try:
            form      = Model.objects.select_related('user').get(user=user)
            bene_id   = str(form.pk)
            status    = form.status
            # الكافل
            if hasattr(form, 'sponsor') and form.sponsor:
                sp = form.sponsor
                sponsor_name = sp.user.get_full_name() if hasattr(sp, 'user') else str(sp)
                sponsor_id   = str(sp.pk)
                sponsor_reg  = sp.user.registration_number if hasattr(sp, 'user') else None
        except Model.DoesNotExist:
            pass

    # المبلغ الشهري
    today         = timezone.now()
    monthly_pays  = Payment.objects.filter(
        beneficiary=user,
        date__year=today.year,
        date__month=today.month,
    )
    monthly_ils   = monthly_pays.aggregate(s=Sum('amount_ils'))['s'] or 0
    monthly_usd   = monthly_pays.aggregate(s=Sum('amount_usd'))['s'] or 0

    # صورة
    photo = None
    if form and hasattr(form, 'photo') and form.photo:
        photo = form.photo.url

    return {
        'id':                str(user.pk),
        'full_name':         user.get_full_name(),
        'first_name':        user.first_name,
        'father_name':       user.father_name,
        'grand_name':        user.grand_name,
        'family_name':       user.family_name,
        'email':             user.email,
        'phone':             f'{user.phone_country}{user.phone}',
        'phone_country':     user.phone_country,
        'id_number':         user.id_number or '',
        'nationality':       user.nationality or '',
        'gender':            user.gender or '',
        'reg_number':        user.registration_number or '',
        'date_joined':       user.date_joined.isoformat(),
        'last_login':        user.last_login.isoformat() if user.last_login else None,
        'is_active':         user.is_active,
        'photo':             photo,
        'sponsor_name':      sponsor_name,
        'sponsor_id':        sponsor_id,
        'sponsor_reg':       sponsor_reg,
        'has_sponsor':       sponsor_name is not None,
        'bene_id':           bene_id,
        'status':            status,
        'monthly_amount_ils': str(round(monthly_ils, 2)),
        'monthly_amount_usd': str(round(monthly_usd, 2)),
        'allow_comm':        getattr(user, 'allow_direct_comm', False),
    }


# ══════════════════════════════════════════════
#  صفحات العرض
# ══════════════════════════════════════════════

@admin_required
def orphans_list(request):
    notif_count = Notification.objects.filter(recipient=request.user, is_read=False).count()
    return render(request, 'admin_panel/orphans.html', {'notif_count': notif_count})

@admin_required
def families_list(request):
    notif_count = Notification.objects.filter(recipient=request.user, is_read=False).count()
    return render(request, 'admin_panel/families.html', {'notif_count': notif_count})

@admin_required
def specials_list(request):
    notif_count = Notification.objects.filter(recipient=request.user, is_read=False).count()
    return render(request, 'admin_panel/specials.html', {'notif_count': notif_count})


# ══════════════════════════════════════════════
#  API — بيانات الجدول
# ══════════════════════════════════════════════

@admin_required
@require_GET
def beneficiaries_data(request):
    bene_type = request.GET.get('type', 'orphan')

    users = CustomUser.objects.filter(
        user_type=bene_type, is_approved=True
    ).order_by('-date_joined')

    today = date.today()
    data  = [_build_user_data(u, bene_type) for u in users]

    # إحصائيات
    total       = len(data)
    sponsored   = sum(1 for d in data if d['has_sponsor'])
    unsponsored = total - sponsored
    disabled    = sum(1 for d in data if not d['is_active'])
    this_month  = users.filter(date_joined__year=today.year, date_joined__month=today.month).count()

    rates = get_exchange_rates()
    from core.models import Payment, Aid

    total_payments = Payment.objects.filter(
        beneficiary__user_type=bene_type
    ).count()

    total_aids = Aid.objects.filter(
        beneficiary__user_type=bene_type
    ).count()

    return JsonResponse({
        'users': data,
        'exchange_rate': rates.get('USD_TO_ILS', 3.7),

        'stats': {
            'total': total,
            'sponsored': sponsored,
            'unsponsored': unsponsored,
            'disabled': disabled,
            'this_month': this_month,
            'total_payments': total_payments,
            'total_aids': total_aids,
        }
    })


# ══════════════════════════════════════════════
#  API — تفاصيل مستخدم
# ══════════════════════════════════════════════

@admin_required
@require_GET
def beneficiary_detail(request):
    user_id   = request.GET.get('id', '').strip()
    bene_type = request.GET.get('type', 'orphan')

    if not user_id:
        return JsonResponse({'status': 'error', 'message': 'معرف غير صالح'})

    try:
        user   = CustomUser.objects.get(pk=user_id)
        detail = _user_detail_bene(user)

        # قائمة الكفلاء المتاحين
        from sponsor.models import SponsorProfile
        sponsors = list(SponsorProfile.objects.filter(
            user__is_approved=True, user__is_active=True
        ).select_related('user').values(
            'id', 'user__first_name', 'user__family_name', 'user__registration_number'
        ))
        sponsors_data = [{'id': str(s['id']),
                          'name': f"{s['user__first_name']} {s['user__family_name']}",
                          'reg': s['user__registration_number'] or ''} for s in sponsors]

        return JsonResponse({'status': 'success', 'detail': detail, 'sponsors': sponsors_data})
    except CustomUser.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'المستخدم غير موجود'})


# ══════════════════════════════════════════════
#  تعطيل / تفعيل
# ══════════════════════════════════════════════

@admin_required
@require_POST
@csrf_protect
def toggle_user(request):
    user_id = request.POST.get('user_id', '').strip()
    active  = request.POST.get('active', '1') == '1'

    if not user_id:
        return JsonResponse({'status': 'error', 'message': 'معرف غير صالح'})

    try:
        user           = CustomUser.objects.get(pk=user_id)
        user.is_active = active
        user.save(update_fields=['is_active'])

        action = 'تفعيل' if active else 'تعطيل'
        log_activity(request.user, 'UPDATE',
                     description=f'{action} حساب: {user.get_full_name()}',
                     target_model='CustomUser', target_id=user.pk, request=request)

        create_notification(
            recipient=user, ntype='SECURITY',
            title=f'تم {action} حسابك',
            message=f'تم {action} حسابك من قِبل الإدارة',
            sender=request.user,
        )
        return JsonResponse({'status': 'success', 'message': f'تم {action} الحساب ✅'})
    except CustomUser.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'المستخدم غير موجود'})


# ══════════════════════════════════════════════
#  حذف
# ══════════════════════════════════════════════

@admin_required
@require_POST
@csrf_protect
def delete_user(request):
    user_id = request.POST.get('user_id', '').strip()

    if not user_id:
        return JsonResponse({'status': 'error', 'message': 'معرف غير صالح'})

    try:
        user = CustomUser.objects.get(pk=user_id)
        name = user.get_full_name()
        log_activity(request.user, 'DELETE',
                     description=f'حذف مستخدم: {name}',
                     target_model='CustomUser', target_id=user.pk, request=request)
        user.delete()
        return JsonResponse({'status': 'success', 'message': f'تم حذف {name} نهائياً'})
    except CustomUser.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'المستخدم غير موجود'})


# ══════════════════════════════════════════════
#  الملاحظات
# ══════════════════════════════════════════════

@admin_required
@require_GET
def get_notes(request):
    user_id = request.GET.get('user_id', '').strip()
    if not user_id:
        return JsonResponse({'status': 'error', 'message': 'معرف غير صالح'})

    notes = UserNote.objects.filter(user_id=user_id).select_related('admin').order_by('-created_at')
    return JsonResponse({'status': 'success', 'notes': [{
        'id':         n.pk,
        'note':       n.note,
        'admin': (n.admin.get_full_name() or n.admin.username) if n.admin else 'النظام',
        'created_at': fmt_dt(n.created_at),
    } for n in notes]})


@admin_required
@require_POST
@csrf_protect
def add_note(request):
    user_id = request.POST.get('user_id', '').strip()
    note    = request.POST.get('note', '').strip()

    if not user_id or not note:
        return JsonResponse({'status': 'error', 'message': 'بيانات ناقصة'})

    try:
        user = CustomUser.objects.get(pk=user_id)
        UserNote.objects.create(user=user, admin=request.user, note=note)
        log_activity(request.user, 'UPDATE',
                     description=f'ملاحظة على: {user.get_full_name()}',
                     target_model='CustomUser', target_id=user.pk, request=request)
        return JsonResponse({'status': 'success', 'message': 'تم حفظ الملاحظة 📝'})
    except CustomUser.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'المستخدم غير موجود'})


# ══════════════════════════════════════════════
#  المدفوعات
# ══════════════════════════════════════════════
@admin_required
@require_GET
def get_payments(request):
    user_id = request.GET.get('user_id', '').strip()
    if not user_id:
        return JsonResponse({'status': 'error', 'message': 'معرف غير صالح'})

    pays      = Payment.objects.filter(
        beneficiary_id=user_id
    ).select_related('sponsor', 'created_by').order_by('-date')

    today      = timezone.now()
    total_ils  = pays.aggregate(s=Sum('amount_ils'))['s'] or 0
    month_ils  = pays.filter(
        date__year=today.year, date__month=today.month
    ).aggregate(s=Sum('amount_ils'))['s'] or 0
    last_pay   = pays.first()

    def _party(p):
        if p.paid_by == 'sponsor' and p.sponsor:
            return p.sponsor.get_full_name()
        elif p.paid_by == 'admin' and p.created_by:
            return p.created_by.get_full_name() or p.created_by.username
        elif p.paid_by == 'external':
            return p.paid_by_note or '—'
        return '—'

    return JsonResponse({
        'status': 'success',
        'summary': {
            'total_ils': str(round(total_ils, 2)),
            'month_ils': str(round(month_ils, 2)),
            'last_date': str(last_pay.date) if last_pay else '—',
        },
        'payments': [{
            'id':           p.pk,
            'amount_ils':   str(p.amount_ils),
            'amount_usd':   str(p.amount_usd),
            'date':         str(p.date),
            'status':       p.status,
            'paid_by':      p.paid_by,
            'paid_by_note': p.paid_by_note or '',
            'note':         p.note or '',
            'party':        _party(p),
        } for p in pays],
    })


@admin_required
@require_POST
@csrf_protect
def add_payment(request):
    user_id      = request.POST.get('user_id', '').strip()
    amount_ils   = request.POST.get('amount_ils', 0)
    amount_usd   = request.POST.get('amount_usd', 0)
    pay_date     = request.POST.get('date', str(date.today()))
    status       = request.POST.get('status', 'paid')
    paid_by      = request.POST.get('paid_by', 'sponsor')
    paid_by_note = request.POST.get('paid_by_note', '')
    note         = request.POST.get('note', '')

    if not user_id:
        return JsonResponse({'status': 'error', 'message': 'معرف غير صالح'})

    try:
        user         = CustomUser.objects.get(pk=user_id)
        sponsor_user = None

        if paid_by == 'sponsor':
            form = _get_form(user, user.user_type)
            if form and hasattr(form, 'sponsor') and form.sponsor:
                sponsor_user = form.sponsor.user
            else:
                paid_by = 'admin'

        pay = Payment.objects.create(
            beneficiary  = user,
            sponsor      = sponsor_user,
            amount_ils   = amount_ils,
            amount_usd   = amount_usd,
            date         = pay_date,
            status       = status,
            paid_by      = paid_by,
            paid_by_note = paid_by_note,
            note         = note,
            created_by   = request.user,
        )

        today     = timezone.now()
        month_ils = Payment.objects.filter(
            beneficiary=user,
            date__year=today.year, date__month=today.month,
        ).aggregate(s=Sum('amount_ils'))['s'] or 0
        month_usd = Payment.objects.filter(
            beneficiary=user,
            date__year=today.year, date__month=today.month,
        ).aggregate(s=Sum('amount_usd'))['s'] or 0

        log_activity(
            request.user, 'PAYMENT',
            description=f'دفعة لـ {user.get_full_name()}: {amount_ils}₪',
            target_model='Payment', target_id=pay.pk, request=request,
        )

        return JsonResponse({
            'status':    'success',
            'message':   'تم حفظ الدفعة ✅',
            'month_ils': str(round(month_ils, 2)),
            'month_usd': str(round(month_usd, 2)),
        })

    except CustomUser.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'المستخدم غير موجود'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


@admin_required
@require_POST
@csrf_protect
def delete_payment(request):
    pay_id = request.POST.get('pay_id', '').strip()

    if not pay_id:
        return JsonResponse({'status': 'error', 'message': 'معرف غير صالح'})

    try:
        pay = Payment.objects.get(pk=pay_id)

        log_activity(
            request.user, 'DELETE',
            description  = f'حذف دفعة: {pay.amount_ils}₪ — {pay.beneficiary.get_full_name()} — بواسطة: {request.user.get_full_name() or request.user.username}',
            target_model = 'Payment',
            target_id    = str(pay.pk),
            old_value    = {
                'amount_ils': str(pay.amount_ils),
                'amount_usd': str(pay.amount_usd),
                'date':       str(pay.date),
                'status':     pay.status,
            },
            request      = request,
        )

        pay.delete()
        return JsonResponse({'status': 'success', 'message': 'تم حذف الدفعة'})

    except Payment.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'الدفعة غير موجودة'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


@admin_required
@require_GET
def export_payments(request):
    user_id = request.GET.get('user_id', '').strip()
    if not user_id:
        return HttpResponse('معرف غير صالح', status=400)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return HttpResponse('openpyxl غير مثبّتة', status=500)

    try:
        user = CustomUser.objects.get(pk=user_id)
    except CustomUser.DoesNotExist:
        return HttpResponse('المستخدم غير موجود', status=404)

    pays = Payment.objects.filter(
        beneficiary=user
    ).select_related('sponsor', 'created_by').order_by('-date')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'المدفوعات'
    ws.sheet_view.rightToLeft = True

    GREEN    = '1A7A4A'
    thin     = Side(style='thin', color='CCCCCC')
    border   = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)
    HDR_FONT = Font(bold=True, color='FFFFFF')
    HDR_FILL = PatternFill('solid', fgColor=GREEN)
    C_ALIGN  = Alignment(horizontal='center', vertical='center')
    R_ALIGN  = Alignment(horizontal='right',  vertical='center')

    STATUS_MAP  = {'paid': 'مدفوعة', 'pending': 'معلّقة', 'late': 'متأخرة'}
    PAID_BY_MAP = {'sponsor': 'كافل', 'admin': 'إدارة', 'external': 'جهة خارجية'}

    headers = ['التاريخ', 'المبلغ بالشيقل', 'المبلغ بالدولار',
               'مصدر الدفع', 'اسم الشخص / الجهة', 'الحالة', 'ملاحظة']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = C_ALIGN; c.border = border
    ws.row_dimensions[1].height = 26

    for r_idx, p in enumerate(pays, 2):
        # تحديد اسم الشخص / الجهة
        if p.paid_by == 'sponsor' and p.sponsor:
            party = p.sponsor.get_full_name()
        elif p.paid_by == 'admin' and p.created_by:
            party = p.created_by.get_full_name() or p.created_by.username
        elif p.paid_by == 'external':
            party = p.paid_by_note or '—'
        else:
            party = '—'

        row = [
            str(p.date),
            str(p.amount_ils),
            str(p.amount_usd),
            PAID_BY_MAP.get(p.paid_by, p.paid_by),
            party,
            STATUS_MAP.get(p.status, p.status),
            p.note or '',
        ]
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = R_ALIGN
            cell.border    = border
            if r_idx % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='F4F9F6')
        ws.row_dimensions[r_idx].height = 20

    for col, w in zip(['A','B','C','D','E','F','G'], [14, 16, 16, 14, 24, 12, 24]):
        ws.column_dimensions[col].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    name  = user.get_full_name().replace(' ', '_')
    fname = f'payments_{name}.xlsx'
    resp  = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(fname)}"
    return resp


# ══════════════════════════════════════════════
#  المساعدات
# ══════════════════════════════════════════════

@admin_required
@require_GET
def get_aids(request):
    user_id = request.GET.get('user_id', '').strip()
    if not user_id:
        return JsonResponse({'status': 'error', 'message': 'معرف غير صالح'})

    aids = Aid.objects.filter(
        beneficiary_id=user_id
    ).select_related('created_by').order_by('-date')

    return JsonResponse({'status': 'success', 'aids': [{
        'id':         a.pk,
        'name':       a.name,
        'aid_type':   a.aid_type,
        'quantity':   a.quantity,
        'provider':   a.provider,
        'date':       str(a.date),
        'note':       a.note or '',
        'created_by': (a.created_by.get_full_name() or a.created_by.username) if a.created_by else 'النظام',
    } for a in aids]})

@admin_required
@require_POST
@csrf_protect
def add_aid(request):
    user_id  = request.POST.get('user_id', '').strip()
    name     = request.POST.get('name', '').strip()
    aid_type = request.POST.get('aid_type', 'other')
    quantity = request.POST.get('quantity', 1)
    provider = request.POST.get('provider', '').strip()
    aid_date = request.POST.get('date', str(date.today()))
    note     = request.POST.get('note', '')

    if not user_id or not name or not provider:
        return JsonResponse({'status': 'error', 'message': 'اسم المساعدة والجهة مطلوبان'})

    try:
        user = CustomUser.objects.get(pk=user_id)
        Aid.objects.create(
            beneficiary=user, name=name, aid_type=aid_type,
            quantity=quantity, provider=provider,
            date=aid_date, note=note, created_by=request.user,
        )
        log_activity(request.user, 'UPDATE',
                     description=f'مساعدة لـ {user.get_full_name()}: {name}',
                     target_model='Aid', request=request)
        return JsonResponse({'status': 'success', 'message': 'تم حفظ المساعدة 🎁'})
    except CustomUser.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'المستخدم غير موجود'})
@admin_required
@require_POST
@csrf_protect
def delete_aid(request):
    aid_id = request.POST.get('aid_id', '').strip()

    if not aid_id:
        return JsonResponse({'status': 'error', 'message': 'معرف غير صالح'})

    try:
        aid = Aid.objects.get(pk=aid_id)

        log_activity(
            request.user, 'DELETE',
            description  = f'حذف مساعدة: {aid.name} — {aid.beneficiary.get_full_name()} — بواسطة: {request.user.get_full_name() or request.user.username}',
            target_model = 'Aid',
            target_id    = str(aid.pk),
            old_value    = {
                'name':     aid.name,
                'aid_type': aid.aid_type,
                'quantity': aid.quantity,
                'provider': aid.provider,
                'date':     str(aid.date),
            },
            request      = request,
        )

        aid.delete()
        return JsonResponse({'status': 'success', 'message': 'تم الحذف'})

    except Aid.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'المساعدة غير موجودة'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})
@admin_required
@require_POST
@csrf_protect
def edit_aid(request):
    aid_id   = request.POST.get('aid_id',  '').strip()
    name     = request.POST.get('name',    '').strip()
    aid_type = request.POST.get('aid_type','other')
    quantity = request.POST.get('quantity', 1)
    provider = request.POST.get('provider','').strip()
    aid_date = request.POST.get('date', str(date.today()))
    note     = request.POST.get('note', '')

    if not aid_id or not name or not provider:
        return JsonResponse({'status': 'error', 'message': 'بيانات ناقصة'})
    try:
        a          = Aid.objects.get(pk=aid_id)
        a.name     = name
        a.aid_type = aid_type
        a.quantity = quantity
        a.provider = provider
        a.date     = aid_date
        a.note     = note
        a.save()
        log_activity(request.user, 'UPDATE',
                     description=f'تعديل مساعدة لـ {a.beneficiary.get_full_name()}: {name}',
                     target_model='Aid', target_id=a.pk, request=request)
        return JsonResponse({'status': 'success', 'message': 'تم التعديل ✅'})
    except Aid.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'المساعدة غير موجودة'})


@admin_required
@require_GET
def export_aids(request):
    user_id = request.GET.get('user_id', '').strip()
    if not user_id:
        return HttpResponse('معرف غير صالح', status=400)
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return HttpResponse('openpyxl غير مثبّتة', status=500)

    try:
        user = CustomUser.objects.get(pk=user_id)
        aids = Aid.objects.filter(beneficiary=user).order_by('-date')
    except CustomUser.DoesNotExist:
        return HttpResponse('المستخدم غير موجود', status=404)

    AID_TYPES = {
        'food':'غذائية','medical':'طبية','financial':'مالية',
        'clothing':'ملابس','furniture':'أثاث','education':'تعليمية','other':'أخرى'
    }

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = 'المساعدات'
    ws.sheet_view.rightToLeft = True

    GREEN  = 'B45309'
    thin   = Side(style='thin', color='E2E8F0')
    border = openpyxl.styles.Border(left=thin,right=thin,top=thin,bottom=thin)

    headers = ['التاريخ','اسم المساعدة','النوع','الكمية','الجهة المقدمة','معتمد الطلب','ملاحظة']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font      = Font(bold=True, color='FFFFFF')
        c.fill      = PatternFill('solid', fgColor=GREEN)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = border
    ws.row_dimensions[1].height = 26

    for r_idx, a in enumerate(aids, 2):
        created_by = (a.created_by.get_full_name() or a.created_by.username) if a.created_by else 'النظام'
        row = [str(a.date), a.name, AID_TYPES.get(a.aid_type, a.aid_type),
               a.quantity, a.provider, created_by, a.note or '']
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border    = border
            if r_idx % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='FEF3C7')
        ws.row_dimensions[r_idx].height = 20

    for col, w in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G'], [14, 24, 12, 10, 22, 20, 24]):
        ws.column_dimensions[col].width = w

    output = io.BytesIO()
    wb.save(output); output.seek(0)
    fname = f'aids_{user.get_full_name()}.xlsx'.replace(' ','_')
    resp  = HttpResponse(output.read(),
                         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(fname)}"
    return resp

# ══════════════════════════════════════════════
#  تحديث الحالة
# ══════════════════════════════════════════════

@admin_required
@require_POST
@csrf_protect
def update_status(request):
    bene_type  = request.POST.get('bene_type', '').strip()
    bene_id    = request.POST.get('bene_id',   '').strip()
    new_status = request.POST.get('status',    '').strip()

    model_map = _get_model_map()
    Model     = model_map.get(bene_type)
    if not Model:
        return JsonResponse({'status': 'error', 'message': 'نوع غير صالح'})

    try:
        form     = Model.objects.get(pk=bene_id)
        old      = form.status
        form.status = new_status
        form.save(update_fields=['status'])

        create_notification(
            recipient=form.user, ntype='SYSTEM',
            title='تم تحديث حالة استمارتك',
            message=f'تم تغيير الحالة من "{old}" إلى "{new_status}"',
            sender=request.user,
        )
        log_activity(request.user, 'UPDATE',
                     description=f'تحديث حالة: {old} → {new_status}',
                     target_model=type(form).__name__, target_id=form.pk,
                     old_value={'status': old}, new_value={'status': new_status},
                     request=request)
        return JsonResponse({'status': 'success', 'message': 'تم تحديث الحالة ✅'})
    except Exception:
        return JsonResponse({'status': 'error', 'message': 'حدث خطأ'})


# ══════════════════════════════════════════════
#  سجل النشاط
# ══════════════════════════════════════════════

@admin_required
@require_GET
def get_logs(request):
    user_id = request.GET.get('user_id', '').strip()
    if not user_id:
        return JsonResponse({'status': 'error', 'message': 'معرف غير صالح'})

    try:
        user = CustomUser.objects.get(pk=user_id)
    except CustomUser.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'المستخدم غير موجود'})

    # IDs السجلات الموجودة حالياً
    note_ids    = [str(i) for i in UserNote.objects.filter(user_id=user_id).values_list('pk', flat=True)]
    payment_ids = [str(i) for i in Payment.objects.filter(beneficiary_id=user_id).values_list('pk', flat=True)]
    aid_ids     = [str(i) for i in Aid.objects.filter(beneficiary_id=user_id).values_list('pk', flat=True)]

    user_full_name = user.get_full_name()

    logs = ActivityLog.objects.filter(
        # سجلات الأدمن على هذا المستخدم مباشرة
        Q(target_id=user_id, target_model='CustomUser') |
        # سجلات الملاحظات الموجودة
        Q(target_id__in=note_ids,    target_model='UserNote') |
        # سجلات المدفوعات الموجودة
        Q(target_id__in=payment_ids, target_model='Payment') |
        # سجلات المساعدات الموجودة
        Q(target_id__in=aid_ids,     target_model='Aid') |
        # سجلات المحذوفات — البحث باسم المستخدم في الوصف
        Q(target_model='UserNote',  description__contains=user_full_name) |
        Q(target_model='Payment',   description__contains=user_full_name) |
        Q(target_model='Aid',       description__contains=user_full_name)
    ).select_related('user').order_by('-created_at').distinct()[:200]

    return JsonResponse({'status': 'success', 'logs': [{
        'id':          l.pk,
        'action':      l.action,
        'description': l.description,
        'created_at': fmt_dt(l.created_at),
        'admin':       (l.user.get_full_name() or l.user.username) if l.user else 'النظام',
        'ip':          l.ip_address or '',
    } for l in logs]})


@admin_required
@require_GET
def export_logs(request):
    user_id = request.GET.get('user_id', '').strip()
    ids     = request.GET.get('ids', '')

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return HttpResponse('openpyxl غير مثبّتة', status=500)

    if ids and ids != 'all':
        id_list = [i.strip() for i in ids.split(',') if i.strip().isdigit()]
        logs    = ActivityLog.objects.filter(pk__in=id_list).order_by('-created_at')
    else:
        logs = ActivityLog.objects.filter(
            Q(user_id=user_id) | Q(target_id=user_id, target_model='CustomUser')
        ).order_by('-created_at')

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = 'سجل النشاط'
    ws.sheet_view.rightToLeft = True

    GREEN  = '7C3AED'
    thin   = Side(style='thin', color='E2E8F0')
    border = openpyxl.styles.Border(left=thin,right=thin,top=thin,bottom=thin)

    headers = ['التاريخ', 'الإجراء', 'التفاصيل', 'الأدمن', 'IP']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font      = Font(bold=True, color='FFFFFF')
        c.fill      = PatternFill('solid', fgColor=GREEN)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = border

    for r_idx, l in enumerate(logs, 2):
        row = [fmt_dt(l.created_at), l.action,
               l.description, l.user.get_short_name() if l.user else 'النظام',
               l.ip_address or '']
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border    = border
            if r_idx % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='F5F3FF')

    for col, w in zip(['A','B','C','D','E'], [16, 14, 40, 16, 14]):
        ws.column_dimensions[col].width = w

    output = io.BytesIO()
    wb.save(output); output.seek(0)
    resp = HttpResponse(output.read(),
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = "attachment; filename*=UTF-8''activity_log.xlsx"
    return resp


# ══════════════════════════════════════════════
#  التواصل
# ══════════════════════════════════════════════

@admin_required
@require_POST
@csrf_protect
def toggle_comm(request):
    user_id   = request.POST.get('user_id',  '').strip()
    bene_id   = request.POST.get('bene_id',  '').strip()
    bene_type = request.POST.get('bene_type','').strip()

    try:
        user = CustomUser.objects.get(pk=user_id)
        user.allow_direct_comm = not getattr(user, 'allow_direct_comm', False)
        user.save(update_fields=['allow_direct_comm'])

        action = 'تفعيل' if user.allow_direct_comm else 'إلغاء'

        # إشعار الكافل إن وجد
        model_map = _get_model_map()
        Model     = model_map.get(bene_type)
        if Model and bene_id:
            try:
                form = Model.objects.get(pk=bene_id)
                if hasattr(form, 'sponsor') and form.sponsor:
                    create_notification(
                        recipient=form.sponsor.user, ntype='COMM_ALLOWED',
                        title=f'تم {action} التواصل',
                        message=f'تم {action} التواصل مع {user.get_full_name()}',
                        sender=request.user,
                    )
            except Exception:
                pass

        log_activity(request.user, 'UPDATE',
                     description=f'{action} تواصل: {user.get_full_name()}',
                     target_model='CustomUser', target_id=user.pk, request=request)

        return JsonResponse({
            'status':     'success',
            'allow_comm': user.allow_direct_comm,
            'message':    f'تم {action} التواصل ✅',
        })
    except CustomUser.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'المستخدم غير موجود'})


# ══════════════════════════════════════════════
#  تغيير الكافل
# ══════════════════════════════════════════════

@admin_required
@require_POST
@csrf_protect
def change_sponsor(request):
    user_id    = request.POST.get('user_id',   '').strip()
    sponsor_id = request.POST.get('sponsor_id','').strip()
    bene_type  = request.POST.get('bene_type', '').strip()

    try:
        user  = CustomUser.objects.get(pk=user_id)
        form  = _get_form(user, bene_type)
        if not form:
            return JsonResponse({'status': 'error', 'message': 'الاستمارة غير موجودة'})

        from sponsor.models import SponsorProfile
        sponsor = SponsorProfile.objects.get(pk=sponsor_id)
        form.sponsor = sponsor
        form.save(update_fields=['sponsor'])

        log_activity(request.user, 'UPDATE',
                     description=f'تغيير كافل {user.get_full_name()} إلى {sponsor.user.get_full_name()}',
                     target_model=type(form).__name__, target_id=form.pk, request=request)

        return JsonResponse({'status': 'success', 'message': 'تم تغيير الكافل ✅'})
    except (CustomUser.DoesNotExist, Exception) as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


@admin_required
@require_POST
@csrf_protect
def remove_sponsor(request):
    user_id   = request.POST.get('user_id',  '').strip()
    bene_type = request.POST.get('bene_type','').strip()

    try:
        user = CustomUser.objects.get(pk=user_id)
        form = _get_form(user, bene_type)
        if not form:
            return JsonResponse({'status': 'error', 'message': 'الاستمارة غير موجودة'})

        form.sponsor = None
        form.save(update_fields=['sponsor'])

        log_activity(request.user, 'UPDATE',
                     description=f'إزالة كافل من {user.get_full_name()}',
                     target_model=type(form).__name__, target_id=form.pk, request=request)

        return JsonResponse({'status': 'success', 'message': 'تم إزالة الكافل ✅'})
    except (CustomUser.DoesNotExist, Exception) as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


# ══════════════════════════════════════════════
#  تصدير Excel للقائمة
# ══════════════════════════════════════════════

@admin_required
@require_GET
def export_beneficiaries(request):
    bene_type   = request.GET.get('type',   'orphan')
    filter_type = request.GET.get('filter', 'all')

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return HttpResponse('openpyxl غير مثبّتة', status=500)

    users = CustomUser.objects.filter(
        user_type=bene_type, is_approved=True
    ).order_by('-date_joined')

    data = [_build_user_data(u, bene_type) for u in users]

    if filter_type == 'sponsored':
        data = [d for d in data if d['has_sponsor']]
    elif filter_type == 'unsponsored':
        data = [d for d in data if not d['has_sponsor']]

    TYPE_NAMES = {'orphan':'الأيتام', 'family':'الأسر', 'special':'ذوو الاحتياجات'}

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return HttpResponse('openpyxl غير مثبّتة', status=500)

    wb = openpyxl.Workbook()

    GREEN  = '1A7A4A'
    ORANGE = 'B45309'
    thin   = Side(style='thin', color='E2E8F0')
    border = openpyxl.styles.Border(left=thin,right=thin,top=thin,bottom=thin)

    headers = ['رقم الاستمارة','الاسم الكامل','رقم الهوية','الكافل',
               'الحالة','المبلغ الشهري ₪','التواصل','حالة الحساب','تاريخ التسجيل']

    def _write_sheet(ws, rows, color):
        ws.sheet_view.rightToLeft = True
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font      = Font(bold=True, color='FFFFFF')
            c.fill      = PatternFill('solid', fgColor=color)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border    = border
        ws.row_dimensions[1].height = 26

        for r_idx, d in enumerate(rows, 2):
            row = [
                d['reg_number'],
                d['full_name'],
                d['id_number'],
                d['sponsor_name'] or 'غير مكفول',
                d['status'],
                d['monthly_amount_ils'],
                'مفعّل' if d.get('allow_comm') else 'معطّل',
                'نشط' if d['is_active'] else 'معطّل',
                d['date_joined'][:10],
            ]
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.border    = border
                if r_idx % 2 == 0:
                    cell.fill = PatternFill('solid', fgColor='F4F9F6')
            ws.row_dimensions[r_idx].height = 20

        for col, w in zip(['A','B','C','D','E','F','G','H','I'],
                          [18, 26, 14, 22, 16, 14, 10, 10, 14]):
            ws.column_dimensions[col].width = w

    if filter_type == 'all':
        # صفحة المكفولين
        ws1 = wb.active
        ws1.title = f'مكفولون — {TYPE_NAMES.get(bene_type,"")}'
        sponsored = [d for d in data if d['has_sponsor']]
        _write_sheet(ws1, sponsored, GREEN)

        # صفحة غير المكفولين
        ws2 = wb.create_sheet(title=f'غير مكفولين — {TYPE_NAMES.get(bene_type,"")}')
        unsponsored = [d for d in data if not d['has_sponsor']]
        _write_sheet(ws2, unsponsored, ORANGE)
    else:
        ws1 = wb.active
        ws1.title = TYPE_NAMES.get(bene_type, 'المستفيدون')
        _write_sheet(ws1, data, GREEN)

    output = io.BytesIO()
    wb.save(output); output.seek(0)
    fname = f'{bene_type}_{filter_type}.xlsx'
    resp  = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(fname)}"

    log_activity(request.user, 'EXPORT',
                 description=f'تصدير قائمة {bene_type} — {filter_type}',
                 request=request)
    return resp

@admin_required
@require_GET
def export_single_user(request):
    user_id = request.GET.get('id', '').strip()
    if not user_id:
        return HttpResponse('معرف غير صالح', status=400)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return HttpResponse('openpyxl غير مثبّتة', status=500)

    try:
        user = CustomUser.objects.get(pk=user_id)
    except CustomUser.DoesNotExist:
        return HttpResponse('المستخدم غير موجود', status=404)

    detail = _user_detail_bene(user)

    # ── ألوان ومساعدات ──
    GREEN  = '1A7A4A'
    PURPLE = '7C3AED'
    ORANGE = 'B45309'
    BLUE   = '2B6CB0'
    thin   = Side(style='thin', color='E2E8F0')
    border = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)

    def _hdr(ws, headers, color):
        ws.sheet_view.rightToLeft = True
        for i, h in enumerate(headers, 1):
            c           = ws.cell(row=1, column=i, value=h)
            c.font      = Font(bold=True, color='FFFFFF')
            c.fill      = PatternFill('solid', fgColor=color)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border    = border
        ws.row_dimensions[1].height = 26

    def _row(ws, r_idx, values, alt_color='F4F9F6'):
        for c_idx, val in enumerate(values, 1):
            cell           = ws.cell(row=r_idx, column=c_idx, value=str(val) if val is not None else '—')
            cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
            cell.border    = border
            if r_idx % 2 == 0:
                cell.fill = PatternFill('solid', fgColor=alt_color)
        ws.row_dimensions[r_idx].height = 20

    def _kv(ws, rows, alt='F4F9F6'):
        """كتابة صفوف مفتاح/قيمة"""
        for r_idx, (k, v) in enumerate(rows, 2):
            _row(ws, r_idx, [k, v or '—'], alt)
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 35

    wb = openpyxl.Workbook()

    # جلب الكافل
    form         = _get_form(user, user.user_type)
    sponsor_name = ''
    sponsor_reg  = ''
    sponsor_phone= ''
    if form and hasattr(form, 'sponsor') and form.sponsor:
        sp           = form.sponsor.user
        sponsor_name = sp.get_full_name()
        sponsor_reg  = sp.registration_number or '—'
        sponsor_phone= f'{sp.phone_country}{sp.phone}'

    STATUS_MAP  = {'paid': 'مدفوعة', 'pending': 'معلّقة', 'late': 'متأخرة'}
    PAID_BY_MAP = {'sponsor': 'كافل', 'admin': 'إدارة', 'external': 'جهة خارجية'}
    AID_TYPES   = {
        'food': 'غذائية', 'medical': 'طبية', 'financial': 'مالية',
        'clothing': 'ملابس', 'furniture': 'أثاث', 'education': 'تعليمية', 'other': 'أخرى',
    }

    # ══════════════════════════════════════════
    # ورقة 1 — البيانات الأساسية
    # ══════════════════════════════════════════
    ws1       = wb.active
    ws1.title = 'البيانات الأساسية'
    _hdr(ws1, ['البيان', 'القيمة'], PURPLE)
    _kv(ws1, [
        ('رقم الاستمارة',  user.registration_number or '—'),
        ('الاسم الكامل',   user.get_full_name()),
        ('الاسم الأول',    user.first_name),
        ('اسم الأب',       user.father_name),
        ('اسم الجد',       user.grand_name),
        ('اسم العائلة',    user.family_name),
        ('رقم الهوية',     user.id_number or '—'),
        ('البريد الإلكتروني', user.email),
        ('الجوال الأول',   f'{user.phone_country}{user.phone}'),
        ('الجوال الثاني',  f'{getattr(user,"phone2_country","")}{getattr(user,"phone2","")}'.strip() or '—'),
        ('واتساب',         f'{user.whatsapp_country}{user.whatsapp}' if user.whatsapp else '—'),
        ('الجنسية',        user.nationality or '—'),
        ('الجنس',          user.gender or '—'),
        ('الكافل',         sponsor_name or 'غير مكفول'),
        ('التواصل المباشر', 'مفعّل' if getattr(user, 'allow_direct_comm', False) else 'معطّل'),
        ('حالة الحساب',    'نشط' if user.is_active else 'معطّل'),
        ('تاريخ التسجيل',  str(user.date_joined.date())),
        ('آخر دخول',       str(user.last_login.date()) if user.last_login else '—'),
    ], 'F5F3FF')

    # ══════════════════════════════════════════
    # ورقة 2 — حسب نوع المستفيد
    # ══════════════════════════════════════════
    ut = user.user_type

    if ut == 'orphan' and detail.get('orphan'):
        o        = detail['orphan']
        ws2      = wb.create_sheet(title='بيانات اليتيم')
        _hdr(ws2, ['البيان', 'القيمة'], GREEN)
        _kv(ws2, [
            ('تاريخ الميلاد',    o.get('birth_date','')),
            ('نوع اليتم',        o.get('orphan_type','')),
            ('الحالة الصحية',    o.get('health_status','')),
            ('المستوى التعليمي', o.get('education_level','')),
            ('الصف الدراسي',     o.get('school_grade','')),
            ('اسم المدرسة',      o.get('school_name','')),
            ('نوع السكن',        o.get('housing_type','')),
            ('ملكية السكن',      o.get('housing_ownership','')),
            ('مبلغ الإيجار ₪',  o.get('monthly_rent','')),
            ('المدينة الحالية',  o.get('current_city','')),
            ('الشارع الحالي',    o.get('current_street','')),
            ('أقرب معلم حالي',  o.get('current_landmark','')),
            ('المدينة السابقة',  o.get('previous_city','')),
            ('الشارع السابق',    o.get('previous_street','')),
            ('أقرب معلم سابق',  o.get('previous_landmark','')),
            ('قصة اليتيم',       o.get('story','')),
        ], 'F0FDF4')

        # الأم
        if detail.get('mother'):
            m        = detail['mother']
            ws_m     = wb.create_sheet(title='بيانات الأم')
            _hdr(ws_m, ['البيان', 'القيمة'], GREEN)
            _kv(ws_m, [
                ('الاسم الكامل',   m.get('full_name','')),
                ('رقم الهوية',     m.get('id_number','')),
                ('تاريخ الميلاد',  m.get('birth_date','')),
                ('الحالة',         'على قيد الحياة' if m.get('is_alive') else 'متوفية'),
                ('تاريخ الوفاة',   m.get('death_date','')),
                ('سبب الوفاة',     m.get('death_reason','')),
                ('الحالة الصحية',  m.get('health_status','')),
                ('المستوى التعليمي', m.get('education_level','')),
                ('المهنة',         m.get('job','')),
                ('الدخل الشهري ₪', m.get('monthly_income','')),
            ], 'F0FDF4')

        # الأب
        if detail.get('father'):
            f        = detail['father']
            ws_f     = wb.create_sheet(title='بيانات الأب')
            _hdr(ws_f, ['البيان', 'القيمة'], GREEN)
            _kv(ws_f, [
                ('الاسم الكامل',    f.get('full_name','')),
                ('رقم الهوية',      f.get('id_number','')),
                ('تاريخ الميلاد',   f.get('birth_date','')),
                ('الحالة',          'على قيد الحياة' if f.get('is_alive') else 'متوفي'),
                ('تاريخ الوفاة',    f.get('death_date','')),
                ('سبب الوفاة',      f.get('death_reason','')),
                ('الحالة الصحية',   f.get('health_status','')),
                ('المستوى التعليمي',f.get('education_level','')),
                ('المهنة',          f.get('job','')),
                ('عدد الأبناء',     f.get('children_count','')),
                ('الدخل قبل الوفاة ₪', f.get('income_before','')),
                ('المعاش بعد الوفاة ₪', f.get('pension_after','')),
            ], 'F0FDF4')

    elif ut == 'special' and detail.get('special'):
        s        = detail['special']
        ws2      = wb.create_sheet(title='بيانات ذوو الاحتياجات')
        _hdr(ws2, ['البيان', 'القيمة'], ORANGE)
        _kv(ws2, [
            ('تاريخ الميلاد',    s.get('birth_date','')),
            ('الحالة الصحية',    s.get('health_status','')),
            ('المستوى التعليمي', s.get('education_level','')),
            ('الصف الدراسي',     s.get('school_grade','')),
            ('اسم المدرسة',      s.get('school_name','')),
            ('نوع السكن',        s.get('housing_type','')),
            ('ملكية السكن',      s.get('housing_ownership','')),
            ('مبلغ الإيجار ₪',  s.get('monthly_rent','')),
            ('المدينة الحالية',  s.get('current_city','')),
            ('الشارع الحالي',    s.get('current_street','')),
            ('أقرب معلم حالي',  s.get('current_landmark','')),
            ('المدينة السابقة',  s.get('previous_city','')),
            ('الشارع السابق',    s.get('previous_street','')),
            ('تفاصيل الحالة',    s.get('case_details','')),
        ], 'FEF3C7')

    elif ut == 'family' and detail.get('family'):
        f        = detail['family']
        ws2      = wb.create_sheet(title='بيانات الأسرة')
        _hdr(ws2, ['البيان', 'القيمة'], BLUE)
        _kv(ws2, [
            ('رقم الهوية',         f.get('id_number','')),
            ('تاريخ الميلاد',      f.get('birth_date','')),
            ('الحالة الاجتماعية',  f.get('marital_status','')),
            ('الحالة الصحية',      f.get('health_status','')),
            ('المستوى التعليمي',   f.get('education_level','')),
            ('المهنة',             f.get('job','')),
            ('نوع السكن',          f.get('housing_type','')),
            ('ملكية السكن',        f.get('housing_ownership','')),
            ('مبلغ الإيجار ₪',    f.get('monthly_rent','')),
            ('المدينة الحالية',    f.get('current_city','')),
            ('الشارع الحالي',      f.get('current_street','')),
            ('أقرب معلم حالي',    f.get('current_landmark','')),
            ('المدينة السابقة',    f.get('previous_city','')),
            ('الشارع السابق',      f.get('previous_street','')),
            ('عدد أفراد الأسرة',   f.get('members_count','')),
            ('عدد المرضى',         f.get('sick_count','')),
            ('الوضع العام',        f.get('general_status','')),
        ], 'EFF6FF')

        if detail.get('wife'):
            w        = detail['wife']
            ws_w     = wb.create_sheet(title='بيانات الزوجة')
            _hdr(ws_w, ['البيان', 'القيمة'], BLUE)
            _kv(ws_w, [
                ('الاسم الكامل',    w.get('full_name','')),
                ('رقم الهوية',      w.get('id_number','')),
                ('تاريخ الميلاد',   w.get('birth_date','')),
                ('الحالة الصحية',   w.get('health_status','')),
                ('المستوى التعليمي',w.get('education_level','')),
            ], 'EFF6FF')

    elif ut == 'sponsor' and detail.get('sponsor'):
        sp       = detail['sponsor']
        ws2      = wb.create_sheet(title='بيانات الكافل')
        _hdr(ws2, ['البيان', 'القيمة'], BLUE)
        _kv(ws2, [
            ('المهنة',  sp.get('job','')),
            ('الدولة',  sp.get('country','')),
            ('المدينة', sp.get('city','')),
        ], 'EFF6FF')

    # ══════════════════════════════════════════
    # ورقة — المعيل
    # ══════════════════════════════════════════
    if detail.get('guardian'):
        g        = detail['guardian']
        ws_g     = wb.create_sheet(title='المعيل الحالي')
        _hdr(ws_g, ['البيان', 'القيمة'], PURPLE)
        _kv(ws_g, [
            ('الاسم الكامل',    g.get('full_name','')),
            ('رقم الهوية',      g.get('id_number','')),
            ('الجنس',           g.get('gender','')),
            ('صلة القرابة',     g.get('relation','')),
            ('المهنة',          g.get('job','')),
            ('الحالة الصحية',   g.get('health_status','')),
            ('المستوى التعليمي',g.get('education_level','')),
            ('الدخل الشهري ₪',  g.get('monthly_income','')),
            ('عدد من يعيلهم',   g.get('dependents','')),
        ], 'F5F3FF')

    # ══════════════════════════════════════════
    # ورقة — أفراد الأسرة
    # ══════════════════════════════════════════
    if detail.get('members'):
        ws_mem       = wb.create_sheet(title='أفراد الأسرة')
        _hdr(ws_mem, ['#','الاسم الكامل','رقم الهوية','الجنس','تاريخ الميلاد',
                      'الحالة الاجتماعية','صلة القرابة','الحالة الصحية','التعليم'], PURPLE)
        for r_idx, m in enumerate(detail['members'], 2):
            _row(ws_mem, r_idx, [
                r_idx - 1,
                m.get('full_name',''), m.get('id_number',''), m.get('gender',''),
                m.get('birth_date',''), m.get('marital_status',''), m.get('relation',''),
                m.get('health_status',''), m.get('education_level',''),
            ], 'F5F3FF')
        for col, w in zip(['A','B','C','D','E','F','G','H','I'],
                          [5, 24, 14, 8, 14, 16, 14, 14, 14]):
            ws_mem.column_dimensions[col].width = w

    # ══════════════════════════════════════════
    # ورقة — الكافل (للمستفيدين)
    # ══════════════════════════════════════════
    if ut != 'sponsor' and sponsor_name:
        ws_sp       = wb.create_sheet(title='بيانات الكافل')
        _hdr(ws_sp, ['البيان', 'القيمة'], BLUE)
        _kv(ws_sp, [
            ('اسم الكافل',      sponsor_name),
            ('رقم الاستمارة',   sponsor_reg),
            ('رقم الجوال',      sponsor_phone),
        ], 'EFF6FF')

    # ══════════════════════════════════════════
    # ورقة — المدفوعات
    # ══════════════════════════════════════════
    ws_pay       = wb.create_sheet(title='المدفوعات')
    pays         = Payment.objects.filter(
        beneficiary=user
    ).select_related('sponsor', 'created_by').order_by('-date')

    _hdr(ws_pay, ['التاريخ','المبلغ ₪','المبلغ $','مصدر الدفع',
                  'اسم الشخص/الجهة','الحالة','ملاحظة'], GREEN)
    if pays.exists():
        for r_idx, p in enumerate(pays, 2):
            if p.paid_by == 'sponsor' and p.sponsor:
                party = p.sponsor.get_full_name()
            elif p.paid_by == 'admin' and p.created_by:
                party = p.created_by.get_full_name() or p.created_by.username
            elif p.paid_by == 'external':
                party = p.paid_by_note or '—'
            else:
                party = '—'
            _row(ws_pay, r_idx, [
                str(p.date), str(p.amount_ils), str(p.amount_usd),
                PAID_BY_MAP.get(p.paid_by, p.paid_by), party,
                STATUS_MAP.get(p.status, p.status), p.note or '',
            ], 'F0FDF4')
    else:
        ws_pay.cell(row=2, column=1, value='لا توجد مدفوعات')
    for col, w in zip(['A','B','C','D','E','F','G'], [14,14,14,14,24,12,24]):
        ws_pay.column_dimensions[col].width = w

    # ══════════════════════════════════════════
    # ورقة — المساعدات
    # ══════════════════════════════════════════
    ws_aid       = wb.create_sheet(title='المساعدات')
    aids         = Aid.objects.filter(
        beneficiary=user
    ).select_related('created_by').order_by('-date')

    _hdr(ws_aid, ['التاريخ','اسم المساعدة','النوع','الكمية',
                  'الجهة المقدمة','معتمد الطلب','ملاحظة'], ORANGE)
    if aids.exists():
        for r_idx, a in enumerate(aids, 2):
            created_by = (a.created_by.get_full_name() or a.created_by.username) if a.created_by else 'النظام'
            _row(ws_aid, r_idx, [
                str(a.date), a.name, AID_TYPES.get(a.aid_type, a.aid_type),
                a.quantity, a.provider, created_by, a.note or '',
            ], 'FEF3C7')
    else:
        ws_aid.cell(row=2, column=1, value='لا توجد مساعدات')
    for col, w in zip(['A','B','C','D','E','F','G'], [14,24,12,10,22,20,24]):
        ws_aid.column_dimensions[col].width = w

    # ══════════════════════════════════════════
    # حفظ وإرسال
    # ══════════════════════════════════════════
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    name   = user.get_full_name().replace(' ', '_')
    id_num = user.id_number or str(user.pk)[:8]
    fname  = f'{name}_{id_num}.xlsx'

    resp = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(fname)}"

    log_activity(
        request.user, 'EXPORT',
        description=f'تصدير Excel كامل: {user.get_full_name()}',
        request=request,
    )
    return resp
@admin_required
@require_POST
@csrf_protect
def edit_note(request):
    note_id = request.POST.get('note_id', '').strip()
    note    = request.POST.get('note', '').strip()

    if not note_id or not note:
        return JsonResponse({'status': 'error', 'message': 'بيانات ناقصة'})

    try:
        n      = UserNote.objects.get(pk=note_id)
        old    = n.note
        n.note = note
        n.save(update_fields=['note'])

        log_activity(
            request.user, 'UPDATE',
            description  = f'تعديل ملاحظة على: {n.user.get_full_name()} — بواسطة: {request.user.get_full_name() or request.user.username}',
            target_model = 'UserNote',
            target_id    = str(n.pk),
            old_value    = {'note': old},
            new_value    = {'note': note},
            request      = request,
        )
        return JsonResponse({'status': 'success', 'message': 'تم التعديل ✅'})

    except UserNote.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'الملاحظة غير موجودة'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


@admin_required
@require_POST
@csrf_protect
def delete_note(request):
    note_id = request.POST.get('note_id', '').strip()
    try:
        n = UserNote.objects.get(pk=note_id)
        log_activity(
            request.user, 'DELETE',
            description  = f'حذف ملاحظة على: {n.user.get_full_name()} — بواسطة: {request.user.get_full_name() or request.user.username}',
            target_model = 'UserNote',
            target_id    = str(n.pk),
            old_value    = {'note': n.note},
            request      = request,
        )
        n.delete()
        return JsonResponse({'status': 'success', 'message': 'تم الحذف'})
    except UserNote.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'الملاحظة غير موجودة'})


@admin_required
@require_GET
def export_notes(request):
    user_id = request.GET.get('user_id', '').strip()
    if not user_id:
        return HttpResponse('معرف غير صالح', status=400)
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return HttpResponse('openpyxl غير مثبّتة', status=500)

    try:
        user  = CustomUser.objects.get(pk=user_id)
        notes = UserNote.objects.filter(user=user).select_related('admin').order_by('-created_at')
    except CustomUser.DoesNotExist:
        return HttpResponse('المستخدم غير موجود', status=404)

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = 'الملاحظات'
    ws.sheet_view.rightToLeft = True

    GREEN  = '7C3AED'
    thin   = Side(style='thin', color='E2E8F0')
    border = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)
    C_ALIGN = Alignment(horizontal='center', vertical='center')
    R_ALIGN = Alignment(horizontal='right',  vertical='center', wrap_text=True)

    headers = ['الوقت', 'التاريخ', 'الأدمن', 'الملاحظة']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font      = Font(bold=True, color='FFFFFF')
        c.fill      = PatternFill('solid', fgColor=GREEN)
        c.alignment = C_ALIGN
        c.border    = border
    ws.row_dimensions[1].height = 26

    for r_idx, n in enumerate(notes, 2):
        # الوقت بصيغة 12 ساعة مع صباحاً/مساءً
        hour   = n.created_at.hour
        minute = n.created_at.strftime('%M')
        if hour == 0:
            time_str = f'12:{minute} صباحاً'
        elif hour < 12:
            time_str = f'{hour}:{minute} صباحاً'
        elif hour == 12:
            time_str = f'12:{minute} مساءً'
        else:
            time_str = f'{hour - 12}:{minute} مساءً'

        # التاريخ
        date_str = n.created_at.strftime('%Y/%m/%d')

        # اسم الأدمن
        if n.admin:
            admin_name = n.admin.get_full_name() or n.admin.username
        else:
            admin_name = 'النظام'

        row = [time_str, date_str, admin_name, n.note]
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = R_ALIGN
            cell.border    = border
            if r_idx % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='F5F3FF')
        ws.row_dimensions[r_idx].height = 20

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 50

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    fname = f'notes_{user.get_full_name()}.xlsx'.replace(' ', '_')
    resp  = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(fname)}"
    return resp

@admin_required
@require_GET
def get_messages(request):
    user_id = request.GET.get('user_id', '').strip()
    if not user_id:
        return JsonResponse({'status': 'error', 'message': 'معرف غير صالح'})

    msgs = Notification.objects.filter(
        recipient_id=user_id,
        ntype='MESSAGE'
    ).select_related('sender').order_by('-created_at')

    return JsonResponse({'status': 'success', 'messages': [{
        'id':         m.pk,
        'title':      m.title,
        'message':    m.message,
        'sender':     (m.sender.get_full_name() or m.sender.username) if m.sender else 'النظام',
        'created_at': m.created_at.isoformat(),
    } for m in msgs]})


@admin_required
@require_GET
def export_messages(request):
    user_id = request.GET.get('user_id', '').strip()
    if not user_id:
        return HttpResponse('معرف غير صالح', status=400)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return HttpResponse('openpyxl غير مثبّتة', status=500)

    try:
        user = CustomUser.objects.get(pk=user_id)
        msgs = Notification.objects.filter(
            recipient=user, ntype='MESSAGE'
        ).select_related('sender').order_by('-created_at')
    except CustomUser.DoesNotExist:
        return HttpResponse('المستخدم غير موجود', status=404)

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = 'الرسائل'
    ws.sheet_view.rightToLeft = True

    PURPLE = '7C3AED'
    thin   = Side(style='thin', color='E2E8F0')
    border = openpyxl.styles.Border(left=thin,right=thin,top=thin,bottom=thin)

    headers = ['التاريخ', 'الموضوع', 'الرسالة', 'المُرسِل']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font      = Font(bold=True, color='FFFFFF')
        c.fill      = PatternFill('solid', fgColor=PURPLE)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = border
    ws.row_dimensions[1].height = 26

    for r_idx, m in enumerate(msgs, 2):
        sender = (m.sender.get_full_name() or m.sender.username) if m.sender else 'النظام'
        row = [fmt_dt(m.created_at), m.title, m.message, sender]
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
            cell.border    = border
            if r_idx % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='F5F3FF')
        ws.row_dimensions[r_idx].height = 20

    for col, w in zip(['A','B','C','D'], [18, 24, 50, 20]):
        ws.column_dimensions[col].width = w

    output = io.BytesIO()
    wb.save(output); output.seek(0)
    fname = f'messages_{user.get_full_name()}.xlsx'.replace(' ','_')
    resp  = HttpResponse(output.read(),
                         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(fname)}"
    return resp

@admin_required
@require_POST
@csrf_protect
def edit_beneficiary(request):
    user_id   = request.POST.get('user_id',   '').strip()
    user_type = request.POST.get('user_type', '').strip()

    if not user_id:
        return JsonResponse({'status': 'error', 'message': 'معرف غير صالح'})

    try:
        user = CustomUser.objects.get(pk=user_id)
    except CustomUser.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'المستخدم غير موجود'})

    errors = {}

    # ── حقول CustomUser ──
    USER_FIELDS = [
        'first_name', 'father_name', 'grand_name', 'family_name',
        'id_number', 'nationality', 'gender', 'email',
        'phone_country', 'phone', 'whatsapp_country', 'whatsapp',
    ]

    for field in USER_FIELDS:
        if field not in request.POST:
            continue
        val = request.POST.get(field, '').strip()

        # تحقق التكرار
        if field in ['id_number', 'email', 'phone']:
            if val and CustomUser.objects.filter(**{field: val}).exclude(pk=user_id).exists():
                label = {'id_number': 'رقم الهوية', 'email': 'البريد', 'phone': 'الجوال'}[field]
                errors[field] = f'هذا {label} مستخدم مسبقاً'
                continue

        # تحقق المطلوب
        if field in ['first_name', 'father_name', 'family_name', 'email'] and not val:
            errors[field] = 'هذا الحقل مطلوب'
            continue

        setattr(user, field, val)

    # ── كلمة المرور ──
    new_pass     = request.POST.get('new_password',     '').strip()
    confirm_pass = request.POST.get('confirm_password', '').strip()
    if new_pass:
        if len(new_pass) < 8:
            errors['new_password'] = 'كلمة المرور يجب أن تكون 8 أحرف على الأقل'
        elif new_pass != confirm_pass:
            errors['confirm_password'] = 'كلمة المرور غير متطابقة'
        else:
            user.set_password(new_pass)

    if errors:
        return JsonResponse({'status': 'error', 'errors': errors})

    user.save()

    # ── حقول النموذج حسب النوع ──
    from beneficiary.models import OrphanForm, FamilyForm, SpecialNeedsForm

    MODEL_MAP = {
        'orphan':  OrphanForm,
        'family':  FamilyForm,
        'special': SpecialNeedsForm,
    }
    FORM_FIELDS = {
        'orphan': [
            'birth_date', 'orphan_type', 'health_status', 'education_level',
            'school_grade', 'school_name', 'housing_type', 'housing_ownership',
            'monthly_rent', 'current_city', 'current_street', 'current_landmark',
            'story', 'phone1', 'phone1_country', 'phone2', 'phone2_country',
            'whatsapp', 'whatsapp_country',
        ],
        'family': [
            'birth_date', 'marital_status', 'health_status', 'education_level',
            'job', 'housing_type', 'housing_ownership', 'monthly_rent',
            'family_members_count', 'sick_members_count',
            'current_city', 'current_street', 'current_landmark',
            'general_status', 'phone1', 'phone1_country', 'phone2', 'phone2_country',
            'whatsapp', 'whatsapp_country',
        ],
        'special': [
            'birth_date', 'health_status', 'education_level',
            'school_grade', 'school_name', 'housing_type', 'housing_ownership',
            'monthly_rent', 'current_city', 'current_street', 'current_landmark',
            'case_details', 'phone1', 'phone1_country', 'phone2', 'phone2_country',
            'whatsapp', 'whatsapp_country',
        ],
        'sponsor': ['job', 'country', 'city'],
    }

    Model = MODEL_MAP.get(user_type)
    if Model:
        try:
            form = Model.objects.get(user=user)
            for field in FORM_FIELDS.get(user_type, []):
                if field in request.POST:
                    setattr(form, field, request.POST.get(field, '').strip())
            form.save()
        except Model.DoesNotExist:
            pass

    elif user_type == 'sponsor':
        try:
            from sponsor.models import SponsorProfile
            sp = SponsorProfile.objects.get(user=user)
            for field in FORM_FIELDS['sponsor']:
                if field in request.POST:
                    setattr(sp, field, request.POST.get(field, '').strip())
            sp.save()
        except Exception:
            pass

    # ── المعيل ──
    GUARDIAN_FIELDS = {
        'guardian_full_name':  'full_name',
        'guardian_id':         'id_number',
        'guardian_relation':   'relation',
        'guardian_job':        'job',
        'guardian_income':     'monthly_income',
        'guardian_dependents': 'dependents',
    }

    try:
        # جلب نموذج المستفيد للحصول على المعيل
        if Model:
            form_obj = Model.objects.filter(user=user).first()
            if form_obj and hasattr(form_obj, 'guardian'):
                g       = form_obj.guardian
                changed = False
                for post_field, model_field in GUARDIAN_FIELDS.items():
                    if post_field in request.POST:
                        setattr(g, model_field, request.POST.get(post_field, '').strip())
                        changed = True
                if changed:
                    g.save()
    except Exception:
        pass

    # ── تسجيل النشاط ──
    try:
        log_activity(
            request.user, 'UPDATE',
            description=f'تعديل بيانات: {user.get_full_name()} — بواسطة: {request.user.get_full_name() or request.user.username}',
            target_model='CustomUser',
            target_id=str(user.pk),
            request=request,
        )
    except Exception:
        pass

    return JsonResponse({'status': 'success', 'message': 'تم الحفظ ✅'})