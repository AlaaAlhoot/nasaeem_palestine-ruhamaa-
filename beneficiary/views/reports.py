"""
beneficiary/views/reports.py
صفحة التقارير للمستفيد + APIs + تصدير Excel/PDF
"""
import os
import hashlib
from io import BytesIO
from datetime import datetime, timedelta, date
from decimal import Decimal
from urllib.parse import quote
from collections import OrderedDict

from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_GET
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, Avg, Max, Q
from django.utils import timezone

from core.models import Payment, Aid, CustomUser, Notification
from core.utils import log_activity
from beneficiary.models import OrphanForm, SpecialNeedsForm, FamilyForm

ALLOWED_BENE_TYPES = ('orphan', 'special', 'family')

PAID_BY_LABELS = dict(Payment.PAID_BY)
AID_TYPE_LABELS = dict(Aid.AID_TYPES)

ARABIC_MONTHS = ['يناير', 'فبراير', 'مارس', 'أبريل', 'مايو', 'يونيو',
                 'يوليو', 'أغسطس', 'سبتمبر', 'أكتوبر', 'نوفمبر', 'ديسمبر']


# ════════════════════════════════════════════
#  Helpers
# ════════════════════════════════════════════

def _bene_required(request):
    return request.user.is_authenticated and request.user.user_type in ALLOWED_BENE_TYPES


def _get_form(user):
    """جلب استمارة المستفيد."""
    pair = {
        'orphan':  OrphanForm,
        'special': SpecialNeedsForm,
        'family':  FamilyForm,
    }.get(user.user_type)
    if not pair:
        return None
    try:
        return pair.objects.select_related('sponsor__user').get(user=user)
    except pair.DoesNotExist:
        return None


def _get_period_range(period):
    """حدّد نطاق التاريخ حسب الفترة."""
    now = timezone.now()
    today = now.date()

    if period == 'day':
        return today, today
    elif period == 'week':
        start = today - timedelta(days=today.weekday() + 1 if today.weekday() != 6 else 0)
        return start, today
    elif period == '2weeks':
        return today - timedelta(days=14), today
    elif period == 'month':
        return today.replace(day=1), today
    elif period == 'year':
        return today - timedelta(days=365), today
    else:  # 'all'
        return None, None


def _filter_by_period(qs, period, date_field='date'):
    """تطبيق الفلتر الزمني على queryset."""
    start, end = _get_period_range(period)
    if start and end:
        return qs.filter(**{f'{date_field}__gte': start, f'{date_field}__lte': end})
    return qs


def _format_arabic_date(d):
    """تنسيق التاريخ بالعربي: '15 يناير 2024'."""
    if not d:
        return ''
    return f'{d.day} {ARABIC_MONTHS[d.month - 1]} {d.year}'


def _build_sponsor_data(form, user):
    """جلب بيانات الكافل من الاستمارة + إحصائياته."""
    sp = form.sponsor if form else None
    if not sp:
        return None

    sp_user = sp.user
    start_date = form.sponsorship_date if form else None

    duration_str = ''
    if start_date:
        days = (timezone.now().date() - start_date).days
        years = days // 365
        months = (days % 365) // 30
        if years > 0 and months > 0:
            duration_str = f'{years} سنة و {months} شهر'
        elif years > 0:
            duration_str = f'{years} سنة'
        elif months > 0:
            duration_str = f'{months} شهر'
        else:
            duration_str = f'{days} يوم'

    sponsor_payments = Payment.objects.filter(beneficiary=user, sponsor=sp_user)
    total_paid = sponsor_payments.aggregate(s=Sum('amount_ils'))['s'] or Decimal('0')

    return {
        'full_name':      sp_user.get_full_name(),
        'photo':          sp.photo.url if sp.photo else None,
        'country':        sp.country or '',
        'city':           sp.city or '',
        'start_date':     _format_arabic_date(start_date),
        'duration':       duration_str,
        'total_paid':     str(total_paid),
        'payments_count': sponsor_payments.count(),
    }


def _build_timeline_period_label(period, dt):
    if period == 'day':
        return dt.strftime('%H:00')
    elif period in ('week', '2weeks'):
        return f'{dt.day} {ARABIC_MONTHS[dt.month - 1][:3]}'
    elif period == 'month':
        return f'{dt.day}'
    elif period == 'year':
        return ARABIC_MONTHS[dt.month - 1][:3]
    else:
        return f'{ARABIC_MONTHS[dt.month - 1][:3]} {dt.year}'


def _build_payments_timeline(user, period):
    qs = Payment.objects.filter(beneficiary=user)
    qs = _filter_by_period(qs, period).order_by('date')

    if not qs.exists():
        return {'labels': [], 'ils': []}

    if period == 'day':
        groups = OrderedDict()
        for p in qs:
            key = p.date.strftime('%Y-%m-%d')
            groups[key] = groups.get(key, Decimal('0')) + p.amount_ils
        labels = [_format_arabic_date(p.date) for p in qs[:10]]
        values = [float(p.amount_ils) for p in qs[:10]]
        return {'labels': labels, 'ils': values}

    elif period in ('week', '2weeks', 'month'):
        groups = OrderedDict()
        for p in qs:
            key = p.date
            groups[key] = groups.get(key, Decimal('0')) + p.amount_ils
        return {
            'labels': [_build_timeline_period_label(period, k) for k in groups.keys()],
            'ils': [float(v) for v in groups.values()]
        }

    elif period == 'year':
        groups = OrderedDict()
        for p in qs:
            key = (p.date.year, p.date.month)
            groups[key] = groups.get(key, Decimal('0')) + p.amount_ils
        labels = []
        values = []
        for (y, m), v in groups.items():
            labels.append(f'{ARABIC_MONTHS[m-1][:3]} {str(y)[-2:]}')
            values.append(float(v))
        return {'labels': labels, 'ils': values}

    else:
        groups = OrderedDict()
        for p in qs:
            key = (p.date.year, p.date.month)
            groups[key] = groups.get(key, Decimal('0')) + p.amount_ils
        items = list(groups.items())[-24:]
        labels = [f'{ARABIC_MONTHS[m-1][:3]} {str(y)[-2:]}' for (y, m), _ in items]
        values = [float(v) for _, v in items]
        return {'labels': labels, 'ils': values}


def _build_sources_data(user, period):
    qs = Payment.objects.filter(beneficiary=user)
    qs = _filter_by_period(qs, period)

    labels, values = [], []
    for code, label in Payment.PAID_BY:
        total = qs.filter(paid_by=code).aggregate(s=Sum('amount_ils'))['s'] or Decimal('0')
        labels.append(label)
        values.append(float(total))

    return {'labels': labels, 'values': values}


def _build_aids_by_type(user, period):
    qs = Aid.objects.filter(beneficiary=user)
    qs = _filter_by_period(qs, period)

    labels, values = [], []
    for code, label in Aid.AID_TYPES:
        cnt = qs.filter(aid_type=code).count()
        if cnt > 0:
            labels.append(label)
            values.append(cnt)

    return {'labels': labels, 'values': values}


def _build_heatmap(user):
    today = timezone.now().date()
    months_data = []
    for i in range(11, -1, -1):
        m = today.month - i
        y = today.year
        while m <= 0:
            m += 12
            y -= 1
        months_data.append((y, m))

    counts = []
    for y, m in months_data:
        cnt = (Payment.objects.filter(beneficiary=user, date__year=y, date__month=m).count() +
               Aid.objects.filter(beneficiary=user, date__year=y, date__month=m).count())
        counts.append(cnt)

    max_cnt = max(counts) if counts else 1
    if max_cnt == 0:
        max_cnt = 1

    cells = []
    for (y, m), cnt in zip(months_data, counts):
        ratio = cnt / max_cnt
        if cnt == 0:        level = 0
        elif ratio < 0.25:  level = 1
        elif ratio < 0.5:   level = 2
        elif ratio < 0.75:  level = 3
        else:               level = 4

        cells.append({
            'label': f'{ARABIC_MONTHS[m-1]} {y}',
            'short': ARABIC_MONTHS[m-1][:3],
            'count': cnt,
            'level': level,
        })

    return cells


def _build_timeline(user, period, limit=10):
    events = []

    pq = Payment.objects.filter(beneficiary=user)
    pq = _filter_by_period(pq, period).order_by('-date')[:limit]
    for p in pq:
        events.append({
            'icon':   '💰',
            'title':  f'دفعة {PAID_BY_LABELS.get(p.paid_by, "")}',
            'date':   _format_arabic_date(p.date),
            'amount': str(p.amount_ils),
            'tag':    p.note[:40] if p.note else '',
            '_dt': timezone.make_aware(datetime.combine(p.date, datetime.min.time())) if p.date else timezone.make_aware(datetime.min),
        })

    aq = Aid.objects.filter(beneficiary=user)
    aq = _filter_by_period(aq, period).order_by('-date')[:limit]
    for a in aq:
        events.append({
            'icon':  '🎁',
            'title': f'{a.name}',
            'date':  _format_arabic_date(a.date),
            'amount': '',
            'tag':   AID_TYPE_LABELS.get(a.aid_type, ''),
            '_dt': timezone.make_aware(datetime.combine(a.date, datetime.min.time())) if a.date else timezone.make_aware(datetime.min),
        })

    nq = Notification.objects.filter(recipient=user).order_by('-created_at')[:5]
    for n in nq:
        if n.created_at:
            events.append({
                'icon':  '🔔',
                'title': n.title or 'إشعار',
                'date':  _format_arabic_date(n.created_at.date()),
                'amount': '',
                'tag':   '',
                '_dt': n.created_at if n.created_at else timezone.make_aware(datetime.min),
            })

    events.sort(key=lambda e: e['_dt'], reverse=True)
    for e in events:
        e.pop('_dt', None)
    return events[:limit]


def _build_summary_and_insights(user, period):
    qs_all = Payment.objects.filter(beneficiary=user)
    qs = _filter_by_period(qs_all, period)

    total_ils      = qs.aggregate(s=Sum('amount_ils'))['s'] or Decimal('0')
    total_usd      = qs.aggregate(s=Sum('amount_usd'))['s'] or Decimal('0')
    payments_count = qs.count()
    aids_count     = _filter_by_period(Aid.objects.filter(beneficiary=user), period).count()
    avg_payment    = qs.aggregate(a=Avg('amount_ils'))['a'] or Decimal('0')
    max_payment    = qs.aggregate(m=Max('amount_ils'))['m'] or Decimal('0')

    trend_pct = None
    start, end = _get_period_range(period)
    if start and end:
        delta = end - start
        prev_start = start - delta - timedelta(days=1)
        prev_end   = start - timedelta(days=1)
        prev_total = qs_all.filter(date__gte=prev_start, date__lte=prev_end).aggregate(s=Sum('amount_ils'))['s'] or Decimal('0')
        if prev_total > 0:
            trend_pct = float(((total_ils - prev_total) / prev_total) * 100)

    summary = {
        'total_ils':      str(total_ils),
        'total_usd':      str(total_usd),
        'payments_count': payments_count,
        'aids_count':     aids_count,
        'avg_payment':    str(avg_payment),
        'max_payment':    str(max_payment),
        'trend_pct':      trend_pct,
    }

    form = _get_form(user)
    sponsorship_months = None
    if form and form.sponsorship_date:
        days = (timezone.now().date() - form.sponsorship_date).days
        sponsorship_months = days // 30

    last_p = qs_all.order_by('-date').first()
    last_payment_date  = _format_arabic_date(last_p.date) if last_p else None
    days_since_last    = (timezone.now().date() - last_p.date).days if last_p and last_p.date else None

    best_month_label = None
    if qs_all.exists():
        groups = {}
        for p in qs_all:
            if p.date:
                key = (p.date.year, p.date.month)
                groups[key] = groups.get(key, Decimal('0')) + p.amount_ils
        if groups:
            best = max(groups.items(), key=lambda x: x[1])
            (y, m), amt = best
            best_month_label = f'{ARABIC_MONTHS[m-1]} {y} ({amt:,.0f} ₪)'

    today = timezone.now().date()
    cur_m_total  = qs_all.filter(date__year=today.year, date__month=today.month).aggregate(s=Sum('amount_ils'))['s'] or Decimal('0')
    prev_y, prev_m = (today.year, today.month - 1) if today.month > 1 else (today.year - 1, 12)
    prev_m_total = qs_all.filter(date__year=prev_y, date__month=prev_m).aggregate(s=Sum('amount_ils'))['s'] or Decimal('0')
    vs_prev = None
    if prev_m_total > 0:
        vs_prev = float(((cur_m_total - prev_m_total) / prev_m_total) * 100)

    insights = {
        'sponsorship_months':      sponsorship_months,
        'last_payment_date':       last_payment_date,
        'days_since_last_payment': days_since_last,
        'best_month':              best_month_label,
        'vs_previous_month':       vs_prev,
    }

    return summary, insights


# ════════════════════════════════════════════
#  Page View
# ════════════════════════════════════════════

@login_required(login_url='/login/')
def reports_view(request):
    if not _bene_required(request):
        return redirect('/')
    return render(request, 'beneficiary/reports.html', {})


# ════════════════════════════════════════════
#  Data API
# ════════════════════════════════════════════

@login_required(login_url='/login/')
@require_GET
def reports_data_api(request):
    if not _bene_required(request):
        return JsonResponse({'status': 'error', 'message': 'غير مسموح'}, status=403)

    period = request.GET.get('period', 'all').strip()
    if period not in ('day', 'week', '2weeks', 'month', 'year', 'all'):
        period = 'all'

    user = request.user
    form = _get_form(user)

    summary, insights = _build_summary_and_insights(user, period)

    return JsonResponse({
        'status': 'success',
        'data': {
            'period':            period,
            'sponsor':           _build_sponsor_data(form, user) if form else None,
            'summary':           summary,
            'insights':          insights,
            'payments_timeline': _build_payments_timeline(user, period),
            'sources':           _build_sources_data(user, period),
            'aids_by_type':      _build_aids_by_type(user, period),
            'heatmap':           _build_heatmap(user),
            'timeline':          _build_timeline(user, period, limit=20),
        }
    })


# ════════════════════════════════════════════
#  Export Excel
# ════════════════════════════════════════════

@login_required(login_url='/login/')
@require_GET
def reports_export_excel_api(request):
    if not _bene_required(request):
        return JsonResponse({'status': 'error', 'message': 'غير مسموح'}, status=403)

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    period = request.GET.get('period', 'all').strip()
    user   = request.user
    form   = _get_form(user)

    HDR_FILL = PatternFill('solid', fgColor='1A7A4A')
    HDR_FONT = Font(name='Tajawal', size=12, bold=True, color='FFFFFF')
    LBL_FONT = Font(name='Tajawal', size=11, bold=True, color='1A7A4A')
    VAL_FONT = Font(name='Tajawal', size=11)
    ALIGN_C  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ALIGN_R  = Alignment(horizontal='right', vertical='center', wrap_text=True)
    THIN     = Side(style='thin', color='C8E6CF')
    BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    ALT_FILL = PatternFill('solid', fgColor='F4F9F6')

    wb = Workbook()
    wb.remove(wb.active)

    # ─── 1) الملخص ───
    ws = wb.create_sheet(title='الملخص')
    ws.sheet_view.rightToLeft = True
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 35

    ws.merge_cells('A1:B1')
    c = ws.cell(row=1, column=1, value=f'📊 تقرير المستفيد — {user.get_full_name()}')
    c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = ALIGN_C
    ws.row_dimensions[1].height = 32

    period_label = {'day':'اليوم','week':'الأسبوع','2weeks':'الأسبوعين','month':'الشهر','year':'آخر سنة','all':'كل الوقت'}.get(period, 'كل الوقت')

    summary, insights = _build_summary_and_insights(user, period)

    pairs = [
        ('الفترة',              period_label),
        ('تاريخ التقرير',       timezone.now().strftime('%Y-%m-%d %H:%M')),
        ('—',                   ''),
        ('إجمالي المدفوعات ₪', float(Decimal(summary['total_ils']))),
        ('إجمالي بالدولار $',   float(Decimal(summary['total_usd']))),
        ('عدد الدفعات',         summary['payments_count']),
        ('عدد المساعدات',       summary['aids_count']),
        ('متوسط الدفعة ₪',     float(Decimal(summary['avg_payment']))),
        ('أعلى دفعة ₪',         float(Decimal(summary['max_payment']))),
        ('—',                   ''),
        ('مدة الكفالة (شهر)',   insights['sponsorship_months'] or '—'),
        ('آخر دفعة',            insights['last_payment_date'] or '—'),
        ('أفضل شهر',            insights['best_month'] or '—'),
    ]

    for i, (label, value) in enumerate(pairs):
        row = i + 3
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=value)
        c1.font = LBL_FONT; c1.alignment = ALIGN_R; c1.border = BORDER
        c2.font = VAL_FONT; c2.alignment = ALIGN_R; c2.border = BORDER
        if i % 2 == 1:
            c1.fill = ALT_FILL; c2.fill = ALT_FILL
        ws.row_dimensions[row].height = 22

    # ─── 2) الكافل ───
    ws = wb.create_sheet(title='الكافل')
    ws.sheet_view.rightToLeft = True
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 35
    ws.merge_cells('A1:B1')
    c = ws.cell(row=1, column=1, value='🤝 بيانات الكافل')
    c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = ALIGN_C
    ws.row_dimensions[1].height = 30

    sp = _build_sponsor_data(form, user) if form else None
    if sp:
        sp_pairs = [
            ('الاسم',         sp['full_name']),
            ('الدولة',        sp['country']),
            ('المدينة',       sp['city']),
            ('تاريخ الكفالة', sp['start_date']),
            ('مدة الكفالة',   sp['duration']),
            ('إجمالي ما دفع', float(Decimal(sp['total_paid']))),
            ('عدد الدفعات',   sp['payments_count']),
        ]
        for i, (label, value) in enumerate(sp_pairs):
            row = i + 3
            c1 = ws.cell(row=row, column=1, value=label)
            c2 = ws.cell(row=row, column=2, value=value)
            c1.font = LBL_FONT; c1.alignment = ALIGN_R; c1.border = BORDER
            c2.font = VAL_FONT; c2.alignment = ALIGN_R; c2.border = BORDER
            if i % 2 == 1: c1.fill = ALT_FILL; c2.fill = ALT_FILL
    else:
        ws.merge_cells('A3:B3')
        c = ws.cell(row=3, column=1, value='⏳ لا يوجد كافل حالياً')
        c.font = Font(name='Tajawal', size=12, bold=True, color='B45309')
        c.alignment = ALIGN_C

    # ─── 3) المدفوعات ───
    pq = Payment.objects.filter(beneficiary=user).select_related('sponsor')
    pq = _filter_by_period(pq, period).order_by('-date')

    ws = wb.create_sheet(title='المدفوعات')
    ws.sheet_view.rightToLeft = True
    headers = ['#', 'التاريخ', 'المبلغ ₪', 'المبلغ $', 'المصدر', 'الكافل', 'الحالة', 'الملاحظة']
    widths  = [6, 14, 14, 14, 16, 25, 12, 30]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    c = ws.cell(row=1, column=1, value=f'💰 سجل المدفوعات ({period_label})')
    c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = ALIGN_C
    ws.row_dimensions[1].height = 30

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = LBL_FONT; cell.fill = ALT_FILL
        cell.alignment = ALIGN_C; cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = w

    total_ils = total_usd = Decimal('0')
    status_labels = dict(Payment.STATUS)
    for i, p in enumerate(pq, 1):
        row = i + 2
        source = PAID_BY_LABELS.get(p.paid_by, '')
        if p.paid_by == 'external' and p.paid_by_note:
            source = f'{source} ({p.paid_by_note})'
        sponsor_name = p.sponsor.get_full_name() if p.sponsor else '—'
        values = [
            i,
            p.date.strftime('%Y-%m-%d') if p.date else '—',
            float(p.amount_ils),
            float(p.amount_usd),
            source,
            sponsor_name,
            status_labels.get(p.status, p.status),
            p.note or '—',
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = VAL_FONT
            cell.alignment = ALIGN_C if col != 8 else ALIGN_R
            cell.border = BORDER
            if i % 2 == 0: cell.fill = ALT_FILL
            if col in (3, 4): cell.number_format = '#,##0.00'
        total_ils += p.amount_ils
        total_usd += p.amount_usd

    if pq.exists():
        last = len(pq) + 3
        ws.merge_cells(start_row=last, start_column=1, end_row=last, end_column=2)
        ws.cell(row=last, column=1, value='الإجمالي').font = LBL_FONT
        c3 = ws.cell(row=last, column=3, value=float(total_ils))
        c4 = ws.cell(row=last, column=4, value=float(total_usd))
        for c in (c3, c4):
            c.font = LBL_FONT; c.alignment = ALIGN_C; c.border = BORDER
            c.fill = PatternFill('solid', fgColor='E8F5EE')
            c.number_format = '#,##0.00'

    # ─── 4) المساعدات ───
    aq = Aid.objects.filter(beneficiary=user).select_related('created_by')
    aq = _filter_by_period(aq, period).order_by('-date')

    ws = wb.create_sheet(title='المساعدات')
    ws.sheet_view.rightToLeft = True
    headers = ['#', 'التاريخ', 'الاسم', 'النوع', 'العدد', 'الجهة المقدمة', 'الملاحظة']
    widths  = [6, 14, 25, 14, 10, 25, 30]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    c = ws.cell(row=1, column=1, value=f'🎁 سجل المساعدات ({period_label})')
    c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = ALIGN_C
    ws.row_dimensions[1].height = 30

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = LBL_FONT; cell.fill = ALT_FILL
        cell.alignment = ALIGN_C; cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = w

    for i, a in enumerate(aq, 1):
        row = i + 2
        values = [
            i,
            a.date.strftime('%Y-%m-%d') if a.date else '—',
            a.name,
            AID_TYPE_LABELS.get(a.aid_type, a.aid_type),
            a.quantity,
            a.provider,
            a.note or '—',
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = VAL_FONT
            cell.alignment = ALIGN_C if col != 7 else ALIGN_R
            cell.border = BORDER
            if i % 2 == 0: cell.fill = ALT_FILL

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'report_{user.id_number or user.pk}_{period}_{timezone.now().strftime("%Y%m%d")}.xlsx'
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"

    log_activity(user, 'EXPORT', description=f'تصدير التقرير Excel ({period_label})', request=request)
    return response


# ════════════════════════════════════════════
#  Export PDF — ReportLab (قديم)
# ════════════════════════════════════════════

@login_required(login_url='/login/')
@require_GET
def reports_export_pdf_api(request):
    if not _bene_required(request):
        return JsonResponse({'status': 'error', 'message': 'غير مسموح'}, status=403)

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.pdfgen import canvas

    period = request.GET.get('period', 'all').strip()
    user   = request.user
    form   = _get_form(user)

    font_name = 'Helvetica'
    try:
        from django.conf import settings
        for p in [
            os.path.join(settings.BASE_DIR, 'static', 'fonts', 'Tajawal-Regular.ttf'),
            os.path.join(settings.BASE_DIR, 'static', 'fonts', 'Amiri-Regular.ttf'),
            os.path.join(settings.BASE_DIR, 'static', 'fonts', 'Cairo-Regular.ttf'),
        ]:
            if os.path.exists(p):
                pdfmetrics.registerFont(TTFont('ArabicFont', p))
                font_name = 'ArabicFont'
                break
    except Exception:
        pass

    def reshape(text):
        try:
            import arabic_reshaper
            from bidi.algorithm import get_display
            return get_display(arabic_reshaper.reshape(str(text or '')))
        except ImportError:
            return str(text or '')

    period_label = {'day':'اليوم','week':'الأسبوع','2weeks':'الأسبوعين','month':'الشهر','year':'آخر سنة','all':'كل الوقت'}.get(period, 'كل الوقت')
    serial    = hashlib.sha256(f'{user.pk}-{timezone.now().isoformat()}'.encode()).hexdigest()[:16].upper()
    timestamp = timezone.now().strftime('%Y-%m-%d %H:%M:%S')

    class WatermarkCanvas(canvas.Canvas):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self.pages = []

        def showPage(self):
            self.pages.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            n = len(self.pages)
            for state in self.pages:
                self.__dict__.update(state)
                self._draw_watermark()
                self._draw_footer(self._pageNumber, n)
                super().showPage()
            super().save()

        def _draw_watermark(self):
            self.saveState()
            self.setFont(font_name, 50)
            self.setFillColor(colors.HexColor('#1a7a4a'))
            self.setFillAlpha(0.06)
            text = reshape('رُحَمَاء — تقرير رسمي')
            for y in range(50, 850, 180):
                for x_off in (-100, 200, 500):
                    self.saveState()
                    self.translate(x_off, y); self.rotate(30)
                    self.drawString(0, 0, text)
                    self.restoreState()
            self.restoreState()

        def _draw_footer(self, page_num, total):
            self.saveState()
            self.setFont(font_name, 8)
            self.setFillColor(colors.HexColor('#6b7280'))
            self.setStrokeColor(colors.HexColor('#1a7a4a'))
            self.setLineWidth(0.5)
            self.line(40, 35, A4[0] - 40, 35)
            self.drawRightString(A4[0] - 40, 22, reshape(f'الرقم التسلسلي: {serial}'))
            self.drawCentredString(A4[0] / 2, 22, reshape(f'صفحة {page_num} من {total}'))
            self.drawString(40, 22, f'Generated: {timestamp}')
            self.setFont(font_name, 7)
            self.setFillColor(colors.HexColor('#c53030'))
            self.drawCentredString(A4[0] / 2, 10, reshape('⚠ هذه وثيقة محمية — أي تعديل عليها يُعدّ تزويراً'))
            self.restoreState()

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        rightMargin=1.8*cm, leftMargin=1.8*cm,
        topMargin=1.5*cm,   bottomMargin=2*cm,
        title=f'تقرير - {user.get_full_name()}',
        author='Ruhamaa System',
    )

    styles     = getSampleStyleSheet()
    title_style = ParagraphStyle('TitleAR', parent=styles['Heading1'], fontName=font_name, fontSize=18, alignment=TA_CENTER, textColor=colors.HexColor('#1a7a4a'), spaceAfter=10)
    h2_style   = ParagraphStyle('H2AR',    parent=styles['Heading2'], fontName=font_name, fontSize=13, alignment=TA_RIGHT,  textColor=colors.HexColor('#1a7a4a'), spaceAfter=8, spaceBefore=8)
    label_style = ParagraphStyle('LabelAR', parent=styles['Normal'],  fontName=font_name, fontSize=10, alignment=TA_RIGHT,  textColor=colors.HexColor('#1a7a4a'))
    value_style = ParagraphStyle('ValueAR', parent=styles['Normal'],  fontName=font_name, fontSize=10, alignment=TA_RIGHT,  textColor=colors.HexColor('#1a1a2e'))
    cell_style  = ParagraphStyle('CellAR',  parent=styles['Normal'],  fontName=font_name, fontSize=9,  alignment=TA_CENTER)
    sub_style   = ParagraphStyle('SubAR',   parent=styles['Normal'],  fontName=font_name, fontSize=11, alignment=TA_CENTER, textColor=colors.HexColor('#6b7280'), spaceAfter=6)

    story = []
    story.append(Paragraph(reshape('📊 التقرير الشامل'), title_style))
    story.append(Paragraph(reshape('جمعية رُحَمَاء الخيرية'), sub_style))
    story.append(Paragraph(reshape(f'الفترة: {period_label}'), sub_style))
    story.append(Spacer(1, 0.4*cm))

    info_data = [
        [Paragraph(reshape(user.get_full_name() or '—'), value_style), Paragraph(reshape('الاسم:'), label_style)],
        [Paragraph(reshape(user.id_number or '—'),       value_style), Paragraph(reshape('رقم الهوية:'), label_style)],
        [Paragraph(reshape(user.registration_number or '—'), value_style), Paragraph(reshape('رقم التسجيل:'), label_style)],
        [Paragraph(timestamp, value_style), Paragraph(reshape('تاريخ التقرير:'), label_style)],
    ]
    info_tbl = Table(info_data, colWidths=[10*cm, 5*cm])
    info_tbl.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),(-1,-1), colors.HexColor('#f4f9f6')),
        ('GRID',          (0,0),(-1,-1), 0.4, colors.HexColor('#c8e6cf')),
        ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
        ('LEFTPADDING',   (0,0),(-1,-1), 8),
        ('RIGHTPADDING',  (0,0),(-1,-1), 8),
        ('TOPPADDING',    (0,0),(-1,-1), 6),
        ('BOTTOMPADDING', (0,0),(-1,-1), 6),
    ]))
    story.append(info_tbl)
    story.append(Spacer(1, 0.5*cm))

    summary, insights = _build_summary_and_insights(user, period)
    story.append(Paragraph(reshape('📈 الملخص العام'), h2_style))

    stats_data = [[
        Paragraph(reshape(f'<b>{summary["payments_count"]}</b>'), cell_style),
        Paragraph(reshape(f'<b>{float(Decimal(summary["total_ils"])):,.0f} ₪</b>'), cell_style),
        Paragraph(reshape(f'<b>{float(Decimal(summary["total_usd"])):,.2f} $</b>'), cell_style),
        Paragraph(reshape(f'<b>{summary["aids_count"]}</b>'), cell_style),
        Paragraph(reshape(f'<b>{float(Decimal(summary["avg_payment"])):,.0f}</b>'), cell_style),
        Paragraph(reshape(f'<b>{float(Decimal(summary["max_payment"])):,.0f}</b>'), cell_style),
    ], [
        Paragraph(reshape('عدد الدفعات'), cell_style),
        Paragraph(reshape('الإجمالي ₪'),  cell_style),
        Paragraph(reshape('الإجمالي $'),  cell_style),
        Paragraph(reshape('المساعدات'),   cell_style),
        Paragraph(reshape('متوسط ₪'),     cell_style),
        Paragraph(reshape('أعلى ₪'),      cell_style),
    ]]
    stats_tbl = Table(stats_data, colWidths=[2.7*cm]*6)
    stats_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0,0),(-1,0), colors.HexColor('#e8f5ee')),
        ('BACKGROUND', (0,1),(-1,1), colors.HexColor('#1a7a4a')),
        ('TEXTCOLOR',  (0,1),(-1,1), colors.white),
        ('GRID',       (0,0),(-1,-1), 0.4, colors.HexColor('#1a7a4a')),
        ('ALIGN',      (0,0),(-1,-1), 'CENTER'),
        ('VALIGN',     (0,0),(-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0),(-1,-1), 8),
        ('BOTTOMPADDING',(0,0),(-1,-1), 8),
    ]))
    story.append(stats_tbl)
    story.append(Spacer(1, 0.5*cm))

    sp = _build_sponsor_data(form, user) if form else None
    if sp:
        story.append(Paragraph(reshape('🤝 بيانات الكافل'), h2_style))
        sp_data = [
            [Paragraph(reshape(sp['full_name']), value_style),                               Paragraph(reshape('الاسم:'), label_style)],
            [Paragraph(reshape(f'{sp["country"]} / {sp["city"]}'), value_style),             Paragraph(reshape('الدولة/المدينة:'), label_style)],
            [Paragraph(reshape(sp['start_date']), value_style),                              Paragraph(reshape('تاريخ الكفالة:'), label_style)],
            [Paragraph(reshape(sp['duration']), value_style),                                Paragraph(reshape('مدة الكفالة:'), label_style)],
            [Paragraph(f'{float(Decimal(sp["total_paid"])):,.0f} ₪ ({sp["payments_count"]} دفعة)', value_style), Paragraph(reshape('إجمالي ما دفع:'), label_style)],
        ]
        sp_tbl = Table(sp_data, colWidths=[10*cm, 5*cm])
        sp_tbl.setStyle(TableStyle([
            ('BACKGROUND',    (0,0),(-1,-1), colors.HexColor('#fff8e1')),
            ('GRID',          (0,0),(-1,-1), 0.4, colors.HexColor('#f59e0b')),
            ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
            ('LEFTPADDING',   (0,0),(-1,-1), 8),
            ('RIGHTPADDING',  (0,0),(-1,-1), 8),
            ('TOPPADDING',    (0,0),(-1,-1), 6),
            ('BOTTOMPADDING', (0,0),(-1,-1), 6),
        ]))
        story.append(sp_tbl)
        story.append(Spacer(1, 0.4*cm))

    pq = Payment.objects.filter(beneficiary=user).select_related('sponsor')
    pq = _filter_by_period(pq, period).order_by('-date')

    if pq.exists():
        story.append(PageBreak())
        story.append(Paragraph(reshape('💰 سجل المدفوعات'), h2_style))
        headers    = ['الملاحظة', 'الحالة', 'الكافل', 'المصدر', 'المبلغ $', 'المبلغ ₪', 'التاريخ', '#']
        table_data = [[Paragraph(reshape(h), cell_style) for h in headers]]
        status_labels = dict(Payment.STATUS)
        total_ils = total_usd = Decimal('0')

        for i, p in enumerate(pq, 1):
            sponsor = p.sponsor.get_full_name() if p.sponsor else '—'
            source  = PAID_BY_LABELS.get(p.paid_by, '')
            if p.paid_by == 'external' and p.paid_by_note:
                source = f'{source}: {p.paid_by_note[:15]}'
            row = [
                Paragraph(reshape((p.note or '—')[:35]), cell_style),
                Paragraph(reshape(status_labels.get(p.status, p.status)), cell_style),
                Paragraph(reshape(sponsor[:20]), cell_style),
                Paragraph(reshape(source[:20]), cell_style),
                Paragraph(f'{p.amount_usd:,.2f}', cell_style),
                Paragraph(f'{p.amount_ils:,.0f}', cell_style),
                Paragraph(p.date.strftime('%Y-%m-%d') if p.date else '—', cell_style),
                Paragraph(str(i), cell_style),
            ]
            table_data.append(row)
            total_ils += p.amount_ils
            total_usd += p.amount_usd

        table_data.append([
            Paragraph('', cell_style), Paragraph('', cell_style), Paragraph('', cell_style),
            Paragraph(reshape('<b>الإجمالي</b>'), cell_style),
            Paragraph(f'<b>{total_usd:,.2f}</b>', cell_style),
            Paragraph(f'<b>{total_ils:,.0f}</b>', cell_style),
            Paragraph('', cell_style), Paragraph('', cell_style),
        ])

        col_widths = [3.3*cm, 1.6*cm, 2.3*cm, 2.3*cm, 1.7*cm, 1.7*cm, 1.7*cm, 0.8*cm]
        tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND',    (0,0),(-1,0),  colors.HexColor('#1a7a4a')),
            ('TEXTCOLOR',     (0,0),(-1,0),  colors.white),
            ('FONTSIZE',      (0,0),(-1,0),  10),
            ('TOPPADDING',    (0,0),(-1,0),  7),
            ('BOTTOMPADDING', (0,0),(-1,0),  7),
            ('GRID',          (0,0),(-1,-1), 0.3, colors.HexColor('#c8e6cf')),
            ('ALIGN',         (0,0),(-1,-1), 'CENTER'),
            ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
            ('FONTSIZE',      (0,1),(-1,-1), 9),
            ('TOPPADDING',    (0,1),(-1,-1), 5),
            ('BOTTOMPADDING', (0,1),(-1,-1), 5),
            ('ROWBACKGROUNDS',(0,1),(-1,-2), [colors.white, colors.HexColor('#f4f9f6')]),
            ('BACKGROUND',    (0,-1),(-1,-1), colors.HexColor('#e8f5ee')),
            ('LINEABOVE',     (0,-1),(-1,-1), 1.5, colors.HexColor('#1a7a4a')),
        ]))
        story.append(tbl)

    aq = Aid.objects.filter(beneficiary=user)
    aq = _filter_by_period(aq, period).order_by('-date')

    if aq.exists():
        story.append(Spacer(1, 0.5*cm))
        story.append(Paragraph(reshape('🎁 سجل المساعدات'), h2_style))
        headers    = ['الملاحظة', 'الجهة', 'العدد', 'النوع', 'الاسم', 'التاريخ', '#']
        table_data = [[Paragraph(reshape(h), cell_style) for h in headers]]
        for i, a in enumerate(aq, 1):
            row = [
                Paragraph(reshape((a.note or '—')[:30]), cell_style),
                Paragraph(reshape(a.provider[:25]), cell_style),
                Paragraph(str(a.quantity), cell_style),
                Paragraph(reshape(AID_TYPE_LABELS.get(a.aid_type, '')), cell_style),
                Paragraph(reshape(a.name[:25]), cell_style),
                Paragraph(a.date.strftime('%Y-%m-%d') if a.date else '—', cell_style),
                Paragraph(str(i), cell_style),
            ]
            table_data.append(row)

        col_widths = [3*cm, 3*cm, 1.3*cm, 2*cm, 3*cm, 1.8*cm, 0.8*cm]
        tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND',    (0,0),(-1,0),  colors.HexColor('#7c3aed')),
            ('TEXTCOLOR',     (0,0),(-1,0),  colors.white),
            ('FONTSIZE',      (0,0),(-1,0),  10),
            ('TOPPADDING',    (0,0),(-1,0),  7),
            ('BOTTOMPADDING', (0,0),(-1,0),  7),
            ('GRID',          (0,0),(-1,-1), 0.3, colors.HexColor('#e9d5ff')),
            ('ALIGN',         (0,0),(-1,-1), 'CENTER'),
            ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
            ('FONTSIZE',      (0,1),(-1,-1), 9),
            ('TOPPADDING',    (0,1),(-1,-1), 5),
            ('BOTTOMPADDING', (0,1),(-1,-1), 5),
            ('ROWBACKGROUNDS',(0,1),(-1,-1), [colors.white, colors.HexColor('#faf5ff')]),
        ]))
        story.append(tbl)

    story.append(Spacer(1, 0.6*cm))
    sign_style = ParagraphStyle('Sign', parent=styles['Normal'], fontName=font_name, fontSize=8, alignment=TA_CENTER, textColor=colors.HexColor('#6b7280'))
    sign_data  = [[Paragraph(reshape(
        f'<b>الرقم التسلسلي:</b> {serial}<br/>'
        f'<b>تم التوليد في:</b> {timestamp}<br/>'
        f'<b>للتحقق من صحة الوثيقة، يرجى التواصل مع إدارة الجمعية</b>'
    ), sign_style)]]
    sign_tbl = Table(sign_data, colWidths=[16*cm])
    sign_tbl.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),(-1,-1), colors.HexColor('#fff8e1')),
        ('BOX',           (0,0),(-1,-1), 1, colors.HexColor('#f59e0b')),
        ('TOPPADDING',    (0,0),(-1,-1), 10),
        ('BOTTOMPADDING', (0,0),(-1,-1), 10),
    ]))
    story.append(sign_tbl)

    doc.build(story, canvasmaker=WatermarkCanvas)

    pdf_data = buf.getvalue()
    buf.close()

    try:
        from PyPDF2 import PdfReader, PdfWriter
        reader = PdfReader(BytesIO(pdf_data))
        writer = PdfWriter()
        for page in reader.pages:
            writer.add_page(page)
        password = (user.id_number or '0000')[-4:]
        writer.encrypt(
            user_password=password,
            owner_password=f'RUH-{serial}',
            use_128bit=True,
            permissions_flag=4,
        )
        out = BytesIO()
        writer.write(out)
        pdf_data = out.getvalue()
        out.close()
    except ImportError:
        pass

    filename = f'report_{user.id_number or user.pk}_{period}_{timezone.now().strftime("%Y%m%d")}.pdf'
    response = HttpResponse(pdf_data, content_type='application/pdf')
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"
    response['X-PDF-Password'] = (user.id_number or '0000')[-4:]

    log_activity(user, 'EXPORT', description=f'تصدير التقرير PDF ReportLab ({period_label}) — Serial: {serial}', request=request)
    return response


# ════════════════════════════════════════════
#  Export PDF — HTML → Print (جديد)
# ════════════════════════════════════════════

@login_required(login_url='/login/')
@require_GET
def reports_export_html_pdf(request):
    """
    يفتح صفحة HTML كاملة تُعرض في المتصفح وتُطبع/تُحفظ كـ PDF مباشرة.
    كلمة السر: آخر 4 أرقام من رقم الهوية.
    """
    if not _bene_required(request):
        return redirect('/login/')

    user   = request.user
    period = request.GET.get('period', 'all').strip()
    if period not in ('day', 'week', '2weeks', 'month', 'year', 'all'):
        period = 'all'

    form = _get_form(user)

    # ── البيانات ──
    pq = Payment.objects.filter(beneficiary=user).select_related('sponsor')
    pq = _filter_by_period(pq, period).order_by('-date')

    aq = Aid.objects.filter(beneficiary=user)
    aq = _filter_by_period(aq, period).order_by('-date')

    summary, insights = _build_summary_and_insights(user, period)
    sponsor_info      = _build_sponsor_data(form, user) if form else None

    # ── تنسيق الدفعات للـ template ──
    status_labels = dict(Payment.STATUS)
    payments = []
    for p in pq:
        source = PAID_BY_LABELS.get(p.paid_by, p.paid_by)
        payments.append({
            'date':         p.date.strftime('%Y-%m-%d') if p.date else '—',
            'amount_ils':   p.amount_ils,
            'amount_usd':   p.amount_usd,
            'paid_by':      p.paid_by,
            'paid_by_label': source,
            'paid_by_note': getattr(p, 'paid_by_note', ''),
            'sponsor':      p.sponsor.get_full_name() if p.sponsor else '—',
            'status':       p.status,
            'status_label': status_labels.get(p.status, p.status),
            'note':         p.note or '',
        })

    # ── تنسيق المساعدات للـ template ──
    aids = []
    for a in aq:
        aids.append({
            'date':          a.date.strftime('%Y-%m-%d') if a.date else '—',
            'name':          a.name,
            'aid_type':      a.aid_type,
            'aid_type_label': AID_TYPE_LABELS.get(a.aid_type, a.aid_type),
            'quantity':      a.quantity,
            'provider':      a.provider,
            'note':          a.note or '',
        })

    # ── الرقم التسلسلي وكلمة السر ──
    serial   = hashlib.sha256(f'{user.pk}-{timezone.now().isoformat()}'.encode()).hexdigest()[:16].upper()
    password = (user.id_number or '0000')[-4:]

    def js_hash(s):
        h = 0
        for ch in s:
            h = (31 * h + ord(ch)) & 0xFFFFFFFF
        return format(h, 'x')

    period_label = {'day':'اليوم','week':'الأسبوع','2weeks':'الأسبوعين','month':'الشهر','year':'آخر سنة','all':'كل الوقت'}.get(period, 'كل الوقت')
    report_date  = timezone.now().strftime('%Y-%m-%d %H:%M:%S')

    log_activity(user, 'EXPORT', description=f'تصدير التقرير HTML-PDF ({period_label}) — Serial: {serial}', request=request)

    return render(request, 'beneficiary/report_pdf.html', {
        'user':         user,
        'period':       period,
        'period_label': period_label,
        'summary':      summary,
        'payments':     payments,
        'aids':         aids,
        'sponsor_info': sponsor_info,
        'serial':       serial,
        'pwd_hash':     js_hash(password),
        'report_date':  report_date,
    })

@login_required(login_url='/login/')
@require_GET
def reports_download_pdf(request):
    """تحميل التقرير كـ PDF مباشرة باستخدام WeasyPrint"""
    if not _bene_required(request):
        return redirect('/login/')

    from weasyprint import HTML, CSS
    from django.template.loader import render_to_string

    user   = request.user
    period = request.GET.get('period', 'all').strip()
    if period not in ('day', 'week', '2weeks', 'month', 'year', 'all'):
        period = 'all'

    form = _get_form(user)

    pq = Payment.objects.filter(beneficiary=user).select_related('sponsor')
    pq = _filter_by_period(pq, period).order_by('-date')

    aq = Aid.objects.filter(beneficiary=user)
    aq = _filter_by_period(aq, period).order_by('-date')

    summary, insights = _build_summary_and_insights(user, period)
    sponsor_info      = _build_sponsor_data(form, user) if form else None

    status_labels = dict(Payment.STATUS)
    payments = []
    for p in pq:
        source = PAID_BY_LABELS.get(p.paid_by, p.paid_by)
        payments.append({
            'date':          p.date.strftime('%Y-%m-%d') if p.date else '—',
            'amount_ils':    p.amount_ils,
            'amount_usd':    p.amount_usd,
            'paid_by':       p.paid_by,
            'paid_by_label': source,
            'paid_by_note':  getattr(p, 'paid_by_note', ''),
            'sponsor':       p.sponsor.get_full_name() if p.sponsor else '—',
            'status':        p.status,
            'status_label':  status_labels.get(p.status, p.status),
            'note':          p.note or '',
        })

    aids = []
    for a in aq:
        aids.append({
            'date':           a.date.strftime('%Y-%m-%d') if a.date else '—',
            'name':           a.name,
            'aid_type':       a.aid_type,
            'aid_type_label': AID_TYPE_LABELS.get(a.aid_type, a.aid_type),
            'quantity':       a.quantity,
            'provider':       a.provider,
            'note':           a.note or '',
        })

    serial      = hashlib.sha256(f'{user.pk}-{timezone.now().isoformat()}'.encode()).hexdigest()[:16].upper()
    period_label = {'day':'اليوم','week':'الأسبوع','2weeks':'الأسبوعين','month':'الشهر','year':'آخر سنة','all':'كل الوقت'}.get(period, 'كل الوقت')
    report_date  = timezone.now().strftime('%Y-%m-%d %H:%M:%S')

    # رندر الـ HTML
    html_string = render_to_string('beneficiary/report_pdf.html', {
        'user':         user,
        'period':       period,
        'period_label': period_label,
        'summary':      summary,
        'payments':     payments,
        'aids':         aids,
        'sponsor_info': sponsor_info,
        'serial':       serial,
        'pwd_hash':     '',        # بدون كلمة سر عند التحميل المباشر
        'report_date':  report_date,
        'is_download':  True,      # علامة نخفي بها شاشة كلمة السر
    }, request=request)

    # تحويل لـ PDF
    pdf_file = HTML(
        string=html_string,
        base_url=request.build_absolute_uri('/')
    ).write_pdf()

    filename = f'report_{user.id_number or user.pk}_{period}_{timezone.now().strftime("%Y%m%d")}.pdf'
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"

    log_activity(user, 'EXPORT', description=f'تحميل التقرير PDF ({period_label}) — Serial: {serial}', request=request)
    return response