"""
beneficiary/views/profile.py
صفحة الملف الشخصي للمستفيد (يتيم / ذوو احتياجات / أسرة)
"""
import os
import re
from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_GET, require_POST
from django.views.decorators.csrf import csrf_protect
from django.contrib.auth.decorators import login_required
from django.db import transaction

from core.models import CustomUser
from core.utils import log_activity, compress_image
from core.validators import validate_id_number, validate_arabic_text, validate_upload_file
from beneficiary.models import (
    OrphanForm, OrphanMother, OrphanFather,
    SpecialNeedsForm, FamilyForm, FamilyWife,
    CurrentGuardian, FamilyMember,
    OrphanDocument, SpecialDocument, FamilyDocument,
    JOBS, CITIES, HEALTH_STATUS, EDUCATION_LEVEL, SCHOOL_GRADE,
    HOUSING_TYPE, HOUSING_OWNERSHIP, ORPHAN_TYPE, MARITAL_STATUS,
    GUARDIAN_RELATION, FAMILY_RELATION, DEATH_REASON,
)
from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ════════════════════════════════════════════
#  Helpers
# ════════════════════════════════════════════

MODEL_MAP = {
    'orphan':  (OrphanForm, OrphanDocument),
    'special': (SpecialNeedsForm, SpecialDocument),
    'family':  (FamilyForm, FamilyDocument),
}

ALLOWED_BENE_TYPES = ('orphan', 'special', 'family')


def _get_form(user):
    """جلب استمارة المستفيد المسجل دخوله."""
    pair = MODEL_MAP.get(user.user_type)
    if not pair:
        return None, None
    Model, DocModel = pair
    try:
        return Model.objects.select_related('sponsor__user').get(user=user), DocModel
    except Model.DoesNotExist:
        return None, DocModel


def _serialize_person(obj, fields):
    """تحويل model instance إلى dict من حقول معينة."""
    if not obj:
        return None
    out = {}
    for f in fields:
        v = getattr(obj, f, None)
        if hasattr(v, 'isoformat'):
            v = v.isoformat() if v else ''
        elif v is not None and not isinstance(v, (str, int, float, bool)):
            v = str(v)
        out[f] = v if v is not None else ''
    if hasattr(obj, 'get_full_name'):
        out['full_name'] = obj.get_full_name()
    return out


def _build_profile_data(user):
    """بناء كامل بيانات المستفيد للعرض."""
    form, DocModel = _get_form(user)
    if not form:
        return None

    bene_type = user.user_type
    data = {
        'user_type': bene_type,
        'user': {
            'id':               str(user.pk),
            'username':         user.username,
            'first_name':       user.first_name,
            'father_name':      user.father_name,
            'grand_name':       user.grand_name,
            'family_name':      user.family_name,
            'full_name':        user.get_full_name(),
            'email':            user.email,
            'id_number':        user.id_number or '',
            'nationality':      user.nationality or '',
            'nationality_code': user.nationality_code or 'PS',
            'gender':           user.gender or '',
            'phone':            user.phone or '',
            'phone_country':    user.phone_country or '+970',
            'whatsapp':         user.whatsapp or '',
            'whatsapp_country': user.whatsapp_country or '+970',
            'reg_number':       user.registration_number or '',
            'date_joined':      user.date_joined.isoformat(),
            'last_login':       user.last_login.isoformat() if user.last_login else None,
        },
        'form': {
            'id':                 form.pk,
            'form_number':        form.form_number,

            'status':             form.status,
            'birth_date':         form.birth_date.isoformat() if form.birth_date else '',
            'health_status':      form.health_status,
            'education_level':    form.education_level,
            'school_grade':       getattr(form, 'school_grade', '') or '',
            'school_name':        getattr(form, 'school_name', '') or '',
            'photo':              form.photo.url if form.photo else None,
            'phone1':             form.phone1,
            'phone1_country':     form.phone1_country,
            'phone2':             form.phone2 or '',
            'phone2_country':     form.phone2_country,
            'whatsapp':           form.whatsapp,
            'whatsapp_country':   form.whatsapp_country,
            'current_city':       form.current_city,
            'current_street':     form.current_street,
            'current_landmark':   form.current_landmark,
            'previous_city':      form.previous_city or '',
            'previous_street':    form.previous_street or '',
            'previous_landmark':  form.previous_landmark or '',
            'housing_type':       form.housing_type,
            'housing_ownership':  form.housing_ownership,
            'monthly_rent':       str(form.monthly_rent) if form.monthly_rent else '',
            'allow_direct_comm':  form.allow_direct_comm,
            'sponsorship_date':   form.sponsorship_date.isoformat() if form.sponsorship_date else None,
        },
    }

    # حقول خاصة بكل نوع
    if bene_type == 'orphan':
        data['form'].update({
            'orphan_type': form.orphan_type,
            'story':       form.story,
        })
        data['mother'] = _serialize_person(
            getattr(form, 'mother', None),
            ['first_name','father_name','grand_name','family_name','birth_date',
             'id_number','nationality','nationality_code','gender','is_alive',
             'death_date','death_reason','health_status','education_level',
             'job','monthly_income']
        )
        data['father'] = _serialize_person(
            getattr(form, 'father', None),
            ['first_name','father_name','grand_name','family_name','birth_date',
             'id_number','nationality','nationality_code','gender','is_alive',
             'death_date','death_reason','health_status','education_level',
             'job','children_count','income_before','pension_after']
        )

    elif bene_type == 'special':
        data['form']['case_details'] = form.case_details

    elif bene_type == 'family':
        data['form'].update({
            'gender':               form.gender,
            'is_alive':             form.is_alive,
            'marital_status':       form.marital_status,
            'job':                  form.job,
            'family_members_count': form.family_members_count,
            'sick_members_count':   form.sick_members_count,
            'general_status':       form.general_status,
        })
        data['wife'] = _serialize_person(
            getattr(form, 'wife', None),
            ['first_name','father_name','grand_name','family_name','birth_date',
             'id_number','nationality','nationality_code','health_status',
             'education_level']
        )

    # المعيل
    g = CurrentGuardian.objects.filter(form_type=bene_type, form_id=form.pk).first()
    data['guardian'] = _serialize_person(g, [
        'first_name','father_name','grand_name','family_name','birth_date',
        'id_number','nationality','nationality_code','gender','health_status',
        'education_level','job','monthly_income','relation','dependents'
    ])

    # أفراد الأسرة
    members = FamilyMember.objects.filter(form_type=bene_type, form_id=form.pk)
    data['members'] = [
        _serialize_person(m, [
            'first_name','father_name','grand_name','family_name','birth_date',
            'id_number','nationality','gender','health_status','education_level',
            'school_grade','marital_status','relation'
        ]) for m in members
    ]

    # الكافل
    sp = form.sponsor
    if sp:
        data['sponsor'] = {
            'id':         str(sp.pk),
            'full_name':  sp.user.get_full_name(),
            'reg_number': sp.user.registration_number or '',
            'job':        sp.job or '',
            'country':    sp.country or '',
            'city':       sp.city or '',
            'photo':      sp.photo.url if sp.photo else None,
            'date':       form.sponsorship_date.isoformat() if form.sponsorship_date else None,
        }
    else:
        data['sponsor'] = None

    # المستندات
    DocModel = MODEL_MAP[bene_type][1]
    docs = DocModel.objects.filter(form=form)
    doc_choices = dict(DocModel.DOC_TYPES)
    required_set = set(getattr(DocModel, 'REQUIRED', []))
    data['documents'] = [
        {
            'id':       d.pk,
            'doc_type': d.doc_type,
            'label':    doc_choices.get(d.doc_type, d.doc_type),
            'required': d.doc_type in required_set,
            'url':      d.file.url if d.file else None,
            'name':     os.path.basename(d.file.name) if d.file else '',
            'ext':      os.path.splitext(d.file.name)[1].lower().lstrip('.') if d.file else '',
        }
        for d in docs
    ]
    # المستندات المتاحة (لإضافة جديدة)
    existing_types = {d.doc_type for d in docs}
    data['available_doc_types'] = [
        {'value': t, 'label': l, 'required': t in required_set}
        for t, l in DocModel.DOC_TYPES if t not in existing_types
    ]

    # الـ choices اللازمة في الفرونت
    data['choices'] = {
        'cities':            [list(c) for c in CITIES],
        'health':            [list(c) for c in HEALTH_STATUS],
        'education':         [list(c) for c in EDUCATION_LEVEL],
        'grades':            [list(c) for c in SCHOOL_GRADE],
        'housing_type':      [list(c) for c in HOUSING_TYPE],
        'housing_ownership': [list(c) for c in HOUSING_OWNERSHIP],
        'orphan_type':       [list(c) for c in ORPHAN_TYPE],
        'marital':           [list(c) for c in MARITAL_STATUS],
        'guardian_relation': [list(c) for c in GUARDIAN_RELATION],
        'family_relation':   [list(c) for c in FAMILY_RELATION],
        'death_reason':      [list(c) for c in DEATH_REASON],
        'jobs':              [list(c) for c in JOBS],
    }

    return data


def _id_exists(value, exclude_user_id=None):
    """تحقق من تكرار رقم الهوية في كل النظام (مع استثناء المستخدم الحالي)."""
    qs_user = CustomUser.objects.filter(id_number=value)
    if exclude_user_id:
        qs_user = qs_user.exclude(pk=exclude_user_id)
    if qs_user.exists():
        return True
    for Model in (OrphanForm, OrphanMother, OrphanFather,
                   SpecialNeedsForm, FamilyForm, FamilyWife,
                   CurrentGuardian, FamilyMember):
        qs = Model.objects.filter(id_number=value)
        if exclude_user_id:
            qs = qs.exclude(user_id=exclude_user_id) if hasattr(Model, 'user') else qs
        if qs.exists():
            return True
    return False


def _phone_exists(value, exclude_user_id=None):
    """تحقق من تكرار رقم الجوال في النظام."""
    q1 = CustomUser.objects.filter(phone=value)
    if exclude_user_id:
        q1 = q1.exclude(pk=exclude_user_id)
    if q1.exists():
        return True
    for Model in (OrphanForm, SpecialNeedsForm, FamilyForm):
        if Model.objects.filter(phone1=value).exclude(user_id=exclude_user_id or 0).exists():
            return True
        if Model.objects.filter(phone2=value).exclude(user_id=exclude_user_id or 0).exists():
            return True
    return False


def _whatsapp_exists(value, exclude_user_id=None):
    q = CustomUser.objects.filter(whatsapp=value)
    if exclude_user_id:
        q = q.exclude(pk=exclude_user_id)
    if q.exists():
        return True
    for Model in (OrphanForm, SpecialNeedsForm, FamilyForm):
        if Model.objects.filter(whatsapp=value).exclude(user_id=exclude_user_id or 0).exists():
            return True
    return False


def _email_exists(value, exclude_user_id=None):
    q = CustomUser.objects.filter(email=value.lower())
    if exclude_user_id:
        q = q.exclude(pk=exclude_user_id)
    return q.exists()


# ════════════════════════════════════════════
#  Views
# ════════════════════════════════════════════

@login_required(login_url='/login/')
def profile_view(request):
    """صفحة الملف الشخصي."""
    if request.user.user_type not in ALLOWED_BENE_TYPES:
        return redirect('/')
    return render(request, 'beneficiary/profile.html', {
        'jobs': JOBS,
    })


@login_required(login_url='/login/')
@require_GET
def profile_data_api(request):
    """API: جلب بيانات المستخدم الحالي."""
    if request.user.user_type not in ALLOWED_BENE_TYPES:
        return JsonResponse({'status': 'error', 'message': 'غير مسموح'}, status=403)

    data = _build_profile_data(request.user)
    if not data:
        return JsonResponse({'status': 'error', 'message': 'لم يتم العثور على الاستمارة'})
    return JsonResponse({'status': 'success', 'data': data})


@login_required(login_url='/login/')
@require_POST
@csrf_protect
def profile_check_field_api(request):
    """API: تحقق فوري من تفرّد حقل."""
    field = request.POST.get('field', '').strip()
    value = request.POST.get('value', '').strip()
    if not field or not value:
        return JsonResponse({'available': True})

    # حماية من XSS
    if re.search(r'[<>\'";]', value):
        return JsonResponse({'available': False, 'message': 'مدخل غير صالح'})

    uid = request.user.pk
    available, msg = True, ''

    if field == 'id_number':
        # رقم الهوية مقفل - لا يجب أن يصل لهنا
        return JsonResponse({'available': True})
    elif field == 'email':
        if _email_exists(value, exclude_user_id=uid):
            available, msg = False, 'البريد الإلكتروني مستخدم مسبقاً'
    elif field == 'phone':
        if not re.match(r'^\d{7,15}$', value):
            available, msg = False, 'رقم الجوال غير صالح'
        elif _phone_exists(value, exclude_user_id=uid):
            available, msg = False, 'رقم الجوال مستخدم مسبقاً'
    elif field == 'whatsapp':
        if not re.match(r'^\d{7,15}$', value):
            available, msg = False, 'رقم الواتساب غير صالح'
        elif _whatsapp_exists(value, exclude_user_id=uid):
            available, msg = False, 'رقم الواتساب مستخدم مسبقاً'

    return JsonResponse({'available': available, 'message': msg})


@login_required(login_url='/login/')
@require_POST
@csrf_protect
def profile_edit_api(request):
    """API: حفظ تعديلات قسم معين."""
    if request.user.user_type not in ALLOWED_BENE_TYPES:
        return JsonResponse({'status': 'error', 'message': 'غير مسموح'}, status=403)

    section = request.POST.get('section', '').strip()
    user    = request.user
    form, _ = _get_form(user)
    if not form:
        return JsonResponse({'status': 'error', 'message': 'الاستمارة غير موجودة'})

    errors = {}
    bene_type = user.user_type

    try:
        with transaction.atomic():

            # ─── الأساسي (CustomUser) ───
            if section == 'personal':
                fields = ['first_name', 'father_name', 'grand_name', 'family_name', 'gender']
                for f in fields:
                    if f in request.POST:
                        v = request.POST.get(f, '').strip()
                        if f in ('first_name', 'father_name', 'family_name') and not v:
                            errors[f] = 'هذا الحقل مطلوب'
                            continue
                        if f != 'gender':
                            try:
                                if v: validate_arabic_text(v)
                            except Exception as e:
                                errors[f] = str(e)
                                continue
                        setattr(user, f, v)
                # الجنسية
                nat = request.POST.get('nationality', '').strip()
                code = request.POST.get('nationality_code', '').strip()
                if nat:  user.nationality = nat
                if code: user.nationality_code = code
                # رقم الهوية - مقفل، نتجاهل أي محاولة تعديل
                if errors:
                    return JsonResponse({'status': 'error', 'errors': errors})
                user.save()

            # ─── التواصل ───
            elif section == 'contact':
                em = request.POST.get('email', '').strip().lower()
                if em:
                    if not re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', em):
                        errors['email'] = 'البريد غير صالح'
                    elif _email_exists(em, exclude_user_id=user.pk):
                        errors['email'] = 'البريد مستخدم مسبقاً'
                ph = request.POST.get('phone', '').strip()
                if ph:
                    if not re.match(r'^\d{7,15}$', ph):
                        errors['phone'] = 'رقم الجوال غير صالح'
                    elif _phone_exists(ph, exclude_user_id=user.pk):
                        errors['phone'] = 'رقم الجوال مستخدم مسبقاً'
                wa = request.POST.get('whatsapp', '').strip()
                if wa:
                    if not re.match(r'^\d{7,15}$', wa):
                        errors['whatsapp'] = 'رقم الواتساب غير صالح'
                    elif _whatsapp_exists(wa, exclude_user_id=user.pk):
                        errors['whatsapp'] = 'رقم الواتساب مستخدم مسبقاً'
                ph2 = request.POST.get('phone2', '').strip()
                if ph2 and ph2 == ph:
                    errors['phone2'] = 'يجب أن يختلف عن الجوال الأول'

                if errors:
                    return JsonResponse({'status': 'error', 'errors': errors})

                user.email = em or user.email
                user.phone = ph or user.phone
                user.phone_country = request.POST.get('phone_country', user.phone_country)
                user.whatsapp = wa or user.whatsapp
                user.whatsapp_country = request.POST.get('whatsapp_country', user.whatsapp_country)
                user.save()
                # نزامن مع الفورم
                form.phone1 = user.phone
                form.phone1_country = user.phone_country
                form.phone2 = ph2
                form.phone2_country = request.POST.get('phone2_country', form.phone2_country)
                form.whatsapp = user.whatsapp
                form.whatsapp_country = user.whatsapp_country
                form.save()

            # ─── بيانات المستفيد الأساسية ───
            elif section == 'main_data':
                form_fields = ['birth_date', 'health_status', 'education_level',
                               'school_grade', 'school_name', 'housing_type',
                               'housing_ownership', 'monthly_rent',
                               'current_city', 'current_street', 'current_landmark',
                               'previous_city', 'previous_street', 'previous_landmark']
                if bene_type == 'orphan':
                    form_fields += ['orphan_type', 'story']
                elif bene_type == 'special':
                    form_fields += ['case_details']
                elif bene_type == 'family':
                    form_fields += ['marital_status', 'job', 'family_members_count',
                                    'sick_members_count', 'general_status']
                for f in form_fields:
                    if f in request.POST:
                        v = request.POST.get(f, '').strip()
                        if hasattr(form, f):
                            setattr(form, f, v if v else (None if f == 'monthly_rent' else getattr(form, f)))
                form.save()

            # ─── الأم (لليتيم فقط) ───
            elif section == 'mother' and bene_type == 'orphan':
                mother = getattr(form, 'mother', None)
                if not mother:
                    return JsonResponse({'status': 'error', 'message': 'لا توجد بيانات أم'})
                _update_person(mother, request.POST, errors, exclude_uid=user.pk,
                               check_id=True, allow_id_edit=True)
                if errors:
                    return JsonResponse({'status': 'error', 'errors': errors})
                mother.save()

            # ─── الأب (لليتيم فقط) ───
            elif section == 'father' and bene_type == 'orphan':
                father = getattr(form, 'father', None)
                if not father:
                    return JsonResponse({'status': 'error', 'message': 'لا توجد بيانات أب'})
                _update_person(father, request.POST, errors, exclude_uid=user.pk,
                               check_id=True, allow_id_edit=True)
                if errors:
                    return JsonResponse({'status': 'error', 'errors': errors})
                father.save()

            # ─── الزوجة (للأسرة فقط) ───
            elif section == 'wife' and bene_type == 'family':
                wife = getattr(form, 'wife', None)
                if not wife:
                    return JsonResponse({'status': 'error', 'message': 'لا توجد بيانات زوجة'})
                _update_person(wife, request.POST, errors, exclude_uid=user.pk,
                               check_id=True, allow_id_edit=True)
                if errors:
                    return JsonResponse({'status': 'error', 'errors': errors})
                wife.save()

            # ─── المعيل ───
            elif section == 'guardian':
                g = CurrentGuardian.objects.filter(form_type=bene_type, form_id=form.pk).first()
                if not g:
                    return JsonResponse({'status': 'error', 'message': 'لا يوجد معيل'})
                _update_person(g, request.POST, errors, exclude_uid=user.pk,
                               check_id=True, allow_id_edit=True)
                # حقول إضافية للمعيل
                for f in ['relation', 'job', 'monthly_income', 'dependents']:
                    if f in request.POST:
                        setattr(g, f, request.POST.get(f, '').strip() or 0 if f in ('monthly_income','dependents') else request.POST.get(f, '').strip())
                if errors:
                    return JsonResponse({'status': 'error', 'errors': errors})
                g.save()

            # ─── صورة الملف الشخصي ───
            elif section == 'photo':
                if 'photo' not in request.FILES:
                    return JsonResponse({'status': 'error', 'message': 'لم يتم رفع صورة'})
                img = request.FILES['photo']
                try:
                    validate_upload_file(img)
                except Exception as e:
                    return JsonResponse({'status': 'error', 'message': str(e)})
                form.photo.save(img.name, compress_image(img), save=True)

            else:
                return JsonResponse({'status': 'error', 'message': 'قسم غير صالح'})

            log_activity(
                user, 'UPDATE',
                description=f'تعديل قسم "{section}" في الملف الشخصي',
                target_model=form.__class__.__name__,
                target_id=form.pk,
                request=request,
            )
            return JsonResponse({'status': 'success', 'message': 'تم الحفظ ✅'})

    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'status': 'error', 'message': f'حدث خطأ: {str(e)}'})


def _update_person(obj, data, errors, exclude_uid=None, check_id=False, allow_id_edit=False):
    """مساعدة لتحديث Person-like model (Mother/Father/Wife/Guardian)."""
    fields = ['first_name', 'father_name', 'grand_name', 'family_name',
              'birth_date', 'gender', 'health_status', 'education_level',
              'nationality', 'nationality_code']
    for f in fields:
        if f in data:
            v = data.get(f, '').strip()
            if hasattr(obj, f):
                setattr(obj, f, v)

    # رقم الهوية - مع تحقق التفرّد إن سُمح بالتعديل
    if check_id and 'id_number' in data and allow_id_edit:
        new_id = data.get('id_number', '').strip()
        if new_id and new_id != getattr(obj, 'id_number', ''):
            try:
                validate_id_number(new_id)
                if _id_exists(new_id, exclude_user_id=exclude_uid):
                    errors['id_number'] = 'رقم الهوية مسجّل مسبقاً'
                else:
                    obj.id_number = new_id
            except Exception as e:
                errors['id_number'] = str(e)

    # حقول وفاة (لو كانت موجودة)
    for f in ['is_alive', 'death_date', 'death_reason']:
        if f in data and hasattr(obj, f):
            v = data.get(f, '').strip()
            if f == 'is_alive':
                setattr(obj, f, v.lower() in ('true', '1', 'on', 'yes'))
            else:
                setattr(obj, f, v if v else None)

    # حقول مالية / عددية
    for f in ['monthly_income', 'children_count', 'income_before', 'pension_after']:
        if f in data and hasattr(obj, f):
            v = data.get(f, '').strip()
            try:
                setattr(obj, f, float(v) if v else 0)
            except ValueError:
                pass


# ════════════════════════════════════════════
#  Documents APIs
# ════════════════════════════════════════════

@login_required(login_url='/login/')
@require_POST
@csrf_protect
def document_upload_api(request):
    """API: رفع/استبدال مستند."""
    if request.user.user_type not in ALLOWED_BENE_TYPES:
        return JsonResponse({'status': 'error', 'message': 'غير مسموح'}, status=403)

    form, DocModel = _get_form(request.user)
    if not form:
        return JsonResponse({'status': 'error', 'message': 'الاستمارة غير موجودة'})

    doc_type = request.POST.get('doc_type', '').strip()
    valid_types = dict(DocModel.DOC_TYPES)
    if doc_type not in valid_types:
        return JsonResponse({'status': 'error', 'message': 'نوع المستند غير صالح'})

    if 'file' not in request.FILES:
        return JsonResponse({'status': 'error', 'message': 'لم يتم رفع ملف'})

    f = request.FILES['file']
    try:
        validate_upload_file(f)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

    try:
        # استبدال إن وُجد، وإلا إنشاء
        doc, created = DocModel.objects.get_or_create(form=form, doc_type=doc_type)
        if not created and doc.file:
            try: doc.file.delete(save=False)
            except Exception: pass
        # اسم منظم
        full_name = form.get_full_name().replace(' ', '_')
        ext = os.path.splitext(f.name)[1].lower()
        new_name = f'{full_name}_{form.id_number}_{doc_type}{ext}'
        new_name = re.sub(r'[<>:"/\\|?*\x00-\x1F]', '_', new_name)
        doc.file.save(new_name, f, save=True)

        log_activity(request.user, 'UPDATE',
                     description=f'{"رفع" if created else "استبدال"} مستند: {valid_types[doc_type]}',
                     target_model=DocModel.__name__, target_id=doc.pk, request=request)

        return JsonResponse({
            'status':  'success',
            'message': 'تم الرفع ✅',
            'doc': {
                'id':       doc.pk,
                'doc_type': doc.doc_type,
                'label':    valid_types[doc_type],
                'required': doc_type in getattr(DocModel, 'REQUIRED', []),
                'url':      doc.file.url,
                'name':     os.path.basename(doc.file.name),
                'ext':      ext.lstrip('.'),
            }
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


@login_required(login_url='/login/')
@require_POST
@csrf_protect
def document_delete_api(request):
    """API: حذف مستند (الإلزامي ممنوع حذفه)."""
    if request.user.user_type not in ALLOWED_BENE_TYPES:
        return JsonResponse({'status': 'error', 'message': 'غير مسموح'}, status=403)

    form, DocModel = _get_form(request.user)
    if not form:
        return JsonResponse({'status': 'error', 'message': 'الاستمارة غير موجودة'})

    doc_id = request.POST.get('doc_id', '').strip()
    try:
        doc = DocModel.objects.get(pk=doc_id, form=form)
    except DocModel.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'المستند غير موجود'})

    if doc.doc_type in getattr(DocModel, 'REQUIRED', []):
        return JsonResponse({
            'status':  'error',
            'message': 'لا يمكن حذف المستند الإلزامي — يمكنك استبداله فقط'
        })

    try:
        try: doc.file.delete(save=False)
        except Exception: pass
        doc_label = dict(DocModel.DOC_TYPES).get(doc.doc_type, doc.doc_type)
        doc.delete()
        log_activity(request.user, 'DELETE',
                     description=f'حذف مستند: {doc_label}',
                     target_model=DocModel.__name__, request=request)
        return JsonResponse({'status': 'success', 'message': 'تم الحذف ✅'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})
# ════════════════════════════════════════════
#  Export Excel
# ════════════════════════════════════════════

# ════════════════════════════════════════════
#  Export Excel
# ════════════════════════════════════════════

@login_required(login_url='/login/')
@require_GET
def export_excel_api(request):
    """تصدير الملف الشخصي كاملاً بصفحات منفصلة لكل تاب."""
    if request.user.user_type not in ALLOWED_BENE_TYPES:
        return JsonResponse({'status': 'error', 'message': 'غير مسموح'}, status=403)

    data = _build_profile_data(request.user)
    if not data:
        return JsonResponse({'status': 'error', 'message': 'لا توجد بيانات'})

    # ─── تنسيقات موحدة ───
    HDR_FILL  = PatternFill('solid', fgColor='1A7A4A')
    HDR_FONT  = Font(name='Tajawal', size=12, bold=True, color='FFFFFF')
    LBL_FONT  = Font(name='Tajawal', size=11, bold=True, color='1A7A4A')
    VAL_FONT  = Font(name='Tajawal', size=11)
    ALIGN_R   = Alignment(horizontal='right', vertical='center', wrap_text=True)
    ALIGN_C   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    THIN      = Side(style='thin', color='C8E6CF')
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    ALT_FILL  = PatternFill('solid', fgColor='F4F9F6')

    wb = Workbook()
    wb.remove(wb.active)

    def _make_sheet(title):
        ws = wb.create_sheet(title=title[:31])
        ws.sheet_view.rightToLeft = True
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 45
        return ws

    def _write_header(ws, text):
        ws.merge_cells('A1:B1')
        c = ws.cell(row=1, column=1, value=text)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = ALIGN_C
        ws.row_dimensions[1].height = 30

    def _write_pairs(ws, pairs, start_row=3):
        for i, (label, value) in enumerate(pairs):
            row = start_row + i
            c1 = ws.cell(row=row, column=1, value=label)
            c2 = ws.cell(row=row, column=2, value=str(value) if value not in (None, '') else '—')
            c1.font = LBL_FONT; c1.alignment = ALIGN_R; c1.border = BORDER
            c2.font = VAL_FONT; c2.alignment = ALIGN_R; c2.border = BORDER
            if i % 2 == 1:
                c1.fill = ALT_FILL; c2.fill = ALT_FILL
            ws.row_dimensions[row].height = 22

    u = data['user']
    f = data['form']

    # ─── 1) الأساسي ───
    ws = _make_sheet('الأساسي')
    _write_header(ws, f"البيانات الأساسية — {u['full_name']}")
    _write_pairs(ws, [
        ('الاسم الكامل',       u['full_name']),
        ('اسم المستخدم',       u['username']),
        ('رقم الهوية',         u['id_number']),
        ('الجنسية',            u['nationality']),
        ('الجنس',              u['gender']),
        ('البريد الإلكتروني',  u['email']),
        ('رقم الجوال',         f"{u['phone_country']} {u['phone']}"),
        ('رقم الجوال 2',       f"{f['phone2_country']} {f['phone2']}" if f['phone2'] else ''),
        ('رقم الواتساب',       f"{u['whatsapp_country']} {u['whatsapp']}"),
        ('رقم الاستمارة',      f['form_number']),
        ('رقم التسجيل',        u.get('reg_number', '')),
        ('حالة الاستمارة',     f['status']),
        ('تاريخ التسجيل',      u['date_joined'][:10]),
    ])

    # ─── 2) البيانات الرئيسية ───
    ws = _make_sheet('البيانات')
    _write_header(ws, '📋 البيانات الرئيسية')
    pairs = [
        ('تاريخ الميلاد',   f['birth_date']),
        ('الحالة الصحية',  f['health_status']),
        ('المستوى التعليمي', f['education_level']),
        ('الصف',            f.get('school_grade', '')),
        ('اسم المدرسة',     f.get('school_name', '')),
    ]
    if data['user_type'] == 'orphan':
        pairs += [('نوع اليتم', f.get('orphan_type', '')), ('قصة اليتيم', f.get('story', ''))]
    elif data['user_type'] == 'special':
        pairs += [('تفاصيل الحالة', f.get('case_details', ''))]
    elif data['user_type'] == 'family':
        pairs += [
            ('الحالة الاجتماعية',  f.get('marital_status', '')),
            ('المهنة',             f.get('job', '')),
            ('عدد أفراد الأسرة',   f.get('family_members_count', '')),
            ('عدد المرضى',         f.get('sick_members_count', '')),
            ('الوضع العام',        f.get('general_status', '')),
        ]
    _write_pairs(ws, pairs)

    # ─── 3) الأم (لليتيم) ───
    if data['user_type'] == 'orphan' and data.get('mother'):
        m = data['mother']
        ws = _make_sheet('الأم')
        _write_header(ws, '👩 بيانات الأم')
        _write_pairs(ws, [
            ('الاسم الكامل',     m['full_name']),
            ('رقم الهوية',       m['id_number']),
            ('تاريخ الميلاد',    m['birth_date']),
            ('الجنسية',          m['nationality']),
            ('الحالة الصحية',   m['health_status']),
            ('المستوى التعليمي', m['education_level']),
            ('المهنة',           m['job']),
            ('الدخل الشهري ₪',   m['monthly_income']),
            ('الحالة',           'على قيد الحياة' if m['is_alive'] else 'متوفاة'),
            ('تاريخ الوفاة',     m.get('death_date', '')),
            ('سبب الوفاة',       m.get('death_reason', '')),
        ])

    # ─── 4) الأب (لليتيم) ───
    if data['user_type'] == 'orphan' and data.get('father'):
        fa = data['father']
        ws = _make_sheet('الأب')
        _write_header(ws, '👨 بيانات الأب')
        _write_pairs(ws, [
            ('الاسم الكامل',     fa['full_name']),
            ('رقم الهوية',       fa['id_number']),
            ('تاريخ الميلاد',    fa['birth_date']),
            ('الجنسية',          fa['nationality']),
            ('الحالة الصحية',   fa['health_status']),
            ('المستوى التعليمي', fa['education_level']),
            ('المهنة',           fa['job']),
            ('عدد الأبناء',      fa['children_count']),
            ('الدخل قبل الوفاة', fa.get('income_before', '')),
            ('المعاش بعد الوفاة', fa.get('pension_after', '')),
            ('الحالة',           'على قيد الحياة' if fa['is_alive'] else 'متوفى'),
            ('تاريخ الوفاة',     fa.get('death_date', '')),
            ('سبب الوفاة',       fa.get('death_reason', '')),
        ])

    # ─── 5) الزوجة (للأسرة) ───
    if data['user_type'] == 'family' and data.get('wife'):
        w = data['wife']
        ws = _make_sheet('الزوجة')
        _write_header(ws, '💍 بيانات الزوجة')
        _write_pairs(ws, [
            ('الاسم الكامل',     w['full_name']),
            ('رقم الهوية',       w['id_number']),
            ('تاريخ الميلاد',    w['birth_date']),
            ('الجنسية',          w['nationality']),
            ('الحالة الصحية',   w['health_status']),
            ('المستوى التعليمي', w['education_level']),
        ])

    # ─── 6) المعيل ───
    if data.get('guardian'):
        g = data['guardian']
        ws = _make_sheet('المعيل')
        _write_header(ws, '🧑 المعيل الحالي')
        _write_pairs(ws, [
            ('الاسم الكامل',     g['full_name']),
            ('رقم الهوية',       g['id_number']),
            ('تاريخ الميلاد',    g['birth_date']),
            ('الجنسية',          g['nationality']),
            ('الجنس',            g['gender']),
            ('الحالة الصحية',   g['health_status']),
            ('المستوى التعليمي', g['education_level']),
            ('المهنة',           g['job']),
            ('الدخل الشهري ₪',   g['monthly_income']),
            ('صلة القرابة',      g['relation']),
            ('عدد المعالين',     g['dependents']),
        ])

    # ─── 7) أفراد الأسرة ───
    members = data.get('members', [])
    if members:
        ws = wb.create_sheet(title='أفراد الأسرة')
        ws.sheet_view.rightToLeft = True
        headers = ['#', 'الاسم الكامل', 'رقم الهوية', 'تاريخ الميلاد', 'الجنس',
                   'صلة القرابة', 'الحالة الصحية', 'المستوى التعليمي', 'الحالة الاجتماعية']
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        c = ws.cell(row=1, column=1, value=f'👥 أفراد الأسرة ({len(members)})')
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = ALIGN_C
        ws.row_dimensions[1].height = 30

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = LBL_FONT; cell.fill = ALT_FILL
            cell.alignment = ALIGN_C; cell.border = BORDER
            ws.column_dimensions[get_column_letter(col)].width = 18

        for i, m in enumerate(members, 1):
            row = i + 2
            values = [i, m['full_name'], m['id_number'], m['birth_date'], m['gender'],
                      m['relation'], m['health_status'], m['education_level'], m['marital_status']]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=str(val) if val else '—')
                cell.font = VAL_FONT; cell.alignment = ALIGN_C; cell.border = BORDER
                if i % 2 == 0: cell.fill = ALT_FILL

    # ─── 8) العنوان والسكن ───
    ws = _make_sheet('العنوان والسكن')
    _write_header(ws, '📍 العنوان والسكن')
    _write_pairs(ws, [
        ('— العنوان الحالي —', ''),
        ('المدينة',         f['current_city']),
        ('الشارع',          f['current_street']),
        ('أقرب معلم',       f['current_landmark']),
        ('— العنوان السابق —', ''),
        ('المدينة',         f['previous_city']),
        ('الشارع',          f['previous_street']),
        ('أقرب معلم',       f['previous_landmark']),
        ('— السكن —',         ''),
        ('نوع السكن',       f['housing_type']),
        ('ملكية السكن',     f['housing_ownership']),
        ('الإيجار الشهري ₪', f['monthly_rent']),
    ])

    # ─── 9) الكافل ───
    sp = data.get('sponsor')
    ws = _make_sheet('الكافل')
    _write_header(ws, '🤝 بيانات الكافل')
    if sp:
        _write_pairs(ws, [
            ('الاسم الكامل',  sp['full_name']),
            ('رقم التسجيل',   sp.get('reg_number', '')),
            ('المهنة',        sp.get('job', '')),
            ('الدولة',        sp.get('country', '')),
            ('المدينة',       sp.get('city', '')),
            ('تاريخ الكفالة', sp.get('date', '')),
        ])
    else:
        ws.merge_cells('A3:B3')
        c = ws.cell(row=3, column=1, value='⏳ لا يوجد كافل حالياً')
        c.font = Font(name='Tajawal', size=12, bold=True, color='B45309')
        c.alignment = ALIGN_C
        ws.row_dimensions[3].height = 40

    # ─── 10) المرفقات ───
    docs = data.get('documents', [])
    ws = wb.create_sheet(title='المرفقات')
    ws.sheet_view.rightToLeft = True
    headers = ['#', 'اسم المستند', 'النوع', 'إلزامي؟', 'اسم الملف']
    widths  = [6, 30, 12, 12, 40]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    c = ws.cell(row=1, column=1, value=f'📎 المرفقات ({len(docs)})')
    c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = ALIGN_C
    ws.row_dimensions[1].height = 30
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = LBL_FONT; cell.fill = ALT_FILL
        cell.alignment = ALIGN_C; cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = w

    for i, d in enumerate(docs, 1):
        row = i + 2
        values = [i, d['label'], d['ext'].upper(), 'نعم' if d['required'] else 'لا', d['name']]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=str(val))
            cell.font = VAL_FONT; cell.alignment = ALIGN_C; cell.border = BORDER
            if i % 2 == 0: cell.fill = ALT_FILL

    # ─── الحفظ والإرسال ───
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    from urllib.parse import quote
    filename = f"{u['full_name']}_{f['form_number']}.xlsx".replace(' ', '_')

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"

    log_activity(request.user, 'EXPORT',
                 description='تصدير الملف الشخصي كـ Excel',
                 target_model=type(_get_form(request.user)[0]).__name__,
                 request=request)
    return response


# ════════════════════════════════════════════
#  Family Members API
# ════════════════════════════════════════════

@login_required(login_url='/login/')
@require_POST
@csrf_protect
def member_save_api(request):
    """API: إضافة / تعديل / حذف فرد عائلة."""
    if request.user.user_type not in ALLOWED_BENE_TYPES:
        return JsonResponse({'status': 'error', 'message': 'غير مسموح'}, status=403)

    form, _ = _get_form(request.user)
    if not form:
        return JsonResponse({'status': 'error', 'message': 'الاستمارة غير موجودة'})

    action = request.POST.get('action', '').strip()
    bene_type = request.user.user_type

    try:
        # ─── الحذف ───
        if action == 'delete':
            try:
                idx = int(request.POST.get('index', -1))
            except ValueError:
                return JsonResponse({'status': 'error', 'message': 'فهرس غير صالح'})

            members = list(FamilyMember.objects.filter(form_type=bene_type, form_id=form.pk).order_by('id'))
            if idx < 0 or idx >= len(members):
                return JsonResponse({'status': 'error', 'message': 'الفرد غير موجود'})

            member = members[idx]
            member_name = member.get_full_name()
            member.delete()

            log_activity(request.user, 'DELETE',
                         description=f'حذف فرد عائلة: {member_name}',
                         target_model='FamilyMember', request=request)
            return JsonResponse({'status': 'success', 'message': 'تم الحذف ✅'})

        # ─── إضافة / تعديل ───
        if action not in ('add', 'edit'):
            return JsonResponse({'status': 'error', 'message': 'إجراء غير صالح'})

        # جلب الحقول
        data = {
            'first_name':      request.POST.get('first_name', '').strip(),
            'father_name':     request.POST.get('father_name', '').strip(),
            'grand_name':      request.POST.get('grand_name', '').strip(),
            'family_name':     request.POST.get('family_name', '').strip(),
            'id_number':       request.POST.get('id_number', '').strip(),
            'birth_date':      request.POST.get('birth_date', '').strip(),
            'gender':          request.POST.get('gender', '').strip(),
            'relation':        request.POST.get('relation', '').strip(),
            'health_status':   request.POST.get('health_status', '').strip(),
            'education_level': request.POST.get('education_level', '').strip(),
            'school_grade':    request.POST.get('school_grade', '').strip() or None,
            'marital_status':  request.POST.get('marital_status', '').strip(),
            'nationality':     'فلسطينية',
            'nationality_code': 'PS',
        }

        # تحقق
        required = ['first_name', 'father_name', 'family_name', 'id_number',
                    'birth_date', 'gender', 'relation', 'health_status',
                    'education_level', 'marital_status']
        for f in required:
            if not data[f]:
                return JsonResponse({'status': 'error', 'message': f'الحقل {f} مطلوب'})

        if not re.match(r'^[9847]\d{8}$', data['id_number']):
            return JsonResponse({'status': 'error', 'message': 'رقم هوية غير صالح'})

        with transaction.atomic():
            if action == 'add':
                # تحقق التفرّد
                if _id_exists(data['id_number']):
                    return JsonResponse({'status': 'error', 'message': 'رقم الهوية مسجّل مسبقاً'})

                FamilyMember.objects.create(
                    form_type=bene_type, form_id=form.pk, **data
                )
                log_activity(request.user, 'CREATE',
                             description=f"إضافة فرد عائلة: {data['first_name']} {data['family_name']}",
                             target_model='FamilyMember', request=request)
                return JsonResponse({'status': 'success', 'message': 'تمت الإضافة ✅'})

            else:  # edit
                try:
                    idx = int(request.POST.get('index', -1))
                except ValueError:
                    return JsonResponse({'status': 'error', 'message': 'فهرس غير صالح'})

                members = list(FamilyMember.objects.filter(form_type=bene_type, form_id=form.pk).order_by('id'))
                if idx < 0 or idx >= len(members):
                    return JsonResponse({'status': 'error', 'message': 'الفرد غير موجود'})

                member = members[idx]

                # تحقق التفرّد لو تغيّر رقم الهوية
                if data['id_number'] != member.id_number:
                    if _id_exists(data['id_number']):
                        return JsonResponse({'status': 'error', 'message': 'رقم الهوية مسجّل مسبقاً'})

                for k, v in data.items():
                    setattr(member, k, v)
                member.save()

                log_activity(request.user, 'UPDATE',
                             description=f'تعديل فرد عائلة: {member.get_full_name()}',
                             target_model='FamilyMember', target_id=member.pk, request=request)
                return JsonResponse({'status': 'success', 'message': 'تم التعديل ✅'})

    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'status': 'error', 'message': f'حدث خطأ: {str(e)}'})