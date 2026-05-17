import io
import os
from datetime import datetime, timedelta, date
from decimal import Decimal

from django.utils import timezone

from django.conf import settings

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from projects.models import Project, ProjectCategory, ProjectImage, ProjectVideo
from contact.models import ContactMessage, Newsletter, SocialMediaContact, FAQ, ContactInfo
from main.models import (
    Statistic, Goal, BoardMember, Partner,
    HomeSlider
)
from dashboard.models import ActivityLog, ReportLog
from django.contrib.auth import get_user_model
User = get_user_model()

def get_current_month_range():
    """الحصول على نطاق الشهر الحالي"""
    now = timezone.now()
    start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    # آخر يوم في الشهر
    if now.month == 12:
        end = start.replace(year=now.year + 1, month=1, day=1) - timedelta(days=1)
    else:
        end = start.replace(month=now.month + 1, day=1) - timedelta(days=1)

    end = end.replace(hour=23, minute=59, second=59)
    return start, end


def get_current_year_range():
    """الحصول على نطاق السنة الحالية"""
    now = timezone.now()
    start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    end = now.replace(month=12, day=31, hour=23, minute=59, second=59)
    return start, end


def style_header(ws, row_num, columns_count):
    """تنسيق رأس الجدول"""
    header_fill = PatternFill(start_color="6B8E23", end_color="6B8E23", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12, name='Arial')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col in range(1, columns_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border


def style_data_row(ws, row_num, columns_count, is_odd=False):
    """تنسيق صف البيانات"""
    fill_color = "F8F9FA" if is_odd else "FFFFFF"
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col in range(1, columns_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)


def auto_adjust_column_width(ws):
    """ضبط عرض الأعمدة تلقائياً"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def save_report_file(output, filename, user, report_type, period_type,
                     date_from=None, date_to=None, user_filter=None, records_count=0):
    """دالة موحدة لحفظ التقارير"""
    # الحصول على الحجم قبل أي عملية
    current_position = output.tell()
    output.seek(0, 2)  # الانتقال لنهاية الملف
    file_size = output.tell()
    output.seek(current_position)  # العودة للموضع الأصلي

    # حفظ في المجلد
    reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
    os.makedirs(reports_dir, exist_ok=True)

    file_path = os.path.join(reports_dir, filename)
    with open(file_path, 'wb') as f:
        output.seek(0)
        f.write(output.read())

    output.seek(0)

    # حفظ السجل
    ReportLog.objects.create(
        report_type=report_type,
        period_type=period_type,
        generated_by=user,
        user_filter=user_filter,
        date_from=date_from,
        date_to=date_to,
        file_name=filename,
        file_path=f'reports/{filename}',
        file_size=file_size,
        records_count=records_count
    )

    return file_size


# ========================================
# التقرير الشهري
# ========================================
def generate_monthly_report(user):
    """توليد التقرير الشهري"""
    wb = Workbook()
    start_date, end_date = get_current_month_range()

    # صفحة المشاريع
    ws_projects = wb.active
    ws_projects.title = "المشاريع"
    ws_projects.append(['#', 'اسم المشروع', 'الفئة', 'الحالة', 'المبلغ المستهدف',
                        'المبلغ المجمع', 'المستفيدين', 'تاريخ الإنشاء'])
    style_header(ws_projects, 1, 8)

    projects = Project.objects.filter(
        created_at__range=[start_date, end_date]
    ).select_related('category')

    for idx, p in enumerate(projects, 2):
        ws_projects.append([
            idx - 1, p.title_ar, p.category.name_ar, p.get_status_display(),
            float(p.target_amount), float(p.raised_amount), p.beneficiaries_count,
            p.created_at.strftime('%Y-%m-%d')
        ])
        style_data_row(ws_projects, idx, 8, idx % 2 == 0)

    auto_adjust_column_width(ws_projects)

    # صفحة المستخدمين
    ws_users = wb.create_sheet("المستخدمين")
    ws_users.append(['#', 'اسم المستخدم', 'الاسم الكامل', 'البريد', 'الجوال',
                     'الحالة', 'تاريخ التسجيل', 'آخر دخول'])
    style_header(ws_users, 1, 8)

    users = User.objects.filter(
        date_joined__range=[start_date, end_date]
    ).select_related('profile')

    for idx, u in enumerate(users, 2):
        ws_users.append([
            idx - 1, u.username, u.get_full_name(), u.email,
            u.profile.phone if hasattr(u, 'profile') else '',
            'نشط' if u.is_active else 'غير نشط',
            u.date_joined.strftime('%Y-%m-%d'),
            u.last_login.strftime('%Y-%m-%d %H:%M') if u.last_login else 'لم يسجل'
        ])
        style_data_row(ws_users, idx, 8, idx % 2 == 0)

    auto_adjust_column_width(ws_users)

    # صفحة نشاط المستخدمين
    ws_activity = wb.create_sheet("نشاط المستخدمين")
    ws_activity.append(['#', 'المستخدم', 'الإجراء', 'العنوان', 'الوصف', 'المستوى', 'التاريخ'])
    style_header(ws_activity, 1, 7)

    activities = ActivityLog.objects.filter(
        timestamp__range=[start_date, end_date]
    ).order_by('-timestamp')

    for idx, log in enumerate(activities, 2):
        ws_activity.append([
            idx - 1, log.username, log.get_action_display(), log.title,
            log.description, log.get_level_display(), log.timestamp.strftime('%Y-%m-%d %H:%M')
        ])
        style_data_row(ws_activity, idx, 7, idx % 2 == 0)

    auto_adjust_column_width(ws_activity)

    # صفحة الرسائل
    ws_messages = wb.create_sheet("الرسائل")
    ws_messages.append(['#', 'الاسم', 'البريد', 'الموضوع', 'الحالة', 'الأولوية', 'التاريخ'])
    style_header(ws_messages, 1, 7)

    messages = ContactMessage.objects.filter(
        created_at__range=[start_date, end_date]
    )

    for idx, msg in enumerate(messages, 2):
        ws_messages.append([
            idx - 1, msg.name, msg.email, msg.subject,
            msg.get_status_display(), msg.get_priority_display(),
            msg.created_at.strftime('%Y-%m-%d %H:%M')
        ])
        style_data_row(ws_messages, idx, 7, idx % 2 == 0)

    auto_adjust_column_width(ws_messages)

    # صفحة الاشتراكات
    ws_newsletter = wb.create_sheet("الاشتراكات")
    ws_newsletter.append(['#', 'البريد الإلكتروني', 'الاسم', 'التكرار', 'مفعل', 'تاريخ الاشتراك'])
    style_header(ws_newsletter, 1, 6)

    newsletters = Newsletter.objects.filter(
        subscribed_at__range=[start_date, end_date]
    )

    for idx, nl in enumerate(newsletters, 2):
        ws_newsletter.append([
            idx - 1, nl.email, nl.name or '',
            nl.get_frequency_display(),
            'نعم' if nl.is_active else 'لا',
            nl.subscribed_at.strftime('%Y-%m-%d')
        ])
        style_data_row(ws_newsletter, idx, 6, idx % 2 == 0)

    auto_adjust_column_width(ws_newsletter)

    # حفظ
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"monthly_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    save_report_file(
        output, filename, user, 'monthly_summary', 'monthly',
        start_date.date(), end_date.date(),
        None, projects.count() + users.count() + messages.count()
    )

    return output, filename


# ========================================
# تقرير المشاريع
# ========================================
def generate_projects_report(user, period_type='all', date_from=None, date_to=None, category_id=None):
    """تقرير شامل للمشاريع"""
    wb = Workbook()
    ws = wb.active
    ws.title = "المشاريع"

    # تحديد الفترة
    if period_type == 'monthly':
        start_date, end_date = get_current_month_range()
    elif period_type == 'yearly':
        start_date, end_date = get_current_year_range()
    elif period_type == 'custom' and date_from and date_to:
        start_date = timezone.make_aware(datetime.combine(date_from, datetime.min.time()))
        end_date = timezone.make_aware(datetime.combine(date_to, datetime.max.time()))
    else:
        start_date = None
        end_date = None

    ws.append(['#', 'اسم المشروع', 'الفئة', 'الحالة', 'الأولوية', 'المبلغ المستهدف',
               'المبلغ المجمع', 'نسبة الإنجاز', 'المستفيدين', 'المستفيدين المستهدفين',
               'الموقع', 'المشاهدات', 'الإعجابات', 'مميز', 'مفعل', 'تاريخ الإنشاء'])
    style_header(ws, 1, 16)


    projects = Project.objects.all().select_related('category')

    # تصفية حسب الفترة
    if start_date and end_date:
        projects = projects.filter(created_at__range=[start_date, end_date])

    # تصفية حسب التصنيف
    if category_id and category_id != 'all':
        projects = projects.filter(category_id=category_id)

    projects = projects.order_by('-created_at')

    for idx, p in enumerate(projects, 2):
        ws.append([
            idx - 1,
            p.title_ar,
            p.category.name_ar,
            p.get_status_display(),
            p.get_priority_display(),
            float(p.target_amount),
            float(p.raised_amount),
            f"{p.get_progress_percentage()}%",
            p.beneficiaries_count,
            p.target_beneficiaries,
            p.location_ar or '',
            p.views_count,
            p.likes_count,
            'نعم' if p.is_featured else 'لا',
            'نعم' if p.is_active else 'لا',
            p.created_at.strftime('%Y-%m-%d')
        ])
        style_data_row(ws, idx, 16, idx % 2 == 0)

    auto_adjust_column_width(ws)

    # حفظ
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"projects_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    save_report_file(
        output, filename, user, 'projects_report', period_type,
        start_date.date() if start_date else None,
        end_date.date() if end_date else None,
        None, projects.count()
    )

    return output, filename


# ========================================
# تقرير نشاط المستخدمين
# ========================================
def generate_users_activity_report(user, user_filter=None, period_type='monthly',
                                   date_from=None, date_to=None):
    """تقرير نشاط المستخدمين"""
    wb = Workbook()
    ws = wb.active
    ws.title = "نشاط المستخدمين"

    # تحديد الفترة
    if period_type == 'monthly':
        start_date, end_date = get_current_month_range()
    elif period_type == 'yearly':
        start_date, end_date = get_current_year_range()
    elif period_type == 'custom' and date_from and date_to:
        start_date = timezone.make_aware(datetime.combine(date_from, datetime.min.time()))
        end_date = timezone.make_aware(datetime.combine(date_to, datetime.max.time()))
    else:
        start_date = None
        end_date = None

    ws.append(['#', 'المستخدم', 'الإجراء', 'العنوان', 'الوصف', 'المستوى',
               'عنوان IP', 'المتصفح', 'التاريخ'])
    style_header(ws, 1, 9)

    activities = ActivityLog.objects.all()

    if user_filter:
        activities = activities.filter(user=user_filter)

    if start_date and end_date:
        activities = activities.filter(timestamp__range=[start_date, end_date])

    activities = activities.order_by('-timestamp')

    for idx, log in enumerate(activities, 2):
        ws.append([
            idx - 1,
            log.username,
            log.get_action_display(),
            log.title,
            log.description[:200] if log.description else '',
            log.get_level_display(),
            log.ip_address or '',
            log.user_agent[:100] if log.user_agent else '',
            log.timestamp.strftime('%Y-%m-%d %H:%M')
        ])
        style_data_row(ws, idx, 9, idx % 2 == 0)

    auto_adjust_column_width(ws)

    # حفظ
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"users_activity_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    save_report_file(
        output, filename, user, 'users_activity', period_type,
        start_date.date() if start_date else None,
        end_date.date() if end_date else None,
        user_filter, activities.count()
    )

    return output, filename


# ========================================
# تقرير الرسائل
# ========================================
def generate_messages_report(user, period_type='monthly', date_from=None, date_to=None):
    """تقرير الرسائل"""
    wb = Workbook()
    ws = wb.active
    ws.title = "الرسائل"

    # تحديد الفترة
    if period_type == 'monthly':
        start_date, end_date = get_current_month_range()
    elif period_type == 'yearly':
        start_date, end_date = get_current_year_range()
    elif period_type == 'custom' and date_from and date_to:
        start_date = timezone.make_aware(datetime.combine(date_from, datetime.min.time()))
        end_date = timezone.make_aware(datetime.combine(date_to, datetime.max.time()))
    else:
        start_date = None
        end_date = None

    ws.append(['#', 'الاسم', 'البريد الإلكتروني', 'رقم الجوال', 'الموضوع',
               'الرسالة', 'الحالة', 'الأولوية', 'عنوان IP', 'تاريخ الإرسال'])
    style_header(ws, 1, 10)

    messages = ContactMessage.objects.all()
    if start_date and end_date:
        messages = messages.filter(created_at__range=[start_date, end_date])

    messages = messages.order_by('-created_at')

    for idx, msg in enumerate(messages, 2):
        ws.append([
            idx - 1, msg.name, msg.email, str(msg.phone) if msg.phone else '',
            msg.subject, msg.message[:200], msg.get_status_display(),
            msg.get_priority_display(), msg.ip_address or '',
            msg.created_at.strftime('%Y-%m-%d %H:%M')
        ])
        style_data_row(ws, idx, 10, idx % 2 == 0)

    auto_adjust_column_width(ws)

    # حفظ
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"messages_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    save_report_file(
        output, filename, user, 'messages_report', period_type,
        start_date.date() if start_date else None,
        end_date.date() if end_date else None,
        None, messages.count()
    )

    return output, filename


# ========================================
# التقرير المخصص
# ========================================
def generate_custom_report(user, report_type, period_type='monthly',
                           date_from=None, date_to=None):
    """تقرير مخصص"""
    wb = Workbook()
    ws = wb.active

    # تحديد الفترة
    if period_type == 'monthly':
        start_date, end_date = get_current_month_range()
    elif period_type == 'yearly':
        start_date, end_date = get_current_year_range()
    elif period_type == 'custom' and date_from and date_to:
        start_date = timezone.make_aware(datetime.combine(date_from, datetime.min.time()))
        end_date = timezone.make_aware(datetime.combine(date_to, datetime.max.time()))
    else:
        start_date = None
        end_date = None

    # حسب نوع التقرير
    if report_type == 'statistics':
        ws.title = "الإحصائيات"
        ws.append(['#', 'العنوان', 'الرقم', 'اللاحقة', 'الأيقونة',
                   'اللون', 'مفعل', 'تاريخ الإنشاء'])
        style_header(ws, 1, 8)

        items = Statistic.objects.all().order_by('order')
        for idx, item in enumerate(items, 2):
            ws.append([
                idx - 1, item.title_ar, item.number, item.suffix_ar or '',
                item.icon, item.color, 'نعم' if item.is_active else 'لا',
                item.created_at.strftime('%Y-%m-%d')
            ])
            style_data_row(ws, idx, 8, idx % 2 == 0)

    elif report_type == 'goals':
        ws.title = "الأهداف"
        ws.append(['#', 'الهدف', 'الوصف', 'الأيقونة', 'مفعل', 'تاريخ الإنشاء'])
        style_header(ws, 1, 6)

        items = Goal.objects.all().order_by('order')
        for idx, item in enumerate(items, 2):
            ws.append([
                idx - 1, item.title_ar, item.description_ar or '',
                item.icon, 'نعم' if item.is_active else 'لا',
                item.created_at.strftime('%Y-%m-%d')
            ])
            style_data_row(ws, idx, 6, idx % 2 == 0)

    elif report_type == 'board_members':
        ws.title = "مجلس الإدارة"
        ws.append(['#', 'الاسم', 'المنصب', 'البريد', 'الجوال',
                   'مفعل', 'تاريخ الإضافة'])
        style_header(ws, 1, 7)

        items = BoardMember.objects.all().order_by('order')
        for idx, item in enumerate(items, 2):
            ws.append([
                idx - 1, item.name_ar, item.position_type_ar,
                item.email or '', item.phone or '',
                'نعم' if item.is_active else 'لا',
                item.created_at.strftime('%Y-%m-%d')
            ])
            style_data_row(ws, idx, 7, idx % 2 == 0)

    elif report_type == 'faqs':
        ws.title = "الأسئلة الشائعة"
        ws.append(['#', 'السؤال', 'الإجابة', 'التصنيف', 'العلامات',
                   'الترتيب', 'المشاهدات', 'تصويتات مفيد', 'مفعل', 'تاريخ الإضافة'])
        style_header(ws, 1, 10)

        items = FAQ.objects.all().select_related('category')
        if start_date and end_date:
            items = items.filter(created_at__range=[start_date, end_date])

        items = items.order_by('order', '-helpful_votes')
        for idx, item in enumerate(items, 2):
            ws.append([
                idx - 1,
                item.question_ar,
                item.answer_ar[:200] if len(item.answer_ar) > 200 else item.answer_ar,
                item.category.category_ar if item.category else '-',
                item.tags_ar or '-',
                item.order,
                item.views_count,
                item.helpful_votes,
                'نعم' if item.is_active else 'لا',
                item.created_at.strftime('%Y-%m-%d')
            ])
            style_data_row(ws, idx, 10, idx % 2 == 0)

    elif report_type == 'partners':
        ws.title = "الشركاء"
        ws.append(['#', 'الاسم', 'الوصف', 'تاريخ الشراكة', 'عدد المشاريع',
                   'الموقع', 'البريد', 'الجوال', 'مفعل', 'تاريخ الإضافة'])
        style_header(ws, 1, 10)

        items = Partner.objects.all().order_by('order')
        for idx, item in enumerate(items, 2):
            ws.append([
                idx - 1, item.name_ar, item.description_ar or '',
                item.partnership_date.strftime('%Y-%m-%d'), item.projects_count,
                item.website or '', item.email or '', item.phone or '',
                'نعم' if item.is_active else 'لا',
                item.created_at.strftime('%Y-%m-%d')
            ])
            style_data_row(ws, idx, 10, idx % 2 == 0)

    elif report_type == 'newsletters':
        ws.title = "الاشتراكات البريدية"
        ws.append(['#', 'البريد الإلكتروني', 'الاسم', 'التكرار',
                   'مفعل', 'مؤكد', 'تاريخ الاشتراك'])
        style_header(ws, 1, 7)

        items = Newsletter.objects.all()
        if start_date and end_date:
            items = items.filter(subscribed_at__range=[start_date, end_date])

        items = items.order_by('-subscribed_at')
        for idx, item in enumerate(items, 2):
            ws.append([
                idx - 1, item.email, item.name or '',
                item.get_frequency_display(),
                'نعم' if item.is_active else 'لا',
                'نعم' if item.is_confirmed else 'لا',
                item.subscribed_at.strftime('%Y-%m-%d')
            ])
            style_data_row(ws, idx, 7, idx % 2 == 0)

    elif report_type == 'social_media':
        ws.title = "وسائل التواصل"
        ws.append(['#', 'المنصة', 'اسم المستخدم', 'الرابط',
                   'عدد النقرات', 'مفعل', 'تاريخ الإضافة'])
        style_header(ws, 1, 7)

        items = SocialMediaContact.objects.all().order_by('order')
        for idx, item in enumerate(items, 2):
            ws.append([
                idx - 1, item.get_platform_display(), item.username,
                item.url, item.clicks_count,
                'نعم' if item.is_active else 'لا',
                item.created_at.strftime('%Y-%m-%d')
            ])
            style_data_row(ws, idx, 7, idx % 2 == 0)

    elif report_type == 'sliders':
        ws.title = "السلايدر"
        ws.append(['#', 'العنوان', 'العنوان الفرعي', 'نص الزر',
                   'رابط الزر', 'مفعل', 'تاريخ الإضافة'])
        style_header(ws, 1, 7)

        items = HomeSlider.objects.all().order_by('order')
        for idx, item in enumerate(items, 2):
            ws.append([
                idx - 1, item.title_ar, item.subtitle_ar or '',
                item.button_text_ar or '', item.button_url or '',
                'نعم' if item.is_active else 'لا',
                item.created_at.strftime('%Y-%m-%d')
            ])
            style_data_row(ws, idx, 7, idx % 2 == 0)

    elif report_type == 'contact_info':
        ws.title = "معلومات التواصل"
        ws.append(['#', 'النوع', 'القيمة بالعربية', 'القيمة بالإنجليزية',
                   'الأيقونة', 'مفعل', 'عرض بالتذييل', 'تاريخ الإضافة'])
        style_header(ws, 1, 8)

        items = ContactInfo.objects.all().order_by('order')
        for idx, item in enumerate(items, 2):
            ws.append([
                idx - 1, item.type_ar, item.value_ar, item.value_en or '',
                item.icon_class,
                'نعم' if item.is_active else 'لا',
                'نعم' if item.show_in_footer else 'لا',
                item.created_at.strftime('%Y-%m-%d')
            ])
            style_data_row(ws, idx, 8, idx % 2 == 0)

    auto_adjust_column_width(ws)

    # حفظ
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    save_report_file(
        output, filename, user, report_type, period_type,
        start_date.date() if start_date else None,
        end_date.date() if end_date else None,
        None, items.count()
    )

    return output, filename


# ========================================
# التقرير السنوي
# ========================================
def generate_yearly_report(user):
    """توليد التقرير السنوي (آخر 365 يوم)"""
    wb = Workbook()

    # حساب الفترة - آخر سنة
    end_date = timezone.now()
    start_date = end_date - timedelta(days=365)

    # صفحة المشاريع
    ws_projects = wb.active
    ws_projects.title = "المشاريع"
    ws_projects.append(['#', 'اسم المشروع', 'الفئة', 'الحالة', 'المبلغ المستهدف',
                        'المبلغ المجمع', 'المستفيدين', 'تاريخ الإنشاء'])
    style_header(ws_projects, 1, 8)

    projects = Project.objects.filter(
        created_at__range=[start_date, end_date]
    ).select_related('category')

    for idx, p in enumerate(projects, 2):
        ws_projects.append([
            idx - 1, p.title_ar, p.category.name_ar, p.get_status_display(),
            float(p.target_amount), float(p.raised_amount), p.beneficiaries_count,
            p.created_at.strftime('%Y-%m-%d')
        ])
        style_data_row(ws_projects, idx, 8, idx % 2 == 0)

    auto_adjust_column_width(ws_projects)

    # صفحة المستخدمين
    ws_users = wb.create_sheet("المستخدمين")
    ws_users.append(['#', 'اسم المستخدم', 'الاسم الكامل', 'البريد', 'الجوال',
                     'الحالة', 'تاريخ التسجيل', 'آخر دخول'])
    style_header(ws_users, 1, 8)

    users = User.objects.filter(
        date_joined__range=[start_date, end_date]
    ).select_related('profile')

    for idx, u in enumerate(users, 2):
        ws_users.append([
            idx - 1, u.username, u.get_full_name(), u.email,
            u.profile.phone if hasattr(u, 'profile') else '',
            'نشط' if u.is_active else 'غير نشط',
            u.date_joined.strftime('%Y-%m-%d'),
            u.last_login.strftime('%Y-%m-%d %H:%M') if u.last_login else 'لم يسجل'
        ])
        style_data_row(ws_users, idx, 8, idx % 2 == 0)

    auto_adjust_column_width(ws_users)

    # صفحة نشاط المستخدمين
    ws_activity = wb.create_sheet("نشاط المستخدمين")
    ws_activity.append(['#', 'المستخدم', 'الإجراء', 'العنوان', 'الوصف', 'المستوى', 'التاريخ'])
    style_header(ws_activity, 1, 7)

    activities = ActivityLog.objects.filter(
        timestamp__range=[start_date, end_date]
    ).order_by('-timestamp')

    for idx, log in enumerate(activities, 2):
        ws_activity.append([
            idx - 1, log.username, log.get_action_display(), log.title,
            log.description, log.get_level_display(), log.timestamp.strftime('%Y-%m-%d %H:%M')
        ])
        style_data_row(ws_activity, idx, 7, idx % 2 == 0)

    auto_adjust_column_width(ws_activity)

    # صفحة الرسائل
    ws_messages = wb.create_sheet("الرسائل")
    ws_messages.append(['#', 'الاسم', 'البريد', 'الموضوع', 'الحالة', 'الأولوية', 'التاريخ'])
    style_header(ws_messages, 1, 7)

    messages = ContactMessage.objects.filter(
        created_at__range=[start_date, end_date]
    )

    for idx, msg in enumerate(messages, 2):
        ws_messages.append([
            idx - 1, msg.name, msg.email, msg.subject,
            msg.get_status_display(), msg.get_priority_display(),
            msg.created_at.strftime('%Y-%m-%d %H:%M')
        ])
        style_data_row(ws_messages, idx, 7, idx % 2 == 0)

    auto_adjust_column_width(ws_messages)

    # صفحة الاشتراكات
    ws_newsletter = wb.create_sheet("الاشتراكات")
    ws_newsletter.append(['#', 'البريد الإلكتروني', 'الاسم', 'التكرار', 'مفعل', 'تاريخ الاشتراك'])
    style_header(ws_newsletter, 1, 6)

    newsletters = Newsletter.objects.filter(
        subscribed_at__range=[start_date, end_date]
    )

    for idx, nl in enumerate(newsletters, 2):
        ws_newsletter.append([
            idx - 1, nl.email, nl.name or '',
            nl.get_frequency_display(),
            'نعم' if nl.is_active else 'لا',
            nl.subscribed_at.strftime('%Y-%m-%d')
        ])
        style_data_row(ws_newsletter, idx, 6, idx % 2 == 0)

    auto_adjust_column_width(ws_newsletter)

    # حفظ
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"yearly_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    save_report_file(
        output, filename, user, 'yearly_summary', 'yearly',
        start_date.date(), end_date.date(),
        None, projects.count() + users.count() + messages.count()
    )

    return output, filename


# ========================================
# التقرير المخصص المتقدم
# ========================================
def generate_advanced_custom_report(user, selected_reports, period_type='monthly',
                                    date_from=None, date_to=None):
    """تقرير مخصص متقدم - عدة تقارير في ملف واحد"""
    wb = Workbook()
    wb.remove(wb.active)  # إزالة الورقة الافتراضية

    # تحديد الفترة
    if period_type == 'monthly':
        start_date, end_date = get_current_month_range()
    elif period_type == 'yearly':
        start_date, end_date = get_current_year_range()
    elif period_type == 'custom' and date_from and date_to:
        start_date = timezone.make_aware(datetime.combine(date_from, datetime.min.time()))
        end_date = timezone.make_aware(datetime.combine(date_to, datetime.max.time()))
    else:
        start_date = None
        end_date = None

    total_records = 0

    # المشاريع
    if 'projects' in selected_reports:
        ws = wb.create_sheet("المشاريع")
        ws.append(['#', 'اسم المشروع', 'الفئة', 'الحالة', 'المبلغ المستهدف',
                   'المبلغ المجمع', 'المستفيدين', 'تاريخ الإنشاء'])
        style_header(ws, 1, 8)

        projects = Project.objects.all().select_related('category')
        if start_date and end_date:
            projects = projects.filter(created_at__range=[start_date, end_date])

        for idx, p in enumerate(projects, 2):
            ws.append([
                idx - 1, p.title_ar, p.category.name_ar, p.get_status_display(),
                float(p.target_amount), float(p.raised_amount), p.beneficiaries_count,
                p.created_at.strftime('%Y-%m-%d')
            ])
            style_data_row(ws, idx, 8, idx % 2 == 0)
        auto_adjust_column_width(ws)
        total_records += projects.count()

    # المستخدمين
    if 'users' in selected_reports:
        ws = wb.create_sheet("المستخدمين")
        ws.append(['#', 'اسم المستخدم', 'الاسم الكامل', 'البريد', 'الحالة', 'تاريخ التسجيل'])
        style_header(ws, 1, 6)

        users = User.objects.all()
        if start_date and end_date:
            users = users.filter(date_joined__range=[start_date, end_date])

        for idx, u in enumerate(users, 2):
            ws.append([
                idx - 1, u.username, u.get_full_name(), u.email,
                'نشط' if u.is_active else 'غير نشط',
                u.date_joined.strftime('%Y-%m-%d')
            ])
            style_data_row(ws, idx, 6, idx % 2 == 0)
        auto_adjust_column_width(ws)
        total_records += users.count()

    # الرسائل
    if 'messages' in selected_reports:
        ws = wb.create_sheet("الرسائل")
        ws.append(['#', 'الاسم', 'البريد', 'الموضوع', 'الحالة', 'التاريخ'])
        style_header(ws, 1, 6)

        messages = ContactMessage.objects.all()
        if start_date and end_date:
            messages = messages.filter(created_at__range=[start_date, end_date])

        for idx, msg in enumerate(messages, 2):
            ws.append([
                idx - 1, msg.name, msg.email, msg.subject,
                msg.get_status_display(), msg.created_at.strftime('%Y-%m-%d')
            ])
            style_data_row(ws, idx, 6, idx % 2 == 0)
        auto_adjust_column_width(ws)
        total_records += messages.count()

    # الاشتراكات
    if 'newsletters' in selected_reports:
        ws = wb.create_sheet("الاشتراكات")
        ws.append(['#', 'البريد', 'الاسم', 'مفعل', 'تاريخ الاشتراك'])
        style_header(ws, 1, 5)

        newsletters = Newsletter.objects.all()
        if start_date and end_date:
            newsletters = newsletters.filter(subscribed_at__range=[start_date, end_date])

        for idx, nl in enumerate(newsletters, 2):
            ws.append([
                idx - 1, nl.email, nl.name or '',
                'نعم' if nl.is_active else 'لا',
                nl.subscribed_at.strftime('%Y-%m-%d')
            ])
            style_data_row(ws, idx, 5, idx % 2 == 0)
        auto_adjust_column_width(ws)
        total_records += newsletters.count()

    # الأسئلة الشائعة
    if 'faqs' in selected_reports:
        ws = wb.create_sheet("الأسئلة الشائعة")
        ws.append(['#', 'السؤال', 'التصنيف', 'المشاهدات', 'مفعل'])
        style_header(ws, 1, 5)

        faqs = FAQ.objects.all().select_related('category')
        if start_date and end_date:
            faqs = faqs.filter(created_at__range=[start_date, end_date])

        for idx, faq in enumerate(faqs, 2):
            ws.append([
                idx - 1, faq.question_ar[:100],
                faq.category.category_ar if faq.category else '-',
                faq.views_count, 'نعم' if faq.is_active else 'لا'
            ])
            style_data_row(ws, idx, 5, idx % 2 == 0)
        auto_adjust_column_width(ws)
        total_records += faqs.count()

    # حفظ
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"advanced_custom_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    save_report_file(
        output, filename, user, 'advanced_custom', period_type,
        start_date.date() if start_date else None,
        end_date.date() if end_date else None,
        None, total_records
    )

    return output, filename