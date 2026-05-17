
from functools import wraps
from datetime import date

from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST, require_GET
from django.views.decorators.csrf import csrf_protect
from django.db.models import Sum
from django.utils import timezone
import re

from core.models import CustomUser, Notification, Payment
from sponsor.models import SponsorProfile
from core.utils import log_activity, get_exchange_rates, compress_image
from core.utils import get_countries
from beneficiary.models import JOBS


# ══ Decorator ══
def sponsor_required(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('/login/')
        if request.user.user_type != 'sponsor':
            return redirect('/login/')
        if not request.user.is_approved:
            return render(request, 'sponsor/pending.html')
        return view_func(request, *args, **kwargs)
    return wrapper


def _get_profile(user):
    try:
        return user.sponsor_profile
    except Exception:
        profile, _ = SponsorProfile.objects.get_or_create(user=user)
        return profile


def _profile_completion(user, profile):
    """حساب نسبة اكتمال الملف الشخصي"""
    fields = [
        user.first_name, user.father_name, user.grand_name, user.family_name,
        user.email, user.phone, user.id_number, user.nationality,
        user.gender, profile.job, profile.country, profile.city,
        profile.whatsapp,
    ]
    photo = profile.photo or user.profile_image
    if photo:
        fields.append(photo)
    filled = sum(1 for f in fields if f)
    return round(filled / len(fields) * 100)


def _get_stats(user, profile):
    from beneficiary.models import OrphanForm, SpecialNeedsForm, FamilyForm
    orphans  = OrphanForm.objects.filter(sponsor=profile).count()
    specials = SpecialNeedsForm.objects.filter(sponsor=profile).count()
    families = FamilyForm.objects.filter(sponsor=profile).count()
    total    = orphans + specials + families

    paid     = Payment.objects.filter(sponsor=user, status='paid')
    total_ils = paid.aggregate(t=Sum('amount_ils'))['t'] or 0
    total_usd = paid.aggregate(t=Sum('amount_usd'))['t'] or 0

    return {
        'total_sponsored': total,
        'orphans_count':   orphans,
        'specials_count':  specials,
        'families_count':  families,
        'total_ils':       round(total_ils, 2),
        'total_usd':       round(total_usd, 2),
        'date_joined':     user.date_joined.strftime('%Y/%m/%d') if user.date_joined else '—',
    }


GENDER_CHOICES = [('ذكر','ذكر'),('أنثى','أنثى')]
JOB_CHOICES    = [
    ('موظف حكومي','موظف حكومي'),('موظف خاص','موظف خاص'),
    ('تاجر','تاجر'),('طبيب','طبيب'),('مهندس','مهندس'),
    ('محامي','محامي'),('معلم','معلم'),('رجل أعمال','رجل أعمال'),
    ('متقاعد','متقاعد'),('أخرى','أخرى'),
]
PHONE_CODES = [
    ('+970','🇵🇸 +970'),('+972','🇮🇱 +972'),('+962','🇯🇴 +962'),
    ('+966','🇸🇦 +966'),('+20','🇪🇬 +20'),('+971','🇦🇪 +971'),
    ('+974','🇶🇦 +974'),('+965','🇰🇼 +965'),('+1','🇺🇸 +1'),
    ('+44','🇬🇧 +44'),('+49','🇩🇪 +49'),('+33','🇫🇷 +33'),
    ('+other','أخرى'),
]


# ══════════════════════════════════════════════════════════
# الصفحة الرئيسية
# ══════════════════════════════════════════════════════════

@sponsor_required
def profile(request):
    user    = request.user
    profile = _get_profile(user)

    notif_count = Notification.objects.filter(recipient=user, is_read=False).count()
    stats       = _get_stats(user, profile)
    completion  = _profile_completion(user, profile)

    # جلب الدول من الملف المحلي
    from core.utils import get_countries
    countries = get_countries()

    # جلب المهن من موديل beneficiary
    try:
        from beneficiary import JOBS
        jobs = JOBS
    except Exception:
        jobs = JOB_CHOICES

    return render(request, 'sponsor/profile.html', {
        'user':           user,
        'profile':        profile,
        'notif_count':    notif_count,
        'stats':          stats,
        'completion':     completion,
        'gender_choices': GENDER_CHOICES,

        'phone_codes':    PHONE_CODES,
        'countries':      countries,
        'job_choices': jobs,
    })


# ══════════════════════════════════════════════════════════
# حفظ البيانات الشخصية
# ══════════════════════════════════════════════════════════

@sponsor_required
@require_POST
@csrf_protect
def update_profile(request):
    user    = request.user
    profile = _get_profile(user)

    # حقول CustomUser
    first_name   = request.POST.get('first_name',   '').strip()
    father_name  = request.POST.get('father_name',  '').strip()
    grand_name   = request.POST.get('grand_name',   '').strip()
    family_name  = request.POST.get('family_name',  '').strip()
    email        = request.POST.get('email',        '').strip().lower()
    phone_country= request.POST.get('phone_country','+970').strip()
    phone        = request.POST.get('phone',        '').strip()
    wa_country   = request.POST.get('whatsapp_country','+970').strip()
    whatsapp     = request.POST.get('whatsapp',     '').strip()
    nationality  = request.POST.get('nationality',  '').strip()
    nat_code     = request.POST.get('nationality_code','').strip()
    gender       = request.POST.get('gender',       '').strip()
    id_number    = request.POST.get('id_number',    '').strip()

    # حقول SponsorProfile
    job          = request.POST.get('job',     '').strip()
    country      = request.POST.get('country', '').strip()
    city         = request.POST.get('city',    '').strip()

    errors = {}

    # XSS prevention
    for val in [first_name, father_name, grand_name, family_name, nationality, country, city]:
        if re.search(r'<script|javascript:', val, re.I):
            return JsonResponse({'status': 'error', 'message': 'مدخل غير صالح'})

    # فحص البريد
    if email and email != user.email:
        if CustomUser.objects.filter(email=email).exclude(pk=user.pk).exists():
            errors['email'] = 'البريد الإلكتروني مستخدم مسبقاً'

    # فحص الجوال
    full_phone = f'{phone_country}{phone}' if phone else ''
    if full_phone and full_phone != f'{user.phone_country}{user.phone}':
        if CustomUser.objects.filter(
            phone_country=phone_country, phone=phone
        ).exclude(pk=user.pk).exists():
            errors['phone'] = 'رقم الجوال مستخدم مسبقاً'

    if errors:
        return JsonResponse({'status': 'error', 'errors': errors})

    # حفظ CustomUser
    user.first_name      = first_name
    user.father_name     = father_name
    user.grand_name      = grand_name
    user.family_name     = family_name
    user.email           = email
    user.phone_country   = phone_country
    user.phone           = phone
    user.whatsapp_country= wa_country
    user.whatsapp        = whatsapp
    user.nationality     = nationality
    user.nationality_code= nat_code
    user.gender          = gender
    user.id_number       = id_number
    user.save()

    # حفظ SponsorProfile
    profile.gender          = gender
    profile.nationality     = nationality
    profile.job             = job
    profile.country         = country
    profile.city            = city
    profile.phone_country   = phone_country
    profile.phone           = phone
    profile.whatsapp_country= wa_country
    profile.whatsapp        = whatsapp

    # الصورة
    if 'photo' in request.FILES:
        img = request.FILES['photo']
        if img.size <= 4 * 1024 * 1024:
            try:
                compressed = compress_image(img)
                profile.photo.save(img.name, compressed, save=False)
            except Exception:
                profile.photo = img

    profile.save()

    log_activity(user, 'UPDATE',
                 description='تحديث البيانات الشخصية',
                 request=request)

    return JsonResponse({
        'status':     'success',
        'message':    'تم تحديث البيانات ✅',
        'completion': _profile_completion(user, profile),
        'full_name':  user.get_full_name(),
    })


# ══════════════════════════════════════════════════════════
# تغيير كلمة المرور
# ══════════════════════════════════════════════════════════

@sponsor_required
@require_POST
@csrf_protect
def change_password(request):
    user      = request.user
    old_pass  = request.POST.get('old_password',  '').strip()
    new_pass  = request.POST.get('new_password',  '').strip()
    new_pass2 = request.POST.get('new_password2', '').strip()

    if not user.check_password(old_pass):
        return JsonResponse({'status': 'error', 'message': 'كلمة المرور الحالية غير صحيحة'})
    if len(new_pass) < 8:
        return JsonResponse({'status': 'error', 'message': 'كلمة المرور الجديدة قصيرة جداً (8 أحرف على الأقل)'})
    if new_pass != new_pass2:
        return JsonResponse({'status': 'error', 'message': 'كلمتا المرور غير متطابقتين'})
    if old_pass == new_pass:
        return JsonResponse({'status': 'error', 'message': 'كلمة المرور الجديدة مطابقة للحالية'})

    user.set_password(new_pass)
    user.save()

    log_activity(user, 'UPDATE',
                 description='تغيير كلمة المرور',
                 request=request)

    return JsonResponse({'status': 'success', 'message': 'تم تغيير كلمة المرور ✅'})


# ══════════════════════════════════════════════════════════
# فحص فوري — البريد
# ══════════════════════════════════════════════════════════

@sponsor_required
@require_GET
def check_email(request):
    email = request.GET.get('email', '').strip().lower()
    user  = request.user
    exists = CustomUser.objects.filter(email=email).exclude(pk=user.pk).exists()
    return JsonResponse({'exists': exists})


# ══════════════════════════════════════════════════════════
# فحص فوري — الجوال
# ══════════════════════════════════════════════════════════

@sponsor_required
@require_GET
def check_phone(request):
    phone_country = request.GET.get('phone_country', '+970').strip()
    phone         = request.GET.get('phone', '').strip()
    user          = request.user
    exists = CustomUser.objects.filter(
        phone_country=phone_country, phone=phone
    ).exclude(pk=user.pk).exists()
    return JsonResponse({'exists': exists})


# ══════════════════════════════════════════════════════════
# تصدير Excel
# ══════════════════════════════════════════════════════════

@sponsor_required
@require_GET
def export_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from urllib.parse import quote
    import io

    user    = request.user
    profile = _get_profile(user)
    stats   = _get_stats(user, profile)

    log_activity(user, 'EXPORT', description='تصدير Excel الملف الشخصي', request=request)

    GREEN = '1A7A4A'; LIGHT = 'E8F5E9'; WHITE = 'FFFFFF'
    thin  = Side(style='thin', color='CCCCCC')
    brd   = Border(left=thin, right=thin, top=thin, bottom=thin)

    def sh(cell, color=GREEN):
        cell.font      = Font(bold=True, color=WHITE, size=10)
        cell.fill      = PatternFill('solid', fgColor=color)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = brd

    def sc(cell, r):
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border    = brd
        if r % 2 == 0:
            cell.fill = PatternFill('solid', fgColor='F4F9F6')

    wb = openpyxl.Workbook()

    # ══ ورقة البيانات الشخصية ══
    ws1 = wb.active
    ws1.title = 'البيانات الشخصية'
    ws1.sheet_view.rightToLeft = True

    ws1.merge_cells('A1:B1')
    c = ws1['A1']
    c.value     = f'البيانات الشخصية — {user.get_full_name()} — {date.today()}'
    c.font      = Font(bold=True, size=12, color=WHITE)
    c.fill      = PatternFill('solid', fgColor=GREEN)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 28

    ws1.cell(2,1,'الحقل').font = Font(bold=True, color=WHITE)
    ws1.cell(2,1).fill = PatternFill('solid', fgColor='2ECC71')
    ws1.cell(2,1).border = brd
    ws1.cell(2,2,'القيمة').font = Font(bold=True, color=WHITE)
    ws1.cell(2,2).fill = PatternFill('solid', fgColor='2ECC71')
    ws1.cell(2,2).border = brd
    ws1.row_dimensions[2].height = 22

    personal_data = [
        ('رقم التسجيل',       user.registration_number or '—'),
        ('الاسم الكامل',      user.get_full_name()),
        ('الاسم الأول',       user.first_name or '—'),
        ('اسم الأب',          user.father_name or '—'),
        ('اسم الجد',          user.grand_name or '—'),
        ('اسم العائلة',       user.family_name or '—'),
        ('اسم المستخدم',      user.username),
        ('البريد الإلكتروني', user.email or '—'),
        ('رقم الجوال',        f'{user.phone_country}{user.phone}' if user.phone else '—'),
        ('الواتساب',          f'{user.whatsapp_country}{user.whatsapp}' if user.whatsapp else '—'),
        ('رقم الهوية',        user.id_number or '—'),
        ('الجنس',             user.gender or '—'),
        ('الجنسية',           user.nationality or '—'),
        ('المهنة',            profile.job or '—'),
        ('الدولة',            profile.country or '—'),
        ('المدينة',           profile.city or '—'),
        ('تاريخ الانضمام',    stats['date_joined']),
    ]

    for r_idx, (field, val) in enumerate(personal_data, start=3):
        ws1.cell(r_idx, 1, field).border = brd
        ws1.cell(r_idx, 1).fill = PatternFill('solid', fgColor=LIGHT) if r_idx%2==0 else PatternFill()
        ws1.cell(r_idx, 2, val).border  = brd
        ws1.row_dimensions[r_idx].height = 20

    ws1.column_dimensions['A'].width = 22
    ws1.column_dimensions['B'].width = 36

    # ══ ورقة الإحصائيات ══
    ws2 = wb.create_sheet('إحصائيات الكفالة')
    ws2.sheet_view.rightToLeft = True

    ws2.merge_cells('A1:B1')
    c2 = ws2['A1']
    c2.value     = f'إحصائيات الكفالة — {user.get_full_name()}'
    c2.font      = Font(bold=True, size=12, color=WHITE)
    c2.fill      = PatternFill('solid', fgColor=GREEN)
    c2.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 28

    stats_data = [
        ('إجمالي المكفولين',      stats['total_sponsored']),
        ('الأيتام',               stats['orphans_count']),
        ('ذوو الاحتياجات',        stats['specials_count']),
        ('الأسر',                 stats['families_count']),
        ('إجمالي المدفوع (₪)',    stats['total_ils']),
        ('إجمالي المدفوع ($)',    stats['total_usd']),
        ('تاريخ الانضمام',       stats['date_joined']),
    ]

    for r_idx, (field, val) in enumerate(stats_data, start=2):
        ws2.cell(r_idx, 1, field).border = brd
        ws2.cell(r_idx, 2, val).border   = brd
        ws2.row_dimensions[r_idx].height = 20
        if r_idx % 2 == 0:
            ws2.cell(r_idx,1).fill = PatternFill('solid', fgColor=LIGHT)
            ws2.cell(r_idx,2).fill = PatternFill('solid', fgColor=LIGHT)

    ws2.column_dimensions['A'].width = 24
    ws2.column_dimensions['B'].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    fname = f'ملف_كافل_{user.get_full_name()}_{date.today()}.xlsx'
    resp  = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(fname)}"
    return resp


# ══════════════════════════════════════════════════════════
# تصدير PDF — يستخدم sponsor_report_pdf.html
# ══════════════════════════════════════════════════════════

@sponsor_required
def export_pdf(request):
    from django.template.loader import render_to_string
    from urllib.parse import quote
    import weasyprint, logging
    from django.utils import timezone as tz

    logging.getLogger('weasyprint').setLevel(logging.ERROR)
    logging.getLogger('fontTools').setLevel(logging.ERROR)

    user    = request.user
    profile = _get_profile(user)
    stats   = _get_stats(user, profile)

    log_activity(user, 'EXPORT', description='تصدير PDF الملف الشخصي', request=request)

    from beneficiary.models import OrphanForm, SpecialNeedsForm, FamilyForm
    from core.models import Payment, Aid

    today = tz.now()

    def build_bene_list(qs, type_label):
        result = []
        for f in qs:
            pays = Payment.objects.filter(beneficiary=f.user, sponsor=user)
            aids = Aid.objects.filter(beneficiary=f.user)
            result.append({
                'full_name':   f.get_full_name(),
                'reg_number':  f.user.registration_number or f.form_number,
                'type_label':  type_label,
                'is_active':   f.user.is_active,
                'month_pays':  str(pays.filter(date__year=today.year, date__month=today.month).aggregate(t=__import__('django.db.models',fromlist=['Sum']).Sum('amount_ils'))['t'] or 0),
                'total_pays':  str(pays.filter(status='paid').aggregate(t=__import__('django.db.models',fromlist=['Sum']).Sum('amount_ils'))['t'] or 0),
                'pays_count':  pays.filter(status='paid').count(),
                'aids_count':  aids.count(),
                'last_pay':    str(pays.filter(status='paid').order_by('-date').values_list('date',flat=True).first() or '—'),
            })
        return result

    orphans  = OrphanForm.objects.filter(sponsor=profile).select_related('user')
    specials = SpecialNeedsForm.objects.filter(sponsor=profile).select_related('user')
    families = FamilyForm.objects.filter(sponsor=profile).select_related('user')

    beneficiaries = (
        build_bene_list(orphans,  'يتيم') +
        build_bene_list(specials, 'ذو احتياج') +
        build_bene_list(families, 'أسرة')
    )

    all_pays = Payment.objects.filter(sponsor=user, status='paid')
    month_pays = all_pays.filter(date__year=today.year, date__month=today.month)

    from django.db.models import Sum as DSum
    context = {
        'sponsor':          profile,
        'period_label':     today.strftime('%B %Y'),
        'bene_count':       len(beneficiaries),
        'beneficiaries':    beneficiaries,
        'month_total_ils':  str(month_pays.aggregate(t=DSum('amount_ils'))['t'] or 0),
        'all_total_ils':    str(all_pays.aggregate(t=DSum('amount_ils'))['t'] or 0),
        'total_aids':       Aid.objects.filter(beneficiary__in=[f.user for f in list(orphans)+list(specials)+list(families)]).count(),
        'printed_by':       user.get_full_name(),
        'request':          request,
    }

    html_content = render_to_string('sponsor/sponsor_profile_pdf.html', context)
    pdf_file     = weasyprint.HTML(string=html_content, base_url=request.build_absolute_uri('/')).write_pdf()

    fname = f'تقرير_{user.get_full_name()}_{date.today()}.pdf'
    resp  = HttpResponse(pdf_file, content_type='application/pdf')
    resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(fname)}"
    return resp


# ══════════════════════════════════════════════════════════
# شهادة الكفالة PDF
# ══════════════════════════════════════════════════════════

@sponsor_required
def export_certificate(request):
    from django.template.loader import render_to_string
    from urllib.parse import quote
    import weasyprint, logging, sys, traceback

    logging.getLogger('weasyprint').setLevel(logging.ERROR)
    logging.getLogger('fontTools').setLevel(logging.ERROR)

    user    = request.user
    profile = _get_profile(user)

    try:
        from beneficiary.models import OrphanForm, SpecialNeedsForm, FamilyForm

        benes = []
        for Model, label in [
            (OrphanForm,       'يتيم'),
            (SpecialNeedsForm, 'ذو احتياج'),
            (FamilyForm,       'أسرة'),
        ]:
            qs = Model.objects.filter(sponsor=profile).select_related('user')

            for f in qs:
                benes.append({
                    'full_name':        f.get_full_name(),
                    'reg_number':       f.user.registration_number or f.form_number,
                    'type_label':       label,
                    'sponsorship_date': str(f.sponsorship_date) if f.sponsorship_date else '—',
                    'city':             f.current_city or '—',
                })

        print(f"[CERT] total benes : {len(benes)}", file=sys.stderr)

        context = {
            'user':       user,
            'profile':    profile,
            'benes':      benes,
            'total':      len(benes),
            'issue_date': date.today().strftime('%Y/%m/%d'),
            'join_date':  user.date_joined.strftime('%Y/%m/%d') if user.date_joined else '—',
            'request':    request,
        }

        print(f"[CERT] rendering template…", file=sys.stderr)
        html_content = render_to_string('sponsor/certificate.html', context)
        print(f"[CERT] html length : {len(html_content)}", file=sys.stderr)

        print(f"[CERT] generating PDF…", file=sys.stderr)
        pdf_file = weasyprint.HTML(
            string=html_content,
            base_url=request.build_absolute_uri('/')
        ).write_pdf()
        print(f"[CERT] pdf size    : {len(pdf_file) if pdf_file else 0} bytes", file=sys.stderr)

        if not pdf_file:

            return HttpResponse('فشل إنشاء PDF', status=500)

        fname = f'شهادة_كفالة_{user.get_full_name()}_{date.today()}.pdf'

        resp = HttpResponse(pdf_file, content_type='application/pdf')
        resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(fname)}"
        resp['Content-Length'] = str(len(pdf_file))
        return resp


    except Exception as e:
        print(f"[CERT] EXCEPTION: {e}", file=sys.stderr)
        traceback.print_exc(file=sys.stderr)
        return HttpResponse(f'خطأ: {str(e)}', status=500)
@sponsor_required
def certificate_blob(request):
    """endpoint للـ fetch فقط — يرجع PDF بدون attachment"""
    from django.template.loader import render_to_string
    from urllib.parse import quote
    import weasyprint, logging

    logging.getLogger('weasyprint').setLevel(logging.ERROR)
    logging.getLogger('fontTools').setLevel(logging.ERROR)

    user    = request.user
    profile = _get_profile(user)

    try:
        from beneficiary.models import OrphanForm, SpecialNeedsForm, FamilyForm

        benes = []
        for Model, label in [
            (OrphanForm,'يتيم'),
            (SpecialNeedsForm,'ذو احتياج'),
            (FamilyForm,'أسرة'),
        ]:
            for f in Model.objects.filter(sponsor=profile).select_related('user'):
                benes.append({
                    'full_name':        f.get_full_name(),
                    'reg_number':       f.user.registration_number or f.form_number,
                    'type_label':       label,
                    'sponsorship_date': str(f.sponsorship_date) if f.sponsorship_date else '—',
                    'city':             f.current_city or '—',
                })

        context = {
            'user':       user,
            'profile':    profile,
            'benes':      benes,
            'total':      len(benes),
            'issue_date': date.today().strftime('%Y/%m/%d'),
            'join_date':  user.date_joined.strftime('%Y/%m/%d') if user.date_joined else '—',
            'request':    request,
        }

        html_content = render_to_string('sponsor/certificate.html', context)
        pdf_file     = weasyprint.HTML(
            string=html_content,
            base_url=request.build_absolute_uri('/')
        ).write_pdf()

        # بدون attachment — يرجع inline للـ fetch
        resp = HttpResponse(pdf_file, content_type='application/pdf')
        resp['Content-Length'] = str(len(pdf_file))
        resp['Cache-Control']  = 'no-store'
        return resp

    except Exception as e:
        return HttpResponse(f'خطأ: {str(e)}', status=500)
@sponsor_required
def export_pdf_blob(request):
    """endpoint للـ fetch فقط — يرجع PDF بدون attachment"""
    from django.template.loader import render_to_string
    import weasyprint, logging

    logging.getLogger('weasyprint').setLevel(logging.ERROR)
    logging.getLogger('fontTools').setLevel(logging.ERROR)

    user    = request.user
    profile = _get_profile(user)
    stats   = _get_stats(user, profile)

    try:
        from beneficiary.models import OrphanForm, SpecialNeedsForm, FamilyForm
        from sponsor.models import PaymentReceipt

        beneficiaries = []
        for Model, label in [
            (OrphanForm,       'يتيم'),
            (SpecialNeedsForm, 'ذو احتياج'),
            (FamilyForm,       'أسرة'),
        ]:
            for f in Model.objects.filter(sponsor=profile).select_related('user'):
                from core.models import Payment
                from django.db.models import Sum
                pays = Payment.objects.filter(beneficiary=f.user, sponsor=user)
                beneficiaries.append({
                    'full_name':        f.get_full_name(),
                    'reg_number':       f.user.registration_number or f.form_number,
                    'type_label':       label,
                    'city':             f.current_city or '—',
                    'sponsorship_date': str(f.sponsorship_date) if f.sponsorship_date else '—',
                    'total_pays':       str(pays.filter(status='paid').aggregate(t=Sum('amount_ils'))['t'] or 0),
                    'pays_count':       pays.filter(status='paid').count(),
                })

        receipts = PaymentReceipt.objects.filter(
            sponsor=profile
        ).select_related('beneficiary').order_by('-submitted_at')[:30]

        context = {
            'sponsor':       profile,
            'beneficiaries': beneficiaries,
            'receipts':      receipts,
            'stats':         stats,
            'printed_by':    user.get_full_name(),
            'request':       request,
        }

        html_content = render_to_string('sponsor/sponsor_profile_pdf.html', context)
        pdf_file     = weasyprint.HTML(
            string=html_content,
            base_url=request.build_absolute_uri('/')
        ).write_pdf()

        resp = HttpResponse(pdf_file, content_type='application/pdf')
        resp['Content-Length'] = str(len(pdf_file))
        resp['Cache-Control']  = 'no-store'
        return resp

    except Exception as e:
        import traceback
        traceback.print_exc()
        return HttpResponse(f'خطأ: {str(e)}', status=500)

@sponsor_required
@require_GET
def check_id_number(request):
    id_number = request.GET.get('id_number', '').strip()
    user      = request.user
    exists    = CustomUser.objects.filter(
        id_number=id_number
    ).exclude(pk=user.pk).exists()
    return JsonResponse({'exists': exists})