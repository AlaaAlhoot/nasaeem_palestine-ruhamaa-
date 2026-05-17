from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_protect
from django.utils import timezone

from core.models import CustomUser, Notification, ActivityLog
from core.utils import log_activity, get_client_ip, get_exchange_rates
from ..models import OrphanForm, SpecialNeedsForm, FamilyForm, Payment
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_protect
from django.utils import timezone

from core.models import (CustomUser, Notification, ActivityLog,
                         DirectMessage)          # ← أضفنا DirectMessage هنا
from core.utils import log_activity, get_client_ip, get_exchange_rates
from ..models import OrphanForm, SpecialNeedsForm, FamilyForm, Payment


# ==================== Decorator مشترك ====================

def beneficiary_required(view_func):
    """يتحقق أن المستخدم مستفيد موافق عليه"""
    from functools import wraps
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('/login/')
        if request.user.user_type not in ['orphan', 'family', 'special']:
            return redirect('/login/')
        if not request.user.is_approved:
            return render(request, 'beneficiary/pending.html')
        return view_func(request, *args, **kwargs)
    return wrapper


def _get_form(user):
    """يرجع استمارة المستخدم حسب نوعه"""
    try:
        if user.user_type == 'orphan':
            return user.orphan_form, 'orphan'
        elif user.user_type == 'special':
            return user.special_form, 'special'
        elif user.user_type == 'family':
            return user.family_form, 'family'
    except Exception:
        pass
    return None, None


def _get_status_steps(status):
    """يرجع خطوات شريط التقدم"""
    steps = [
        'مسودة',
        'مرسلة',
        'قيد المراجعة',
        'مقبولة',
        'تم التدقيق',
        'تم التكفل',
    ]
    current = steps.index(status) if status in steps else 0
    return [
        {'label': s, 'done': i < current, 'active': i == current}
        for i, s in enumerate(steps)
    ]


# ==================== الرئيسية ====================


# ==================== الملف الشخصي ====================

# ==================== beneficiary/views.py ====================
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST, require_http_methods
from django.views.decorators.csrf import csrf_protect
from django.utils import timezone
from django.db import transaction
import json

from core.models import CustomUser, Notification, ActivityLog
from core.utils import log_activity, get_client_ip, get_exchange_rates, compress_image
from ..models import OrphanForm, SpecialNeedsForm, FamilyForm, Payment
from core.validators import validate_id_number, validate_arabic_text


# ==================== Decorator مشترك ====================

def beneficiary_required(view_func):
    """يتحقق أن المستخدم مستفيد موافق عليه"""
    from functools import wraps
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('/login/')
        if request.user.user_type not in ['orphan', 'family', 'special']:
            return redirect('/login/')
        if not request.user.is_approved:
            return render(request, 'beneficiary/pending.html')
        return view_func(request, *args, **kwargs)

    return wrapper


def _get_form(user):
    """يرجع استمارة المستخدم حسب نوعه"""
    try:
        if user.user_type == 'orphan':
            return user.orphan_form, 'orphan'
        elif user.user_type == 'special':
            return user.special_form, 'special'
        elif user.user_type == 'family':
            return user.family_form, 'family'
    except Exception:
        pass
    return None, None


# ==================== الملف الشخصي - عرض ====================
# ==================== beneficiary/views.py ====================
# الجزء الخاص بالملف الشخصي فقط
# انسخ هذه الدوال وأضفها لملفك الموجود

import json
import logging
import re
import datetime
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.views.decorators.http import require_POST, require_http_methods
from django.views.decorators.csrf import csrf_protect
from django.db import transaction

from core.models import CustomUser
from core.utils import log_activity, compress_image
from core.validators import validate_id_number
from ..models import (
    OrphanForm, OrphanMother, OrphanFather,
    SpecialNeedsForm, FamilyForm, FamilyWife,
    CurrentGuardian, FamilyMember, JOBS,
)

logger = logging.getLogger(__name__)


# ============================================================
# Decorator مشترك (موجود أصلاً عندك)
# ============================================================

def beneficiary_required(view_func):
    """يتحقق أن المستخدم مستفيد موافق عليه"""
    from functools import wraps
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('/login/')
        if request.user.user_type not in ['orphan', 'family', 'special']:
            return redirect('/login/')
        if not request.user.is_approved:
            return render(request, 'beneficiary/pending.html')
        return view_func(request, *args, **kwargs)
    return wrapper


def _get_form(user):
    """يرجع استمارة المستخدم حسب نوعه"""
    try:
        if user.user_type == 'orphan':
            return user.orphan_form, 'orphan'
        elif user.user_type == 'special':
            return user.special_form, 'special'
        elif user.user_type == 'family':
            return user.family_form, 'family'
    except Exception:
        pass
    return None, None


# ============================================================




def _check_id_duplicate(id_number, ftype, current_form_id, field_name, section, user):
    """تحقق من تكرار رقم الهوية"""
    # الهوية الرئيسية للنموذج نفسه
    if section == 'main' or field_name == 'id_number':
        # مقارنة مع الاستمارات الأخرى
        if ftype == 'orphan':
            return OrphanForm.objects.filter(id_number=id_number).exclude(id=current_form_id).exists()
        elif ftype == 'special':
            return SpecialNeedsForm.objects.filter(id_number=id_number).exclude(id=current_form_id).exists()
        elif ftype == 'family':
            return FamilyForm.objects.filter(id_number=id_number).exclude(id=current_form_id).exists()
    return False


# ============================================================
# 3. تحديث الملف الشخصي (AJAX) - الدالة الرئيسية
# ============================================================



# ============================================================
# دوال تحديث منفصلة لكل نوع
# ============================================================

def _update_orphan(form, data, files):
    """تحديث بيانات اليتيم + الأم + الأب"""

    # ── البيانات الأساسية ──
    form.father_name      = data.get('father_name', form.father_name)
    form.grand_name       = data.get('grand_name', form.grand_name)
    form.family_name      = data.get('family_name', form.family_name)
    form.birth_date       = data.get('birth_date', form.birth_date)
    form.gender           = data.get('gender', form.gender)
    form.orphan_type      = data.get('orphan_type', form.orphan_type)
    form.nationality      = data.get('nationality', form.nationality)
    form.nationality_code = data.get('nationality_code', form.nationality_code)

    # ── الحالة الصحية والتعليم ──
    form.health_status   = data.get('health_status', form.health_status)
    form.education_level = data.get('education_level', form.education_level)
    form.school_grade    = data.get('school_grade', '') or None
    form.school_name     = data.get('school_name', form.school_name)

    # ── العنوان ──
    form.current_city      = data.get('current_city', form.current_city)
    form.current_street    = data.get('current_street', form.current_street)
    form.current_landmark  = data.get('current_landmark', form.current_landmark)
    form.previous_city     = data.get('previous_city', '') or ''
    form.previous_street   = data.get('previous_street', '') or ''
    form.previous_landmark = data.get('previous_landmark', '') or ''

    # ── السكن ──
    form.housing_type      = data.get('housing_type', form.housing_type)
    form.housing_ownership = data.get('housing_ownership', form.housing_ownership)
    form.monthly_rent      = data.get('monthly_rent', '') or None

    # ── التواصل ──
    form.phone1           = data.get('phone1', form.phone1)
    form.phone1_country   = data.get('phone1_country', form.phone1_country)
    form.phone2           = data.get('phone2', '') or ''
    form.phone2_country   = data.get('phone2_country', form.phone2_country)
    form.whatsapp         = data.get('whatsapp', form.whatsapp)
    form.whatsapp_country = data.get('whatsapp_country', form.whatsapp_country)

    # ── القصة ──
    form.story = data.get('story', form.story)

    # ── الصورة ──
    if 'photo' in files:
        compressed = compress_image(files['photo'])
        form.photo.save(files['photo'].name, compressed, save=False)

    form.save()

    # ── تحديث الأم ──
    if form.orphan_type in ['يتيم الأم', 'يتيم الأبوين']:
        mother = OrphanMother.objects.filter(form=form).first()
        if mother and data.get('mother_id_number'):
            mother.first_name      = data.get('mother_first_name', mother.first_name)
            mother.father_name     = data.get('mother_father_name', mother.father_name)
            mother.grand_name      = data.get('mother_grand_name', mother.grand_name)
            mother.family_name     = data.get('mother_family_name', mother.family_name)
            mother.birth_date      = data.get('mother_birth_date', mother.birth_date)
            mother.id_number       = data.get('mother_id_number', mother.id_number)
            mother.nationality     = data.get('mother_nationality', mother.nationality)
            mother.health_status   = data.get('mother_health_status', mother.health_status)
            mother.education_level = data.get('mother_education_level', mother.education_level)
            mother.save()

    # ── تحديث الأب ──
    if form.orphan_type in ['يتيم الأب', 'يتيم الأبوين']:
        father = OrphanFather.objects.filter(form=form).first()
        if father and data.get('father_id_number'):
            father.first_name      = data.get('father_first_name', father.first_name)
            father.father_name     = data.get('father_father_name', father.father_name)
            father.grand_name      = data.get('father_grand_name', father.grand_name)
            father.family_name     = data.get('father_family_name', father.family_name)
            father.birth_date      = data.get('father_birth_date', father.birth_date)
            father.id_number       = data.get('father_id_number', father.id_number)
            father.nationality     = data.get('father_nationality', father.nationality)
            father.health_status   = data.get('father_health_status', father.health_status)
            father.education_level = data.get('father_education_level', father.education_level)
            father.job             = data.get('father_job', father.job)
            father.save()


def _update_special(form, data, files):
    """تحديث بيانات ذوي الاحتياجات الخاصة"""
    form.father_name  = data.get('father_name', form.father_name)
    form.grand_name   = data.get('grand_name', form.grand_name)
    form.family_name  = data.get('family_name', form.family_name)
    form.birth_date   = data.get('birth_date', form.birth_date)
    form.gender       = data.get('gender', form.gender)
    form.nationality  = data.get('nationality', form.nationality)

    form.health_status   = data.get('health_status', form.health_status)
    form.education_level = data.get('education_level', form.education_level)
    form.school_grade    = data.get('school_grade', '') or None
    form.school_name     = data.get('school_name', '')

    form.current_city      = data.get('current_city', form.current_city)
    form.current_street    = data.get('current_street', form.current_street)
    form.current_landmark  = data.get('current_landmark', form.current_landmark)
    form.previous_city     = data.get('previous_city', '') or ''
    form.previous_street   = data.get('previous_street', '') or ''
    form.previous_landmark = data.get('previous_landmark', '') or ''

    form.housing_type      = data.get('housing_type', form.housing_type)
    form.housing_ownership = data.get('housing_ownership', form.housing_ownership)
    form.monthly_rent      = data.get('monthly_rent', '') or None

    form.phone1           = data.get('phone1', form.phone1)
    form.phone1_country   = data.get('phone1_country', form.phone1_country)
    form.phone2           = data.get('phone2', '') or ''
    form.phone2_country   = data.get('phone2_country', form.phone2_country)
    form.whatsapp         = data.get('whatsapp', form.whatsapp)
    form.whatsapp_country = data.get('whatsapp_country', form.whatsapp_country)

    form.case_details = data.get('case_details', form.case_details)

    if 'photo' in files:
        compressed = compress_image(files['photo'])
        form.photo.save(files['photo'].name, compressed, save=False)

    form.save()


def _update_family(form, data, files):
    """تحديث بيانات الأسرة + الزوجة"""
    form.father_name = data.get('father_name', form.father_name)
    form.grand_name  = data.get('grand_name', form.grand_name)
    form.family_name = data.get('family_name', form.family_name)
    form.birth_date  = data.get('birth_date', form.birth_date)
    form.gender      = data.get('gender', form.gender)
    form.nationality = data.get('nationality', form.nationality)
    form.is_alive    = data.get('is_alive', 'true') == 'true'

    form.marital_status  = data.get('marital_status', form.marital_status)
    form.health_status   = data.get('health_status', form.health_status)
    form.education_level = data.get('education_level', form.education_level)
    form.job             = data.get('job', form.job)

    form.current_city      = data.get('current_city', form.current_city)
    form.current_street    = data.get('current_street', form.current_street)
    form.current_landmark  = data.get('current_landmark', form.current_landmark)
    form.previous_city     = data.get('previous_city', '') or ''
    form.previous_street   = data.get('previous_street', '') or ''
    form.previous_landmark = data.get('previous_landmark', '') or ''

    form.housing_type      = data.get('housing_type', form.housing_type)
    form.housing_ownership = data.get('housing_ownership', form.housing_ownership)
    form.monthly_rent      = data.get('monthly_rent', '') or None

    form.phone1           = data.get('phone1', form.phone1)
    form.phone1_country   = data.get('phone1_country', form.phone1_country)
    form.phone2           = data.get('phone2', '') or ''
    form.phone2_country   = data.get('phone2_country', form.phone2_country)
    form.whatsapp         = data.get('whatsapp', form.whatsapp)
    form.whatsapp_country = data.get('whatsapp_country', form.whatsapp_country)

    form.family_members_count = data.get('family_members_count', form.family_members_count) or 1
    form.sick_members_count   = data.get('sick_members_count', 0) or 0
    form.general_status       = data.get('general_status', form.general_status)

    if 'photo' in files:
        compressed = compress_image(files['photo'])
        form.photo.save(files['photo'].name, compressed, save=False)

    form.save()

    # ── تحديث/إنشاء/حذف الزوجة ──
    if form.marital_status == 'متزوج' and data.get('wife_id_number'):
        wife, _ = FamilyWife.objects.get_or_create(form=form)
        wife.first_name      = data.get('wife_first_name', '')
        wife.father_name     = data.get('wife_father_name', '')
        wife.grand_name      = data.get('wife_grand_name', '')
        wife.family_name     = data.get('wife_family_name', '')
        wife.birth_date      = data.get('wife_birth_date', '2000-01-01')
        wife.id_number       = data.get('wife_id_number', '')
        wife.nationality     = data.get('wife_nationality', 'فلسطينية')
        wife.gender          = 'أنثى'
        wife.health_status   = data.get('wife_health_status', 'سليم')
        wife.education_level = data.get('wife_education_level', 'غير متعلم')

        if 'wife_photo' in files:
            compressed = compress_image(files['wife_photo'])
            wife.photo.save(files['wife_photo'].name, compressed, save=False)

        wife.save()
    else:
        # حذف الزوجة عند تغيير الحالة
        FamilyWife.objects.filter(form=form).delete()


# ============================================================
# المعيل (مشترك بين الكل)
# ============================================================

def _update_guardian(form, ftype, data):
    """تحديث بيانات المعيل"""
    if not data.get('guardian_id_number'):
        return

    guardian, _ = CurrentGuardian.objects.get_or_create(
        form_type=ftype, form_id=form.id
    )
    guardian.first_name       = data.get('guardian_first_name', '')
    guardian.father_name      = data.get('guardian_father_name', '')
    guardian.grand_name       = data.get('guardian_grand_name', '')
    guardian.family_name      = data.get('guardian_family_name', '')
    guardian.birth_date       = data.get('guardian_birth_date', '2000-01-01')
    guardian.id_number        = data.get('guardian_id_number', '')
    guardian.nationality      = data.get('guardian_nationality', 'فلسطينية')
    guardian.nationality_code = data.get('guardian_nationality_code', 'PS')
    guardian.gender           = data.get('guardian_gender', 'ذكر')
    guardian.health_status    = data.get('guardian_health_status', 'سليم')
    guardian.education_level  = data.get('guardian_education_level', 'غير متعلم')
    guardian.job              = data.get('guardian_job', 'أخرى')
    guardian.monthly_income   = data.get('guardian_monthly_income', 0) or 0
    guardian.relation         = data.get('guardian_relation', 'أخرى')
    guardian.dependents       = data.get('guardian_dependents', 1) or 1
    guardian.save()


# ============================================================
# أفراد الأسرة (مشترك بين special و family)
# ============================================================

def _update_members(form, ftype, data):
    """تحديث أفراد الأسرة من JSON"""
    members_json = data.get('family_members_json', '[]')
    try:
        members = json.loads(members_json)
    except Exception:
        return

    # حذف القديم وإضافة الجديد
    FamilyMember.objects.filter(form_type=ftype, form_id=form.id).delete()

    for m in members:
        if not m.get('id_number'):
            continue
        FamilyMember.objects.create(
            form_type        = ftype,
            form_id          = form.id,
            first_name       = m.get('first_name', ''),
            father_name      = m.get('father_name', ''),
            grand_name       = m.get('grand_name', ''),
            family_name      = m.get('family_name', ''),
            birth_date       = m.get('birth_date', '2000-01-01'),
            id_number        = m.get('id_number', ''),
            nationality      = m.get('nationality', 'فلسطينية'),
            nationality_code = m.get('nationality_code', 'PS'),
            gender           = m.get('gender', 'ذكر'),
            health_status    = m.get('health_status', 'سليم'),
            education_level  = m.get('education_level', 'غير متعلم'),
            school_grade     = m.get('school_grade', '') or None,
            marital_status   = m.get('marital_status', 'أعزب'),
            relation         = m.get('relation', 'أخرى'),
        )



# ==================== الدفعات ====================



# ==================== التواصل ====================






# ==================== التقارير ====================

@beneficiary_required
def reports(request):
    user        = request.user
    form, ftype = _get_form(user)
    rates       = get_exchange_rates()

    all_payments = Payment.objects.filter(
        beneficiary_type = ftype,
        beneficiary_id   = form.id if form else 0,
    ).order_by('-payment_date')

    context = {
        'form':     form,
        'ftype':    ftype,
        'payments': all_payments,
        'rates':    rates,
    }
    return render(request, 'beneficiary/reports.html', context)


@beneficiary_required
def export_report(request):
    """تصدير تقرير PDF أو Excel"""
    user        = request.user
    form, ftype = _get_form(user)
    export_type = request.GET.get('type', 'excel')
    rates       = get_exchange_rates()

    all_payments = Payment.objects.filter(
        beneficiary_type = ftype,
        beneficiary_id   = form.id if form else 0,
    ).order_by('-payment_date')

    log_activity(user, 'EXPORT',
                 description=f'تصدير تقرير {export_type}',
                 request=request)

    if export_type == 'excel':
        return _export_excel(form, ftype, all_payments, rates)
    else:
        return _export_pdf(form, ftype, all_payments, rates)


def _export_excel(form, ftype, payments, rates):
    """تصدير Excel احترافي"""
    try:
        import openpyxl
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                     Border, Side, GradientFill)
        from openpyxl.utils import get_column_letter
        import io
        from datetime import date

        wb = openpyxl.Workbook()

        # ===== ورقة 1: البيانات الشخصية =====
        ws1 = wb.active
        ws1.title = 'البيانات الشخصية'
        ws1.sheet_view.rightToLeft = True

        # ألوان
        GREEN_DARK  = '1A7A4A'
        GREEN_LIGHT = 'E8F5E9'
        WHITE       = 'FFFFFF'
        GRAY        = 'F4F9F6'

        # Header
        ws1.merge_cells('A1:D1')
        title_cell = ws1['A1']
        title_cell.value         = 'منصة رُحَمَاء — جمعية نسائم فلسطين الخيرية'
        title_cell.font          = Font(bold=True, size=14, color=WHITE)
        title_cell.fill          = PatternFill('solid', fgColor=GREEN_DARK)
        title_cell.alignment     = Alignment(horizontal='center', vertical='center')
        ws1.row_dimensions[1].height = 35

        ws1.merge_cells('A2:D2')
        date_cell = ws1['A2']
        date_cell.value      = f'تاريخ التصدير: {date.today().strftime("%Y-%m-%d")}'
        date_cell.font       = Font(size=10, color=GREEN_DARK)
        date_cell.fill       = PatternFill('solid', fgColor=GREEN_LIGHT)
        date_cell.alignment  = Alignment(horizontal='center')
        ws1.row_dimensions[2].height = 20

        # بيانات
        type_labels = {'orphan':'يتيم', 'special':'ذوو احتياجات', 'family':'أسرة'}
        rows = []
        if form:
            rows = [
                ('الحقل', 'القيمة'),
                ('نوع الحساب',     type_labels.get(ftype, '')),
                ('الاسم الكامل',   form.get_full_name()),
                ('رقم الهوية',     form.id_number or '—'),
                ('تاريخ الميلاد',  str(form.birth_date) if form.birth_date else '—'),
                ('المدينة الحالية',form.current_city or '—'),
                ('الشارع',         form.current_street or '—'),
                ('رقم التواصل',    form.phone1 or '—'),
                ('حالة الاستمارة', form.status or '—'),
            ]

        header_fill   = PatternFill('solid', fgColor=GREEN_DARK)
        alt_fill      = PatternFill('solid', fgColor=GRAY)
        thin          = Side(style='thin', color='CCCCCC')
        border        = Border(left=thin, right=thin, top=thin, bottom=thin)

        for r_idx, row in enumerate(rows, start=3):
            for c_idx, val in enumerate(row, start=1):
                cell            = ws1.cell(row=r_idx, column=c_idx, value=val)
                cell.border     = border
                cell.alignment  = Alignment(horizontal='right', vertical='center')
                ws1.row_dimensions[r_idx].height = 22
                if r_idx == 3:
                    cell.font = Font(bold=True, color=WHITE)
                    cell.fill = header_fill
                elif r_idx % 2 == 0:
                    cell.fill = alt_fill

        ws1.column_dimensions['A'].width = 20
        ws1.column_dimensions['B'].width = 35

        # ===== ورقة 2: الدفعات =====
        ws2 = wb.create_sheet('الدفعات الشهرية')
        ws2.sheet_view.rightToLeft = True

        ws2.merge_cells('A1:E1')
        t2 = ws2['A1']
        t2.value     = 'سجل الدفعات الشهرية'
        t2.font      = Font(bold=True, size=13, color=WHITE)
        t2.fill      = PatternFill('solid', fgColor=GREEN_DARK)
        t2.alignment = Alignment(horizontal='center', vertical='center')
        ws2.row_dimensions[1].height = 32

        headers = ['التاريخ', 'المبلغ (شيقل)', 'المبلغ (دولار)', 'الوصف', 'الحالة']
        for c_idx, h in enumerate(headers, start=1):
            cell            = ws2.cell(row=2, column=c_idx, value=h)
            cell.font       = Font(bold=True, color=WHITE)
            cell.fill       = PatternFill('solid', fgColor='2ECC71')
            cell.alignment  = Alignment(horizontal='center', vertical='center')
            cell.border     = border
            ws2.row_dimensions[2].height = 24

        total_shekel = 0
        for r_idx, p in enumerate(payments, start=3):
            dollar = round(float(p.amount_shekel) * rates['ILS_TO_USD'], 2)
            row_data = [
                str(p.payment_date),
                float(p.amount_shekel),
                dollar,
                p.description or '—',
                p.status,
            ]
            total_shekel += float(p.amount_shekel)
            for c_idx, val in enumerate(row_data, start=1):
                cell           = ws2.cell(row=r_idx, column=c_idx, value=val)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border    = border
                ws2.row_dimensions[r_idx].height = 20
                if r_idx % 2 == 0:
                    cell.fill = alt_fill

        # صف الإجمالي
        total_row = len(payments) + 3
        ws2.cell(total_row, 1, 'الإجمالي').font = Font(bold=True)
        ws2.cell(total_row, 2, total_shekel).font = Font(bold=True, color=GREEN_DARK)
        ws2.cell(total_row, 3, round(total_shekel * rates['ILS_TO_USD'], 2)).font = Font(bold=True, color=GREEN_DARK)
        for c in range(1, 6):
            ws2.cell(total_row, c).fill   = PatternFill('solid', fgColor=GREEN_LIGHT)
            ws2.cell(total_row, c).border = border

        for col in ['A','B','C','D','E']:
            ws2.column_dimensions[col].width = 18

        # حفظ
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        name = form.get_full_name() if form else 'تقرير'
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{name}_report.xlsx"'
        return response

    except ImportError:
        return HttpResponse('مكتبة openpyxl غير مثبّتة', status=500)


def _export_pdf(form, ftype, payments, rates):
    """تصدير PDF بسيط"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                        Paragraph, Spacer)
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import io
        from django.conf import settings
        import os

        buffer = io.BytesIO()
        doc    = SimpleDocTemplate(buffer, pagesize=A4,
                                   rightMargin=40, leftMargin=40,
                                   topMargin=40, bottomMargin=40)

        story  = []
        styles = getSampleStyleSheet()
        green  = colors.HexColor('#1A7A4A')

        # العنوان
        title_style = ParagraphStyle('title', fontSize=18, textColor=green,
                                     alignment=1, fontName='Helvetica-Bold')
        story.append(Paragraph('منصة رُحَمَاء — تقرير شخصي', title_style))
        story.append(Spacer(1, 20))

        # بيانات
        if form:
            data = [
                ['الحقل', 'القيمة'],
                ['الاسم', form.get_full_name()],
                ['رقم الهوية', form.id_number or '—'],
                ['المدينة', form.current_city or '—'],
                ['الحالة', form.status or '—'],
            ]
            table = Table(data, colWidths=[150, 300])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), green),
                ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
                ('ALIGN',      (0,0), (-1,-1), 'RIGHT'),
                ('GRID',       (0,0), (-1,-1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0,1), (-1,-1),
                 [colors.HexColor('#F4F9F6'), colors.white]),
            ]))
            story.append(table)
            story.append(Spacer(1, 20))

        # الدفعات
        if payments:
            pay_data = [['التاريخ', 'المبلغ (₪)', 'المبلغ ($)', 'الحالة']]
            total = 0
            for p in payments:
                dollar = round(float(p.amount_shekel) * rates['ILS_TO_USD'], 2)
                pay_data.append([
                    str(p.payment_date),
                    str(p.amount_shekel),
                    str(dollar),
                    p.status,
                ])
                total += float(p.amount_shekel)
            pay_data.append(['الإجمالي', str(total),
                             str(round(total * rates['ILS_TO_USD'], 2)), ''])
            table2 = Table(pay_data, colWidths=[100, 100, 100, 150])
            table2.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2ECC71')),
                ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
                ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
                ('GRID',       (0,0), (-1,-1), 0.5, colors.grey),
                ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#E8F5E9')),
                ('FONTNAME',   (0,-1), (-1,-1), 'Helvetica-Bold'),
            ]))
            story.append(table2)

        doc.build(story)
        buffer.seek(0)

        name = form.get_full_name() if form else 'تقرير'
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{name}_report.pdf"'
        return response

    except ImportError:
        return HttpResponse('مكتبة reportlab غير مثبّتة', status=500)

@beneficiary_required
@require_POST
def mark_notifications_read(request):
    Notification.objects.filter(recipient=request.user, is_read=False).update(is_read=True)
    return JsonResponse({'status': 'success'})